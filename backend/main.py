import csv
import io
import json
import re
import os
import math
import logging
import secrets
from datetime import date, time, timedelta, datetime, timezone
from zoneinfo import ZoneInfo
from typing import Optional, List
from collections import defaultdict

from fastapi import FastAPI, Depends, Header, HTTPException, Query, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.security import HTTPAuthorizationCredentials
from sqlalchemy import or_, func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

import models, schemas, raporty, rezerwacje, sprzatanie, rozliczenia, ical_import, ratelimit, prawo_pracy, seating, platnosci, uprawnienia
import reservation_access
import reservation_allocator
import reservation_audit
import reservation_demand
import reservation_rules
import reservation_service
import reservation_communication
import reservation_payments
import reservation_payment_worker
import workstation_auth
import integracje
import szyfrowanie
import maintenance
import crm_governance
from crm_identity import (
    hash_key as _hash_klucz_crm,
    identity_hash as _identity_hash_crm,
    identity_key as _identity_key_crm,
    reservation_fallback_hash as _reservation_fallback_hash,
)
from database import get_db, init_db, SessionLocal
from algorithm import auto_assign as _auto_assign, przelicz_imprezy_na_wymagania

import jwt
from auth import (
    get_current_user, hash_password, verify_password,
    create_access_token, resolve_request_user, SECRET_KEY, ALGORITHM,
)
from validators import sprawdz_login, sprawdz_haslo, sprawdz_email
from push import wyslij_push, wyslij_push_do_pracownika, wyslij_push_do_adminow

import openpyxl

import settings as app_settings
from deps import get_subskrypcja, subskrypcja_aktywna, utcnow_naive, get_lokal_config, token_agenta_ok, unikalny_login_z_emaila, modul_aktywny, synchronizuj_subskrypcje, rewir_dla_pracownika as _rewir_dla_pracownika
# Helpery współdzielone z routerami (wyniesione do deps.py — dekompozycja main, audyt CTO):
from deps import (
    ROZLICZENIA_START, _napiwki_podzial, _norm_nazwa, _przypisz_odbicia_do_pracownika,
    _sala_stanowisko_ids, _teraz_lokalnie, _user_out, _zamowienie_out, _zbuduj_rozliczenie,
)
from routers.instancja import router as instancja_router
from routers.lokal import router as lokal_router
from routers.platnosci import router as platnosci_router
from routers.crm import router as crm_router
from routers.analityka_rezerwacji import router as analityka_rezerwacji_router
from routers.gielda import router as gielda_router
from routers.plan_sali import (
    operational_plan_payload,
    router as plan_sali_router,
)
from routers.ogloszenia import router as ogloszenia_router
from routers.zgodnosc import router as zgodnosc_router
from routers.imprezy_ai import router as imprezy_ai_router
from routers.portal_imprezy import router as portal_imprezy_router
from routers.antyfraud import router as antyfraud_router
from routers.portfel import router as portfel_router
from routers.moje import router as moje_router
from routers.kadry import router as kadry_router
from routers.zaproszenia import router as zaproszenia_router
from routers.flota import router as flota_router
from routers.pos import router as pos_router
from routers.rodo import (
    eksport_komunikacji_operacyjnej,
    router as rodo_router,
    usun_outbox_przed_usunieciem_pii,
    usun_powiazane_pii_rezerwacji,
    usun_powiazane_publiczne_sekrety,
    wyczysc_notatki_kontekstu_nadpisan,
)
from routers.reguly_rezerwacji import (
    router as reguly_rezerwacji_router,
    _waliduj_serwis as _waliduj_serwis_r3,
)
from routers.workstations import router as workstations_router
from routers.reservation_ops import router as reservation_ops_router
import provisioning

logger = logging.getLogger(__name__)
# Swagger UI + schemat OpenAPI wyłączone w produkcji (nie ujawniaj pełnej mapy API — tras
# płatności, floty, ingestu, schematu karty). W dev/test dostępne do pracy (CWE-200).
app = FastAPI(
    title="Lokalo API",
    docs_url="/docs" if app_settings.IS_DEV else None,
    redoc_url="/redoc" if app_settings.IS_DEV else None,
    openapi_url="/openapi.json" if app_settings.IS_DEV else None,
)


@app.exception_handler(workstation_auth.WorkstationReauthRequired)
async def workstation_reauth_required_handler(
    _request: Request,
    _exc: workstation_auth.WorkstationReauthRequired,
):
    return JSONResponse(
        status_code=428,
        content={
            "detail": "Ta operacja wymaga ponownego potwierdzenia PIN-em operatora.",
            "code": "WORKSTATION_REAUTH_REQUIRED",
        },
        headers={"Cache-Control": "private, no-store"},
    )


@app.exception_handler(reservation_communication.CommunicationDeliveryInProgress)
async def communication_erasure_conflict_handler(
    _request: Request,
    _exc: reservation_communication.CommunicationDeliveryInProgress,
):
    """Fail closed without exposing recipient, content or provider diagnostics."""
    return JSONResponse(
        status_code=409,
        content={
            "detail": (
                "Trwa dostarczanie wiadomości związanej z tymi danymi. "
                "Ponów usunięcie za chwilę."
            ),
            "code": reservation_communication.CommunicationDeliveryInProgress.code,
        },
        headers={"Cache-Control": "no-store"},
    )


@app.exception_handler(reservation_service.ReservationError)
async def reservation_error_handler(request: Request, exc: reservation_service.ReservationError):
    """Stabilny kontrakt konfliktu bez łamania istniejącego klienta czytającego ``detail``."""
    availability = exc.availability.to_dict()
    if request.url.path == "/api/online" or request.url.path.startswith("/api/online/"):
        # Widget gościa dostaje decyzję i czytelny komunikat, nigdy wewnętrzny
        # zapas coverów, identyfikatory reguł/sal ani pełną konfigurację polityki.
        availability["checks"] = []
        availability["applied_rules"] = []
        availability["candidates"] = []
        availability["alternatives"] = []
        availability.pop("buffer_min", None)
        availability["violations"] = [
            {
                "code": item.get("code"),
                "rule": item.get("rule"),
                "message": item.get("message"),
            }
            for item in (availability.get("violations") or [])
        ]
        availability["can_override"] = False
        service = availability.get("service")
        if isinstance(service, dict):
            availability["service"] = {
                key: service.get(key) for key in ("name", "godz_od", "godz_do")
            }
    return JSONResponse(
        {
            "detail": exc.message,
            "code": exc.code,
            "availability": availability,
        },
        status_code=exc.status_code,
    )


@app.exception_handler(reservation_payments.PaymentDomainError)
async def payment_domain_error_handler(
    _request: Request,
    exc: reservation_payments.PaymentDomainError,
):
    """Stabilny, nieujawniający danych providera kontrakt błędu R5c."""
    client_errors = {
        "PAYMENT_RETRY_NOT_ALLOWED",
        "PAYMENT_RESERVATION_INACTIVE",
        "PREAUTH_TOO_EARLY",
        "PAYMENT_NOT_AUTHORIZED",
        "PAYMENT_NOT_CAPTURED",
        "PAYMENT_CANNOT_CANCEL",
        "PARTIAL_REFUND_UNSUPPORTED",
        "PAYMENT_OPERATION_KEY_REUSED",
        "PAYMENT_SETTLEMENT_REQUIRED_BEFORE_EDIT",
        "PAYMENT_OPERATION_ALREADY_PENDING",
    }
    return JSONResponse(
        status_code=409 if exc.code in client_errors else 400,
        content={"detail": exc.message, "code": exc.code},
        headers={"Cache-Control": "no-store"},
    )


@app.exception_handler(integracje.PaymentProviderConfigurationError)
async def payment_provider_configuration_error_handler(
    _request: Request,
    exc: integracje.PaymentProviderConfigurationError,
):
    """Wymagana płatność nigdy nie degraduje się po cichu do produkcyjnego demo."""
    return JSONResponse(
        status_code=503,
        content={"detail": exc.message, "code": exc.code},
        headers={"Cache-Control": "no-store"},
    )
app.include_router(instancja_router)   # subskrypcja/licencja, audyt, status integracji (Rec#5: dekompozycja main)
app.include_router(lokal_router)       # konfiguracja lokalu / branding (Rec#5: dekompozycja main)
app.include_router(platnosci_router)   # płatności zadatków online (Rec#7)
app.include_router(crm_router)         # CRM gości / scoring no-show (roadmapa v1.5)
app.include_router(analityka_rezerwacji_router)   # analityka rezerwacji (covery/no-show/szczyty)
app.include_router(rodo_router)        # RODO — eksport/anonimizacja gościa + retencja (audyt bezp.)
app.include_router(gielda_router)      # giełda wymiany zmian (roadmapa v1.5)
app.include_router(plan_sali_router)   # plan sali — rozmieszczenie stolików + status (roadmapa v1.5)
app.include_router(ogloszenia_router)  # ogłoszenia zespołowe — tablica manager→pracownicy (roadmapa v1.5)
app.include_router(zgodnosc_router)    # zgodność lokalu — badania załogi + terminy (roadmapa v2, oś B)
app.include_router(imprezy_ai_router)  # skrzynka zapytań o imprezy — ekstrakcja+dostępność+szkic (roadmapa v2, oś A)
app.include_router(portal_imprezy_router)  # portal klienta imprezy — tokenowa strona + wątek ustaleń (roadmapa v2, oś A)
app.include_router(antyfraud_router)   # antyfraud POS — storna/rabaty per kelner + flagi (roadmapa v2, oś B)
app.include_router(portfel_router)     # portfel pracownika — zarobek na żywo + zaliczki (roadmapa v2, oś C)
app.include_router(moje_router)        # „Moje" /api/me/* — samoobsługa pracownika (dekompozycja main — audyt CTO)
app.include_router(kadry_router)       # kadry i konta zespołu — users/pracownicy/stanowiska/dyspozycje/urlopy (dekompozycja main — audyt CTO)
app.include_router(zaproszenia_router) # zaproszenia pracowników do kont — jedyna ścieżka rejestracji (feedback UX)
app.include_router(flota_router)       # samoobsługowe zakładanie lokali + panel floty (feedback: zero ręcznej pracy)
app.include_router(pos_router)         # uniwersalne API danych POS: utarg dnia + heartbeat agenta (tor A integracji)
app.include_router(reguly_rezerwacji_router)  # R3 — konfiguracja i symulator reguł dostępności
app.include_router(workstations_router)  # R6a — imienne sesje PIN współdzielonego stanowiska
app.include_router(reservation_ops_router)  # R8 — PII-free gotowość produkcyjna rezerwacji

# CORS „secure by default": w produkcji domyślnie tylko same-origin (backend serwuje
# frontend z tego samego adresu), w dev lokalne origins. Pełna logika w settings.cors_origins().
ALLOWED_ORIGINS = app_settings.cors_origins()
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    # Publiczny widget może używać jawnie dozwolonego API_BASE. HttpOnly
    # capability cookie wymaga credentialed CORS; wildcard jest odrzucany w prod.
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Role nadzorcze i ich dozwolone ścieżki GET (poza /api/me/* dostępnym dla każdego zalogowanego).
# Wszystko spoza tych prefiksów = 403. Zapisy (POST/PUT/DELETE) zarezerwowane dla admina.
OVERSIGHT_GET = {
    # Szef kuchni: godziny kuchni (bez wypłat), podgląd stołów na żywo, rezerwacje.
    "szef_kuchni": (
        "/api/szefkuchni/", "/api/gastro/stoly", "/api/rezerwacje",
    ),
    # Pracownik kuchni: podgląd rezerwacji (zagregowane liczby + rozbicie godzinowe,
    # bez danych klienta) do planowania pracy kuchni. Imprezy widzi przez
    # /api/me/imprezy (bez nazwy klienta — patrz preferencja prywatności).
    "kuchnia": (
        "/api/rezerwacje",
    ),
    # Pracownik obsługi (employee): jak kuchnia — rezerwacje (liczby + rozbicie godzinowe,
    # bez danych klienta) do planowania pracy na sali. Imprezy przez /api/me/imprezy (bez klienta).
    # Reszta panelu admina dalej 403 (poza /api/me/* dostępnym dla każdego zalogowanego).
    "employee": (
        "/api/rezerwacje",
    ),
}

# Szef korzysta z dokładnej mapy endpoint → efektywne uprawnienie konta. Celowo nie
# używamy tu prefiksów: `imprezy.podglad` ma otwierać listę wydarzeń, ale nie finansowe
# `/api/imprezy/rozliczenia`. Pozostałe role zachowują historyczną, segmentową allowlistę.
SZEF_GET_PERMISSION = {
    "/api/raporty/godziny": "raporty.podglad",
    "/api/szef/grafik": "grafik.podglad",
    "/api/pracownicy": "grafik.podglad",
    "/api/stanowiska": "grafik.podglad",
    "/api/gastro/stoly": "grafik.podglad",
    "/api/gastro/stoly-historia": "grafik.podglad",
    "/api/rezerwacje": "rezerwacje.podglad",
    "/api/imprezy": "imprezy.podglad",
    "/api/szef/rozliczenie": "zeszyt.podglad",
    "/api/szef/zeszyt": "zeszyt.podglad",
    "/api/pulpit": "pulpit.podglad",
    "/api/alerty-kasowe": "zeszyt.podglad",
}

SZEF_ME_GET_PERMISSION = {
    "/api/me/imprezy": "imprezy.podglad",
    "/api/me/rezerwacje": "rezerwacje.podglad",
}


def _sciezka_na_whitelist(path: str, prefiksy) -> bool:
    """Dopasowanie whitelisty ODPORNE na kolizję prefiksów. Zwykłe `startswith` przepuszczało
    „/api/rezerwacje-stolik" (pełne PII gości) na wpisie „/api/rezerwacje" (zagregowany, bez PII) —
    bo to prefiks tekstowy, nie segmentowy. Dopasowujemy po GRANICY segmentu: dokładnie albo z „/".
    Dzięki temu „/api/rezerwacje" łapie „/api/rezerwacje" i „/api/rezerwacje/...", ale NIE „...-stolik"."""
    for p in prefiksy:
        pp = p.rstrip("/")
        if path == pp or path.startswith(pp + "/"):
            return True
    return False


# ── Tabela tras autoryzacji ───────────────────────────────────────────────────
# Audyt CTO: „kruchość role_guard — długi warunek z wieloma wyjątkami; łatwo o
# regresję przy dodawaniu endpointów". Zamiast łańcucha `not path.startswith(...)`
# — deklaratywne tabele. Dopasowanie prefiksów jest SEGMENTOWE (dokładnie albo
# granica „/"), więc wpis „/api/online" nie przepuści „/api/online-cokolwiek".
# Ścieżki, które dotąd jechały na tekstowym prefiksie (stoly → stoly-historia!),
# są wypisane JAWNIE — tabela jest jednocześnie dokumentacją allowlisty.

# Publiczna przestrzeń ``/api/online`` jest fail-closed: każdy handler musi mieć
# osobną parę metoda + pełny template. Parametr w nawiasach klamrowych zajmuje
# dokładnie jeden niepusty segment. Dodanie przyszłego handlera bez dopisania go
# tutaj pozostawi go za JWT zamiast przypadkiem otworzyć cały prefiks.
ONLINE_PUBLIC_ROUTE_TEMPLATES = (
    ("GET", "/api/online/widget-config"),
    ("POST", "/api/online/hold"),
    ("DELETE", "/api/online/hold"),
    ("GET", "/api/online/dostepnosc"),
    ("GET", "/api/online/alternatywy"),
    ("GET", "/api/online/najblizszy-termin"),
    ("POST", "/api/online/rezerwacja"),
    ("POST", "/api/online/popyt/odrzucony"),
    ("GET", "/api/online/zarzadzanie/rezerwacja"),
    ("GET", "/api/online/zarzadzanie/platnosc"),
    ("POST", "/api/online/zarzadzanie/platnosc/ponow"),
    ("POST", "/api/online/zarzadzanie/potwierdz"),
    ("POST", "/api/online/zarzadzanie/odwolaj"),
    ("POST", "/api/online/zarzadzanie/edytuj"),
    ("GET", "/api/online/zarzadzanie/dane"),
    ("POST", "/api/online/zarzadzanie/dane/usun"),
    ("POST", "/api/online/lista-oczekujacych"),
    ("GET", "/api/online/nowy-lokal/status"),
    ("POST", "/api/online/nowy-lokal"),
    ("POST", "/api/online/rejestracja"),
    ("POST", "/api/online/rejestracja/{external_id}/oplac"),
    ("GET", "/api/online/rejestracja/{external_id}"),
    ("GET", "/api/online/zaproszenie/{token}"),
    ("POST", "/api/online/zaproszenie/{token}/rejestracja"),
    ("GET", "/api/online/imprezy/{token}"),
    ("PUT", "/api/online/imprezy/{token}/goscie"),
    ("POST", "/api/online/imprezy/{token}/wiadomosci"),
    ("POST", "/api/online/imprezy/{token}/menu"),
    ("POST", "/api/online/platnosci/stripe/webhook"),
)

_PUBLIC_ROUTE_PARAMETER = re.compile(r"^\{[A-Za-z_][A-Za-z0-9_]*\}$")


def _compile_public_route_template(template: str) -> re.Pattern:
    segments = []
    for segment in template.split("/"):
        if _PUBLIC_ROUTE_PARAMETER.fullmatch(segment):
            segments.append(r"[^/]+")
        else:
            if "{" in segment or "}" in segment:
                raise ValueError(f"Niepoprawny template trasy publicznej: {template}")
            segments.append(re.escape(segment))
    return re.compile(r"\A" + "/".join(segments) + r"\Z")


_ONLINE_PUBLIC_ROUTE_PATTERNS = tuple(
    (method, _compile_public_route_template(template))
    for method, template in ONLINE_PUBLIC_ROUTE_TEMPLATES
)


# Pozostałe historyczne przestrzenie publiczne — bez JWT. None = każda metoda,
# inaczej krotka metod. Ingest agenta POS autoryzuje się stałym tokenem
# (X-RCP-Token) już wewnątrz endpointów.
TRASY_PUBLICZNE = (
    ("/api/auth", None),
    ("/api/onboarding", None),                   # status + jednorazowy bootstrap (guard 409 w środku)
    ("/api/health", None),
    ("/api/lokal/branding", None),               # white-label dla ekranu logowania
    ("/api/rcp/ingest", None),
    ("/api/gastro/stoly", ("POST",)),            # agent: stan stołów na żywo
    ("/api/gastro/stoly-historia", ("POST",)),   # agent: historia stolików
    ("/api/gastro/rozliczenia", ("POST",)),      # agent: rozliczenia kelnerów
    ("/api/gastro/zadatki", ("POST",)),          # agent: zadatki KP
    ("/api/gastro/storna", ("POST",)),           # agent: storna/rabaty (antyfraud)
    ("/api/pos/utarg-dnia", ("POST",)),          # uniwersalny utarg dnia (agent/CSV/ręczny) — auth w handlerze
    ("/api/pos/heartbeat", ("POST",)),           # telemetria agenta — auth w handlerze
    ("/api/instancja/puls", ("GET",)),           # panel floty: matka odpytuje dzieci (token FLEET_TOKEN w handlerze)
)

# Dokładne trasy uwierzytelniane cookie zarejestrowanego urządzenia, ale bez JWT.
# Nie są prefiksami: przyszła podtrasa nie może odziedziczyć publicznego statusu.
WORKSTATION_DEVICE_PUBLIC_ROUTES = frozenset({
    ("GET", "/api/reservation-workstations/operators"),
    ("POST", "/api/reservation-workstations/unlock"),
    ("POST", "/api/reservation-workstations/forget-device"),
})

# Wyjątki od degradacji READ_ONLY — zapis dozwolony mimo nieaktywnej subskrypcji:
# logowanie, kreator pierwszej konfiguracji, przedłużenie subskrypcji, health.
READ_ONLY_WYJATKI = (
    "/api/auth",
    "/api/onboarding",
    "/api/subskrypcja",
    "/api/health",
    "/api/reservation-workstations",
    "/api/me/reservation-workstation",
)
# Odczytowe komendy POST mają dokładne dopasowanie metoda+trasa. Nie dodajemy ich do
# prefiksowej listy powyżej, żeby przyszła podtrasa zapisu nie odziedziczyła wyjątku.
READ_ONLY_POST_ODCZYT = frozenset({
    "/api/rezerwacje-stolik/wyszukaj",
    "/api/crm/goscie/wyszukaj",
    "/api/rezerwacje/reguly/symuluj",
    "/api/rodo/eksport-gosc",
})
# Prawa osoby i obowiązek retencji muszą działać również po wygaśnięciu
# subskrypcji, ale przyszła podtrasa /api/rodo nie dziedziczy tego automatycznie.
READ_ONLY_DOKLADNE_WYJATKI = frozenset({
    "/api/rodo/anonimizuj-gosc",
    "/api/rodo/retencja",
    "/api/online/zarzadzanie/odwolaj",
    "/api/online/zarzadzanie/dane/usun",
    "/api/online/platnosci/stripe/webhook",
})
_READ_ONLY_OPERATOR_PAYMENT_EXCEPTION = re.compile(
    r"^/api/platnosci/[1-9]\d*/(?:anuluj-autoryzacje|zwrot|reconcile)$"
)


def _dozwolony_zapis_read_only(method: str, path: str) -> bool:
    """Wąska allowlista zapisów koniecznych do zamknięcia zobowiązań i praw gościa."""
    return method == "POST" and (
        path in READ_ONLY_DOKLADNE_WYJATKI
        or _READ_ONLY_OPERATOR_PAYMENT_EXCEPTION.fullmatch(path) is not None
    )

# Odpowiedzi tych przestrzeni mogą zawierać PII gościa. Segmentowe dopasowanie
# zapobiega przypadkowemu objęciu podobnie nazwanej przyszłej trasy.
PII_NO_STORE_PREFIXES = (
    "/api/online",
    "/api/crm",
    "/api/plan-sali",
    "/api/sale-rezerwacyjne",
    "/api/rezerwacje-stolik",
    "/api/rezerwacje",
    "/api/me/rezerwacje",
    "/api/host",
    "/api/lista-oczekujacych",
    "/api/rodo",
    "/api/terminy",
    "/api/reservation-workstations",
    "/api/me/reservation-workstation",
)
PII_NO_STORE_EXACT = frozenset({
    # ``UserOut`` zawiera imię, nazwisko i e-mail operatora. Trasa jest dostępna
    # także dla sesji PIN, więc jej snapshot nie może zostać w pamięci HTTP po
    # automatycznej blokadzie współdzielonego stanowiska.
    "/api/auth/me",
})

# Przestrzenie, w których rola nadzorcza ma PEŁNY dostęp (też zapisy). Każdy taki
# endpoint sam pilnuje, że dotyczy wyłącznie swojej domeny (np. grafik kuchni).
ROLA_PELNA_PRZESTRZEN = {"szef_kuchni": ("/api/szefkuchni",)}


def _trasa_publiczna(path: str, metoda: str) -> bool:
    if (metoda, path) in WORKSTATION_DEVICE_PUBLIC_ROUTES:
        return True
    if any(
        metoda == dozwolona_metoda and pattern.fullmatch(path)
        for dozwolona_metoda, pattern in _ONLINE_PUBLIC_ROUTE_PATTERNS
    ):
        return True
    for prefiks, metody in TRASY_PUBLICZNE:
        if _sciezka_na_whitelist(path, (prefiks,)) and (metody is None or metoda in metody):
            return True
    return False


# Centralna ochrona API: tabele wyżej + reguły ról. /api/me/* dla każdego
# zalogowanego, całość /api/* dla admina, role nadzorcze wg OVERSIGHT_GET.
# Statyczny frontend jest publiczny.
@app.middleware("http")
async def role_guard(request: Request, call_next):
    path, metoda = request.url.path, request.method
    if metoda == "OPTIONS" or not path.startswith("/api/"):
        return await call_next(request)

    # 1) Degradacja READ_ONLY: nieaktywna subskrypcja → zapisy zwracają 402.
    #    Celowo PRZED autoryzacją (zachowanie historyczne: 402 także bez tokenu).
    if (
        metoda in ("POST", "PUT", "DELETE", "PATCH")
        and not (metoda == "POST" and path in READ_ONLY_POST_ODCZYT)
        and not _dozwolony_zapis_read_only(metoda, path)
        and not _sciezka_na_whitelist(path, READ_ONLY_WYJATKI)
    ):
        _db = SessionLocal()
        try:
            synchronizuj_subskrypcje(_db)   # trial→Free po wygaśnięciu, ZANIM ocenimy READ_ONLY
            if not subskrypcja_aktywna(_db):
                return JSONResponse(
                    {"detail": "Subskrypcja nieaktywna — instancja działa w trybie tylko do odczytu. "
                               "Przedłuż subskrypcję, aby zapisywać zmiany."},
                    status_code=402)
        finally:
            _db.close()

    # 2) Trasy publiczne — bez JWT.
    if _trasa_publiczna(path, metoda):
        return await call_next(request)

    # 3) Wszystko inne wymaga JWT.
    header = request.headers.get("authorization", "")
    token = header[7:] if header.lower().startswith("bearer ") else ""
    credentials = (
        HTTPAuthorizationCredentials(scheme="Bearer", credentials=token)
        if token else None
    )
    using_workstation = not token and workstation_auth.workstation_request(request)
    if not token and not using_workstation:
        return JSONResponse(
            {"detail": "Wymagane logowanie."},
            status_code=401,
        )
    # Egzekwuj BIEŻĄCY stan konta z bazy — NIE ufaj roli wmrożonej w token (CWE-613). Dzięki temu
    # dezaktywacja (User.aktywny=False) i degradacja roli działają natychmiast, a nie dopiero po
    # wygaśnięciu tokenu. Większość endpointów admina jest chroniona wyłącznie tym middlewarem.
    _db = SessionLocal()
    try:
        _user = resolve_request_user(
            request,
            _db,
            credentials,
            require_workstation_csrf=(
                using_workstation and metoda in {"POST", "PUT", "PATCH", "DELETE"}
            ),
            touch_workstation=(
                using_workstation and metoda in {"POST", "PUT", "PATCH", "DELETE"}
            ),
        )
    except HTTPException as exc:
        response = JSONResponse(
            {"detail": exc.detail},
            status_code=exc.status_code,
            headers=exc.headers,
        )
        if exc.status_code == 423:
            response.headers["Cache-Control"] = "private, no-store"
            response.delete_cookie(
                workstation_auth.SESSION_COOKIE,
                path="/api",
                secure=not app_settings.IS_DEV,
                httponly=True,
                samesite="strict",
            )
            response.delete_cookie(
                workstation_auth.CSRF_COOKIE,
                path="/",
                secure=not app_settings.IS_DEV,
                httponly=False,
                samesite="strict",
            )
        return response
    finally:
        _db.close()
    rola = _user.rola
    wymaganie_rezerwacji = reservation_access.requirement_for(metoda, path)
    if using_workstation:
        dozwolone_sesji = path in {
            "/api/me/uprawnienia",
            "/api/me/reservation-workstation",
            "/api/me/reservation-workstation/touch",
            "/api/me/reservation-workstation/lock",
            "/api/me/reservation-workstation/reauthorize",
        }
        if dozwolone_sesji:
            return await call_next(request)
        if (
            wymaganie_rezerwacji is not None
            and reservation_access.user_satisfies(_user, wymaganie_rezerwacji)
        ):
            return await call_next(request)
        return JSONResponse(
            {"detail": "Brak uprawnień sesji stanowiska."},
            status_code=403,
        )
    if rola == "admin":
        return await call_next(request)

    # Granularne prawa rezerwacji są sprawdzane przed ogólnymi regułami ról.
    # Nowa trasa w chronionej przestrzeni jest domyślnie admin-only, dopóki nie
    # zostanie jawnie dodana do deklaratywnej polityki.
    if wymaganie_rezerwacji is not None:
        if reservation_access.user_satisfies(_user, wymaganie_rezerwacji):
            return await call_next(request)
        return JSONResponse({"detail": "Brak uprawnień."}, status_code=403)

    if rola == "szef" and path in SZEF_ME_GET_PERMISSION:
        wymagane = SZEF_ME_GET_PERMISSION[path]
        if metoda == "GET" and uprawnienia.ma_user(_user, wymagane):
            return await call_next(request)
        return JSONResponse({"detail": "Brak uprawnień."}, status_code=403)
    if path.startswith("/api/me/"):
        return await call_next(request)
    if _sciezka_na_whitelist(path, ROLA_PELNA_PRZESTRZEN.get(rola, ())):
        return await call_next(request)
    if rola == "szef" and metoda == "GET":
        wymagane = SZEF_GET_PERMISSION.get(path)
        if wymagane and uprawnienia.ma_user(_user, wymagane):
            return await call_next(request)
        return JSONResponse({"detail": "Brak uprawnień."}, status_code=403)
    if metoda == "GET" and _sciezka_na_whitelist(path, OVERSIGHT_GET.get(rola, ())):
        return await call_next(request)
    return JSONResponse({"detail": "Brak uprawnień."}, status_code=403)


# Nagłówki bezpieczeństwa na KAŻDEJ odpowiedzi (też statyki SPA i odmowy z role_guard).
# Zarejestrowany PO role_guard → jest warstwą zewnętrzną, więc obejmuje jego wczesne 401/403/402.
# HSTS tylko w produkcji (dev bywa po http). CSP celowo pominięta tu (łamie SPA Vite/three/gsap) —
# antyclickjacking daje X-Frame-Options; CSP wprowadzać osobno w trybie report-only (CWE-693/1021).
@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "no-referrer")
    if response.status_code == 423 and workstation_auth.workstation_request(request):
        response.headers["Cache-Control"] = "private, no-store"
        response.delete_cookie(
            workstation_auth.SESSION_COOKIE,
            path="/api",
            secure=not app_settings.IS_DEV,
            httponly=True,
            samesite="strict",
        )
        response.delete_cookie(
            workstation_auth.CSRF_COOKIE,
            path="/",
            secure=not app_settings.IS_DEV,
            httponly=False,
            samesite="strict",
        )
    if (
        request.url.path in PII_NO_STORE_EXACT
        or _sciezka_na_whitelist(request.url.path, PII_NO_STORE_PREFIXES)
    ):
        response.headers["Cache-Control"] = "private, no-store"
        vary = [part.strip() for part in response.headers.get("Vary", "").split(",") if part.strip()]
        for credential_header in ("Authorization", "Cookie"):
            if not any(part.casefold() == credential_header.casefold() for part in vary):
                vary.append(credential_header)
        response.headers["Vary"] = ", ".join(vary)
    if not app_settings.IS_DEV:
        response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    return response


# Nazwa „ukrytego" stanowiska, na które trafiają zmiany z grafiku KUCHNI. Pracownik kuchni
# nie wybiera stanowiska — wszystkie jego zmiany idą na to jedno stanowisko, a stawkę ustawia
# się per osoba (StawkaPracownika na tym stanowisku). Dzięki temu reszta logiki (RCP×grafik,
# wypłaty) działa bez zmian i bez ryzykownej migracji „stanowisko_id NULL".
KUCHNIA_STANOWISKO = "Kuchnia"
# Stanowisko dla pracowników technicznych — pełne godziny RCP × stawka (bez grafiku). Jak kuchnia.
TECHNICZNY_STANOWISKO = "Techniczny"


def _stanowisko_wg_roli(db, rola: str, nazwa_legacy: str, utworz: bool = True):
    """Znajduje (lub tworzy) ukryte stanowisko po ROLI. Niezależne od nazwy — nowy klient może
    nazwać je dowolnie, byle miało właściwą rolę. Fallback + samo-naprawa: jeśli istnieje
    stanowisko o nazwie legacy bez ustawionej roli, dostaje rolę (adopcja istniejących danych)."""
    s = db.query(models.Stanowisko).filter_by(rola=rola).first()
    if s:
        return s
    s = db.query(models.Stanowisko).filter_by(nazwa=nazwa_legacy).first()
    if s:
        if s.rola != rola:
            s.rola = rola; db.commit(); db.refresh(s)
        return s
    if not utworz:
        return None
    s = models.Stanowisko(nazwa=nazwa_legacy, rola=rola)
    db.add(s); db.commit(); db.refresh(s)
    return s


def _kuchnia_stanowisko(db) -> models.Stanowisko:
    return _stanowisko_wg_roli(db, "kuchnia", KUCHNIA_STANOWISKO)


def _techniczny_stanowisko(db) -> models.Stanowisko:
    return _stanowisko_wg_roli(db, "techniczny", TECHNICZNY_STANOWISKO)


def zapisz_audyt(db, user, akcja, *, zasob=None, pracownik_id=None, request=None, szczegoly=None):
    """Zapisuje wpis dziennika audytu dostępu do danych wrażliwych (RODO). Best-effort —
    błąd audytu NIGDY nie przerywa właściwej operacji. `login` denormalizowany (rozliczalność).
    Znacznik czasu jako naiwny UTC (spójność SQLite/Postgres)."""
    try:
        ip = request.client.host if (request is not None and request.client) else None
        db.add(models.AuditLog(
            ts=datetime.now(timezone.utc).replace(tzinfo=None),
            user_id=getattr(user, "id", None),
            login=getattr(user, "login", None),
            akcja=akcja, zasob=zasob, pracownik_id=pracownik_id, ip=ip, szczegoly=szczegoly,
        ))
        db.commit()
    except Exception:
        db.rollback()


def _przelicz_zamykajacego(db, dzien: date):
    """„zamyka lokal" dla danego dnia. RĘCZNE NADPISANIE ma pierwszeństwo: jeśli ktoś tego dnia
    ma zamyka_reczny=True, to ON zamyka i automat go nie zmienia. W innym wypadku AUTO wybiera
    osobę z NAJPÓŹNIEJSZYM godz_od na parkiecie (Sala*) — kandydaci muszą mieć godz_od. Reszta
    dnia ma zamyka=False. Commit tylko gdy coś realnie się zmienia. Zwraca zamykającego/None."""
    rows = db.query(models.PrzydzialZmiany).filter(models.PrzydzialZmiany.data == dzien).all()
    reczny = next((a for a in rows if a.zamyka_reczny), None)
    if reczny is not None:
        zamykajacy = reczny                      # ręczne nadpisanie — automat nie rusza
    else:
        sala_ids = _sala_stanowisko_ids(db)
        kandydaci = [a for a in rows if a.stanowisko_id in sala_ids]
        # Najpóźniejszy start zamyka. Brak godziny = traktujemy jak najwcześniejszą (00:00),
        # więc osoba z wpisaną godziną wygrywa; gdy NIKT na Sali nie ma godziny — i tak wybieramy
        # ostatnio dodanego (najwyższe id), żeby ZAWSZE ktoś był oznaczony jako zamykający.
        zamykajacy = max(kandydaci, key=lambda a: (a.godz_od or time.min, a.id)) if kandydaci else None
    zmienione = False
    for a in rows:
        powinien = zamykajacy is not None and a.id == zamykajacy.id
        if bool(a.zamyka) != powinien:
            a.zamyka = powinien
            zmienione = True
    if zmienione:
        db.commit()
    return zamykajacy


@app.on_event("startup")
def startup():
    # Fail-fast: w produkcji odmawiamy startu przy niebezpiecznych sekretach domyślnych.
    app_settings.validate_critical_secrets()
    init_db()
    maintenance.start_maintenance()
    reservation_communication.start_worker()
    try:
        payment_interval = float(os.environ.get("RESERVATION_PAYMENT_INTERVAL_SECONDS", "2"))
    except (TypeError, ValueError):
        payment_interval = 2.0
    reservation_payment_worker.start_worker(interval_seconds=payment_interval)
    # Instancja-matka z włączoną samoobsługą: podnieś instancje floty, których proces
    # nie przeżył restartu hosta (best-effort; provisioning.py, feedback: zero ręcznej pracy).
    if provisioning.wlaczony():
        try:
            provisioning.wskrzes_flote()
        except Exception:
            logger.exception("Wskrzeszanie floty nie powiodło się (kontynuuję start).")
    # Stanowisko kuchni tworzymy LENIWIE (endpoint /api/grafik/kuchnia-stanowisko), nie na starcie —
    # żeby nie zaśmiecać bazy/testów dodatkowym stanowiskiem, gdy grafik kuchni nie jest używany.
    # Uwaga: konto administratora NIE jest już tworzone z pliku konfiguracyjnego (.env).
    # Admina zakłada się wyłącznie w bazie skryptem: python create_admin.py


@app.on_event("shutdown")
def shutdown():
    reservation_payment_worker.stop_worker(timeout_seconds=2.0)
    reservation_communication.stop_worker()
    maintenance.stop_maintenance()


@app.get("/api/health")
def health():
    """Publiczny status backendu (m.in. Electron sprawdza tu gotowość serwera)."""
    return {"status": "ok"}


# ... [Funkcja parse_date pozostaje bez zmian] ...
def parse_date(s: str) -> date:
    s = s.strip()
    if re.match(r"\d{4}-\d{2}-\d{2}", s): return date.fromisoformat(s)
    if re.match(r"\d{2}\.\d{2}\.\d{4}", s):
        d, m, y = s.split(".")
        return date(int(y), int(m), int(d))
    raise ValueError(f"Nieznany format daty: {s}")


# ═══════════════════════════════════════════════════════════════════════════
# AUTORYZACJA / UŻYTKOWNICY
# ═══════════════════════════════════════════════════════════════════════════

_DUMMY_HASH = None


def _dummy_hash() -> str:
    """Stały bcrypt-hash do wyrównania czasu logowania, gdy konto nie istnieje (liczony raz)."""
    global _DUMMY_HASH
    if _DUMMY_HASH is None:
        _DUMMY_HASH = hash_password("wyrownanie-czasu-logowania")
    return _DUMMY_HASH


@app.post("/api/auth/login", response_model=schemas.TokenOut)
def login(dane: schemas.LoginIn, request: Request, db: Session = Depends(get_db)):
    # Logowanie e-mailem (nowe konta) lub loginem (stare konta bez e-maila — fallback).
    # Ochrona przed brute-force: limit prób + czasowy lockout (per identyfikator+IP oraz per IP).
    ip = request.client.host if request.client else "?"
    ident = ((dane.email or dane.login) or "").strip()
    if not ident:
        raise HTTPException(400, "Podaj e-mail i hasło.")
    klucze = [f"login:{ident.lower()}|ip:{ip}", f"ip:{ip}"]
    for k in klucze:
        blok = ratelimit.pozostala_blokada(k)
        if blok:
            raise HTTPException(429, f"Za dużo prób logowania. Spróbuj ponownie za {blok} s.",
                                headers={"Retry-After": str(blok)})
    if dane.email:
        user = db.query(models.User).filter(func.lower(models.User.email) == ident.lower()).first()
    else:
        user = db.query(models.User).filter(models.User.login == ident).first()
    # Stały czas odpowiedzi niezależnie od istnienia konta — inaczej brak bcrypt dla nieznanego
    # e-maila daje rzędowy skok czasu = enumeracja kont timingiem (CWE-208). Zawsze liczymy bcrypt.
    if user and user.aktywny:
        ok = verify_password(dane.haslo, user.haslo_hash)
    else:
        verify_password(dane.haslo, _dummy_hash())
        ok = False
    if not ok:
        for k in klucze:
            ratelimit.zarejestruj_porazke(k)
        raise HTTPException(401, "Nieprawidłowy e-mail lub hasło.")
    for k in klucze:
        ratelimit.zarejestruj_sukces(k)
    return schemas.TokenOut(access_token=create_access_token(user), user=_user_out(user))

@app.post("/api/auth/register", response_model=schemas.TokenOut, status_code=201)
def register(dane: schemas.RegisterIn, db: Session = Depends(get_db)):
    """Samodzielna rejestracja pracownika. DOMYŚLNIE WYŁĄCZONA (rejestracja_otwarta=False):
    konto zakłada się z linku-zaproszenia od managera (routers/zaproszenia.py); flaga
    w konfiguracji lokalu pozwala świadomie wrócić do otwartej rejestracji.
    Tworzy Pracownika + konto (rola employee) i od razu loguje (zwraca token)."""
    if not get_lokal_config(db).rejestracja_otwarta:
        raise HTTPException(
            403, "Samodzielna rejestracja jest wyłączona — poproś managera o link z zaproszeniem.")
    email = sprawdz_email(dane.email)        # kanał logowania
    sprawdz_haslo(dane.haslo)                 # min 8, litera+cyfra+znak specjalny, ASCII
    imie = (dane.imie or "").strip()
    nazwisko = (dane.nazwisko or "").strip()
    if not imie or not nazwisko:
        raise HTTPException(400, "Podaj imię i nazwisko.")
    if db.query(models.User).filter(func.lower(models.User.email) == email).first():
        raise HTTPException(400, "Ten e-mail jest już zajęty.")

    # Nie duplikuj pracownika: jesli istnieje juz ktos o tym samym (znormalizowanym) imieniu
    # i nazwisku BEZ konta (np. zalozony wczesniej albo dopasowany przez RCP) — podepnij konto
    # pod niego, zeby od razu mialo jego godziny. Inaczej tworzymy nowego pracownika.
    norm = _norm_nazwa(f"{imie} {nazwisko}")
    zajete = {u.pracownik_id for u in db.query(models.User).all() if u.pracownik_id}
    prac = next(
        (p for p in db.query(models.Pracownik).all()
         if p.id not in zajete and _norm_nazwa(f"{p.imie} {p.nazwisko}") == norm),
        None,
    )
    if prac is None:
        prac = models.Pracownik(imie=imie, nazwisko=nazwisko, aktywny=True)
        db.add(prac)
        db.flush()  # nadaje prac.id bez commita
    user = models.User(
        login=unikalny_login_z_emaila(db, email), email=email,
        haslo_hash=hash_password(dane.haslo),
        rola="employee", pracownik_id=prac.id,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    _przypisz_odbicia_do_pracownika(db, prac)  # podlinkuj zalegle odbicia RCP do nowego konta
    return schemas.TokenOut(access_token=create_access_token(user), user=_user_out(user))


# ── ONBOARDING (samoobsługowa pierwsza konfiguracja instancji) ────────────────

@app.get("/api/onboarding/status")
def onboarding_status(db: Session = Depends(get_db)):
    """Publiczny: czy instancja wymaga pierwszej konfiguracji (brak jakiegokolwiek użytkownika)."""
    potrzebny = db.query(models.User).first() is None
    return {"potrzebny": potrzebny, "nazwa_lokalu": get_lokal_config(db).nazwa_lokalu}


@app.post("/api/onboarding/bootstrap", response_model=schemas.TokenOut, status_code=201)
def onboarding_bootstrap(dane: schemas.OnboardingIn, db: Session = Depends(get_db)):
    """Publiczny, JEDNORAZOWY: na świeżej instancji (0 użytkowników) tworzy pierwszego
    administratora i ustawia nazwę lokalu, po czym od razu loguje (zwraca token). Jeśli
    jakikolwiek użytkownik już istnieje → 409 (ochrona przed przejęciem działającej instancji)."""
    if db.query(models.User).first() is not None:
        raise HTTPException(409, "Onboarding już wykonany — instancja ma administratora.")
    email = sprawdz_email(dane.email)
    sprawdz_haslo(dane.haslo)
    admin = models.User(
        login=unikalny_login_z_emaila(db, email), email=email,
        haslo_hash=hash_password(dane.haslo), rola="admin",
    )
    db.add(admin)
    cfg = get_lokal_config(db)
    if dane.nazwa_lokalu and dane.nazwa_lokalu.strip():
        cfg.nazwa_lokalu = dane.nazwa_lokalu.strip()
    db.commit(); db.refresh(admin)
    return schemas.TokenOut(access_token=create_access_token(admin), user=_user_out(admin))


@app.get("/api/auth/me", response_model=schemas.UserOut)
def auth_me(user: models.User = Depends(get_current_user)):
    return _user_out(user)


# --- Zarządzanie kontami (/api/users) → routers/kadry.py (dekompozycja main — audyt CTO) ---


# --- GRAFIK SPRZĄTANIA (część admina; /api/me/* w routers/moje.py) ---

@app.get("/api/sprzatanie")
def sprzatanie_admin(start: date = Query(...), end: date = Query(...), db: Session = Depends(get_db)):
    return {"pozycje": sprzatanie.generuj(db, start, end), "sale": list(sprzatanie.sale_lokalu(db))}


@app.post("/api/sprzatanie/korekty", status_code=204)
def korekta_sprzatania(dane: schemas.SprzatanieKorektaIn, db: Session = Depends(get_db)):
    """Dodaj/usuń pozycję sprzątania. Przeciwna akcja do istniejącej korekty KASUJE ją
    (powrót do automatu) — dzięki temu przyciski w UI działają jak przełącznik."""
    if dane.sala not in sprzatanie.sale_lokalu(db):
        raise HTTPException(400, "Nieznana sala.")
    if dane.akcja not in ("dodaj", "usun"):
        raise HTTPException(400, "Akcja musi być 'dodaj' albo 'usun'.")
    istn = db.query(models.SprzatanieKorekta).filter_by(data=dane.data, sala=dane.sala).first()
    if istn:
        if istn.akcja != dane.akcja:
            db.delete(istn)   # przeciwna korekta = cofnięcie poprzedniej
        # ta sama akcja -> idempotentnie nic
    else:
        db.add(models.SprzatanieKorekta(data=dane.data, sala=dane.sala, akcja=dane.akcja))
    db.commit()


# --- ZAMÓWIENIA SPRZĄTACZKI (dział techniczny) ---


@app.get("/api/zamowienia")
def lista_zamowien(db: Session = Depends(get_db)):
    """Wszystkie zamówienia (admin) — od najnowszych."""
    rows = (db.query(models.ZamowienieSprzataczki)
            .order_by(models.ZamowienieSprzataczki.utworzono_at.desc()).all())
    prac_map = {p.id: f"{p.imie} {p.nazwisko}" for p in db.query(models.Pracownik).all()}
    return {"zamowienia": [_zamowienie_out(z, prac_map) for z in rows]}


@app.get("/api/zamowienia/{zid}/zdjecie")
def zamowienie_zdjecie(zid: int, db: Session = Depends(get_db)):
    z = db.get(models.ZamowienieSprzataczki, zid)
    if not z:
        raise HTTPException(404, "Nie znaleziono.")
    return {"zdjecie": z.zdjecie}


@app.put("/api/zamowienia/{zid}/status", status_code=204)
def zmien_status_zamowienia(zid: int, dane: schemas.ZamowienieStatusIn, db: Session = Depends(get_db)):
    """Admin: 'odczytane' albo 'zamowione' → push do autorki."""
    z = db.get(models.ZamowienieSprzataczki, zid)
    if not z:
        raise HTTPException(404, "Nie znaleziono zamówienia.")
    if dane.status not in ("odczytane", "zamowione"):
        raise HTTPException(400, "Status musi być 'odczytane' albo 'zamowione'.")
    teraz = utcnow_naive()
    z.status = dane.status
    if dane.status == "odczytane" and not z.odczytano_at:
        z.odczytano_at = teraz
    if dane.status == "zamowione" and not z.zamowiono_at:
        z.zamowiono_at = teraz
    db.commit()
    tresc = ("Twoje zamówienie zostało odczytane." if dane.status == "odczytane"
             else f"Zamówiono: {z.nazwa}.")
    wyslij_push_do_pracownika(db, z.pracownik_id, "Zamówienie", tresc, url="/")


# --- URLOPY (obsługa) ---


@app.get("/api/urlopy")
def lista_urlopow(db: Session = Depends(get_db)):
    """Wszystkie wnioski (admin) — oczekujące najpierw, potem wg daty startu malejąco."""
    rows = db.query(models.Urlop).all()
    prac_map = {p.id: f"{p.imie} {p.nazwisko}" for p in db.query(models.Pracownik).all()}
    rows.sort(key=lambda u: (u.status != "oczekuje", -u.start.toordinal()))
    return {"urlopy": [_urlop_out(u, prac_map) for u in rows]}


@app.put("/api/urlopy/{uid}/status", status_code=204)
def rozpatrz_urlop(uid: int, dane: schemas.UrlopStatusIn, db: Session = Depends(get_db)):
    """Admin: 'zaakceptowany' / 'odrzucony' → push do pracownika."""
    u = db.get(models.Urlop, uid)
    if not u:
        raise HTTPException(404, "Nie znaleziono wniosku.")
    if dane.status not in ("zaakceptowany", "odrzucony"):
        raise HTTPException(400, "Status musi być 'zaakceptowany' albo 'odrzucony'.")
    u.status = dane.status
    u.rozpatrzono_at = utcnow_naive()
    db.commit()
    slowo = "zaakceptowany" if dane.status == "zaakceptowany" else "odrzucony"
    wyslij_push_do_pracownika(db, u.pracownik_id, "Urlop",
                              f"Twój wniosek urlopowy ({u.start.strftime('%d.%m')}–{u.koniec.strftime('%d.%m')}) został {slowo}.", url="/")


# --- ROZLICZANIE IMPREZ (osoba wyznaczona w grafiku) ---


def _wymagaj_modul_imprezy(db: Session = Depends(get_db)):
    """Bramka modułu: endpointy imprez działają tylko gdy moduł włączony I odblokowany w planie
    (tier-gating: Premium/Enterprise lub trial). Chroni backend niezależnie od ukrywania zakładek."""
    if not modul_aktywny(db, "modul_imprezy"):
        raise HTTPException(403, "Moduł imprez jest niedostępny w tym planie — odblokujesz go w pakiecie Premium.")


def imp_dla_dnia(db, data: date) -> dict:
    """Kwoty IMP dla rozliczenia dnia (D2). Gotówka SFISKALIZOWANA z imprez → minus w kasach;
    karta z imprez → minus w terminalach i kasach. Gotówka niesfiskalizowana i przelew NIE wchodzą.
    Lokal bez osobnego rozliczania imprez (impreza_osobne_rozliczenie=False): IMP zawsze 0 —
    sprzedaż imprezowa siedzi w zwykłym obrocie sali."""
    if not get_lokal_config(db).impreza_osobne_rozliczenie:
        return {"gotowka_sfiskalizowana": 0.0, "karta": 0.0}
    gotowka_sfisk = karta = 0.0
    for r in db.query(models.RozliczenieImprezy).filter_by(data=data).all():
        for p in r.pozycje:
            if p.forma == "gotowka" and p.sfiskalizowane:
                gotowka_sfisk += p.kwota or 0
            elif p.forma == "karta":
                karta += p.kwota or 0
    return {"gotowka_sfiskalizowana": round(gotowka_sfisk, 2), "karta": round(karta, 2)}


@app.get("/api/imprezy/rozliczenia", dependencies=[Depends(_wymagaj_modul_imprezy)])
def rejestr_imprez(start: date = Query(...), end: date = Query(...), db: Session = Depends(get_db)):
    """Rejestr rozliczeń imprez (admin) — per impreza, z pozycjami i sumami per forma."""
    rows = (db.query(models.RozliczenieImprezy)
            .filter(models.RozliczenieImprezy.data >= start, models.RozliczenieImprezy.data <= end)
            .order_by(models.RozliczenieImprezy.data.desc()).all())
    prac_map = {p.id: f"{p.imie} {p.nazwisko}" for p in db.query(models.Pracownik).all()}
    out = []
    for r in rows:
        poz = [{"forma": p.forma, "kwota": p.kwota, "sfiskalizowane": p.sfiskalizowane} for p in r.pozycje]
        out.append({
            "id": r.id, "data": str(r.data), "pracownik": prac_map.get(r.pracownik_id), "opis": r.opis,
            "pozycje": poz,
            "suma_gotowka": round(sum(p["kwota"] for p in poz if p["forma"] == "gotowka"), 2),
            "suma_karta": round(sum(p["kwota"] for p in poz if p["forma"] == "karta"), 2),
            "suma_przelew": round(sum(p["kwota"] for p in poz if p["forma"] == "przelew"), 2),
        })
    return {"rozliczenia": out, "razem": {k: round(sum(o[k] for o in out), 2) for k in ("suma_gotowka", "suma_karta", "suma_przelew")}}


# --- ROZLICZENIE DNIA (sala) ---


def _kp_dla_dnia(db, data: date) -> float:
    """Σ zadatków (KP „Kasa przyjęła") z Gastro dla dnia (po dacie wystawienia dokumentu kasowego).
    Dokumenty KP idą z agenta do tabeli kp_zadatki. Gotówkowe (KP = dokument kasowy)."""
    rows = db.query(models.KpZadatek).filter(models.KpZadatek.data == data).all()
    return round(sum(z.kwota or 0 for z in rows), 2)


def _imp_wynikowe(db, roz: models.RozliczenieDnia) -> dict:
    """IMP do liczenia: ręczne nadpisanie (gdy imp_reczny) albo automat z rozliczeń imprez."""
    if roz.imp_reczny:
        return {"gotowka_sfiskalizowana": roz.imp_gotowka or 0, "karta": roz.imp_karta or 0}
    return imp_dla_dnia(db, roz.data)


def _tryb_efektywny(db, roz: models.RozliczenieDnia) -> str:
    """Skąd czytać utarg dnia. Tryb 'pula'/'indywidualnie' z konfiguracji jest GLOBALNY i zmienny,
    a dane dnia leżą tam, gdzie były zapisane. Żeby zmiana trybu nie wyzerowała utargu na starych
    dniach — źródło podąża za DANYMI (bieżący tryb rozstrzyga tylko, gdy oba źródła są puste)."""
    tryb = get_lokal_config(db).rozliczenia_tryb_kelnera or "indywidualnie"
    ma_pule = bool(roz.pula_gotowka or roz.pula_karta or roz.pula_fv or roz.pula_kw)
    ma_kelnerow = any(k.gotowka or k.karta or k.fv or k.kw for k in roz.kelnerzy)
    if tryb == "pula":
        return "pula" if (ma_pule or not ma_kelnerow) else "indywidualnie"
    return "pula" if (ma_pule and not ma_kelnerow) else "indywidualnie"


def _wynik_rozliczenia(db, roz: models.RozliczenieDnia) -> dict:
    # Tryb 'pula': jeden syntetyczny „kelner" (zbiorcze G/T) — silnik policz_dzien bez zmian.
    if _tryb_efektywny(db, roz) == "pula":
        kelnerzy = [{"gotowka": roz.pula_gotowka or 0, "karta": roz.pula_karta or 0}]
        fv = roz.pula_fv or 0
        kw = roz.pula_kw or 0
    else:
        kelnerzy = [{"gotowka": k.gotowka, "karta": k.karta} for k in roz.kelnerzy]
        fv = sum(k.fv for k in roz.kelnerzy)
        kw = sum(k.kw for k in roz.kelnerzy)
    terminale = [p.get("kwota", 0) for p in (roz.terminale or [])]
    kasy = [p.get("kwota", 0) for p in (roz.kasy or [])]
    return rozliczenia.policz_dzien(kelnerzy=kelnerzy, fv=fv, terminale=terminale, kasy=kasy,
                                    zadatek_gotowka=roz.zadatek_gotowka or 0,
                                    zadatek_karta=roz.zadatek_karta or 0,
                                    kw=kw, imp=_imp_wynikowe(db, roz))


def _rozliczenie_out(db, roz: models.RozliczenieDnia) -> dict:
    pm = {p.id: f"{p.imie} {p.nazwisko}" for p in db.query(models.Pracownik).all()}
    # Front renderuje wg EFEKTYWNEGO źródła (nie surowej flagi), więc po zmianie trybu widać
    # realne dane starego dnia (a nie pustą kartę puli / pustą tabelę).
    tryb = _tryb_efektywny(db, roz)
    return {
        "data": str(roz.data), "status": roz.status,
        "tryb_kelnera": tryb,
        "zadatek_gotowka": roz.zadatek_gotowka or 0, "zadatek_karta": roz.zadatek_karta or 0,
        "kp_baza": _kp_dla_dnia(db, roz.data),    # Σ KP z Gastro (podpowiedź do rozbicia)
        "imp_reczny": roz.imp_reczny, "imp_gotowka": roz.imp_gotowka or 0, "imp_karta": roz.imp_karta or 0,
        "przelew": roz.przelew or 0,
        "kelnerzy": [{"pracownik_id": k.pracownik_id, "pracownik": pm.get(k.pracownik_id),
                      "gotowka": k.gotowka, "karta": k.karta, "fv": k.fv, "kw": k.kw} for k in roz.kelnerzy],
        "pula": {"gotowka": roz.pula_gotowka or 0, "karta": roz.pula_karta or 0,
                 "fv": roz.pula_fv or 0, "kw": roz.pula_kw or 0},
        "terminale": roz.terminale or [], "kasy": roz.kasy or [],
        "wynik": _wynik_rozliczenia(db, roz),
    }


@app.get("/api/rozliczenie")
def get_rozliczenie(data: date = Query(...), db: Session = Depends(get_db)):
    return _rozliczenie_out(db, _zbuduj_rozliczenie(db, data))


@app.put("/api/rozliczenie")
def zapisz_rozliczenie(dane: schemas.RozliczenieDniaIn, data: date = Query(...), db: Session = Depends(get_db)):
    roz = _zbuduj_rozliczenie(db, data)
    roz.zadatek_gotowka = float(dane.zadatek_gotowka or 0)
    roz.zadatek_karta = float(dane.zadatek_karta or 0)
    roz.imp_reczny = bool(dane.imp_reczny)
    roz.imp_gotowka = float(dane.imp_gotowka or 0)
    roz.imp_karta = float(dane.imp_karta or 0)
    roz.przelew = float(dane.przelew or 0)
    roz.terminale = [p.model_dump() for p in dane.terminale]
    roz.kasy = [p.model_dump() for p in dane.kasy]
    # Zbiorcza pula sali (tryb 'pula') — zapisywana, gdy front ją przysłał; nieszkodliwa w trybie
    # indywidualnym (pola puli są tam ignorowane przy liczeniu).
    if dane.pula_gotowka is not None: roz.pula_gotowka = float(dane.pula_gotowka)
    if dane.pula_karta is not None: roz.pula_karta = float(dane.pula_karta)
    if dane.pula_fv is not None: roz.pula_fv = float(dane.pula_fv)
    if dane.pula_kw is not None: roz.pula_kw = float(dane.pula_kw)
    by_pid = {k.pracownik_id: k for k in roz.kelnerzy}
    for kin in dane.kelnerzy:
        k = by_pid.get(kin.pracownik_id)
        if k is None:
            k = models.RozliczenieKelner(pracownik_id=kin.pracownik_id); roz.kelnerzy.append(k)
        k.gotowka = kin.gotowka; k.karta = kin.karta; k.fv = kin.fv; k.kw = kin.kw
    db.commit(); db.refresh(roz)
    return _rozliczenie_out(db, roz)


@app.put("/api/rozliczenie/przelew", status_code=204)
def ustaw_przelew(data: date = Query(...), przelew: float = Query(0.0), db: Session = Depends(get_db)):
    """Przelew dnia z palca (admin) — szybki zapis (używany też z Zeszytu na wierszu SALA)."""
    roz = _zbuduj_rozliczenie(db, data)
    roz.przelew = float(przelew or 0)
    db.commit()


@app.post("/api/rozliczenie/przekaz-szef", status_code=204)
def przekaz_szef(data: date = Query(...), db: Session = Depends(get_db)):
    roz = db.query(models.RozliczenieDnia).filter_by(data=data).first()
    if not roz:
        raise HTTPException(404, "Brak rozliczenia tego dnia.")
    roz.status = "u_szefa"; roz.przekazano_szef_at = utcnow_naive(); db.commit()


@app.get("/api/szef/rozliczenie")
def szef_rozliczenie(data: date = Query(...), db: Session = Depends(get_db)):
    """Szef — tylko utarg SALI (G+T), zadatki osobno, braki/nadwyżki. Bez imprez i bez FV."""
    roz = db.query(models.RozliczenieDnia).filter_by(data=data).first()
    if not roz:
        return {"data": str(data), "status": "brak", "utarg": None}
    w = _wynik_rozliczenia(db, roz)
    return {"data": str(data), "status": roz.status,
            "utarg": {"gotowka": w["suma_szef"]["gotowka"], "karta": w["suma_szef"]["karta"],
                      "fv": w["fv"], "razem": round(w["suma_szef"]["razem"] + w["fv"], 2)},   # utarg sali Z FV
            "zadatek": w["zadatek"],
            "roznica_karty": w["terminale"]["roznica_karty"], "roznica_calosc": w["kasy"]["roznica"]}


# ── ZESZYT KASOWY ─────────────────────────────────────────────────────────────
# Kasa dzienna: PRZYCHÓD (SALA z rozliczenia + imprezy) − ROZCHÓD (ręczne wpisy) → STAN
# (saldo gotówkowe narastająco od „stanu początkowego"). Liczone tylko z gotówki.

def _zeszyt_dane(db, start: date, end: date) -> dict:
    cfg = db.query(models.ZeszytConfig).first()
    stan0 = float(cfg.stan_poczatkowy) if cfg else 0.0
    anchor = cfg.stan_poczatkowy_data if (cfg and cfg.stan_poczatkowy_data) else start
    if anchor > start:
        anchor = start                       # licz zawsze od początku okna, gdyby data startowa była później
    rozl = {r.data: r for r in db.query(models.RozliczenieDnia)
            .filter(models.RozliczenieDnia.data >= anchor, models.RozliczenieDnia.data <= end).all()}
    imp_by_day = {}
    # Wiersze imprez tylko w lokalach z osobnym rozliczaniem imprez — w trybie „imprezy
    # w ogólnym obrocie" zeszyt widzi wyłącznie SALĘ i wpisy ręczne.
    if get_lokal_config(db).impreza_osobne_rozliczenie:
        for im in (db.query(models.RozliczenieImprezy)
                   .filter(models.RozliczenieImprezy.data >= anchor, models.RozliczenieImprezy.data <= end).all()):
            imp_by_day.setdefault(im.data, []).append(im)
    poz_by_day = {}
    for p in (db.query(models.ZeszytPozycja)
              .filter(models.ZeszytPozycja.data >= anchor, models.ZeszytPozycja.data <= end).all()):
        poz_by_day.setdefault(p.data, []).append(p)
    przy_by_day = {}
    for p in (db.query(models.ZeszytPrzychod)
              .filter(models.ZeszytPrzychod.data >= anchor, models.ZeszytPrzychod.data <= end).all()):
        przy_by_day.setdefault(p.data, []).append(p)

    dni, stan = [], stan0
    for ordv in range(anchor.toordinal(), end.toordinal() + 1):
        d = date.fromordinal(ordv)
        wiersze, cash_in = [], 0.0
        r = rozl.get(d)
        if r:                          # SALA z rozliczenia (także wersji roboczej) trafia do zeszytu
            w = _wynik_rozliczenia(db, r)
            sg, sk = w["suma_zeszyt"]["gotowka"], w["suma_zeszyt"]["karta"]
            # suma_zeszyt.gotowka = Σ gotówka + KW (rekonstrukcja UTARGU do porównania z kasą fiskalną).
            # Do FIZYCZNEGO salda szuflady wchodzi utarg BEZ KW — KW to gotówka WYPŁACONA (np. zwrot
            # kaucji), więc jej doliczanie zawyżało saldo kasy/Pulpit o kwotę KW każdego takiego dnia.
            kw_dnia = w["kw"]
            pz = float(r.przelew or 0)
            if sg or sk or pz:
                wiersze.append({"zrodlo": "SALA", "gotowka": sg, "terminal": sk, "przelew": pz, "impreza": 0.0,
                                "manualny": False, "sala_id": r.id})   # sala_id → edycja przelewu z palca w zeszycie
                cash_in += sg - kw_dnia
        for im in imp_by_day.get(d, []):
            g_sf = round(sum(p.kwota for p in im.pozycje if p.forma == "gotowka" and p.sfiskalizowane), 2)
            g_ns = round(sum(p.kwota for p in im.pozycje if p.forma == "gotowka" and not p.sfiskalizowane), 2)
            kt = round(sum(p.kwota for p in im.pozycje if p.forma == "karta"), 2)
            pz = round(sum(p.kwota for p in im.pozycje if p.forma == "przelew"), 2)
            wiersze.append({"zrodlo": im.opis or "Impreza", "gotowka": g_sf, "terminal": kt, "przelew": pz, "impreza": g_ns, "manualny": False})
            cash_in += g_sf + g_ns
        for p in przy_by_day.get(d, []):
            wiersze.append({"id": p.id, "zrodlo": p.zrodlo or "—", "gotowka": p.gotowka, "terminal": p.terminal,
                            "przelew": p.przelew, "impreza": p.impreza, "manualny": True})
            cash_in += (p.gotowka or 0) + (p.impreza or 0)
        rozchod = [{"id": p.id, "kolumna": p.kolumna, "opis": p.opis, "kwota": p.kwota} for p in poz_by_day.get(d, [])]
        rozchod_suma = round(sum(p.kwota for p in poz_by_day.get(d, [])), 2)
        stan = round(stan + cash_in - rozchod_suma, 2)
        if d >= start:
            dni.append({"data": str(d), "wiersze": wiersze, "rozchod": rozchod,
                        "przychod_gotowka": round(cash_in, 2), "rozchod_suma": rozchod_suma, "stan": stan})
    return {"stan_poczatkowy": stan0,
            "stan_poczatkowy_data": str(cfg.stan_poczatkowy_data) if (cfg and cfg.stan_poczatkowy_data) else None,
            "dni": dni}


@app.get("/api/zeszyt")
def get_zeszyt(start: date = Query(...), end: date = Query(...), db: Session = Depends(get_db)):
    return _zeszyt_dane(db, start, end)


def _zeszyt_bez_pozycji_wyplat(dane: dict) -> dict:
    """Ukrywa wypłaty oraz agregaty, z których można by odtworzyć ich kwotę.

    Rzeczywiste saldo uwzględnia wypłaty, więc nie może być zwrócone osobie
    bez ``wyplaty.podglad``. Suma rozchodu obejmuje wyłącznie widoczne pozycje.
    """
    dni = []
    for dzien in dane.get("dni", []):
        widoczny_rozchod = [
            p for p in dzien.get("rozchod", []) if p.get("kolumna") != "wyplaty"
        ]
        dni.append({
            **dzien,
            "rozchod": widoczny_rozchod,
            "rozchod_suma": round(sum(float(p.get("kwota") or 0) for p in widoczny_rozchod), 2),
            "stan": None,
        })
    return {
        **dane,
        "stan_poczatkowy": None,
        "stan_poczatkowy_data": None,
        "dane_czesciowo_ukryte": True,
        "dni": dni,
    }


@app.get("/api/szef/zeszyt")
def szef_zeszyt(start: date = Query(...), end: date = Query(...),
                 user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    dane = _zeszyt_dane(db, start, end)
    if not uprawnienia.ma_user(user, "wyplaty.podglad"):
        return _zeszyt_bez_pozycji_wyplat(dane)
    return dane


@app.post("/api/zeszyt/pozycja", status_code=201)
def dodaj_zeszyt_pozycja(dane: schemas.ZeszytPozycjaIn, db: Session = Depends(get_db)):
    if dane.kolumna not in ("towar", "koszty", "wyplaty", "inne"):
        raise HTTPException(400, "Nieznana kolumna rozchodu.")
    p = models.ZeszytPozycja(data=dane.data, kolumna=dane.kolumna, opis=dane.opis, kwota=float(dane.kwota or 0))
    db.add(p); db.commit(); db.refresh(p)
    return {"id": p.id}


@app.delete("/api/zeszyt/pozycja/{poz_id}", status_code=204)
def usun_zeszyt_pozycja(poz_id: int, db: Session = Depends(get_db)):
    p = db.get(models.ZeszytPozycja, poz_id)
    if p:
        db.delete(p); db.commit()


@app.post("/api/zeszyt/przychod", status_code=201)
def dodaj_zeszyt_przychod(dane: schemas.ZeszytPrzychodIn, db: Session = Depends(get_db)):
    p = models.ZeszytPrzychod(data=dane.data, zrodlo=dane.zrodlo, gotowka=float(dane.gotowka or 0),
                              terminal=float(dane.terminal or 0), przelew=float(dane.przelew or 0),
                              impreza=float(dane.impreza or 0))
    db.add(p); db.commit(); db.refresh(p)
    return {"id": p.id}


@app.delete("/api/zeszyt/przychod/{poz_id}", status_code=204)
def usun_zeszyt_przychod(poz_id: int, db: Session = Depends(get_db)):
    p = db.get(models.ZeszytPrzychod, poz_id)
    if p:
        db.delete(p); db.commit()


@app.get("/api/zeszyt/config")
def get_zeszyt_config(db: Session = Depends(get_db)):
    cfg = db.query(models.ZeszytConfig).first()
    return {"stan_poczatkowy": float(cfg.stan_poczatkowy) if cfg else 0.0,
            "stan_poczatkowy_data": str(cfg.stan_poczatkowy_data) if (cfg and cfg.stan_poczatkowy_data) else None}


@app.put("/api/zeszyt/config")
def set_zeszyt_config(dane: schemas.ZeszytConfigIn, db: Session = Depends(get_db)):
    cfg = db.query(models.ZeszytConfig).first()
    if cfg is None:
        cfg = models.ZeszytConfig(id=1); db.add(cfg)
    cfg.stan_poczatkowy = float(dane.stan_poczatkowy or 0)
    cfg.stan_poczatkowy_data = dane.stan_poczatkowy_data
    db.commit()
    return {"stan_poczatkowy": cfg.stan_poczatkowy,
            "stan_poczatkowy_data": str(cfg.stan_poczatkowy_data) if cfg.stan_poczatkowy_data else None}


# ═══════════════════════════════════════════════════════════════════════════
# PULPIT WŁAŚCICIELA (KPI cockpit) — agregacja istniejących danych, zero nowych tabel
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/pulpit")
def pulpit(start: date = Query(...), end: date = Query(...),
           user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Zbiorcze KPI lokalu za okres [start, end]: przychód (z zeszytu kasowego), saldo kasy,
    rozchód, ruch (rachunki z POS), rezerwacje (moduł stolików) oraz koszt pracy miesiąca
    końca okresu. Czysta agregacja — nie zapisuje nic do bazy."""
    if end < start:
        raise HTTPException(400, "Koniec okresu przed początkiem.")

    # — Przychód, rozchód i saldo kasy: z zeszytu kasowego (SALA + imprezy + ręczne wpisy) —
    z = _zeszyt_dane(db, start, end)
    dni = z["dni"]
    got = term = przel = impr = przychod_total = rozchod_total = 0.0
    przychod_dzienny = []
    for d in dni:
        dsum = 0.0
        for w in d["wiersze"]:
            g = float(w.get("gotowka") or 0); k = float(w.get("terminal") or 0)
            p = float(w.get("przelew") or 0); i = float(w.get("impreza") or 0)
            got += g; term += k; przel += p; impr += i; dsum += g + k + p + i
        rozchod_total += float(d["rozchod_suma"] or 0)
        przychod_total += dsum
        przychod_dzienny.append({"data": d["data"], "przychod": round(dsum, 2),
                                 "rozchod": round(float(d["rozchod_suma"] or 0), 2)})
    saldo = dni[-1]["stan"] if dni else z["stan_poczatkowy"]

    # — Ruch (liczba rachunków z POS) —
    ruch_rows = sorted(db.query(models.StolikiHistoria).filter(
        models.StolikiHistoria.data >= start, models.StolikiHistoria.data <= end).all(),
        key=lambda r: r.data)
    ruch_total = sum(int(r.liczba or 0) for r in ruch_rows)
    ruch_dzienny = [{"data": str(r.data), "liczba": int(r.liczba or 0)} for r in ruch_rows]

    # — Rezerwacje (moduł stolików) —
    rez = db.query(models.Termin).filter(
        models.Termin.rodzaj == "stolik",
        models.Termin.data >= start, models.Termin.data <= end).all()
    rez_status, rez_goscie = {}, 0
    for r in rez:
        rez_status[r.status] = rez_status.get(r.status, 0) + 1
        if r.status in ("rezerwacja", "potwierdzona", "odbyla"):
            rez_goscie += int(r.liczba_osob or 0)

    # — Koszt pracy: miesiąc końca okresu (z raportu godzin × stawki) —
    rap = raporty.raport_godzin_miesiac(db, end.year, end.month)
    koszt_pracy = round(sum(float(p["do_wyplaty"] or 0) for p in rap["pracownicy"]), 2)
    alerty = _alerty_kasowe(db, start, end)

    n = max(1, len(przychod_dzienny))
    wynik = {
        "okres": {"start": str(start), "end": str(end), "dni": len(dni)},
        "przychod": {
            "razem": round(przychod_total, 2), "gotowka": round(got, 2), "karta": round(term, 2),
            "przelew": round(przel, 2), "impreza": round(impr, 2),
            "srednia_dzienna": round(przychod_total / n, 2), "dzienny": przychod_dzienny,
        },
        "rozchod": {"razem": round(rozchod_total, 2)},
        "wynik": round(przychod_total - rozchod_total - koszt_pracy, 2),   # poglądowo: przychód − rozchód − koszt pracy
        "saldo_kasy": saldo,
        "ruch": {"rachunki": ruch_total, "dzienny": ruch_dzienny,
                 "srednia_dzienna": round(ruch_total / max(1, len(ruch_dzienny)), 1) if ruch_dzienny else 0},
        "rezerwacje": {"razem": len(rez), "wg_statusu": rez_status, "goscie": rez_goscie},
        "koszt_pracy_miesiac": {"rok": end.year, "miesiac": end.month, "kwota": koszt_pracy},
        "alerty_kasowe": {"dni_z_anomalia": alerty["dni_z_anomalia"],
                          "suma_braki": alerty["suma_braki"], "suma_nadwyzki": alerty["suma_nadwyzki"]},
    }
    if user.rola == "szef":
        if not uprawnienia.ma_user(user, "wyplaty.podglad"):
            for pole in ("koszt_pracy_miesiac", "rozchod", "saldo_kasy", "wynik"):
                wynik.pop(pole, None)
            for dzien in wynik.get("przychod", {}).get("dzienny", []):
                dzien.pop("rozchod", None)
        if not uprawnienia.ma_user(user, "zeszyt.podglad"):
            for pole in ("przychod", "rozchod", "saldo_kasy", "alerty_kasowe", "wynik"):
                wynik.pop(pole, None)
        if not uprawnienia.ma_user(user, "rezerwacje.podglad"):
            wynik.pop("rezerwacje", None)
    return wynik


# ── ALERTY KASOWE (różnice w rozliczeniu: brak/nadwyżka na kartach lub w kasie) ──

def _alerty_kasowe(db, start: date, end: date, prog: float = 1.0) -> dict:
    """Skanuje rozliczenia dnia w okresie i zwraca dni z różnicą ponad próg [zł].
    Źródło: policz_dzien → terminale.roznica_karty (karty) i kasy.roznica (fiskalizacja)."""
    rozl = db.query(models.RozliczenieDnia).filter(
        models.RozliczenieDnia.data >= start, models.RozliczenieDnia.data <= end).all()
    alerty = []
    for r in sorted(rozl, key=lambda x: x.data):
        w = _wynik_rozliczenia(db, r)
        rk = float(w["terminale"]["roznica_karty"])
        rc = float(w["kasy"]["roznica"])
        problemy = []
        if abs(rk) >= prog:
            problemy.append({"typ": "karty", "roznica": rk, "etykieta": w["terminale"]["etykieta"]})
        if abs(rc) >= prog:
            problemy.append({"typ": "kasa", "roznica": rc, "etykieta": w["kasy"]["etykieta"]})
        if problemy:
            alerty.append({
                "data": str(r.data), "status": r.status, "problemy": problemy,
                "braki": round(sum(p["roznica"] for p in problemy if p["roznica"] < 0), 2),
                "nadwyzki": round(sum(p["roznica"] for p in problemy if p["roznica"] > 0), 2),
            })
    return {"prog": prog, "alerty": alerty, "dni_z_anomalia": len(alerty),
            "suma_braki": round(sum(a["braki"] for a in alerty), 2),
            "suma_nadwyzki": round(sum(a["nadwyzki"] for a in alerty), 2)}


@app.get("/api/alerty-kasowe")
def alerty_kasowe(start: date = Query(...), end: date = Query(...), prog: float = 1.0,
                  db: Session = Depends(get_db)):
    """Dni z anomalią kasową (różnica ≥ prog zł) w okresie. Admin/szef."""
    if end < start:
        raise HTTPException(400, "Koniec okresu przed początkiem.")
    return _alerty_kasowe(db, start, end, max(0.0, float(prog)))


def _alerty_obsady_okres(db: Session, start: date, end: date) -> dict:
    """Niedobory obsady w jawnym okresie; wspólny rdzeń pulpitu i widoku szefa."""
    wymagane = defaultdict(int)
    for w in db.query(models.WymaganiaDnia).filter(
            models.WymaganiaDnia.data >= start, models.WymaganiaDnia.data <= end).all():
        wymagane[(w.data, w.stanowisko_id)] += (w.liczba_osob or 0)
    obsadzone = defaultdict(int)
    for p in db.query(models.PrzydzialZmiany).filter(
            models.PrzydzialZmiany.data >= start, models.PrzydzialZmiany.data <= end).all():
        obsadzone[(p.data, p.stanowisko_id)] += 1
    stan_nazwa = {s.id: s.nazwa for s in db.query(models.Stanowisko).all()}
    alerty = []
    for (d, sid), wym in wymagane.items():
        obs = obsadzone.get((d, sid), 0)
        if obs < wym:
            alerty.append({"data": str(d), "stanowisko": stan_nazwa.get(sid, "?"),
                           "wymagane": wym, "obsadzone": obs, "brakuje": wym - obs})
    alerty.sort(key=lambda x: (x["data"], x["stanowisko"]))
    return {"alerty": alerty, "razem_brakuje": sum(a["brakuje"] for a in alerty)}


@app.get("/api/alerty-obsady")
def alerty_obsady(dni: int = 14, db: Session = Depends(get_db)):
    """Nadchodzące dni z NIEDOBOREM obsady: wymagana liczba osób (WymaganiaDnia) > przydzielona
    (PrzydzialZmiany) na danym stanowisku. Wsparcie decyzji „gdzie brakuje ludzi". Admin."""
    dni = max(1, min(int(dni), 60))
    today = date.today()
    wynik = _alerty_obsady_okres(db, today, today + timedelta(days=dni))
    return {**wynik, "dni": dni}


@app.get("/api/szef/grafik")
def szef_grafik(start: date = Query(...), end: date = Query(...), db: Session = Depends(get_db)):
    """Published-only grafik szefa z bezpiecznym alertem obsady na bieżący dzień."""
    if end < start:
        raise HTTPException(400, "Koniec okresu przed początkiem.")
    if (end - start).days > 62:
        raise HTTPException(400, "Zakres grafiku może obejmować maksymalnie 63 dni.")

    publikacja = db.query(models.PublikacjaGrafiku).filter_by(start=start, koniec=end).first()
    if not publikacja:
        return {
            "opublikowany": False,
            "opublikowano_at": None,
            "przydzialy": [],
            "alerty_dzis": [],
            "razem_brakuje_dzis": 0,
        }

    przydzialy = db.query(models.PrzydzialZmiany).filter(
        models.PrzydzialZmiany.data >= start,
        models.PrzydzialZmiany.data <= end,
    ).all()
    today = date.today()
    alerty_dzis = (
        _alerty_obsady_okres(db, today, today)
        if start <= today <= end
        else {"alerty": [], "razem_brakuje": 0}
    )
    return {
        "opublikowany": True,
        "opublikowano_at": publikacja.opublikowano_at.isoformat(),
        "przydzialy": [
            schemas.PrzydzialOut.model_validate(p).model_dump(mode="json")
            for p in przydzialy
        ],
        "alerty_dzis": alerty_dzis["alerty"],
        "razem_brakuje_dzis": alerty_dzis["razem_brakuje"],
    }


# ── NAPIWKI (admin; helpery podziału w deps.py, /api/me/napiwki w routers/moje.py) ──

@app.get("/api/napiwki")
def get_napiwki(data: date = Query(...), db: Session = Depends(get_db)):
    """Manager: podział napiwków dnia na obsługę sali (wg zapisanej kwoty/sposobu)."""
    return _napiwki_podzial(db, data)


@app.put("/api/napiwki")
def zapisz_napiwki(dane: schemas.NapiwkiIn, data: date = Query(...), db: Session = Depends(get_db)):
    """Manager: ustaw pulę napiwków dnia i sposób podziału ('godziny'|'rowno')."""
    if (dane.kwota or 0) < 0:
        raise HTTPException(400, "Kwota napiwków nie może być ujemna.")
    sposob = dane.sposob if dane.sposob in ("godziny", "rowno") else "godziny"
    rec = db.query(models.NapiwkiDnia).filter_by(data=data).first()
    if rec is None:
        rec = models.NapiwkiDnia(data=data, utworzono_at=utcnow_naive())
        db.add(rec)
    rec.kwota = float(dane.kwota or 0); rec.sposob = sposob
    db.commit()
    return _napiwki_podzial(db, data)


# ── PROGNOZA RUCHU (z historii StolikiHistoria — wsparcie decyzji o obsadzie) ──

_DNI_TYG = ["Poniedziałek", "Wtorek", "Środa", "Czwartek", "Piątek", "Sobota", "Niedziela"]


def _zabukowane_7dni(db, today):
    """Zabukowane rezerwacje stolikowe na najbliższe 7 dni (per data): liczba (≈ rachunki) i covery
    (osoby). Zasila prognozę realnymi bukowaniami — floor przez max(), nie sumę (historia już
    zawiera zrealizowane rezerwacje → sumowanie zawyżałoby = double counting)."""
    rez = db.query(models.Termin).filter(
        models.Termin.rodzaj == "stolik", models.Termin.status.in_(REZ_AKTYWNE),
        models.Termin.data > today, models.Termin.data <= today + timedelta(days=7)).all()
    agg = defaultdict(lambda: {"rezerwacje": 0, "covery": 0})
    for t in rez:
        agg[t.data]["rezerwacje"] += 1
        agg[t.data]["covery"] += (t.liczba_osob or 0)
    return agg


@app.get("/api/prognoza-ruchu")
def prognoza_ruchu(dni: int = 90, db: Session = Depends(get_db)):
    """Prognoza ruchu (liczba rachunków) z historii: średnia per dzień tygodnia, trend
    28d vs poprzednie 28d, projekcja na najbliższe 7 dni. Czysto z danych StolikiHistoria."""
    dni = max(7, min(int(dni), 365))
    today = date.today()
    od = today - timedelta(days=dni)
    rows = db.query(models.StolikiHistoria).filter(models.StolikiHistoria.data >= od,
                                                   models.StolikiHistoria.data <= today).all()
    wg_dnia = defaultdict(list)
    for r in rows:
        wg_dnia[r.data.weekday()].append(int(r.liczba or 0))
    per_dzien = []
    for w in range(7):
        vals = wg_dnia.get(w, [])
        per_dzien.append({"dzien": w, "nazwa": _DNI_TYG[w],
                          "srednia": round(sum(vals) / len(vals), 1) if vals else 0,
                          "max": max(vals) if vals else 0, "probek": len(vals)})
    # Trend: ostatnie 28 dni vs poprzednie 28 dni — liczone na WŁASNYM, PEŁNYM oknie 56 dni,
    # niezależnie od parametru `dni`. Inaczej wybór krótszego okna (np. 30 dni) zostawiał
    # poprzednie 28 dni prawie puste → absurdalny %. Okna SYMETRYCZNE (28 vs 28 dni).
    trend_rows = rows if dni >= 55 else db.query(models.StolikiHistoria).filter(
        models.StolikiHistoria.data >= today - timedelta(days=55),
        models.StolikiHistoria.data <= today).all()
    def _suma(start, end):
        return sum(int(r.liczba or 0) for r in trend_rows if start <= r.data < end)
    okno1 = _suma(today - timedelta(days=27), today + timedelta(days=1))    # 28 dni: today-27..today
    okno0 = _suma(today - timedelta(days=55), today - timedelta(days=27))   # 28 dni: today-55..today-28
    trend = round((okno1 - okno0) / okno0 * 100, 1) if okno0 else None
    # Projekcja na 7 dni wg średniej dnia tygodnia + sugerowana obsada (wg parametrów lokalu).
    srednie = {p["dzien"]: p["srednia"] for p in per_dzien}
    cfg = get_lokal_config(db)
    na_osobe = max(1, int(cfg.obsada_rachunki_na_osobe or 20))
    obsada_min = max(1, int(cfg.obsada_min or 1))
    zab = _zabukowane_7dni(db, today)
    projekcja = []
    for i in range(1, 8):
        d = today + timedelta(days=i)
        prog = srednie.get(d.weekday(), 0)
        z = zab.get(d, {"rezerwacje": 0, "covery": 0})
        prog_eff = max(prog, z["rezerwacje"])   # rezerwacja ≈ rachunek; max = brak double-count z historią
        obsada = max(obsada_min, math.ceil(prog_eff / na_osobe)) if prog_eff else obsada_min
        projekcja.append({"data": str(d), "nazwa": _DNI_TYG[d.weekday()], "prognoza": prog,
                          "covery_zabukowane": z["covery"], "rezerwacje_zabukowane": z["rezerwacje"],
                          "zrodlo": ("rezerwacje" if z["rezerwacje"] > prog else "historia"),
                          "sugerowana_obsada": obsada})
    return {"okres_dni": dni, "probek": len(rows),
            "srednia_dzienna": round(sum(int(r.liczba or 0) for r in rows) / len(rows), 1) if rows else 0,
            "trend_28d_proc": trend,
            "parametry_obsady": {"rachunki_na_osobe": na_osobe, "min": obsada_min},
            "per_dzien_tygodnia": per_dzien,
            "projekcja_7dni": projekcja}


def _projekcja_obsady(db):
    """Projekcja 7 dni: prognoza (śr. rachunki wg dnia tygodnia z ostatnich 90 dni) + sugerowana
    obsada (ceil(prognoza / rachunki_na_osobe), min obsada_min z LokalConfig). Współdzielona baza
    dla /api/prognoza-ruchu (informacyjnie) i /api/wymagania/z-prognozy (auto-obsada)."""
    today = date.today()
    rows = db.query(models.StolikiHistoria).filter(
        models.StolikiHistoria.data >= today - timedelta(days=90),
        models.StolikiHistoria.data <= today).all()
    wg = defaultdict(list)
    for r in rows:
        wg[r.data.weekday()].append(int(r.liczba or 0))
    srednie = {w: (round(sum(v) / len(v), 1) if v else 0) for w, v in wg.items()}
    cfg = get_lokal_config(db)
    na_osobe = max(1, int(cfg.obsada_rachunki_na_osobe or 20))
    obsada_min = max(1, int(cfg.obsada_min or 1))
    zab = _zabukowane_7dni(db, today)
    out = []
    for i in range(1, 8):
        d = today + timedelta(days=i)
        prog = srednie.get(d.weekday(), 0)
        z = zab.get(d, {"rezerwacje": 0, "covery": 0})
        prog_eff = max(prog, z["rezerwacje"])   # floor rezerwacjami (max, nie suma — bez double-count)
        obsada = max(obsada_min, math.ceil(prog_eff / na_osobe)) if prog_eff else obsada_min
        out.append({"data": str(d), "nazwa": _DNI_TYG[d.weekday()], "prognoza": prog,
                    "covery_zabukowane": z["covery"], "rezerwacje_zabukowane": z["rezerwacje"],
                    "zrodlo": ("rezerwacje" if z["rezerwacje"] > prog else "historia"),
                    "sugerowana_obsada": obsada})
    return out


@app.post("/api/wymagania/z-prognozy", status_code=200)
def wymagania_z_prognozy(dane: schemas.WymaganiaZPrognozy, db: Session = Depends(get_db)):
    """Auto-obsada: tworzy/aktualizuje wymagania na najbliższe 7 dni z sugerowanej obsady prognozy,
    na wskazanym stanowisku (zwykle Sala). Upsert po (data, stanowisko, godz_od=None, rewir=None)."""
    stan = db.get(models.Stanowisko, dane.stanowisko_id)
    if stan is None:
        raise HTTPException(404, "Stanowisko nie istnieje.")
    projekcja = _projekcja_obsady(db)
    zastosowano = 0
    for p in projekcja:
        d = date.fromisoformat(p["data"])
        osob = int(p["sugerowana_obsada"])
        if osob < 1:
            continue
        istn = db.query(models.WymaganiaDnia).filter_by(
            data=d, stanowisko_id=stan.id, godz_od=None, rewir=None).first()
        if istn:
            istn.liczba_osob = osob
        else:
            db.add(models.WymaganiaDnia(data=d, stanowisko_id=stan.id, godz_od=None,
                                        rewir=None, liczba_osob=osob))
        zastosowano += 1
    db.commit()
    return {"zastosowano": zastosowano, "stanowisko": stan.nazwa, "projekcja": projekcja}


# ── KALENDARZ IMPREZ ──────────────────────────────────────────────────────────

def _termin_out(t: models.Termin, zadatek_kp: float = 0.0) -> dict:
    return {"id": t.id, "data": str(t.data), "nazwisko": t.nazwisko, "typ": t.typ,
            "liczba_osob": t.liczba_osob, "telefon": t.telefon, "sala": t.sala,
            "notatka": t.notatka, "status": t.status, "zadatek": t.zadatek or 0,
            "zadatek_kp": round(zadatek_kp, 2)}   # suma zadatków KP przypisanych do terminu


@app.get("/api/terminy", dependencies=[Depends(_wymagaj_modul_imprezy)])
def get_terminy(start: date = Query(...), end: date = Query(...), db: Session = Depends(get_db)):
    rows = (db.query(models.Termin)
            .filter(models.Termin.data >= start, models.Termin.data <= end,
                    models.Termin.rodzaj != "stolik")     # kalendarz imprez ≠ rezerwacje stolików (osobne byty)
            .order_by(models.Termin.data.asc(), models.Termin.id.asc()).all())
    kp_sum = {}
    ids = [t.id for t in rows]
    if ids:
        for z in db.query(models.KpZadatek).filter(models.KpZadatek.termin_id.in_(ids)).all():
            kp_sum[z.termin_id] = kp_sum.get(z.termin_id, 0) + (z.kwota or 0)
    return {"terminy": [_termin_out(t, kp_sum.get(t.id, 0)) for t in rows]}


@app.post("/api/terminy", status_code=201, dependencies=[Depends(_wymagaj_modul_imprezy)])
def dodaj_termin(dane: schemas.TerminIn, db: Session = Depends(get_db)):
    t = models.Termin(data=dane.data, nazwisko=dane.nazwisko.strip(), typ=dane.typ,
                      liczba_osob=dane.liczba_osob, telefon=dane.telefon, sala=dane.sala,
                      notatka=dane.notatka, status=dane.status or "rezerwacja", rodzaj="impreza",
                      zadatek=float(dane.zadatek or 0), utworzono_at=utcnow_naive())
    db.add(t); db.commit(); db.refresh(t)
    return _termin_out(t)


@app.put("/api/terminy/{termin_id}", dependencies=[Depends(_wymagaj_modul_imprezy)])
def edytuj_termin(termin_id: int, dane: schemas.TerminIn, db: Session = Depends(get_db)):
    t = db.get(models.Termin, termin_id)
    if not t or t.rodzaj == "stolik":            # rezerwacji stolika nie edytuje się przez API imprez
        raise HTTPException(404, "Brak terminu.")
    crm_identity_before = _identity_key_crm(t)
    stara_data = t.data
    t.data = dane.data; t.nazwisko = dane.nazwisko.strip(); t.typ = dane.typ
    t.liczba_osob = dane.liczba_osob; t.telefon = dane.telefon; t.sala = dane.sala
    t.notatka = dane.notatka; t.status = dane.status or "rezerwacja"; t.zadatek = float(dane.zadatek or 0)
    # Termin z iCloud ma sparowaną Imprezę (obsada) — synchronizujemy ją z ręczną edycją,
    # żeby korekta liczby osób / daty / sali zmieniła wymagania obsady.
    if t.ical_uid:
        imp = db.query(models.Impreza).filter(models.Impreza.sciezka_pliku == f"ical:{t.ical_uid}").first()
        if imp is not None:
            imp.data = t.data; imp.klient = t.nazwisko
            imp.liczba_osob = (t.liczba_osob or 0); imp.sala = (t.sala or "Brak")
    if t.rodzaj == "sala":
        _migruj_osierocony_profil_po_zmianie_tozsamosci(db, t, crm_identity_before)
        crm_governance.revert_orphaned_identity_merges_after_contact_change(
            db,
            reservation_id=t.id,
            previous_identity_key=crm_identity_before,
            actor=None,
        )
    db.commit(); db.refresh(t)
    if t.ical_uid:
        _odswiez_wymagania_imprez(db, min(stara_data, t.data), max(stara_data, t.data))
    return _termin_out(t)


@app.delete("/api/terminy/{termin_id}", status_code=204, dependencies=[Depends(_wymagaj_modul_imprezy)])
def usun_termin(termin_id: int, db: Session = Depends(get_db)):
    t = db.get(models.Termin, termin_id)
    if t and t.rodzaj == "stolik":               # rezerwacji stolika nie kasuje się przez API imprez
        raise HTTPException(404, "Brak terminu.")
    if t:
        for z in db.query(models.KpZadatek).filter_by(termin_id=termin_id).all():
            z.termin_id = None       # odepnij zadatki (zostają w skrzynce)
        uid, data_t = t.ical_uid, t.data
        _usun_profil_fallbacku_rezerwacji(db, t.id)
        db.delete(t)
        # Usuń też sparowaną Imprezę z iCloud (obsada) i odśwież wymagania na ten dzień.
        if uid:
            imp = db.query(models.Impreza).filter(models.Impreza.sciezka_pliku == f"ical:{uid}").first()
            if imp is not None:
                db.delete(imp)
        db.commit()
        if uid:
            _odswiez_wymagania_imprez(db, data_t, data_t)


# ═══════════════════════════════════════════════════════════════════════════
# MODUŁ REZERWACJI (stoliki + godziny otwarcia + rezerwacje na encji Termin)
# ═══════════════════════════════════════════════════════════════════════════

REZ_STATUSY = ("rezerwacja", "potwierdzona", "odbyla", "no_show", "odwolana")
# Dozwolone przejścia statusu rezerwacji (rezerwacja/potwierdzona = aktywne, reszta terminalna).
REZ_PRZEJSCIA = {
    "rezerwacja":   {"potwierdzona", "odbyla", "no_show", "odwolana"},
    "potwierdzona": {"odbyla", "no_show", "odwolana"},
    "odbyla": set(), "no_show": set(), "odwolana": set(),
}
REZ_AKTYWNE = ("rezerwacja", "potwierdzona")   # blokują stolik (liczą się do kolizji)
DOMYSLNY_SLOT_MIN = 120
MAX_REZERWACJE_SEARCH_DAYS = 366

# Faza operacyjna hosta (obok status księgowego). NULL = jeszcze nie przyszedł.
HOST_FAZY = ("przybyl", "posadzony", "rachunek", "oplacony", "wyszedl")
HOST_PRZEJSCIA = {
    None:         {"przybyl", "posadzony"},           # z niczego: przyjście lub od razu posadzenie
    "przybyl":    {"posadzony", "wyszedl"},
    "posadzony":  {"rachunek", "oplacony", "wyszedl"},
    "rachunek":   {"oplacony", "wyszedl"},
    "oplacony":   {"wyszedl"},
    "wyszedl":    set(),
}
HOST_NA_SALI = ("posadzony", "rachunek", "oplacony")


def _wymagaj_modul_rezerwacje(db: Session = Depends(get_db)):
    if not modul_aktywny(db, "modul_rezerwacje"):
        raise HTTPException(403, "Moduł rezerwacji jest niedostępny w tym planie — odblokujesz go w pakiecie Pro.")


def _dodaj_minuty(t: time, minuty: int) -> time:
    """t + minuty, obcięte do tej samej doby [00:00, 23:59] (bez przejścia przez północ w MVP;
    ujemne minuty dozwolone — dla bufora cofającego początek okna, clamp do 00:00)."""
    total = min(max(0, t.hour * 60 + t.minute + minuty), 23 * 60 + 59)
    return time(total // 60, total % 60)


def _serwisy_dnia(db, data: date):
    """Kompatybilny adapter do kanonicznego resolvera serwisów R3."""
    return list(reservation_rules.serwisy_dnia(db, data))


def _jest_blackout(db, data) -> bool:
    """Czy dzień ma JAWNY blackout (WyjatekKalendarza). Puste GodzinyOtwarcia ≠ blackout — brak
    skonfigurowanych godzin to historycznie 'otwarte' (rezerwacja dozwolona z DOMYSLNY turn-time)."""
    return db.query(models.WyjatekKalendarza).filter_by(data=data, typ="blackout").first() is not None


def _serwis_dla_godziny(db, data, godz_od):
    """Adapter legacy; zapisy używają ścisłego evaluatora R3 bez fallbacku."""
    if godz_od is None:
        return None
    return reservation_rules.serwis_dla_godziny(
        db, data, godz_od, strict=False,
    )


def _turn_time(serwis, liczba_osob) -> int:
    return reservation_rules.turn_time(serwis, liczba_osob)


def _dlugosc_dla(db, data, godz_od, liczba_osob) -> int:
    """Długość zasiadku dla rezerwacji: turn-time serwisu obejmującego godz_od wg wielkości grupy."""
    return _turn_time(_serwis_dla_godziny(db, data, godz_od), liczba_osob)


def _ids_stolikow(wartosci) -> set[int]:
    """Normalizuje JSON-ową listę id; stare/uszkodzone elementy nie mogą wywrócić operacji ochronnej."""
    if not isinstance(wartosci, (list, tuple, set)):
        return set()
    ids = set()
    for wartosc in (wartosci or []):
        try:
            ids.add(int(wartosc))
        except (TypeError, ValueError):
            continue
    return ids


def _stoly_terminu(t) -> set:
    """Wszystkie stoły zajmowane przez rezerwację: wiodący (stolik_id) + składowe kombinacji."""
    stoly = _ids_stolikow(t.stoliki_dodatkowe)
    if t.stolik_id:
        stoly.add(t.stolik_id)
    return stoly


def _waliduj_przydzial_rezerwacji(
    db, data, godz_od, godz_do, stoliki, liczba_osob, pomin_id=None, pojemnosc_override=None,
    zachowaj_nieaktywny_przydzial=False,
):
    """Waliduje cały przydział (pojedynczy stół albo zachowaną kombinację) i zwraca godz_do."""
    ids = _ids_stolikow(stoliki)
    if not ids:
        return godz_do
    rekordy = reservation_service.lock_tables(db, ids)
    runtime_by_id = {stolik["id"]: stolik for stolik in _stoly_do_seating(db)}
    if zachowaj_nieaktywny_przydzial:
        # Wyłączenie sali blokuje nowe przydziały, ale nie może uniemożliwić
        # obsługi istniejącej rezerwacji, która zachowuje dokładnie ten sam zestaw.
        for stolik in _stoliki_do_odczytu(db):
            if stolik["id"] in ids and stolik["id"] not in runtime_by_id:
                runtime_by_id[stolik["id"]] = stolik
    if len(rekordy) != len(ids) or not ids <= set(runtime_by_id):
        raise HTTPException(400, "Nieznany lub nieaktywny stolik.")
    pojemnosc_fizyczna = sum(runtime_by_id[sid]["pojemnosc"] for sid in ids)
    pojemnosc = pojemnosc_override if pojemnosc_override is not None else pojemnosc_fizyczna
    if liczba_osob and liczba_osob > pojemnosc:
        nazwa = " + ".join(runtime_by_id[sid]["nazwa"] for sid in sorted(ids))
        raise HTTPException(400, f"Przydział „{nazwa}” mieści {pojemnosc} os. (próba: {liczba_osob}).")
    if godz_od is None:
        return godz_do
    if godz_do is None:
        godz_do = _dodaj_minuty(godz_od, _dlugosc_dla(db, data, godz_od, liczba_osob))
    zajete = reservation_service.occupied_table_ids(
        db,
        data=data,
        start=godz_od,
        end=godz_do,
        # Szybka kontrola konfliktu fizycznego. Wlasciwy, typowany bufor R3
        # (serwis/sala/wyjatek, lacznie z jawnym zerem) egzekwuje atomowo ledger.
        buffer_min=0,
        exclude_termin_id=pomin_id,
        now=utcnow_naive(),
    )
    if ids & zajete:
        raise reservation_service.ReservationError(
            409,
            "TABLE_CONFLICT",
            "Stolik zajęty w tym czasie.",
            rule="table",
        )
    return godz_do


def _waliduj_rezerwacje(db, data, godz_od, godz_do, stolik_id, liczba_osob, pomin_id=None):
    """Kompatybilna walidacja ręcznego przydziału pojedynczego stolika."""
    stoliki = [stolik_id] if stolik_id is not None else []
    return _waliduj_przydzial_rezerwacji(
        db, data, godz_od, godz_do, stoliki, liczba_osob, pomin_id=pomin_id)


def _parametry_pacingu(db, data, godz_od):
    serwis = _serwis_dla_godziny(db, data, godz_od) if godz_od else None
    def dodatni_limit(value):
        return value if value is not None and value > 0 else None
    return {
        # Starsze bazy mogły zawierać zero lub wartość ujemną. Obie oznaczają brak limitu;
        # nowe wartości ujemne odrzuca już schema wejściowa.
        "max_reservations": dodatni_limit(
            getattr(serwis, "pacing_max_rez", None) if serwis else None
        ),
        "max_covers": dodatni_limit(
            getattr(serwis, "pacing_max_osob", None) if serwis else None
        ),
        "pacing_window_min": (
            getattr(serwis, "pacing_okno_min", None)
            or getattr(serwis, "dlugosc_slotu_min", None)
            or DOMYSLNY_SLOT_MIN
        ),
    }


def _sala_dla_stolikow(db, stoliki) -> Optional[int]:
    """Rozwiązuje salę przydziału także dla legacy stolików mapowanych po strefie."""
    ids = _ids_stolikow(stoliki)
    if not ids:
        return None
    rows = db.query(models.Stolik).filter(models.Stolik.id.in_(ids)).all()
    if len(rows) != len(ids):
        # Walidacja przydziału zwróci właściwy błąd zasobu; evaluator może nadal
        # bezpiecznie sprawdzić limity globalne i kanałowe.
        return None
    room_by_name = {
        (room.nazwa or "").strip().casefold(): room.id
        for room in db.query(models.SalaRezerwacyjna).all()
    }
    sale = {
        row.sala_id
        if row.sala_id is not None
        else room_by_name.get((row.strefa or "").strip().casefold())
        for row in rows
    }
    niepuste = {sala_id for sala_id in sale if sala_id is not None}
    if len(niepuste) > 1 or (niepuste and None in sale):
        raise HTTPException(400, "Stoły jednego przydziału muszą należeć do tej samej sali.")
    return next(iter(niepuste)) if niepuste else None


def _ocen_reguly_slotu(
    db,
    *,
    data,
    godz_od,
    liczba_osob,
    kanal,
    godz_do=None,
    intent="create",
    sala_id=None,
    existing_termin_id=None,
    preserve_existing_room_access=False,
    preserve_explicit_interval=False,
    now=None,
):
    if godz_od is None:
        return None
    return reservation_rules.evaluate_reservation_rules(
        db,
        reservation_rules.RuleRequest(
            data=data,
            godz_od=godz_od,
            godz_do=godz_do,
            liczba_osob=max(1, int(liczba_osob or 1)),
            kanal=reservation_service.normalise_reservation_channel(kanal),
            sala_id=sala_id,
            existing_termin_id=existing_termin_id,
            intent=intent,
            preserve_existing_room_access=preserve_existing_room_access,
            preserve_explicit_interval=preserve_explicit_interval,
        ),
        # Reguły biznesowe (okno wyprzedzenia/cutoff/DST) operują w czasie
        # lokalu. Lifecycle holdów i claimów pozostaje osobno w naive UTC.
        now=(
            now.replace(tzinfo=timezone.utc).astimezone(
                ZoneInfo("Europe/Warsaw"),
            )
            if now is not None and now.tzinfo is None
            else (
                now.astimezone(ZoneInfo("Europe/Warsaw"))
                if now is not None
                else (_teraz_lokalnie() or datetime.now(timezone.utc))
            )
        ),
    )


def _ocen_reguly_terminu(
    db,
    t,
    *,
    stoliki=None,
    intent="create",
    sala_id=None,
    preserve_existing_room_access=False,
):
    ids = _ids_stolikow(stoliki if stoliki is not None else _stoly_terminu(t))
    room_id = sala_id if sala_id is not None else _sala_dla_stolikow(db, ids)
    return _ocen_reguly_slotu(
        db,
        data=t.data,
        godz_od=t.godz_od,
        godz_do=t.godz_do,
        liczba_osob=t.liczba_osob,
        kanal=t.kanal,
        sala_id=room_id,
        existing_termin_id=(t.id if getattr(t, "id", None) else None),
        intent=intent,
        preserve_existing_room_access=preserve_existing_room_access,
    )


def _zastap_ledger_terminu(
    db,
    t,
    *,
    stoliki=None,
    enforce_pacing=False,
    candidates=(),
    alternatives=(),
    zachowaj_nieaktywny_przydzial=False,
    evaluation=None,
    override=False,
    intent="create",
    buffer_override=None,
    preserve_interval=False,
    now=None,
):
    """Aktualizuje ledger w tej samej transakcji co projekcję ``Termin``."""
    if t.status not in REZ_AKTYWNE:
        reservation_service.release_termin_allocation(db, t.id)
        return reservation_service.AvailabilityResult(available=True)
    ids = _ids_stolikow(stoliki if stoliki is not None else _stoly_terminu(t))
    locked_tables = reservation_service.lock_tables(db, ids)
    runtime_ids = {stolik["id"] for stolik in _stoly_do_seating(db)}
    if len(locked_tables) != len(ids) or (
        not zachowaj_nieaktywny_przydzial and not ids <= runtime_ids
    ):
        raise HTTPException(400, "Nieznany lub nieaktywny stolik.")
    room_id = _sala_dla_stolikow(db, ids)
    evaluation = evaluation or _ocen_reguly_terminu(
        db,
        t,
        stoliki=ids,
        intent=intent,
        sala_id=room_id,
        preserve_existing_room_access=zachowaj_nieaktywny_przydzial,
    )
    if (
        evaluation is not None
        and evaluation.godz_do is not None
        and not preserve_interval
    ):
        t.godz_do = evaluation.godz_do
    if evaluation is not None and enforce_pacing:
        reservation_rules.enforce_rule_evaluation(
            evaluation,
            override=bool(override),
            can_override=bool(override),
        )
    pacing = _parametry_pacingu(db, t.data, t.godz_od)
    has_typed_buffer = bool(
        evaluation is not None
        and any(
            rule.get("key") == "buffer_min"
            for rule in evaluation.applied_rules
        )
    )
    buffer_min = (
        max(0, int(buffer_override))
        if buffer_override is not None
        else (
            evaluation.buffer_min
            if has_typed_buffer
            else (get_lokal_config(db).rez_bufor_min or 0)
        )
    )
    return reservation_service.replace_termin_allocation(
        db,
        termin_id=t.id,
        data=t.data,
        start=t.godz_od,
        end=t.godz_do,
        table_ids=ids,
        party_size=t.liczba_osob or 1,
        buffer_min=buffer_min,
        # Kanoniczny evaluator wykonał kontrolę już pod blokadą dnia. Stary
        # checker pozostaje dla bezpośrednich klientów warstwy service, lecz tutaj
        # nie może nadpisać typowanych wyjątków R3 ustawieniami legacy.
        enforce_pacing=False,
        override=bool(override),
        room_id=room_id,
        channel=reservation_service.normalise_reservation_channel(t.kanal),
        now=(now if now is not None else utcnow_naive()),
        candidates=candidates,
        alternatives=alternatives,
        **pacing,
    )


def _commit_zapis_rezerwacji(db, guards):
    reservation_service.touch_days(guards)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise reservation_service.translate_integrity_error(exc) from exc


def _zablokuj_termin(db, rid, dodatkowe_dni=()):
    """Blokuje aktualny dzień rekordu, odpornie na równoległe przeniesienie rezerwacji."""
    t = db.get(models.Termin, rid)
    if t is None:
        return None, ()
    dni = {t.data, *(d for d in dodatkowe_dni if d is not None)}
    while True:
        guards = reservation_service.begin_locked_write(db, dni)
        t = db.get(models.Termin, rid)
        if t is None:
            db.rollback()
            return None, ()
        if t.data in dni:
            return t, guards
        dni.add(t.data)


def _zablokuj_rezerwacje_online(db, token, dodatkowe_dni=()):
    """Blokuje rezerwację po rekordzie hash-only, także przy idempotentnym replayu."""
    token_record = reservation_service.lookup_management_token(
        db, token, secret=SECRET_KEY,
    )
    if token_record is None:
        return None, ()
    t = db.get(models.Termin, token_record.termin_id)
    if t is None or t.kanal != "online":
        return None, ()
    dni = {t.data, *(d for d in dodatkowe_dni if d is not None)}
    while True:
        guards = reservation_service.begin_locked_write(db, dni)
        token_record = reservation_service.lookup_management_token(
            db, token, secret=SECRET_KEY,
        )
        if token_record is None:
            db.rollback()
            return None, ()
        t = db.get(models.Termin, token_record.termin_id)
        if t is None or t.kanal != "online":
            db.rollback()
            return None, ()
        if t.data in dni:
            return t, guards
        dni.add(t.data)


def _zablokuj_waitliste(db, wid):
    """Blokuje dzień wpisu waitlisty przed zmianą jego holda lub realizacją."""
    w = db.get(models.ListaOczekujacych, wid)
    if w is None:
        return None, ()
    dni = {w.data}
    while True:
        guards = reservation_service.begin_locked_write(db, dni)
        w = db.get(models.ListaOczekujacych, wid)
        if w is None:
            db.rollback()
            return None, ()
        if w.data in dni:
            return w, guards
        dni.add(w.data)


def _jawne_kombinacje_dla_zestawu(db, stoliki):
    ids = _ids_stolikow(stoliki)
    return [
        combination for combination in _kombinacje_do_seating(db)
        if _ids_stolikow(combination["stoliki"]) == ids
    ]


def _zamrozona_kombinacja_terminu(db, termin, stoliki):
    """Zwraca historyczny snapshot kombinacji tylko dla dokładnie tego przydziału."""
    if termin is None or not termin.przydzial_kombinacja_planu_id:
        return None
    version_id = termin.przydzial_wersja_planu_id
    combination = db.query(models.KombinacjaStolowPlanu).filter_by(
        id=termin.przydzial_kombinacja_planu_id,
        wersja_id=version_id,
    ).first()
    if combination is None:
        return None
    member_ids = {
        table_id for (table_id,) in db.query(
            models.SkladnikKombinacjiPlanu.stolik_id,
        ).filter_by(
            kombinacja_id=combination.id,
            wersja_id=version_id,
        ).all()
    }
    if member_ids != _ids_stolikow(stoliki):
        return None
    return combination


def _pojemnosc_zachowanego_zestawu(db, stoliki, termin=None):
    ids = _ids_stolikow(stoliki)
    frozen = _zamrozona_kombinacja_terminu(db, termin, ids)
    if frozen is not None:
        return frozen.pojemnosc_max
    jawne = _jawne_kombinacje_dla_zestawu(db, ids)
    if not jawne:
        return None
    runtime_by_id = {stolik["id"]: stolik for stolik in _stoly_do_seating(db)}
    if not ids <= set(runtime_by_id):
        # Właściwa walidacja przydziału zwróci czytelne 400; nie wywracaj starego,
        # osieroconego JSON-u wyjątkiem AttributeError przed tą walidacją.
        return None
    fizyczna = sum(runtime_by_id[sid]["pojemnosc"] for sid in ids)
    return max((combination["pojemnosc_max"] or fizyczna) for combination in jawne)


def _waliduj_rozmiar_zachowanej_kombinacji(
    db, stoliki, liczba_osob, termin=None,
):
    """Zmiana liczby gości musi pozostawić istniejący zestaw legalnym kandydatem silnika."""
    ids = _ids_stolikow(stoliki)
    if len(ids) < 2:
        return
    osoby = max(1, int(liczba_osob or 1))
    frozen = _zamrozona_kombinacja_terminu(db, termin, ids)
    if termin is not None and termin.przydzial_kombinacja_planu_id and frozen is None:
        raise HTTPException(409, "Historyczna proweniencja przydziału jest niespójna.")
    jawne = _jawne_kombinacje_dla_zestawu(db, ids)
    if frozen is not None:
        dozwolone = frozen.pojemnosc_min <= osoby <= frozen.pojemnosc_max
    elif jawne:
        runtime_by_id = {stolik["id"]: stolik for stolik in _stoly_do_seating(db)}
        pojemnosc_fizyczna = sum(
            runtime_by_id[sid]["pojemnosc"] for sid in ids
            if sid in runtime_by_id
        )
        dozwolone = any(
            (combination["pojemnosc_min"] or 1)
            <= osoby
            <= (combination["pojemnosc_max"] or pojemnosc_fizyczna)
            for combination in jawne
        )
    elif ids & _wersjonowane_stoliki_ids(db):
        # Starszy przydział sprzed R2.2b nie ma wiarygodnej proweniencji. Zachowujemy
        # kompatybilność bez zgadywania wersji; nowe przydziały zawsze zapisują snapshot.
        dozwolone = True
    else:
        kandydaci = seating.kandydaci(
            osoby, _stoly_do_seating(db), [], sasiedztwo=_sasiedztwo_do_seating(db))
        dozwolone = any(_ids_stolikow(k["stoliki"]) == ids for k in kandydaci)
    if not dozwolone:
        raise HTTPException(
            400,
            "Ta liczba gości nie mieści się w dozwolonym zakresie przypisanej kombinacji stołów.",
        )


def _hm(t):
    return t.strftime("%H:%M") if t else None


def _ma_dostep_rezerwacji(user, permission: str) -> bool:
    """``None`` oznacza zaufany kontekst systemowy; odpowiedzi HTTP zawsze przekazują aktora."""
    return user is None or uprawnienia.ma_user(user, permission)


def _jawne_nadpisanie_limitow(user, requested: bool, confirmation=None) -> Optional[dict]:
    """Normalizuje stare potwierdzenie bool i nowy, audytowalny kontrakt R3."""
    if not requested and confirmation is None:
        return None
    if not _ma_dostep_rezerwacji(user, "rezerwacje.nadpisuj_limity"):
        raise HTTPException(403, "Brak uprawnienia do przekraczania limitów rezerwacji.")
    if confirmation is not None:
        return {
            "reason_code": confirmation.powod,
            "note": confirmation.notatka,
            "legacy": False,
        }
    return {
        "reason_code": "legacy_confirmation",
        "note": None,
        "legacy": True,
    }


def _wymagaj_reautoryzacji_nadpisania(
    db: Session,
    request: Request,
    user: models.User,
    override_active: bool,
) -> None:
    """PIN session needs a fresh one-use proof only for a real R3 override."""
    if override_active:
        workstation_auth.consume_reauth_grant(
            db,
            request=request,
            user=user,
            scope=workstation_auth.REAUTH_SCOPE_RESERVATION_OVERRIDE,
        )


def _szczegoly_przekroczenia_pacingu(db, t: models.Termin) -> Optional[dict]:
    """Zwraca wyłącznie faktycznie przekroczone, nieosobowe parametry limitu."""
    if t.status not in REZ_AKTYWNE or t.godz_od is None:
        return None
    params = _parametry_pacingu(db, t.data, t.godz_od)
    status = reservation_service.pacing_status(
        db,
        data=t.data,
        start=t.godz_od,
        window_min=params["pacing_window_min"],
        party_size=t.liczba_osob or 1,
        max_reservations=params["max_reservations"],
        max_covers=params["max_covers"],
        exclude_termin_id=t.id,
    )
    violations = []
    if status["reservation_full"]:
        violations.append({
            "rule": "pacing_reservations",
            "observed": status["reservations"],
            "limit": status["max_reservations"],
            "projected": status["would_reservations"],
        })
    if status["covers_full"]:
        violations.append({
            "rule": "pacing_covers",
            "observed": status["covers"],
            "limit": status["max_covers"],
            "projected": status["would_covers"],
        })
    return {"violations": violations} if violations else None


def _szczegoly_nadpisania_r3(evaluation, *, legacy=False) -> Optional[dict]:
    """Buduje wyłącznie bezpieczne, typowane metadane faktycznie złamanych reguł."""
    if evaluation is None or evaluation.decision != "override_required":
        return None
    violations = []
    for item in evaluation.violations:
        row = {
            "rule": item.rule,
            "observed": item.observed,
            "limit": item.limit,
            "projected": item.projected,
        }
        if not legacy:
            row["code"] = item.code
            row["scope"] = item.scope
            row["source"] = item.source
        violations.append(row)
    return {"violations": violations} if violations else None


def _powod_audytu_nadpisania(evaluation) -> str:
    rules = {item.rule for item in (evaluation.violations if evaluation else ())}
    if rules & {"concurrent_reservations", "concurrent_covers"}:
        return "capacity_override"
    if rules & {"pacing_reservations", "pacing_covers"}:
        return "pacing_override"
    # Ogolne nadpisania dostepnosci korzystaja z istniejacej wartosci CHECK.
    # Szczegolowy typ reguly i kod powodu operatora sa zapisane osobno w
    # bezpiecznych metadanych R3 oraz szyfrowanym kontekscie override.
    return "other"


def _wartosci_pii_rezerwacji(t: models.Termin) -> dict:
    """Chwilowy stan do porównania; wartości nigdy nie trafiają do dziennika audytu."""
    return {
        field: getattr(t, field, None)
        for field in reservation_audit.PII_FIELDS
    }


def _zmienione_pii(before: dict, after: models.Termin) -> set[str]:
    return {
        field
        for field, old_value in before.items()
        if old_value != getattr(after, field, None)
    }


def _utworzone_pii(t: models.Termin) -> set[str]:
    return {
        field
        for field in reservation_audit.PII_FIELDS
        if getattr(t, field, None) not in (None, "")
    }


def _termin_z_replay_idempotencji(db, decision) -> models.Termin:
    """Replay zwraca bieżący rekord, nigdy historyczny JSON z dawnymi uprawnieniami."""
    termin_id = getattr(getattr(decision, "record", None), "termin_id", None)
    db.rollback()
    termin = db.get(models.Termin, termin_id) if termin_id else None
    if termin is None or termin.rodzaj != "stolik":
        raise reservation_service.ReservationError(
            409,
            "IDEMPOTENCY_TARGET_GONE",
            "Poprzednio utworzona rezerwacja już nie istnieje.",
            rule="idempotency",
        )
    return termin


def _rezerwacja_out(t: models.Termin, user=None) -> dict:
    kontakt = _ma_dostep_rezerwacji(user, "rezerwacje.dane_kontaktowe")
    notatki = _ma_dostep_rezerwacji(user, "rezerwacje.notatki_wewnetrzne")
    finanse = _ma_dostep_rezerwacji(user, "rezerwacje.finanse")
    ukryte = []
    if not kontakt:
        ukryte.extend(("nazwisko", "telefon", "email"))
    if not notatki:
        ukryte.append("notatka")
    if not finanse:
        ukryte.append("zadatek")
    return {
        "id": t.id,
        "data": str(t.data),
        "godz_od": _hm(t.godz_od),
        "godz_do": _hm(t.godz_do),
        "stolik_id": t.stolik_id,
        "stoliki_dodatkowe": t.stoliki_dodatkowe or [],
        "auto_przydzielony": bool(t.auto_przydzielony),
        "przydzial_wersja_planu_id": t.przydzial_wersja_planu_id,
        "przydzial_kombinacja_planu_id": t.przydzial_kombinacja_planu_id,
        "nazwisko": t.nazwisko if kontakt else "Gość",
        "telefon": t.telefon if kontakt else None,
        "email": t.email if kontakt else None,
        "kanal_komunikacji": t.kanal_komunikacji if kontakt else None,
        "liczba_osob": t.liczba_osob,
        "notatka": t.notatka if notatki else None,
        "status": t.status,
        "faza_hosta": t.faza_hosta,
        "zadatek": (t.zadatek or 0) if finanse else None,
        "kanal": t.kanal,
        "ukryte_pola": ukryte,
    }


def _rezerwacje_out_z_komunikacja(db, rows, user) -> list[dict]:
    summaries = reservation_communication.summaries_for_reservations(
        db, (row.id for row in rows),
    )
    output = []
    for row in rows:
        item = _rezerwacja_out(row, user)
        item["communication_summary"] = summaries.get(row.id)
        output.append(item)
    return output


def _waliduj_zakres_bazy_rezerwacji(start: date, end: date) -> None:
    if end < start:
        raise HTTPException(400, "Zakres dat jest odwrócony.")
    dni = (end - start).days + 1
    if dni > MAX_REZERWACJE_SEARCH_DAYS:
        raise HTTPException(
            400,
            f"Baza rezerwacji może obejmować maksymalnie {MAX_REZERWACJE_SEARCH_DAYS} dni.",
        )


def _pasuje_do_wyszukiwania_rezerwacji(t: models.Termin, query: Optional[str]) -> bool:
    if not query:
        return True
    fraza = query.casefold()
    if fraza in (t.nazwisko or "").casefold():
        return True
    # EncryptedString używa niedeterministycznego Ferneta, więc telefonu nie da się
    # bezpiecznie filtrować przez SQL. Porównujemy cyfry dopiero po odszyfrowaniu
    # kandydatów z ograniczonego zakresu dat; minimum 3 cyfry ogranicza przypadkowe trafienia.
    cyfry_query = "".join(ch for ch in query if ch.isdigit())
    if len(cyfry_query) < 3:
        return False
    cyfry_telefonu = "".join(ch for ch in (t.telefon or "") if ch.isdigit())
    return cyfry_query in cyfry_telefonu


def _sortuj_baze_rezerwacji(rows, sort: str):
    def chronologicznie(t):
        return t.data, t.godz_od or time.min, t.id

    if sort == "data_desc":
        return sorted(rows, key=chronologicznie, reverse=True)
    if sort == "nazwisko_asc":
        return sorted(
            rows,
            key=lambda t: ((t.nazwisko or "").casefold(), t.data, t.godz_od or time.min, t.id),
        )
    return sorted(rows, key=chronologicznie)


# ── Stoliki ──────────────────────────────────────────────────────────────────
@app.get("/api/stoliki", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def get_stoliki(db: Session = Depends(get_db)):
    return {"stoliki": _stoliki_do_odczytu(db)}


def _waliduj_parametry_stolika(dane: schemas.StolikIn, db: Session):
    if dane.sala_id is not None:
        sala = db.get(models.SalaRezerwacyjna, dane.sala_id)
        if sala is not None:
            # ``strefa`` pozostaje projekcją dla starszych ekranów i silnika sadzania.
            dane.strefa = sala.nazwa
    else:
        nazwa_strefy = (dane.strefa or "").strip().casefold()
        if nazwa_strefy:
            sala = next(
                (
                    candidate
                    for candidate in db.query(models.SalaRezerwacyjna).all()
                    if (candidate.nazwa or "").strip().casefold() == nazwa_strefy
                ),
                None,
            )
            if sala is not None:
                dane.sala_id = sala.id
                dane.strefa = sala.nazwa
    if dane.pojemnosc_min is not None and dane.pojemnosc_min > dane.pojemnosc:
        raise HTTPException(400, "Minimalna liczba osób nie może przekraczać liczby miejsc stolika.")


    if dane.sala_id is not None and db.get(models.SalaRezerwacyjna, dane.sala_id) is None:
        raise HTTPException(400, "Nieznana sala rezerwacyjna.")


def _sala_ma_wersjonowany_plan(db: Session, sala_id: Optional[int]) -> bool:
    return bool(
        sala_id is not None
        and db.query(models.PlanSali.id).filter_by(sala_id=sala_id).first()
    )


def _stolik_ma_wersjonowany_plan(db: Session, stolik: models.Stolik) -> bool:
    """Uwzględnia kontrolowany fallback ``strefa`` dla niezmigrowanych stolików."""
    if stolik.sala_id is not None:
        return _sala_ma_wersjonowany_plan(db, stolik.sala_id)
    nazwa_strefy = (stolik.strefa or "").strip().casefold()
    if not nazwa_strefy:
        return False
    return any(
        (sala.nazwa or "").strip().casefold() == nazwa_strefy
        for sala in (
            db.query(models.SalaRezerwacyjna)
            .join(models.PlanSali, models.PlanSali.sala_id == models.SalaRezerwacyjna.id)
            .all()
        )
    )


def _wymagaj_legacy_stolikow_poza_planem(db: Session, stoliki_ids) -> None:
    """Legacy CRUD nie może zmieniać konfiguracji należącej do wersjonowanego planu."""
    ids = _ids_stolikow(stoliki_ids)
    if not ids:
        return
    stoliki = db.query(models.Stolik).filter(models.Stolik.id.in_(ids)).all()
    if any(_stolik_ma_wersjonowany_plan(db, stolik) for stolik in stoliki):
        _wymagaj_szkicu_planu(
            "Sąsiedztwo i kombinacje zmieniaj w szkicu sali, a następnie opublikuj plan."
        )


def _wymagaj_szkicu_planu(message: str):
    raise HTTPException(
        409,
        detail={
            "code": "FLOOR_PLAN_VERSIONING_REQUIRED",
            "message": message,
        },
    )


def _blokuj_topologie_aktywnej_oferty(
    db: Session,
    table_ids,
    *,
    now: Optional[datetime] = None,
) -> None:
    """Fence legacy topology writes against current offered table claims.

    Callers first lock the same ``stoliki`` rows as offer creation.  The query
    is intentionally conservative: a short-lived current offer must be
    withdrawn before any physical/topological property of its tables changes.
    """
    ids = tuple(sorted(_ids_stolikow(table_ids)))
    if not ids:
        return
    effective_now = now or reservation_service.lifecycle_now_utc()
    rows = db.query(
        models.RezerwacjaStolikClaim.waitlist_id,
        models.RezerwacjaStolikClaim.stolik_id,
    ).join(
        models.ListaOczekujacych,
        models.ListaOczekujacych.id
        == models.RezerwacjaStolikClaim.waitlist_id,
    ).filter(
        models.RezerwacjaStolikClaim.stolik_id.in_(ids),
        models.RezerwacjaStolikClaim.expires_at > effective_now,
        models.ListaOczekujacych.status == "zaoferowano",
        models.ListaOczekujacych.hold_do > effective_now,
        models.ListaOczekujacych.oferta_wygasa_at
        == models.ListaOczekujacych.hold_do,
    ).all()
    if not rows:
        return
    raise HTTPException(
        409,
        detail={
            "code": "WAITLIST_OFFER_TOPOLOGY_CONFLICT",
            "message": (
                "Te stoliki są objęte aktywną ofertą waitlisty. "
                "Najpierw wycofaj ofertę, a potem zmień konfigurację."
            ),
            "table_ids": sorted({table_id for _, table_id in rows}),
            "waitlist_ids": sorted({waitlist_id for waitlist_id, _ in rows}),
        },
    )


@app.post("/api/stoliki", status_code=201, dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def dodaj_stolik(dane: schemas.StolikIn, db: Session = Depends(get_db)):
    _waliduj_parametry_stolika(dane, db)
    if (
        dane.sala_id is None
        and db.query(models.SalaRezerwacyjna.id).first() is not None
    ):
        _wymagaj_szkicu_planu(
            "Najpierw dodaj salę w konfiguracji, a następnie utwórz w niej stół."
        )
    if _sala_ma_wersjonowany_plan(db, dane.sala_id):
        _wymagaj_szkicu_planu(
            "Dodaj stół w konfiguracji sali. Pozostanie nieaktywny do publikacji planu."
        )
    s = models.Stolik(**dane.model_dump()); db.add(s); db.commit(); db.refresh(s)
    return schemas.StolikOut.model_validate(s).model_dump()


@app.put("/api/stoliki/{sid}", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def edytuj_stolik(sid: int, dane: schemas.StolikIn, db: Session = Depends(get_db)):
    reservation_service.begin_floor_plan_write(db)
    locked = reservation_service.lock_tables(db, [sid])
    s = locked[0] if locked else None
    if not s:
        raise HTTPException(404, "Brak stolika.")
    _waliduj_parametry_stolika(dane, db)
    incoming = dane.model_dump()
    if any(getattr(s, key) != value for key, value in incoming.items()):
        _blokuj_topologie_aktywnej_oferty(db, [sid])
    previous_room_id = s.sala_id
    previous_active = s.aktywny
    if _stolik_ma_wersjonowany_plan(db, s):
        canonical = next(
            (row for row in _stoliki_do_odczytu(db) if row["id"] == sid),
            schemas.StolikOut.model_validate(s).model_dump(),
        )
        guarded_fields = set(schemas.StolikIn.model_fields) - {"rewir_nr"}
        if any(canonical[field] != getattr(dane, field) for field in guarded_fields):
            _wymagaj_szkicu_planu(
                "Właściwości stołu zmieniaj w szkicu sali. Zmiana będzie widoczna po publikacji."
            )
        # Powiązanie POS pozostaje celowo poza snapshotem planu. Nie kopiujemy
        # reszty payloadu do stabilnego rekordu, bo opublikowany snapshot jest
        # jedynym źródłem właściwości operacyjnych.
        s.rewir_nr = dane.rewir_nr
        db.commit()
        canonical["rewir_nr"] = s.rewir_nr
        return canonical
    if (
        previous_room_id != dane.sala_id
        and (
            _sala_ma_wersjonowany_plan(db, previous_room_id)
            or _sala_ma_wersjonowany_plan(db, dane.sala_id)
        )
    ):
        _wymagaj_szkicu_planu(
            "Przeniesienie stołu między salami wymaga zmiany w szkicu planu."
        )
    if (
        previous_active != dane.aktywny
        and _sala_ma_wersjonowany_plan(db, previous_room_id)
    ):
        _wymagaj_szkicu_planu(
            "Włączanie i wyłączanie stołu wymaga zmiany w szkicu planu."
        )
    if s.aktywny and not dane.aktywny:
        dzis_lokalny = (_teraz_lokalnie() or datetime.now()).date()
        przyszle = db.query(models.Termin).filter(
            models.Termin.rodzaj == "stolik",
            models.Termin.data >= dzis_lokalny,
            models.Termin.status.in_(REZ_AKTYWNE),
        ).all()
        if any(sid in _stoly_terminu(termin) for termin in przyszle):
            raise HTTPException(
                409,
                "Stolik ma aktywne lub przyszłe rezerwacje. Najpierw przepnij je na inne stoły.",
            )
        now = utcnow_naive()
        if db.query(models.RezerwacjaStolikClaim.id).filter(
            models.RezerwacjaStolikClaim.stolik_id == sid,
            or_(
                models.RezerwacjaStolikClaim.expires_at.is_(None),
                models.RezerwacjaStolikClaim.expires_at > now,
            ),
        ).first():
            raise HTTPException(
                409,
                "Stolik ma aktywne zajęcie lub hold. Najpierw zwolnij go w rezerwacjach.",
            )
    for k, v in incoming.items():
        setattr(s, k, v)
    db.commit(); db.refresh(s)
    return schemas.StolikOut.model_validate(s).model_dump()


@app.delete("/api/stoliki/{sid}", status_code=204, dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def usun_stolik(sid: int, db: Session = Depends(get_db)):
    # Serializuj usunięcie z producentami claimów na tym samym wierszu stołu.
    reservation_service.begin_floor_plan_write(db)
    locked = reservation_service.lock_tables(db, [sid])
    s = locked[0] if locked else None
    if s:
        # Id stolika jest częścią historii rezerwacji i JSON-owych składów kombinacji. Ciche
        # usunięcie zostawiałoby osierocone identyfikatory, których FK nie potrafi ochronić.
        # Używany stolik można bezpiecznie wyłączyć (`aktywny=false`), ale nie usuwać fizycznie.
        if db.query(models.Termin.id).filter(models.Termin.stolik_id == sid).first():
            raise HTTPException(409, "Stolik jest przypisany do rezerwacji. Zamiast usuwać, oznacz go jako nieaktywny.")
        dodatkowe = db.query(models.Termin.stoliki_dodatkowe).filter(
            models.Termin.stoliki_dodatkowe.isnot(None)).all()
        if any(sid in _ids_stolikow(wartosc) for (wartosc,) in dodatkowe):
            raise HTTPException(409, "Stolik należy do przydziału wielostolikowego. Najpierw zmień przypisane rezerwacje.")
        kombinacje = db.query(models.KombinacjaStolow.stoliki).all()
        if any(sid in _ids_stolikow(wartosc) for (wartosc,) in kombinacje):
            raise HTTPException(409, "Stolik należy do kombinacji. Najpierw usuń lub zmień tę kombinację.")
        now = utcnow_naive()
        claim_dates = {
            value for (value,) in db.query(
                models.RezerwacjaStolikClaim.data,
            ).filter(
                models.RezerwacjaStolikClaim.stolik_id == sid,
            ).distinct().all()
        }
        if claim_dates:
            reservation_service.cleanup_expired_holds(
                db, now, dates=claim_dates,
            )
            db.flush()
        if db.query(models.RezerwacjaStolikClaim.id).filter_by(
            stolik_id=sid,
        ).first():
            raise HTTPException(
                409,
                "Stolik ma aktywne zajęcie lub hold. Najpierw zwolnij go w rezerwacjach.",
            )
        if db.query(models.PozycjaStolikaPlanu.id).filter_by(stolik_id=sid).first():
            raise HTTPException(
                409,
                "Stolik należy do historii planu sali. Zamiast usuwać, wyłącz go w nowym szkicu.",
            )
        db.delete(s); db.commit()


# ── Kombinacje stołów (predefiniowane łączenie pod większe grupy) ─────────────
def _waliduj_sklad_kombinacji(db, stoliki):
    """Deduplikuje id stołów, wymaga ≥2 różnych i sprawdza, że istnieją. Zwraca listę id."""
    ids = []
    for x in (stoliki or []):
        xi = int(x)
        if xi not in ids:
            ids.append(xi)
    if len(ids) < 2:
        raise HTTPException(400, "Kombinacja musi łączyć co najmniej 2 różne stoły.")
    for sid in ids:
        if not db.get(models.Stolik, sid):
            raise HTTPException(400, f"Nieznany stolik (id={sid}).")
    _wymagaj_legacy_stolikow_poza_planem(db, ids)
    _sala_dla_stolikow(db, ids)
    nieaktywne = [sid for sid in ids if not db.get(models.Stolik, sid).aktywny]
    if nieaktywne:
        raise HTTPException(
            409,
            detail={
                "code": "TABLE_NOT_OPERATIONAL",
                "message": "Kombinacja może zawierać wyłącznie aktywne stoliki.",
                "table_ids": nieaktywne,
            },
        )
    return ids


def _suma_pojemnosci(db, ids) -> int:
    return sum((db.get(models.Stolik, i).pojemnosc or 0) for i in ids)


def _pojemnosc_kombinacji(db, ids, minimum, maksimum) -> int:
    efektywne_maksimum = maksimum or _suma_pojemnosci(db, ids)
    if efektywne_maksimum < 1:
        raise HTTPException(400, "Kombinacja musi mieć dodatnią pojemność.")
    if minimum is not None and minimum > efektywne_maksimum:
        raise HTTPException(400, "Minimalna liczba osób nie może przekraczać maksymalnej.")
    return efektywne_maksimum


def _zablokuj_legacy_kombinacje(db: Session, kid: int):
    """Lock and refresh one legacy combination before deriving table locks."""
    query = db.query(models.KombinacjaStolow).filter_by(id=kid)
    if db.get_bind().dialect.name == "postgresql":
        query = query.with_for_update()
    return query.populate_existing().first()


@app.get("/api/kombinacje", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def get_kombinacje(db: Session = Depends(get_db)):
    return {"kombinacje": _kombinacje_do_odczytu(db)}


@app.post("/api/kombinacje", status_code=201, dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def dodaj_kombinacje(dane: schemas.KombinacjaStolowIn, db: Session = Depends(get_db)):
    requested_ids = _ids_stolikow(dane.stoliki)
    if len(requested_ids) < 2:
        # Zachowaj dotychczasowy, precyzyjny błąd walidacji przed wejściem
        # w sekcję blokującą zapis planu.
        _waliduj_sklad_kombinacji(db, dane.stoliki)
    reservation_service.begin_floor_plan_write(db)
    locked_tables = reservation_service.lock_tables(db, requested_ids)
    if len(locked_tables) != len(requested_ids):
        raise HTTPException(400, "Nieznany stolik w kombinacji.")
    _blokuj_topologie_aktywnej_oferty(db, requested_ids)
    # Powtórna walidacja pod blokadami chroni przed zmianą aktywności, sali
    # lub wersjonowania stolików pomiędzy odczytem żądania a zapisem definicji.
    ids = _waliduj_sklad_kombinacji(db, dane.stoliki)
    pojemnosc_max = _pojemnosc_kombinacji(
        db, ids, dane.pojemnosc_min, dane.pojemnosc_max)
    k = models.KombinacjaStolow(
        nazwa=dane.nazwa.strip(), stoliki=ids, pojemnosc_min=dane.pojemnosc_min,
        pojemnosc_max=pojemnosc_max,
        aktywna=dane.aktywna, priorytet=dane.priorytet)
    db.add(k); db.commit(); db.refresh(k)
    return schemas.KombinacjaStolowOut.model_validate(k).model_dump()


@app.put("/api/kombinacje/{kid}", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def edytuj_kombinacje(kid: int, dane: schemas.KombinacjaStolowIn, db: Session = Depends(get_db)):
    reservation_service.begin_floor_plan_write(db)
    k = _zablokuj_legacy_kombinacje(db, kid)
    if not k:
        raise HTTPException(404, "Brak kombinacji.")
    touched_ids = _ids_stolikow(k.stoliki) | _ids_stolikow(dane.stoliki)
    reservation_service.lock_tables(db, touched_ids)
    _blokuj_topologie_aktywnej_oferty(db, touched_ids)
    _wymagaj_legacy_stolikow_poza_planem(db, k.stoliki)
    ids = _waliduj_sklad_kombinacji(db, dane.stoliki)
    pojemnosc_max = _pojemnosc_kombinacji(
        db, ids, dane.pojemnosc_min, dane.pojemnosc_max)
    k.nazwa = dane.nazwa.strip(); k.stoliki = ids; k.pojemnosc_min = dane.pojemnosc_min
    k.pojemnosc_max = pojemnosc_max
    k.aktywna = dane.aktywna; k.priorytet = dane.priorytet
    db.commit(); db.refresh(k)
    return schemas.KombinacjaStolowOut.model_validate(k).model_dump()


@app.delete("/api/kombinacje/{kid}", status_code=204, dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def usun_kombinacje(kid: int, db: Session = Depends(get_db)):
    reservation_service.begin_floor_plan_write(db)
    k = _zablokuj_legacy_kombinacje(db, kid)
    if k:
        touched_ids = _ids_stolikow(k.stoliki)
        reservation_service.lock_tables(db, touched_ids)
        _blokuj_topologie_aktywnej_oferty(db, touched_ids)
        _wymagaj_legacy_stolikow_poza_planem(db, k.stoliki)
        db.delete(k); db.commit()


# ── Wyjątki kalendarza (blackouty / godziny specjalne per dzień) ──────────────
_WYJ_TYPY = ("blackout", "godziny_specjalne")


@app.get("/api/wyjatki-kalendarza", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def get_wyjatki_kalendarza(od: date = Query(None), do: date = Query(None), db: Session = Depends(get_db)):
    q = db.query(models.WyjatekKalendarza)
    if od:
        q = q.filter(models.WyjatekKalendarza.data >= od)
    if do:
        q = q.filter(models.WyjatekKalendarza.data <= do)
    rows = q.order_by(models.WyjatekKalendarza.data).all()
    return {"wyjatki": [schemas.WyjatekKalendarzaOut.model_validate(w).model_dump() for w in rows]}


@app.post("/api/wyjatki-kalendarza", status_code=201, dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def dodaj_wyjatek_kalendarza(dane: schemas.WyjatekKalendarzaIn, db: Session = Depends(get_db)):
    typ = (dane.typ or "").strip()
    if typ not in _WYJ_TYPY:
        raise HTTPException(400, "typ musi być 'blackout' lub 'godziny_specjalne'.")
    if typ == "godziny_specjalne":
        if not (dane.godz_od and dane.godz_do):
            raise HTTPException(400, "Godziny specjalne wymagają godz_od i godz_do.")
        if dane.godz_do <= dane.godz_od:
            raise HTTPException(400, "Godziny specjalne muszą kończyć się po rozpoczęciu.")
        if dane.ostatni_zasiadek is not None and not (
            dane.godz_od <= dane.ostatni_zasiadek <= dane.godz_do
        ):
            raise HTTPException(400, "Ostatnie przyjęcie musi mieścić się w godzinach specjalnych.")
    w = models.WyjatekKalendarza(**{**dane.model_dump(), "typ": typ})
    db.add(w); db.commit(); db.refresh(w)
    return schemas.WyjatekKalendarzaOut.model_validate(w).model_dump()


@app.put("/api/wyjatki-kalendarza/{wid}", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def edytuj_wyjatek_kalendarza(
    wid: int,
    dane: schemas.WyjatekKalendarzaIn,
    db: Session = Depends(get_db),
):
    w = db.get(models.WyjatekKalendarza, wid)
    if w is None:
        raise HTTPException(404, "Brak wyjątku kalendarza.")
    typ = (dane.typ or "").strip()
    if typ not in _WYJ_TYPY:
        raise HTTPException(400, "typ musi być 'blackout' lub 'godziny_specjalne'.")
    if typ == "godziny_specjalne":
        if not (dane.godz_od and dane.godz_do):
            raise HTTPException(400, "Godziny specjalne wymagają godz_od i godz_do.")
        if dane.godz_do <= dane.godz_od:
            raise HTTPException(400, "Godziny specjalne muszą kończyć się po rozpoczęciu.")
        if dane.ostatni_zasiadek is not None and not (
            dane.godz_od <= dane.ostatni_zasiadek <= dane.godz_do
        ):
            raise HTTPException(400, "Ostatnie przyjęcie musi mieścić się w godzinach specjalnych.")
    for key, value in dane.model_dump().items():
        setattr(w, key, value)
    w.typ = typ
    db.commit(); db.refresh(w)
    return schemas.WyjatekKalendarzaOut.model_validate(w).model_dump()


@app.delete("/api/wyjatki-kalendarza/{wid}", status_code=204, dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def usun_wyjatek_kalendarza(wid: int, db: Session = Depends(get_db)):
    w = db.get(models.WyjatekKalendarza, wid)
    if w:
        db.delete(w); db.commit()


# ── Graf sąsiedztwa stołów (auto-kombinacje w silniku) ────────────────────────
@app.get("/api/sasiedztwo", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def get_sasiedztwo(db: Session = Depends(get_db)):
    return {"krawedzie": _sasiedztwo_do_odczytu(db)}


@app.post("/api/sasiedztwo", status_code=201, dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def dodaj_sasiedztwo(dane: schemas.SasiedztwoStolowIn, db: Session = Depends(get_db)):
    a, b = sorted((int(dane.stolik_a), int(dane.stolik_b)))     # normalizacja a<b (graf nieskierowany)
    if a == b:
        raise HTTPException(400, "Sąsiedztwo łączy dwa RÓŻNE stoły.")
    if not db.get(models.Stolik, a) or not db.get(models.Stolik, b):
        raise HTTPException(400, "Nieznany stolik.")
    _wymagaj_legacy_stolikow_poza_planem(db, (a, b))
    _sala_dla_stolikow(db, (a, b))
    if db.query(models.SasiedztwoStolow).filter_by(stolik_a=a, stolik_b=b).first():
        raise HTTPException(409, "Ta para stołów już sąsiaduje.")
    k = models.SasiedztwoStolow(stolik_a=a, stolik_b=b)
    db.add(k); db.commit(); db.refresh(k)
    return {"id": k.id, "stolik_a": k.stolik_a, "stolik_b": k.stolik_b}


@app.delete("/api/sasiedztwo/{kid}", status_code=204, dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def usun_sasiedztwo(kid: int, db: Session = Depends(get_db)):
    reservation_service.begin_floor_plan_write(db)
    k = db.get(models.SasiedztwoStolow, kid)
    if k:
        touched_ids = (k.stolik_a, k.stolik_b)
        reservation_service.lock_tables(db, touched_ids)
        _blokuj_topologie_aktywnej_oferty(db, touched_ids)
        _wymagaj_legacy_stolikow_poza_planem(db, touched_ids)
        db.delete(k); db.commit()


# ── Silnik sadzania (best-fit + kombinacje): SUGESTIA + AUTO ──────────────────
@app.get("/api/host/sugestia-stolika", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def host_sugestia_stolika(
    data: date = Query(...),
    godz_od: time = Query(...),
    osoby: int = 2,
    strefa: Optional[str] = None,
    waitlist_id: Optional[int] = Query(None, gt=0),
    db: Session = Depends(get_db),
):
    """Top-3 propozycje stołu/kombinacji dla grupy — host akceptuje jedną (tryb SUGESTIA)."""
    osoby = max(1, osoby)
    target_channel = "wewnetrzna"
    if waitlist_id is not None:
        waitlist = db.get(models.ListaOczekujacych, waitlist_id)
        if waitlist is None or waitlist.status not in reservation_service.WAITLIST_ACTIVE_STATUSES:
            raise HTTPException(404, "Brak aktywnego wpisu waitlisty.")
        if waitlist.data != data or max(1, int(waitlist.liczba_osob or 1)) != osoby:
            raise reservation_service.ReservationError(
                409,
                "WAITLIST_PREVIEW_CONTEXT_MISMATCH",
                "Podgląd musi zachować dzień i liczbę osób z wpisu waitlisty.",
                rule="waitlist_offer",
            )
        target_channel = "online" if waitlist.kanal == "online" else "wewnetrzna"
    result = _ocen_przydzial_rezerwacji(
        db,
        data=data,
        godz_od=godz_od,
        osoby=osoby,
        kanal=target_channel,
        intent="quote",
        preferowana_strefa=strefa,
        alternative_limit=2,
    )
    payload = result.to_dict(expose_exact=True)
    return {
        "data": str(data),
        "godz_od": _hm(godz_od),
        "godz_do": payload.get("visit_end"),
        "osoby": osoby,
        "kandydaci": [
            _kandydat_legacy_z_alokacji(candidate)
            for candidate in result.candidates
        ],
        **payload,
    }


@app.post("/api/rezerwacje-stolik/{rid}/auto-przydziel", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def auto_przydziel_stolik(
    rid: int,
    request: Request,
    dane: Optional[schemas.AutoPrzydzialIn] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Silnik sam dobiera najlepszy stół/kombinację dla rezerwacji (tryb AUTO). 409 gdy brak miejsca."""
    dane = dane or schemas.AutoPrzydzialIn()
    override_context = _jawne_nadpisanie_limitow(
        user, dane.przekrocz_limity, dane.nadpisanie_limitow,
    )
    t, guards = _zablokuj_termin(db, rid)
    if not t or t.rodzaj != "stolik":
        db.rollback()
        raise HTTPException(404, "Brak rezerwacji.")
    if not t.godz_od:
        raise HTTPException(400, "Rezerwacja bez godziny — nie można dobrać stołu.")
    before = reservation_audit.reservation_snapshot(t)
    osoby = max(1, t.liczba_osob or 1)
    result = _ocen_przydzial_rezerwacji(
        db,
        data=t.data,
        godz_od=t.godz_od,
        godz_do=t.godz_do,
        osoby=osoby,
        kanal=t.kanal,
        pomin_id=rid,
        intent="assign",
    )
    override_active = bool(
        override_context and result.evaluation.decision == "override_required"
    )
    _wymagaj_reautoryzacji_nadpisania(
        db, request, user, override_active,
    )
    selected = _wymagaj_dozwolonego_przydzialu(
        result, override=override_active,
    )
    wybrany = _kandydat_legacy_z_alokacji(selected)
    godz_do = result.evaluation.godz_do
    t.stolik_id = wybrany["stoliki"][0]
    t.stoliki_dodatkowe = (wybrany["stoliki"][1:] or None)
    t.godz_do = godz_do
    t.auto_przydzielony = True
    _ustaw_proweniencje_przydzialu(t, wybrany)
    try:
        db.flush()
        _zastap_ledger_terminu(
            db,
            t,
            candidates=[{"stoliki": wybrany["stoliki"]}],
            alternatives=result.to_dict(expose_exact=True).get("alternatives") or (),
            enforce_pacing=True,
            evaluation=result.evaluation,
            override=override_active,
            intent="assign",
        )
        reservation_audit.add_reservation_audit(
            db, termin=t, action="assign", actor=user, before=before, after=t,
        )
        override_details = (
            _szczegoly_nadpisania_r3(
                result.evaluation,
                legacy=bool(override_context and override_context["legacy"]),
            )
            if override_active else None
        )
        if override_details:
            reservation_audit.add_reservation_audit(
                db,
                termin=t,
                action="override",
                actor=user,
                reason=_powod_audytu_nadpisania(result.evaluation),
                before=before,
                after=t,
                override_details=override_details,
                override_reason_code=override_context["reason_code"],
                override_note=override_context["note"],
            )
        _commit_zapis_rezerwacji(db, guards)
    except IntegrityError as exc:
        db.rollback()
        raise reservation_service.translate_integrity_error(exc) from exc
    db.refresh(t)
    payload = result.to_dict(expose_exact=True)
    return {
        "rezerwacja": _rezerwacja_out(t, user),
        "przydzial": wybrany,
        "allocation": {
            **(payload.get("allocation") or {}),
            "state": "assigned",
        },
        "alternatives": payload.get("alternatives") or [],
        "reasons": payload.get("reasons") or [],
    }


# ── Widok hosta: kolejka dnia + fazy operacyjne + przydział stołu ─────────────
def _crm_hash(t) -> str:
    """Wewnętrzny hash tożsamości CRM: kontakt albo dokładna rezerwacja."""
    return _identity_hash_crm(t)


_CRM_PROFILE_FILL_FIELDS = (
    "nazwisko",
    "preferowana_strefa",
    "okazja_typ",
    "okazja_data",
)
_CRM_PROFILE_MERGE_TEXT_FIELDS = ("alergie", "dieta", "notatka")
_CRM_PROFILE_TEXT_SEPARATOR = "\n---\n"


def _scal_tekst_profilu(primary, secondary):
    """Zachowuje obie różne wartości w stabilnej kolejności bez duplikatów."""
    values = []
    seen = set()
    for raw in (primary, secondary):
        if not raw:
            continue
        for part in str(raw).split(_CRM_PROFILE_TEXT_SEPARATOR):
            value = part.strip()
            normalized = value.casefold()
            if value and normalized not in seen:
                seen.add(normalized)
                values.append(value)
    return _CRM_PROFILE_TEXT_SEPARATOR.join(values) or None


def _scal_profile_gosci(target, source) -> None:
    """Scala starszą tożsamość do bieżącego profilu bez utraty danych operacyjnych.

    Istniejący profil docelowy ma pierwszeństwo dla pól pojedynczych, a teksty,
    tagi i flagi łączymy zachowawczo.
    """
    for field in _CRM_PROFILE_FILL_FIELDS:
        if not getattr(target, field, None) and getattr(source, field, None):
            setattr(target, field, getattr(source, field))
    for field in _CRM_PROFILE_MERGE_TEXT_FIELDS:
        setattr(
            target,
            field,
            _scal_tekst_profilu(getattr(target, field, None), getattr(source, field, None)),
        )

    tags = list(target.tagi or [])
    for tag in source.tagi or []:
        if tag not in tags:
            tags.append(tag)
    target.tagi = tags or None
    target.vip = bool(target.vip or source.vip)
    # Dwa scalane profile nie mają wersjonowanej proweniencji zgody. Kolizja
    # zawsze wymaga ponownego, jawnego opt-in; nie wolno wskrzeszać wycofanej zgody.
    target.marketing_zgoda = False
    if source.utworzono_at and (
        target.utworzono_at is None or source.utworzono_at < target.utworzono_at
    ):
        target.utworzono_at = source.utworzono_at
    target.zaktualizowano_at = utcnow_naive()


def _migruj_osierocony_profil_po_zmianie_tozsamosci(
    db, termin, previous_key: str,
) -> None:
    """Przenosi profil tylko wtedy, gdy stary klucz nie opisuje innej wizyty.

    Dzięki temu korekta lub usunięcie kontaktu nie zostawia niedostępnego profilu
    z PII, ale zmiana jednej z wielu wizyt wspólnego gościa nie kradnie profilu
    pozostałej historii.
    """
    current_key = _identity_key_crm(termin)
    if not previous_key or not current_key or current_key == previous_key:
        return

    previous_hash = _hash_klucz_crm(previous_key)
    current_hash = _hash_klucz_crm(current_key)
    source = db.query(models.ProfilGoscia).filter_by(klucz_hash=previous_hash).first()
    if source is None:
        return

    other_rows = db.query(models.Termin).filter(
        models.Termin.rodzaj.in_(("stolik", "sala")),
        models.Termin.id != termin.id,
    ).all()
    if any(_identity_key_crm(other) == previous_key for other in other_rows):
        return

    target = db.query(models.ProfilGoscia).filter_by(klucz_hash=current_hash).first()
    if target is None:
        source.klucz_hash = current_hash
        source.zaktualizowano_at = utcnow_naive()
        return

    # R7.3: collision with an existing identity is not proof that both profiles
    # describe the same person (shared family/company contacts are common).
    # Preserve both encrypted profiles and let the explicit, reversible CRM
    # governance flow present a candidate to an authorised operator.
    return


def _usun_profil_fallbacku_rezerwacji(db, reservation_id: int) -> None:
    """Usuwa profil zależny od ID przed zwolnieniem ID rezerwacji (SQLite reuse)."""
    fallback_hash = _reservation_fallback_hash(reservation_id)
    if not fallback_hash:
        return
    profile = db.query(models.ProfilGoscia).filter_by(klucz_hash=fallback_hash).first()
    if profile is not None:
        db.delete(profile)


def _profile_dla_terminow(db, terminy):
    """Mapa {klucz_hash: ProfilGoscia} dla listy terminów — JEDNO zapytanie po hashach (bez PII w SQL)."""
    hashe = {_crm_hash(t) for t in terminy}
    if not hashe:
        return {}
    return {p.klucz_hash: p for p in db.query(models.ProfilGoscia)
            .filter(models.ProfilGoscia.klucz_hash.in_(hashe)).all()}


def _flagi_profilu(p, user=None):
    """Skrót flag gościa (VIP/alergie/okazja/tagi) dla widoku hosta. Endpoint admin-only → PII (alergie)
    dozwolone, ale zwracane oszczędnie. None gdy brak profilu."""
    if not p:
        return None
    wrazliwe = _ma_dostep_rezerwacji(user, "rezerwacje.dane_wrazliwe")
    kontakt = _ma_dostep_rezerwacji(user, "rezerwacje.dane_kontaktowe")
    return {
        "vip": bool(p.vip),
        "ma_alergie": bool(p.alergie) if wrazliwe else None,
        "alergie": (p.alergie or None) if wrazliwe else None,
        "okazja_typ": p.okazja_typ if kontakt else None,
        "okazja_data": p.okazja_data if kontakt else None,
        "tagi": (p.tagi or []) if wrazliwe else [],
        # Nie ujawniamy nawet faktu istnienia danych zdrowotnych bez osobnego prawa.
        "dane_wrazliwe_ukryte": False,
    }


def _host_out(t: models.Termin, teraz=None, profil=None, user=None) -> dict:
    wpis = _rezerwacja_out(t, user)
    wpis["faza_hosta"] = t.faza_hosta
    if t.faza_hosta in HOST_NA_SALI and t.host_seated_at and teraz:
        wpis["minuty_od_posadzenia"] = max(0, int((teraz - t.host_seated_at).total_seconds() // 60))
    wpis["gosc"] = _flagi_profilu(profil, user)
    return wpis


def _host_waitlist_out(w, *, user, communication_summary=None):
    """Minimal operational projection; never expands contact or note PII."""
    pokaz_kontakt = _ma_dostep_rezerwacji(user, "rezerwacje.dane_kontaktowe")
    summary = dict(communication_summary) if communication_summary else None
    if summary is not None and not pokaz_kontakt:
        summary["channel"] = None
    return {
        "id": w.id,
        "nazwisko": w.nazwisko if pokaz_kontakt else "Gość",
        "godz_od": _hm(w.godz_od),
        "liczba_osob": w.liczba_osob,
        "status": w.status,
        "utworzono_at": w.utworzono_at.isoformat() if w.utworzono_at else None,
        "priorytet": int(w.priorytet or 0),
        "offer_version": int(w.offer_version or 0),
        "offer_auto_przydzielony": w.offer_auto_przydzielony,
        "offer_override_authorized": w.offer_override_authorized,
        "can_queue_communication": bool(
            pokaz_kontakt
            and reservation_communication.available_delivery_channels(w)
        ),
        "zaoferowano_at": w.zaoferowano_at.isoformat() if w.zaoferowano_at else None,
        "oferta_wygasa_at": (
            w.oferta_wygasa_at.isoformat() if w.oferta_wygasa_at else None
        ),
        "hold_stolik_id": w.hold_stolik_id,
        "hold_stoliki_dodatkowe": list(w.hold_stoliki_dodatkowe or []),
        "hold_godz_od": _hm(w.hold_godz_od),
        "hold_godz_do": _hm(w.hold_godz_do),
        "hold_do": w.hold_do.isoformat() if w.hold_do else None,
        "communication_summary": summary,
    }


def _host_waitlist_communication_out(w, *, communication_summary):
    """Minimalny, kontaktowy inbox błędów po zakończeniu waitlisty."""
    return {
        "id": w.id,
        "nazwisko": w.nazwisko,
        "status": w.status,
        "communication_summary": dict(communication_summary),
    }


def _host_kolejka_payload(
    *,
    dzien: date,
    db: Session,
    user: models.User,
    teraz: Optional[datetime] = None,
) -> dict:
    """Buduje jeden PII-safe payload kolejki dla tras hosta."""
    teraz = teraz or utcnow_naive()
    rez = db.query(models.Termin).filter(
        models.Termin.rodzaj == "stolik", models.Termin.data == dzien).order_by(models.Termin.godz_od).all()
    profile = _profile_dla_terminow(db, rez)           # {hash: ProfilGoscia} — flagi VIP/alergie/okazja
    nadchodzace, na_sali, zakonczone = [], [], []
    for t in rez:
        if t.status == "odwolana":
            continue                                   # anulowane nie zaśmiecają widoku hosta
        wpis = _host_out(t, teraz, profile.get(_crm_hash(t)), user)
        if t.faza_hosta == "wyszedl" or t.status in ("odbyla", "no_show"):
            zakonczone.append(wpis)
        elif t.faza_hosta in HOST_NA_SALI:
            na_sali.append(wpis)
        else:
            nadchodzace.append(wpis)
    waitlist_rows = db.query(models.ListaOczekujacych).filter(
        models.ListaOczekujacych.data == dzien,
        models.ListaOczekujacych.status.in_(
            reservation_service.WAITLIST_ACTIVE_STATUSES
        ),
    ).all()
    waitlist_rows = [
        row for row in waitlist_rows
        if row.status == "oczekuje"
        or _czy_kompletny_hold_oferty(db, row, teraz=teraz)
    ]
    waitlist_rows.sort(key=lambda row: (
        0 if row.status == "zaoferowano" else 1,
        row.oferta_wygasa_at or datetime.max,
        -int(row.priorytet or 0),
        row.utworzono_at,
        row.id,
    ))
    summaries = reservation_communication.summaries_for_waitlists(
        db, (row.id for row in waitlist_rows),
    )
    waitlista = [
        _host_waitlist_out(
            row,
            user=user,
            communication_summary=summaries.get(row.id),
        )
        for row in waitlist_rows
    ]
    komunikacja_waitlist = []
    if _ma_dostep_rezerwacji(user, "rezerwacje.dane_kontaktowe"):
        terminal_rows = db.query(models.ListaOczekujacych).filter(
            models.ListaOczekujacych.data == dzien,
            models.ListaOczekujacych.status.in_((
                "zaakceptowano", "wygasla", "anulowano",
            )),
        ).all()
        terminal_summaries = reservation_communication.summaries_for_waitlists(
            db, (row.id for row in terminal_rows),
        )
        terminal_with_attention = [
            (row, terminal_summaries.get(row.id))
            for row in terminal_rows
            if (terminal_summaries.get(row.id) or {}).get("attention_required")
        ]
        terminal_with_attention.sort(key=lambda item: (
            (item[1] or {}).get("last_event_at") or "",
            item[0].id,
        ))
        komunikacja_waitlist = [
            _host_waitlist_communication_out(
                row, communication_summary=summary,
            )
            for row, summary in terminal_with_attention
        ]
    return {
        "data": str(dzien), "nadchodzace": nadchodzace, "na_sali": na_sali,
        "zakonczone": zakonczone, "waitlista": waitlista,
        "komunikacja_waitlist": komunikacja_waitlist,
        "podsumowanie": {"nadchodzace": len(nadchodzace), "na_sali": len(na_sali),
                         "zakonczone": len(zakonczone), "waitlista": len(waitlista),
                         "coverow_na_sali": sum((w.get("liczba_osob") or 0) for w in na_sali)},
    }


@app.get("/api/host/kolejka", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def host_kolejka(
    data: date = Query(None),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Widok hosta na dzień: nadchodzący / na sali (z timerem obrotu) / zakończeni + waitlista."""
    return _host_kolejka_payload(
        dzien=data or date.today(),
        db=db,
        user=user,
    )


def _host_os_czasu_payload(
    *,
    dzien: date,
    db: Session,
    user: models.User,
    teraz: Optional[datetime] = None,
) -> dict:
    """Buduje PII-safe oś czasu z kombinacjami rozbitymi na stoły składowe."""
    rez = db.query(models.Termin).filter(
        models.Termin.rodzaj == "stolik", models.Termin.data == dzien,
        models.Termin.status.in_(REZ_AKTYWNE), models.Termin.godz_od.isnot(None)).order_by(models.Termin.godz_od).all()
    teraz = (
        teraz.astimezone(timezone.utc).replace(tzinfo=None)
        if teraz is not None and teraz.tzinfo is not None
        else (teraz.replace(tzinfo=None) if teraz is not None else utcnow_naive())
    )
    oferty = db.query(models.ListaOczekujacych).filter(
        models.ListaOczekujacych.data == dzien,
        models.ListaOczekujacych.status == "zaoferowano",
        models.ListaOczekujacych.hold_do > teraz,
        models.ListaOczekujacych.hold_stolik_id.isnot(None),
    ).all()
    oferty = [
        oferta for oferta in oferty
        if _czy_kompletny_hold_oferty(db, oferta, teraz=teraz)
    ]
    przypisane_ids = set().union(*(_stoly_terminu(t) for t in rez)) if rez else set()
    for oferta in oferty:
        przypisane_ids.update(_stoliki_zadania(
            oferta.hold_stolik_id,
            oferta.hold_stoliki_dodatkowe,
        ))
    runtime_stoly = {stolik["id"]: stolik for stolik in _stoly_do_seating(db)}
    # Wyłączenie sali zatrzymuje nowe przydziały, ale nie może osierocić paska już
    # przypisanej rezerwacji. Do osi dokładamy więc wyłącznie historycznie użyte
    # stoły z pełnego published/legacy adaptera odczytowego.
    for stolik in _stoliki_do_odczytu(db):
        if stolik["id"] in przypisane_ids and stolik["id"] not in runtime_stoly:
            runtime_stoly[stolik["id"]] = stolik
    stoly = [
        {
            "id": stolik["id"],
            "nazwa": stolik["nazwa"],
            "sekcja": stolik.get("sekcja") or stolik.get("strefa"),
            "strefa": stolik.get("strefa"),
        }
        for stolik in sorted(
            runtime_stoly.values(),
            key=lambda row: (row.get("kolejnosc") or 0, row["id"]),
        )
    ]
    godziny = [_hm(g) for g, _ in _sloty_dnia(db, dzien)]
    zajetosci = []
    for t in rez:
        godz_do = t.godz_do or _dodaj_minuty(t.godz_od, _dlugosc_dla(db, dzien, t.godz_od, t.liczba_osob))
        for sid in sorted(_stoly_terminu(t)):
            zajetosci.append({"stolik_id": sid, "godz_od": _hm(t.godz_od), "godz_do": _hm(godz_do),
                              "rezerwacja_id": t.id,
                              "nazwisko": t.nazwisko if _ma_dostep_rezerwacji(
                                  user, "rezerwacje.dane_kontaktowe") else "Gość",
                              "liczba_osob": t.liczba_osob, "faza_hosta": t.faza_hosta,
                              "typ": "rezerwacja"})
    pokaz_kontakt = _ma_dostep_rezerwacji(user, "rezerwacje.dane_kontaktowe")
    for oferta in oferty:
        for sid in _stoliki_zadania(
            oferta.hold_stolik_id,
            oferta.hold_stoliki_dodatkowe,
        ):
            zajetosci.append({
                "stolik_id": sid,
                "godz_od": _hm(oferta.hold_godz_od),
                "godz_do": _hm(oferta.hold_godz_do),
                "rezerwacja_id": None,
                "waitlist_id": oferta.id,
                "nazwisko": oferta.nazwisko if pokaz_kontakt else "Gość",
                "liczba_osob": oferta.liczba_osob,
                "faza_hosta": None,
                "typ": "oferta",
                "status": oferta.status,
                "offer_version": int(oferta.offer_version or 0),
                "offer_auto_przydzielony": oferta.offer_auto_przydzielony,
                "offer_override_authorized": oferta.offer_override_authorized,
                "oferta_wygasa_at": (
                    oferta.oferta_wygasa_at.isoformat()
                    if oferta.oferta_wygasa_at else None
                ),
                "hold_do": oferta.hold_do.isoformat() if oferta.hold_do else None,
            })
    return {"data": str(dzien), "stoly": stoly, "godziny": godziny, "zajetosci": zajetosci}


@app.get("/api/host/os-czasu", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def host_os_czasu(
    data: date = Query(None),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Oś czasu hosta: stoły × godziny + paski zajętości."""
    return _host_os_czasu_payload(
        dzien=data or date.today(),
        db=db,
        user=user,
    )


HOST_SNAPSHOT_SCHEMA_VERSION = 1


def _host_snapshot_payload(
    *,
    dzien: date,
    db: Session,
    user: models.User,
) -> dict:
    """Buduje wszystkie projekcje w obrębie jednego snapshotu transakcji."""
    generated_at_utc = datetime.now(timezone.utc)
    generated_at = generated_at_utc.isoformat(timespec="milliseconds").replace(
        "+00:00", "Z",
    )
    return {
        "version": generated_at,
        "schema_version": HOST_SNAPSHOT_SCHEMA_VERSION,
        "data": str(dzien),
        "generated_at": generated_at,
        "kolejka": _host_kolejka_payload(
            dzien=dzien,
            db=db,
            user=user,
            teraz=generated_at_utc.replace(tzinfo=None),
        ),
        "os_czasu": _host_os_czasu_payload(
            dzien=dzien,
            db=db,
            user=user,
            teraz=generated_at_utc,
        ),
        "plan_sali": operational_plan_payload(
            dzien=dzien,
            sala_id=None,
            db=db,
            user=user,
            teraz=generated_at_utc,
        ),
    }


@app.get("/api/host/snapshot", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def host_snapshot(
    data: date = Query(...),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Spójny snapshot live: kolejka, oś czasu i published-only plan dnia."""
    bind = db.get_bind()
    if bind.dialect.name != "postgresql":
        return _host_snapshot_payload(dzien=data, db=db, user=user)

    # Auth wykonał już SELECT w sesji requestu, więc nie można bezpiecznie podnieść
    # jej izolacji. Osobna transakcja REPEATABLE READ gwarantuje, że równoległy zapis
    # nie pojawi się tylko w jednej z trzech projekcji hosta.
    snapshot_engine = getattr(bind, "engine", bind)
    with snapshot_engine.connect().execution_options(
        isolation_level="REPEATABLE READ",
    ) as connection:
        with Session(bind=connection, autoflush=False) as snapshot_db:
            with snapshot_db.begin():
                snapshot_user = snapshot_db.get(models.User, user.id)
                if snapshot_user is None or not snapshot_user.aktywny:
                    raise HTTPException(status_code=401, detail="Konto jest nieaktywne.")
                return _host_snapshot_payload(
                    dzien=data,
                    db=snapshot_db,
                    user=snapshot_user,
                )


def _host_exact_contract(dane) -> bool:
    return dane.stoliki is not None and dane.oczekiwane_stoliki is not None


def _host_exact_idempotency_payload(rid, operation, dane, target_ids, expected_ids):
    confirmation = (
        dane.nadpisanie_limitow.model_dump(mode="json")
        if dane.nadpisanie_limitow is not None else None
    )
    return {
        "reservation_id": int(rid),
        "operation": operation,
        "stoliki": list(target_ids),
        "oczekiwane_stoliki": list(expected_ids),
        "przekrocz_limity": bool(dane.przekrocz_limity),
        "nadpisanie_limitow": confirmation,
    }


def _host_assignment_source_version(table_ids) -> str:
    return "tables:" + ",".join(str(value) for value in sorted(table_ids))


def _host_existing_assignment_metadata(db, t, table_ids):
    rows = {
        row.id: row for row in db.query(models.Stolik).filter(
            models.Stolik.id.in_(table_ids),
        ).all()
    } if table_ids else {}
    ordered = [rows[value] for value in table_ids if value in rows]
    room_ids = {row.sala_id for row in ordered if row.sala_id is not None}
    room_id = next(iter(room_ids)) if len(room_ids) == 1 else None
    room = db.get(models.SalaRezerwacyjna, room_id) if room_id is not None else None
    names = [row.nazwa for row in ordered]
    capacity = sum(max(0, int(row.pojemnosc or 0)) for row in ordered)
    przydzial = {
        "stoliki": list(table_ids),
        "sala_id": room_id,
        "nazwa": " + ".join(names) or None,
        "suma_pojemnosci": capacity,
        "nadmiar_miejsc": max(0, capacity - max(1, int(t.liczba_osob or 1))),
        "kombinacja": len(table_ids) > 1,
        "wersja_planu_id": t.przydzial_wersja_planu_id,
        "kombinacja_planu_id": t.przydzial_kombinacja_planu_id,
    }
    allocation = {
        "state": "assigned",
        "visibility": "exact",
        "visit_end": _hm(t.godz_do),
        "kind": "combination" if len(table_ids) > 1 else "single_table",
        "room": (
            {"id": room_id, "name": room.nazwa if room is not None else None}
            if room_id is not None else None
        ),
        "tables": [
            {"id": row.id, "name": row.nazwa} for row in ordered
        ],
        "capacity": capacity,
        "reasons": [],
    }
    return przydzial, allocation


def _host_exact_alternatives(result, target_ids, *, visit_end):
    target = set(target_ids)
    rows = []
    for candidate in result.candidates:
        if set(candidate.table_ids) == target:
            continue
        rows.append({
            "kind": "resource",
            "allocation": candidate.to_display_dict(
                visit_end=visit_end,
                expose_exact=True,
            ),
        })
        if len(rows) == 3:
            break
    return rows


class _HostExactAvailabilityAdapter:
    def __init__(self, evaluation, result, target_ids):
        self.evaluation = evaluation
        self.result = result
        self.target_ids = target_ids

    def to_dict(self):
        payload = self.evaluation.to_dict()
        visit_end = payload.get("visit_end")
        payload.update({
            "candidates": [
                candidate.to_dict(expose_exact=True)
                for candidate in self.result.candidates[:3]
            ],
            "alternatives": _host_exact_alternatives(
                self.result, self.target_ids, visit_end=visit_end,
            ),
            "resource_allocation": "unavailable",
        })
        return payload


def _host_exact_response(t, user, metadata, *, replayed, now=None):
    mutation = dict(metadata.get("mutation") or {})
    mutation["replayed"] = bool(replayed)
    return {
        "rezerwacja": _host_out(t, now or utcnow_naive(), user=user),
        "przydzial": metadata.get("przydzial"),
        "allocation": metadata.get("allocation"),
        "alternatives": list(metadata.get("alternatives") or []),
        "mutation": mutation,
        "undo_command": metadata.get("undo_command"),
    }


def _host_exact_allocation_mutation(
    *,
    rid,
    operation,
    dane,
    request,
    idempotency_key,
    db,
    user,
):
    """R6b.3: jedna atomowa komenda exact-set dla pointera i klawiatury."""
    target_ids = tuple(sorted(int(value) for value in dane.stoliki))
    expected_ids = tuple(sorted(int(value) for value in dane.oczekiwane_stoliki))
    idem_operation = f"host_{operation}_exact"
    idem_payload = _host_exact_idempotency_payload(
        rid, operation, dane, target_ids, expected_ids,
    )
    reservation_service.required_idempotency_identity(
        operation=idem_operation,
        raw_key=idempotency_key,
        payload=idem_payload,
        secret=SECRET_KEY,
    )
    override_context = _jawne_nadpisanie_limitow(
        user, dane.przekrocz_limity, dane.nadpisanie_limitow,
    )
    t, guards = _zablokuj_termin(db, rid)
    if not t or t.rodzaj != "stolik":
        db.rollback()
        raise HTTPException(404, "Brak rezerwacji.")

    now = utcnow_naive()
    idem = reservation_service.begin_idempotency(
        db,
        operation=idem_operation,
        raw_key=idempotency_key,
        payload=idem_payload,
        secret=SECRET_KEY,
        now=now,
    )
    if idem.replayed:
        metadata = dict(idem.response or {})
        replayed = _termin_z_replay_idempotencji(db, idem)
        expected_target = tuple(sorted(
            int(value)
            for value in (metadata.get("mutation") or {}).get("to_stoliki") or ()
        ))
        if (
            tuple(sorted(_stoly_terminu(replayed))) != expected_target
            or replayed.status not in REZ_AKTYWNE
        ):
            raise reservation_service.ReservationError(
                409,
                "IDEMPOTENCY_RESULT_STALE",
                "Pierwotny wynik tej akcji nie jest już aktualny. Odśwież widok przed kolejną zmianą.",
                rule="idempotency",
            )
        return _host_exact_response(
            replayed, user, metadata, replayed=True, now=utcnow_naive(),
        )

    if t.status not in REZ_AKTYWNE:
        raise reservation_service.ReservationError(
            409,
            "RESERVATION_NOT_ACTIVE",
            "Nie można zmienić stołu zakończonej lub anulowanej rezerwacji.",
            rule="reservation_lifecycle",
        )
    if operation == "move" and t.faza_hosta not in HOST_NA_SALI:
        raise reservation_service.ReservationError(
            409,
            "HOST_MOVE_PHASE_INVALID",
            "Przeniesienie jest dostępne dopiero po posadzeniu gości.",
            rule="host_phase",
        )
    if operation == "seat" and "posadzony" not in HOST_PRZEJSCIA.get(t.faza_hosta, set()):
        raise reservation_service.ReservationError(
            409,
            "HOST_SEAT_PHASE_INVALID",
            f"Nie można posadzić rezerwacji w fazie {t.faza_hosta or '—'}.",
            rule="host_phase",
        )

    current_ids = tuple(sorted(_stoly_terminu(t)))
    if current_ids != expected_ids:
        raise reservation_service.ReservationError(
            409,
            "HOST_ASSIGNMENT_CHANGED",
            "Przydział tej rezerwacji zmienił się. Odśwież widok i spróbuj ponownie.",
            rule="reservation_state",
        )

    if operation == "move" and target_ids == current_ids:
        przydzial, allocation = _host_existing_assignment_metadata(
            db, t, target_ids,
        )
        metadata = {
            "termin_id": t.id,
            "przydzial": przydzial,
            "allocation": allocation,
            "alternatives": [],
            "mutation": {
                "operation": "move",
                "changed": False,
                "from_stoliki": list(current_ids),
                "to_stoliki": list(target_ids),
                "replayed": False,
            },
            "undo_command": None,
        }
        reservation_service.complete_idempotency(
            idem.record,
            response=metadata,
            http_status=200,
            termin_id=t.id,
            now=now,
        )
        db.commit()
        db.refresh(t)
        return _host_exact_response(t, user, metadata, replayed=False, now=now)

    if t.godz_od is None:
        raise reservation_service.ReservationError(
            400,
            "INVALID_RESERVATION_INTERVAL",
            "Rezerwacja bez godziny nie może otrzymać stołu.",
            rule="interval",
        )
    godz_do = t.godz_do or _dodaj_minuty(
        t.godz_od, _dlugosc_dla(db, t.data, t.godz_od, t.liczba_osob),
    )
    _waliduj_przydzial_rezerwacji(
        db,
        t.data,
        t.godz_od,
        godz_do,
        target_ids,
        t.liczba_osob,
        pomin_id=t.id,
    )
    result = _ocen_przydzial_rezerwacji(
        db,
        data=t.data,
        godz_od=t.godz_od,
        godz_do=godz_do,
        osoby=t.liczba_osob or 1,
        kanal=t.kanal,
        pomin_id=t.id,
        intent="assign",
        alternative_limit=9999,
        now=now,
    )
    candidate = _kandydat_zestawu(result, target_ids)
    if candidate is None:
        if result.evaluation.decision != "allow" or result.selected is None:
            _wymagaj_dozwolonego_przydzialu(
                result,
                override=bool(
                    override_context
                    and result.evaluation.decision == "override_required"
                ),
            )
        payload = result.to_dict(expose_exact=True)
        raise reservation_service.ReservationError(
            409,
            "INVALID_TABLE_COMBINATION",
            "Wybrany zestaw nie jest aktualną, zatwierdzoną konfiguracją.",
            rule="table",
            candidates=(payload.get("candidates") or ())[:3],
            alternatives=(payload.get("alternatives") or ())[:3],
        )

    evaluation = _ocen_reguly_slotu(
        db,
        data=t.data,
        godz_od=t.godz_od,
        godz_do=godz_do,
        liczba_osob=t.liczba_osob or 1,
        kanal=t.kanal,
        sala_id=candidate.room_id,
        existing_termin_id=t.id,
        intent="assign",
        now=now,
    )
    override_active = bool(
        override_context and evaluation.decision == "override_required"
    )
    _wymagaj_reautoryzacji_nadpisania(
        db, request, user, override_active,
    )
    try:
        reservation_rules.enforce_rule_evaluation(
            evaluation,
            override=override_active,
            can_override=override_active,
        )
    except reservation_service.ReservationError as exc:
        exc.availability = _HostExactAvailabilityAdapter(
            evaluation, result, target_ids,
        )
        raise

    before = reservation_audit.reservation_snapshot(t)
    canonical_ids = tuple(candidate.table_ids)
    t.stolik_id = canonical_ids[0]
    t.stoliki_dodatkowe = list(canonical_ids[1:]) or None
    t.auto_przydzielony = False
    canonical = _kandydat_legacy_z_alokacji(candidate)
    _ustaw_proweniencje_przydzialu(t, canonical)
    t.godz_do = evaluation.godz_do or godz_do
    if operation == "seat":
        t.faza_hosta = "posadzony"
        t.host_seated_at = now
        if t.status == "rezerwacja":
            t.status = "potwierdzona"
            t.potwierdzono_at = now

    alternatives = _host_exact_alternatives(
        result, canonical_ids, visit_end=_hm(t.godz_do),
    )
    przydzial = canonical
    allocation = candidate.to_display_dict(
        visit_end=_hm(t.godz_do),
        expose_exact=True,
        state="assigned",
    )
    undo_command = None
    if operation == "move":
        undo_command = {
            "path": f"/host/rezerwacja/{t.id}/przydziel-stolik",
            "method": "POST",
            "source_version": _host_assignment_source_version(canonical_ids),
            "body": {
                "stoliki": list(current_ids),
                "oczekiwane_stoliki": list(canonical_ids),
            },
        }
    metadata = {
        "termin_id": t.id,
        "przydzial": przydzial,
        "allocation": allocation,
        "alternatives": alternatives,
        "mutation": {
            "operation": operation,
            "changed": True,
            "from_stoliki": list(current_ids),
            "to_stoliki": list(canonical_ids),
            "replayed": False,
        },
        "undo_command": undo_command,
    }
    try:
        db.flush()
        _zastap_ledger_terminu(
            db,
            t,
            stoliki=canonical_ids,
            enforce_pacing=True,
            evaluation=evaluation,
            override=override_active,
            candidates=[{"stoliki": list(canonical_ids)}],
            alternatives=alternatives,
            intent="assign",
        )
        reservation_audit.add_reservation_audit(
            db,
            termin=t,
            action="host" if operation == "seat" else "assign",
            actor=user,
            before=before,
            after=t,
        )
        override_details = (
            _szczegoly_nadpisania_r3(
                evaluation,
                legacy=bool(override_context and override_context["legacy"]),
            )
            if override_active else None
        )
        if override_details:
            reservation_audit.add_reservation_audit(
                db,
                termin=t,
                action="override",
                actor=user,
                reason=_powod_audytu_nadpisania(evaluation),
                before=before,
                after=t,
                override_details=override_details,
                override_reason_code=override_context["reason_code"],
                override_note=override_context["note"],
            )
        reservation_service.complete_idempotency(
            idem.record,
            response=metadata,
            http_status=200,
            termin_id=t.id,
            now=now,
        )
        _commit_zapis_rezerwacji(db, guards)
    except IntegrityError as exc:
        db.rollback()
        raise reservation_service.translate_integrity_error(exc) from exc
    db.refresh(t)
    return _host_exact_response(t, user, metadata, replayed=False, now=now)


@app.post("/api/host/rezerwacja/{rid}/faza", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def host_zmien_faze(
    rid: int,
    dane: schemas.HostFazaIn,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Zmiana fazy operacyjnej (przybył/posadzony/rachunek/opłacony/wyszedł) z walidacją przejść.
    'posadzony' potwierdza rezerwację; 'wyszedł' domyka ją jako odbytą (status księgowy)."""
    t, guards = _zablokuj_termin(db, rid)
    if not t or t.rodzaj != "stolik":
        db.rollback()
        raise HTTPException(404, "Brak rezerwacji.")
    faza = (dane.faza or "").strip()
    if faza not in HOST_FAZY:
        raise HTTPException(400, "Nieznana faza.")
    if faza not in HOST_PRZEJSCIA.get(t.faza_hosta, set()):
        raise HTTPException(409, f"Niedozwolone przejście fazy {t.faza_hosta or '—'} → {faza}.")
    before = reservation_audit.reservation_snapshot(t)
    teraz = utcnow_naive()
    t.faza_hosta = faza
    if faza == "przybyl":
        t.host_arrived_at = teraz
    elif faza == "posadzony":
        t.host_seated_at = teraz
        if t.status == "rezerwacja":
            t.status = "potwierdzona"; t.potwierdzono_at = teraz
    elif faza == "wyszedl":
        t.host_left_at = teraz
        if t.status in REZ_AKTYWNE:
            t.status = "odbyla"
        reservation_demand.mark_waitlist_attended(db, t)
        reservation_service.release_termin_allocation(db, t.id)
        reservation_communication.cancel_pending(db, t.id)
    reservation_audit.add_reservation_audit(
        db, termin=t, action="host", actor=user, before=before, after=t,
    )
    _commit_zapis_rezerwacji(db, guards)
    db.refresh(t)
    if faza == "wyszedl" and t.godz_od:                    # obrót zakończony → stół wolny
        _bezpiecznie_po_zwolnieniu_stolu(db, t.data, t.godz_od, _koniec_okna(db, t))
    return _host_out(t, teraz, user=user)


@app.post("/api/host/rezerwacja/{rid}/przydziel-stolik", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def host_przydziel_stolik(
    rid: int,
    dane: schemas.HostStolikIn,
    request: Request,
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Ręczny przydział/przeniesienie rezerwacji na konkretny stół (walidacja pojemności + kolizji)."""
    if _host_exact_contract(dane):
        return _host_exact_allocation_mutation(
            rid=rid,
            operation="move",
            dane=dane,
            request=request,
            idempotency_key=idempotency_key,
            db=db,
            user=user,
        )
    override_context = _jawne_nadpisanie_limitow(
        user, dane.przekrocz_limity, dane.nadpisanie_limitow,
    )
    t, guards = _zablokuj_termin(db, rid)
    if not t or t.rodzaj != "stolik":
        db.rollback()
        raise HTTPException(404, "Brak rezerwacji.")
    if not t.godz_od:
        raise HTTPException(400, "Rezerwacja bez godziny — nie można przydzielić stołu.")
    before = reservation_audit.reservation_snapshot(t)
    godz_do = t.godz_do or _dodaj_minuty(t.godz_od, _dlugosc_dla(db, t.data, t.godz_od, t.liczba_osob))
    _waliduj_rezerwacje(db, t.data, t.godz_od, godz_do, dane.stolik_id, t.liczba_osob, pomin_id=rid)
    t.stolik_id = dane.stolik_id
    t.stoliki_dodatkowe = None          # ręczny pojedynczy stół kasuje wcześniejszą kombinację
    t.auto_przydzielony = False         # od tej chwili źródłem przydziału jest decyzja operatora
    _ustaw_proweniencje_przydzialu(t)
    t.godz_do = godz_do
    evaluation = _ocen_reguly_terminu(
        db, t, stoliki=[dane.stolik_id], intent="assign",
    )
    override_active = bool(
        override_context and evaluation.decision == "override_required"
    )
    _wymagaj_reautoryzacji_nadpisania(
        db, request, user, override_active,
    )
    try:
        db.flush()
        _zastap_ledger_terminu(
            db,
            t,
            enforce_pacing=True,
            evaluation=evaluation,
            override=override_active,
            intent="assign",
        )
        reservation_audit.add_reservation_audit(
            db, termin=t, action="assign", actor=user, before=before, after=t,
        )
        override_details = (
            _szczegoly_nadpisania_r3(
                evaluation,
                legacy=bool(override_context and override_context["legacy"]),
            )
            if override_active else None
        )
        if override_details:
            reservation_audit.add_reservation_audit(
                db,
                termin=t,
                action="override",
                actor=user,
                reason=_powod_audytu_nadpisania(evaluation),
                before=before,
                after=t,
                override_details=override_details,
                override_reason_code=override_context["reason_code"],
                override_note=override_context["note"],
            )
        _commit_zapis_rezerwacji(db, guards)
    except IntegrityError as exc:
        db.rollback()
        raise reservation_service.translate_integrity_error(exc) from exc
    db.refresh(t)
    return _host_out(t, utcnow_naive(), user=user)


@app.post("/api/host/rezerwacja/{rid}/posadz", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def host_posadz_rezerwacje(
    rid: int,
    dane: schemas.HostPosadzIn,
    request: Request,
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Atomowo dobiera/przypisuje stół i przeprowadza rezerwację do fazy ``posadzony``."""
    if _host_exact_contract(dane):
        return _host_exact_allocation_mutation(
            rid=rid,
            operation="seat",
            dane=dane,
            request=request,
            idempotency_key=idempotency_key,
            db=db,
            user=user,
        )
    override_context = _jawne_nadpisanie_limitow(
        user, dane.przekrocz_limity, dane.nadpisanie_limitow,
    )
    t, guards = _zablokuj_termin(db, rid)
    if not t or t.rodzaj != "stolik":
        db.rollback()
        raise HTTPException(404, "Brak rezerwacji.")
    if not t.godz_od:
        raise HTTPException(400, "Rezerwacja bez godziny — nie można posadzić gości.")
    if "posadzony" not in HOST_PRZEJSCIA.get(t.faza_hosta, set()):
        raise HTTPException(409, f"Nie można posadzić rezerwacji w fazie {t.faza_hosta or '—'}.")

    before = reservation_audit.reservation_snapshot(t)
    osoby = max(1, t.liczba_osob or 1)
    godz_do = t.godz_do or _dodaj_minuty(
        t.godz_od, _dlugosc_dla(db, t.data, t.godz_od, osoby),
    )
    allocation_result = None
    if dane.stolik_id is not None:
        _waliduj_rezerwacje(
            db, t.data, t.godz_od, godz_do, dane.stolik_id, osoby, pomin_id=t.id,
        )
        t.stolik_id = dane.stolik_id
        t.stoliki_dodatkowe = None
        t.auto_przydzielony = False
        _ustaw_proweniencje_przydzialu(t)
    elif not t.stolik_id:
        allocation_result = _ocen_przydzial_rezerwacji(
            db,
            data=t.data,
            godz_od=t.godz_od,
            godz_do=godz_do,
            osoby=osoby,
            kanal=t.kanal,
            pomin_id=t.id,
            intent="assign",
        )
        selected = _wymagaj_dozwolonego_przydzialu(
            allocation_result,
            override=bool(
                override_context
                and allocation_result.evaluation.decision == "override_required"
            ),
        )
        wybrany = _kandydat_legacy_z_alokacji(selected)
        t.stolik_id = wybrany["stoliki"][0]
        t.stoliki_dodatkowe = (wybrany["stoliki"][1:] or None)
        t.auto_przydzielony = True
        _ustaw_proweniencje_przydzialu(t, wybrany)

    evaluation = (
        allocation_result.evaluation
        if allocation_result is not None
        else _ocen_reguly_terminu(db, t, intent="assign")
    )
    override_active = bool(
        override_context and evaluation.decision == "override_required"
    )
    _wymagaj_reautoryzacji_nadpisania(
        db, request, user, override_active,
    )
    godz_do = evaluation.godz_do or godz_do
    teraz = utcnow_naive()
    t.godz_do = godz_do
    t.faza_hosta = "posadzony"
    t.host_seated_at = teraz
    if t.status == "rezerwacja":
        t.status = "potwierdzona"
        t.potwierdzono_at = teraz
    try:
        db.flush()
        _zastap_ledger_terminu(
            db,
            t,
            enforce_pacing=True,
            evaluation=evaluation,
            override=override_active,
            alternatives=(
                allocation_result.to_dict(expose_exact=True).get("alternatives") or ()
                if allocation_result else ()
            ),
            intent="assign",
        )
        reservation_audit.add_reservation_audit(
            db, termin=t, action="host", actor=user, before=before, after=t,
        )
        override_details = (
            _szczegoly_nadpisania_r3(
                evaluation,
                legacy=bool(override_context and override_context["legacy"]),
            )
            if override_active else None
        )
        if override_details:
            reservation_audit.add_reservation_audit(
                db,
                termin=t,
                action="override",
                actor=user,
                reason=_powod_audytu_nadpisania(evaluation),
                before=before,
                after=t,
                override_details=override_details,
                override_reason_code=override_context["reason_code"],
                override_note=override_context["note"],
            )
        _commit_zapis_rezerwacji(db, guards)
    except IntegrityError as exc:
        db.rollback()
        raise reservation_service.translate_integrity_error(exc) from exc
    db.refresh(t)
    response = _host_out(t, teraz, user=user)
    if allocation_result is not None:
        payload = allocation_result.to_dict(expose_exact=True)
        response.update({
            "allocation": {
                **(payload.get("allocation") or {}),
                "state": "assigned",
            },
            "alternatives": payload.get("alternatives") or [],
            "reasons": payload.get("reasons") or [],
        })
    return response


# ── Bezpieczny kontrakt konfiguracji rezerwacji ──────────────────────────────
@app.get("/api/rezerwacje/config", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def get_rezerwacje_config(db: Session = Depends(get_db)):
    """Minimalne ustawienia potrzebne operacyjnemu UI; bez finansów i konfiguracji lokalu."""
    cfg = get_lokal_config(db)
    return {"sale": list(cfg.sale or [])}


# ── Godziny otwarcia ─────────────────────────────────────────────────────────
@app.get("/api/godziny-otwarcia", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def get_godziny_otwarcia(db: Session = Depends(get_db)):
    rows = db.query(models.GodzinyOtwarcia).order_by(models.GodzinyOtwarcia.dzien_tygodnia,
                                                     models.GodzinyOtwarcia.godz_od).all()
    return {"godziny": [schemas.GodzinyOtwarciaOut.model_validate(g).model_dump() for g in rows]}


@app.post("/api/godziny-otwarcia", status_code=201, dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def dodaj_godziny_otwarcia(dane: schemas.GodzinyOtwarciaIn, db: Session = Depends(get_db)):
    if not (0 <= dane.dzien_tygodnia <= 6):
        raise HTTPException(400, "dzien_tygodnia musi być 0–6.")
    _waliduj_serwis_r3(db, dane)
    g = models.GodzinyOtwarcia(**dane.model_dump()); db.add(g); db.commit(); db.refresh(g)
    return schemas.GodzinyOtwarciaOut.model_validate(g).model_dump()


@app.delete("/api/godziny-otwarcia/{gid}", status_code=204, dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def usun_godziny_otwarcia(gid: int, db: Session = Depends(get_db)):
    g = db.get(models.GodzinyOtwarcia, gid)
    if g:
        if db.query(models.ReservationRecommendationReview.id).filter_by(
            service_id=gid,
        ).first():
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RESERVATION_SERVICE_HAS_RECOMMENDATION_HISTORY",
                    "message": (
                        "Serwis ma historię rekomendacji. Wyłącz go zamiast "
                        "usuwać, aby zachować ślad decyzji."
                    ),
                },
            )
        try:
            db.delete(g)
            db.commit()
        except IntegrityError as exc:
            db.rollback()
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "RESERVATION_SERVICE_IN_USE",
                    "message": (
                        "Serwis jest używany przez dane rezerwacji i nie może "
                        "zostać usunięty. Możesz go wyłączyć."
                    ),
                },
            ) from exc


# ── Rezerwacje (rodzaj=stolik na encji Termin) ───────────────────────────────
@app.get("/api/rezerwacje-stolik", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def get_rezerwacje_stolik(start: date = Query(...), end: date = Query(...),
                          status: Optional[str] = None, stolik_id: Optional[int] = None,
                          db: Session = Depends(get_db),
                          user: models.User = Depends(get_current_user)):
    q = db.query(models.Termin).filter(models.Termin.rodzaj == "stolik",
                                       models.Termin.data >= start, models.Termin.data <= end)
    if status:
        q = q.filter(models.Termin.status == status)
    rows = q.order_by(models.Termin.data, models.Termin.godz_od).all()
    if stolik_id:
        rows = [t for t in rows if stolik_id in _stoly_terminu(t)]
    return {"rezerwacje": _rezerwacje_out_z_komunikacja(db, rows, user)}


@app.post("/api/rezerwacje-stolik/wyszukaj", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def wyszukaj_rezerwacje_stolik(
    dane: schemas.RezerwacjeWyszukajIn,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Baza rezerwacji z PII w body, ograniczonym zakresem i deterministyczną paginacją."""
    if not _ma_dostep_rezerwacji(user, "rezerwacje.dane_kontaktowe"):
        # Obrona warstwowa na wypadek przyszłej regresji w centralnej mapie metoda+trasa.
        raise HTTPException(403, "Brak uprawnienia do danych kontaktowych gości.")
    _waliduj_zakres_bazy_rezerwacji(dane.start, dane.end)

    q = db.query(models.Termin).filter(
        models.Termin.rodzaj == "stolik",
        models.Termin.data >= dane.start,
        models.Termin.data <= dane.end,
    )
    if dane.status:
        q = q.filter(models.Termin.status == dane.status)
    rows = [t for t in q.all() if _pasuje_do_wyszukiwania_rezerwacji(t, dane.query)]
    rows = _sortuj_baze_rezerwacji(rows, dane.sort)
    total = len(rows)
    page = rows[dane.offset:dane.offset + dane.limit]
    return {
        "rezerwacje": _rezerwacje_out_z_komunikacja(db, page, user),
        "total": total,
        "offset": dane.offset,
        "limit": dane.limit,
    }


@app.get("/api/rezerwacje-stolik/{rid}", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def get_rezerwacja_stolik(
    rid: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    t = db.get(models.Termin, rid)
    if not t or t.rodzaj != "stolik":
        raise HTTPException(404, "Brak rezerwacji.")
    return _rezerwacje_out_z_komunikacja(db, [t], user)[0]


@app.post("/api/rezerwacje-stolik", status_code=201, dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def dodaj_rezerwacje_stolik(
    dane: schemas.RezerwacjaIn,
    request: Request,
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    override_context = _jawne_nadpisanie_limitow(
        user, dane.przekrocz_limity, dane.nadpisanie_limitow,
    )
    kanal = dane.kanal or "reczna"
    guards = reservation_service.begin_locked_write(db, [dane.data])
    teraz = utcnow_naive()
    try:
        idem = reservation_service.begin_idempotency(
            db,
            operation=(
                "reservation.create.walk_in:v1"
                if kanal == "walk_in" else "reservation.create.manual:v1"
            ),
            raw_key=idempotency_key,
            payload=dane.model_dump(mode="json"),
            secret=SECRET_KEY,
            now=teraz,
        )
        if idem.replayed:
            replayed = _termin_z_replay_idempotencji(db, idem)
            return _rezerwacja_out(replayed, user)
        auto_allocate = bool(
            dane.stolik_id is None and (kanal == "walk_in" or dane.auto_przydziel)
        )
        if auto_allocate and dane.godz_od is None:
            raise HTTPException(
                400,
                "Automatyczny przydział wymaga godziny rezerwacji.",
            )
        allocation_result = None
        selected = None
        if auto_allocate:
            allocation_result = _ocen_przydzial_rezerwacji(
                db,
                data=dane.data,
                godz_od=dane.godz_od,
                godz_do=dane.godz_do,
                osoby=dane.liczba_osob or 1,
                kanal=kanal,
                intent="create",
            )
            if allocation_result.selected is None:
                _wymagaj_dozwolonego_przydzialu(allocation_result)
            evaluation = allocation_result.evaluation
            override_for_selection = bool(
                override_context and evaluation.decision == "override_required"
            )
            if evaluation.decision == "deny" or (
                evaluation.decision == "override_required" and not override_for_selection
            ):
                reservation_rules.enforce_rule_evaluation(evaluation)
            selected = allocation_result.selected
            table_ids = list(selected.table_ids)
            reservation_service.lock_tables(db, table_ids)
            godz_do = evaluation.godz_do
        else:
            table_ids = [dane.stolik_id] if dane.stolik_id is not None else []
            godz_do = _waliduj_rezerwacje(
                db, dane.data, dane.godz_od, dane.godz_do,
                dane.stolik_id, dane.liczba_osob,
            )
            evaluation = None
        t = models.Termin(
            data=dane.data, nazwisko=dane.nazwisko.strip(), telefon=dane.telefon, email=dane.email,
            kanal_komunikacji=dane.kanal_komunikacji,
            liczba_osob=dane.liczba_osob,
            notatka=(dane.notatka if _ma_dostep_rezerwacji(
                user, "rezerwacje.notatki_wewnetrzne") else None),
            status="potwierdzona",
            zadatek=(float(dane.zadatek or 0) if _ma_dostep_rezerwacji(
                user, "rezerwacje.finanse") else 0.0),
            utworzono_at=teraz,
            godz_od=dane.godz_od, godz_do=godz_do,
            stolik_id=(table_ids[0] if table_ids else None),
            stoliki_dodatkowe=(table_ids[1:] or None),
            auto_przydzielony=(True if selected is not None else False),
            rodzaj="stolik", kanal=kanal,
            faza_hosta=("posadzony" if kanal == "walk_in" else None),
            host_arrived_at=(teraz if kanal == "walk_in" else None),
            host_seated_at=(teraz if kanal == "walk_in" else None),
            potwierdzono_at=(teraz if kanal == "walk_in" else None),
        )
        if selected is not None:
            _ustaw_proweniencje_przydzialu(
                t, _kandydat_legacy_z_alokacji(selected),
            )
        db.add(t); db.flush()
        evaluation = evaluation or _ocen_reguly_terminu(db, t, intent="create")
        override_active = bool(
            override_context
            and evaluation is not None
            and evaluation.decision == "override_required"
        )
        _wymagaj_reautoryzacji_nadpisania(
            db, request, user, override_active,
        )
        override_details = (
            _szczegoly_nadpisania_r3(
                evaluation, legacy=bool(override_context and override_context["legacy"]),
            )
            if override_active else None
        )
        _zastap_ledger_terminu(
            db,
            t,
            enforce_pacing=True,
            evaluation=evaluation,
            override=override_active,
            intent="create",
            candidates=([{"stoliki": table_ids}] if table_ids else ()),
            alternatives=(
                allocation_result.to_dict(expose_exact=True).get("alternatives") or ()
                if allocation_result else ()
            ),
        )
        odpowiedz = _rezerwacja_out(t, user)
        reservation_audit.add_reservation_audit(
            db,
            termin=t,
            action="create",
            actor=user,
            after=t,
            pii_changed=_utworzone_pii(t),
        )
        if float(t.zadatek or 0) <= 0:
            payment_policy = reservation_payments.resolve_policy(
                db,
                t.data,
                evaluation.service_id if evaluation is not None else None,
                int(t.liczba_osob or 1),
                "wewnetrzna",
            )
            payment_provider = (
                integracje.provider_platnosci_wymaganej()
                if payment_policy is not None and payment_policy.required
                else "sandbox"
            )
            payment, payment_command = reservation_payments.create_payment_for_reservation(
                db,
                t,
                payment_policy,
                provider=payment_provider,
                now=teraz,
                business_today=(
                    (_teraz_lokalnie() or datetime.now(timezone.utc).replace(tzinfo=None)).date()
                ),
                service_id=(evaluation.service_id if evaluation is not None else None),
                operation_key="initial",
                actor_kind="user",
                actor_user_id=user.id,
            )
            if payment is not None and payment_provider == "sandbox" and payment_command is not None:
                payment.link = "/?platnosc=sandbox&rezerwuj"
                payment_command.stan = "succeeded"
                payment_command.finished_at = teraz
                payment_command.updated_at = teraz
        if override_details:
            reservation_audit.add_reservation_audit(
                db,
                termin=t,
                action="override",
                actor=user,
                reason=_powod_audytu_nadpisania(evaluation),
                after=t,
                override_details=override_details,
                override_reason_code=override_context["reason_code"],
                override_note=override_context["note"],
            )
        if kanal != "walk_in":
            reservation_communication.enqueue_reservation(
                db,
                t,
                "confirmation",
                dedupe_key=(
                    f"reservation:{t.id}:confirmation:create:"
                    f"{secrets.token_hex(16)}"
                ),
                actor=user,
            )
            reservation_communication.schedule_reminder(db, t, actor=user)
        reservation_service.complete_idempotency(
            idem.record, response=odpowiedz, http_status=201, termin_id=t.id, now=teraz,
        )
        _commit_zapis_rezerwacji(db, guards)
    except IntegrityError as exc:
        db.rollback()
        raise reservation_service.translate_integrity_error(exc) from exc
    db.refresh(t)
    wyslij_push_do_adminow(db, "Nowa rezerwacja",
                           f"{t.nazwisko} — {t.data} {_hm(t.godz_od) or ''}".strip(), url="/")
    return odpowiedz


@app.put("/api/rezerwacje-stolik/{rid}", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def edytuj_rezerwacje_stolik(
    rid: int,
    dane: schemas.RezerwacjaIn,
    request: Request,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    override_context = _jawne_nadpisanie_limitow(
        user, dane.przekrocz_limity, dane.nadpisanie_limitow,
    )
    t, guards = _zablokuj_termin(db, rid, [dane.data])
    if not t or t.rodzaj != "stolik":
        db.rollback()
        raise HTTPException(404, "Brak rezerwacji.")
    payment_before_edit = _platnosc_rezerwacji(db, t.id)
    financial_context_changed = (
        (t.data, t.godz_od, int(t.liczba_osob or 1))
        != (dane.data, dane.godz_od, int(dane.liczba_osob or 1))
    )
    if (
        financial_context_changed
        and payment_before_edit is not None
        and payment_before_edit.status in {"oczekuje", "autoryzowana", "oplacona"}
    ):
        db.rollback()
        raise reservation_payments.PaymentDomainError(
            "PAYMENT_SETTLEMENT_REQUIRED_BEFORE_EDIT",
            "Najpierw anuluj, zwolnij lub zwróć aktywną płatność, a potem zmień termin albo liczbę gości.",
        )
    before = reservation_audit.reservation_snapshot(t)
    before_guest_details = (
        t.data, t.godz_od, t.liczba_osob, t.telefon, t.email,
        t.kanal_komunikacji,
    )
    pii_before = _wartosci_pii_rezerwacji(t)
    crm_identity_before = _identity_key_crm(t)
    zachowaj_przydzial = bool(
        dane.stolik_id is not None
        and dane.stolik_id == t.stolik_id
    )
    stoliki = _stoly_terminu(t) if zachowaj_przydzial else [dane.stolik_id]
    pojemnosc_override = (
        _pojemnosc_zachowanego_zestawu(db, stoliki, t)
        if zachowaj_przydzial and len(stoliki) > 1 else None
    )
    godz_do = _waliduj_przydzial_rezerwacji(
        db, dane.data, dane.godz_od, dane.godz_do, stoliki, dane.liczba_osob,
        pomin_id=rid,
        pojemnosc_override=pojemnosc_override,
        zachowaj_nieaktywny_przydzial=zachowaj_przydzial,
    )
    if zachowaj_przydzial and len(stoliki) > 1 and dane.liczba_osob != t.liczba_osob:
        _waliduj_rozmiar_zachowanej_kombinacji(
            db, stoliki, dane.liczba_osob, t,
        )
    t.data = dane.data
    if _ma_dostep_rezerwacji(user, "rezerwacje.dane_kontaktowe"):
        t.nazwisko = dane.nazwisko.strip(); t.telefon = dane.telefon; t.email = dane.email
        if "kanal_komunikacji" in dane.model_fields_set:
            t.kanal_komunikacji = dane.kanal_komunikacji
    t.liczba_osob = dane.liczba_osob
    if _ma_dostep_rezerwacji(user, "rezerwacje.notatki_wewnetrzne"):
        t.notatka = dane.notatka
    if (
        _ma_dostep_rezerwacji(user, "rezerwacje.finanse")
        and payment_before_edit is None
    ):
        t.zadatek = float(dane.zadatek or 0)
    t.godz_od = dane.godz_od; t.godz_do = godz_do
    t.stolik_id = dane.stolik_id
    if not zachowaj_przydzial:
        # Zmiana (także na „bez stolika”) jest ręcznym, pojedynczym przydziałem. Nie wolno
        # zachować niewidocznych składników wcześniejszej kombinacji ani flagi AUTO.
        t.stoliki_dodatkowe = None
        t.auto_przydzielony = False
        _ustaw_proweniencje_przydzialu(t)
    try:
        _migruj_osierocony_profil_po_zmianie_tozsamosci(db, t, crm_identity_before)
        crm_governance.revert_orphaned_identity_merges_after_contact_change(
            db,
            reservation_id=t.id,
            previous_identity_key=crm_identity_before,
            actor=user,
        )
        db.flush()
        evaluation = _ocen_reguly_terminu(
            db,
            t,
            intent="edit",
            preserve_existing_room_access=zachowaj_przydzial,
        )
        override_active = bool(
            override_context
            and evaluation is not None
            and evaluation.decision == "override_required"
        )
        _wymagaj_reautoryzacji_nadpisania(
            db, request, user, override_active,
        )
        override_details = (
            _szczegoly_nadpisania_r3(
                evaluation, legacy=bool(override_context and override_context["legacy"]),
            )
            if override_active else None
        )
        _zastap_ledger_terminu(
            db,
            t,
            enforce_pacing=True,
            zachowaj_nieaktywny_przydzial=zachowaj_przydzial,
            evaluation=evaluation,
            override=override_active,
            intent="edit",
        )
        reservation_audit.add_reservation_audit(
            db,
            termin=t,
            action="edit",
            actor=user,
            before=before,
            after=t,
            pii_changed=_zmienione_pii(pii_before, t),
        )
        if (
            financial_context_changed
            and payment_before_edit is not None
            and payment_before_edit.status in {
                "nieudana", "wygasla", "anulowana", "zwrocona",
            }
        ):
            # Terminalna próba należy do poprzedniego kontekstu finansowego.
            # Zachowujemy ją w globalnym audycie po ``reservation_ref``, ale nie
            # pozwalamy, aby była ponawiana lub pokazywana jako bieżąca po edycji.
            if payment_before_edit.status != "zwrocona":
                reservation_payments.mark_payment_superseded(
                    db,
                    payment_before_edit,
                    now=utcnow_naive(),
                    actor_kind="user",
                    actor_user_id=user.id,
                )
            payment_before_edit.termin_id = None
        if (
            financial_context_changed
            and (
                payment_before_edit is None
                or payment_before_edit.status in {
                    "nieudana", "wygasla", "anulowana", "zwrocona",
                }
            )
            and (
                payment_before_edit is not None
                or float(t.zadatek or 0) <= 0
            )
        ):
            payment_policy = reservation_payments.resolve_policy(
                db,
                t.data,
                evaluation.service_id if evaluation is not None else None,
                int(t.liczba_osob or 1),
                "online" if t.kanal == "online" else "wewnetrzna",
            )
            payment_provider = (
                integracje.provider_platnosci_wymaganej()
                if payment_policy is not None and payment_policy.required
                else "sandbox"
            )
            previous_ref = payment_before_edit.id if payment_before_edit is not None else 0
            payment, payment_command = reservation_payments.create_payment_for_reservation(
                db,
                t,
                payment_policy,
                provider=payment_provider,
                now=utcnow_naive(),
                business_today=(
                    (_teraz_lokalnie() or datetime.now(timezone.utc).replace(tzinfo=None)).date()
                ),
                service_id=(evaluation.service_id if evaluation is not None else None),
                operation_key=f"edit:{previous_ref}:{t.data}:{int(t.liczba_osob or 1)}",
                actor_kind="user",
                actor_user_id=user.id,
            )
            if payment is not None and payment_provider == "sandbox" and payment_command is not None:
                payment.link = "/?platnosc=sandbox&rezerwuj"
                payment_command.stan = "succeeded"
                payment_command.finished_at = utcnow_naive()
                payment_command.updated_at = payment_command.finished_at
        if override_details:
            reservation_audit.add_reservation_audit(
                db,
                termin=t,
                action="override",
                actor=user,
                reason=_powod_audytu_nadpisania(evaluation),
                before=before,
                after=t,
                override_details=override_details,
                override_reason_code=override_context["reason_code"],
                override_note=override_context["note"],
            )
        after_guest_details = (
            t.data, t.godz_od, t.liczba_osob, t.telefon, t.email,
            t.kanal_komunikacji,
        )
        if before_guest_details != after_guest_details:
            reservation_communication.cancel_pending(
                db,
                t.id,
                # Reminder jest mutowany dopiero wewnątrz schedule_reminder,
                # już pod planner lockiem (globalnie: day -> planner -> outbox).
                event_types=("confirmation", "change"),
            )
            if t.status in REZ_AKTYWNE:
                reservation_communication.enqueue_reservation(
                    db, t, "change", actor=user,
                )
            reservation_communication.schedule_reminder(
                db, t, actor=user, force_new=True,
            )
        _commit_zapis_rezerwacji(db, guards)
    except IntegrityError as exc:
        db.rollback()
        raise reservation_service.translate_integrity_error(exc) from exc
    db.refresh(t)
    return _rezerwacja_out(t, user)


def _koniec_okna(db, t) -> time:
    """Godzina końca okna rezerwacji (godz_do jawne albo z długości slotu dla grupy)."""
    return t.godz_do or _dodaj_minuty(t.godz_od, _dlugosc_dla(db, t.data, t.godz_od, t.liczba_osob))


def _po_zwolnieniu_stolu(db, data, godz_od, godz_do) -> dict:
    """Re-optymalizacja po zwolnieniu stołu (odwołanie / no-show / wyjście / usunięcie).

    Auto-przydzielone rezerwacje tego dnia, które JESZCZE nie są na sali i nachodzą na zwolnione
    okno, mogą przeskoczyć na tańszy stół (np. z kombinacji na pojedynczy stół, który się zwolnił).
    Posadzonych gości nie ruszamy. Zwraca też wpisy listy oczekujących pasujące do okna — jako
    propozycję dla hosta (realizacja pozostaje ręczna)."""
    guards = reservation_service.begin_locked_write(db, [data])
    # Reoptymalizacja może po kolei rozważać różne, zachodzące na siebie
    # kombinacje. Na PostgreSQL blokujemy więc cały zbiór kandydatów raz,
    # globalnie po rosnącym ID, zanim pierwsza rezerwacja zacznie zmieniać
    # przydział. Kolejne blokady podzbiorów są już wtedy reentrantne w tej
    # samej transakcji. SQLite nadal polega na BEGIN IMMEDIATE powyżej.
    active_table_ids = sorted(stolik["id"] for stolik in _stoly_do_seating(db))
    reservation_service.lock_tables(db, active_table_ids)
    przesadzone = []
    auto = (db.query(models.Termin).filter(
        models.Termin.rodzaj == "stolik", models.Termin.data == data,
        models.Termin.auto_przydzielony.is_(True),
        models.Termin.status.in_(REZ_AKTYWNE), models.Termin.godz_od.isnot(None)).all())
    for t in auto:
        if t.faza_hosta in HOST_NA_SALI:                 # już posadzony — obrót w toku, nie przenoś
            continue
        t_do = _koniec_okna(db, t)
        if not (t.godz_od < godz_do and godz_od < t_do):  # nie nachodzi na zwolnione okno
            continue
        osoby = max(1, t.liczba_osob or 1)
        obecne = _stoly_terminu(t)
        result = _ocen_przydzial_rezerwacji(
            db,
            data=data,
            godz_od=t.godz_od,
            godz_do=t_do,
            osoby=osoby,
            kanal=t.kanal,
            pomin_id=t.id,
            intent="reoptimize",
            alternative_limit=9999,
        )
        if result.decision != "allow" or result.selected is None:
            continue
        obecny_kandydat = _kandydat_zestawu(result, obecne)
        if (
            obecny_kandydat is not None
            and set(result.selected.table_ids) != obecne
            and result.selected.room_id == obecny_kandydat.room_id
            and len(result.selected.table_ids) == len(obecny_kandydat.table_ids)
            and math.isclose(
                result.selected.ranking_cost,
                obecny_kandydat.ranking_cost,
                rel_tol=0.0,
                abs_tol=1e-9,
            )
        ):
            # Deterministyczny tie-break po ID nie jest realną optymalizacją i nie
            # powinien niepotrzebnie przesadzać przyszłej rezerwacji.
            continue
        najlepszy = _kandydat_legacy_z_alokacji(result.selected)
        if set(najlepszy["stoliki"]) != obecne:
            before = reservation_audit.reservation_snapshot(t)
            t.stolik_id = najlepszy["stoliki"][0]
            t.stoliki_dodatkowe = (najlepszy["stoliki"][1:] or None)
            _ustaw_proweniencje_przydzialu(t, najlepszy)
            _zastap_ledger_terminu(
                db,
                t,
                candidates=[{"stoliki": najlepszy["stoliki"]}],
                enforce_pacing=False,
                evaluation=result.evaluation,
                intent="edit",
            )
            reservation_audit.add_reservation_audit(
                db,
                termin=t,
                action="assign",
                actor_kind="system",
                reason="system_automation",
                before=before,
                after=t,
            )
            przesadzone.append(t.id)
    if przesadzone:
        _commit_zapis_rezerwacji(db, guards)
    else:
        db.rollback()
    propozycje = [w for w in db.query(models.ListaOczekujacych).filter(
        models.ListaOczekujacych.data == data,
        models.ListaOczekujacych.status == "oczekuje").all()
        if w.godz_od is None or (godz_od <= w.godz_od <= godz_do)]
    return {"przesadzone": przesadzone, "propozycje_waitlisty": [_lista_out(w) for w in propozycje]}


def _bezpiecznie_po_zwolnieniu_stolu(db, data, godz_od, godz_do) -> dict:
    """Reoptymalizacja jest skutkiem ubocznym i nie może odwrócić wyniku zatwierdzonej operacji."""
    try:
        return _po_zwolnieniu_stolu(db, data, godz_od, godz_do)
    except Exception:  # noqa: BLE001 — zapis główny został już zatwierdzony
        db.rollback()
        logger.exception("Nie udała się reoptymalizacja po zwolnieniu stołu")
        return {"przesadzone": [], "propozycje_waitlisty": []}


@app.post("/api/host/auto-no-show", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def host_auto_no_show(
    data: date = Query(...),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Oznacza jako no_show rezerwacje, które nie przyszły: minęło godz_od + rez_no_show_po_min,
    a faza_hosta pusta (gość się nie pojawił). Idempotentne. No-op gdy rez_no_show_po_min=0.
    Każde zwolnienie uruchamia re-optymalizację auto-przydziałów."""
    prog = get_lokal_config(db).rez_no_show_po_min or 0
    if prog <= 0:
        return {"oznaczone": []}
    guards = reservation_service.begin_locked_write(db, [data])
    teraz = utcnow_naive()
    oznaczone = []
    for t in (db.query(models.Termin).filter(
            models.Termin.rodzaj == "stolik", models.Termin.data == data,
            models.Termin.status.in_(REZ_AKTYWNE), models.Termin.godz_od.isnot(None),
            models.Termin.faza_hosta.is_(None)).all()):
        if datetime.combine(t.data, _dodaj_minuty(t.godz_od, prog)) < teraz:
            before = reservation_audit.reservation_snapshot(t)
            t.status = "no_show"
            reservation_service.release_termin_allocation(db, t.id)
            reservation_communication.cancel_pending(db, t.id)
            _nalicz_no_show_fee(db, t, commit=False)
            reservation_audit.add_reservation_audit(
                db, termin=t, action="status", actor=user, before=before, after=t,
            )
            oznaczone.append((t.id, t.data, t.godz_od, _koniec_okna(db, t)))
    if oznaczone:
        _commit_zapis_rezerwacji(db, guards)
        for rid, d, od, do in oznaczone:
            _bezpiecznie_po_zwolnieniu_stolu(db, d, od, do)
    else:
        db.rollback()
    return {"oznaczone": [rid for rid, *_ in oznaczone]}


@app.post("/api/rezerwacje-stolik/{rid}/status", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def zmien_status_rezerwacji_stolik(
    rid: int,
    dane: schemas.RezerwacjaStatusIn,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    t, guards = _zablokuj_termin(db, rid)
    if not t or t.rodzaj != "stolik":
        raise HTTPException(404, "Brak rezerwacji.")
    before = reservation_audit.reservation_snapshot(t)
    nowy = (dane.status or "").strip()
    if nowy not in REZ_STATUSY:
        raise HTTPException(400, "Nieznany status.")
    if nowy not in REZ_PRZEJSCIA.get(t.status, set()):
        raise HTTPException(409, f"Niedozwolone przejście {t.status} → {nowy}.")
    status_now = utcnow_naive()
    t.status = nowy
    if nowy == "potwierdzona":
        t.potwierdzono_at = status_now
    elif nowy == "odwolana":
        t.odwolano_at = status_now
        reservation_payments.request_reservation_cancellation_settlement(
            db,
            t,
            now=status_now,
            actor_kind="user",
            actor_user_id=user.id,
        )
    if nowy not in REZ_AKTYWNE:
        reservation_service.release_termin_allocation(db, t.id)
    if nowy == "no_show":
        _nalicz_no_show_fee(db, t, commit=False)
    reservation_audit.add_reservation_audit(
        db,
        termin=t,
        action="cancel" if nowy == "odwolana" else "status",
        actor=user,
        before=before,
        after=t,
    )
    if nowy == "odwolana":
        reservation_communication.cancel_pending(db, t.id)
        reservation_communication.enqueue_reservation(
            db, t, "cancellation", actor=user,
        )
    elif nowy not in REZ_AKTYWNE:
        reservation_communication.cancel_pending(db, t.id)
    elif nowy == "potwierdzona":
        reservation_communication.schedule_reminder(db, t, actor=user)
    _commit_zapis_rezerwacji(db, guards); db.refresh(t)
    if nowy not in REZ_AKTYWNE and t.godz_od:                # stół się zwolnił → re-optymalizacja
        _bezpiecznie_po_zwolnieniu_stolu(db, t.data, t.godz_od, _koniec_okna(db, t))
    return _rezerwacja_out(t, user)


@app.delete("/api/rezerwacje-stolik/{rid}", status_code=204, dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def usun_rezerwacje_stolik(
    rid: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    t, guards = _zablokuj_termin(db, rid)
    if t and t.rodzaj == "stolik":
        okno = (t.data, t.godz_od, _koniec_okna(db, t)) if t.godz_od else None
        before = reservation_audit.reservation_snapshot(t)
        reservation_service.release_termin_allocation(db, t.id)
        reservation_payments.request_reservation_cancellation_settlement(
            db,
            t,
            now=utcnow_naive(),
            actor_kind="user",
            actor_user_id=user.id,
            operation_key=f"reservation-delete:{t.id}",
        )
        audit = reservation_audit.add_reservation_audit(
            db, termin=t, action="delete", actor=user, before=before,
        )
        db.flush()
        wyczysc_notatki_kontekstu_nadpisan(db, [t.id])
        # Jawne odpięcie całej historii nie polega na konfiguracji FK konkretnego silnika.
        db.query(models.ReservationAudit).filter_by(termin_id=t.id).update(
            {models.ReservationAudit.termin_id: None}, synchronize_session=False,
        )
        audit.termin_id = None
        _usun_profil_fallbacku_rezerwacji(db, t.id)
        usun_powiazane_publiczne_sekrety(db, [t.id])
        usun_outbox_przed_usunieciem_pii(db, reservation_ids=[t.id])
        db.delete(t); _commit_zapis_rezerwacji(db, guards)
        if okno:
            _bezpiecznie_po_zwolnieniu_stolu(db, *okno)


@app.post(
    "/api/rezerwacje-stolik/{rid}/wyslij-potwierdzenie",
    response_model=schemas.RezerwacjaKomunikacjaQueueOut,
    dependencies=[Depends(_wymagaj_modul_rezerwacje)],
)
def wyslij_potwierdzenie_stolik(
    rid: int,
    idempotency_key: str = Header(..., alias="Idempotency-Key"),
    confirm_resend: bool = Header(False, alias="X-Confirm-Resend"),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Kolejkuje nowe potwierdzenie bez wykonywania sieci w żądaniu HTTP."""
    t, _guards = _zablokuj_termin(db, rid)
    if not t or t.rodzaj != "stolik":
        raise HTTPException(404, "Brak rezerwacji.")
    if t.status not in REZ_AKTYWNE:
        raise HTTPException(409, "Nie można wysłać potwierdzenia zakończonej rezerwacji.")
    if not idempotency_key.strip():
        raise HTTPException(400, "Niepoprawny nagłówek Idempotency-Key.")
    now = utcnow_naive()
    idem = reservation_service.begin_idempotency(
        db,
        operation="reservation.confirmation.manual:v1",
        raw_key=idempotency_key,
        payload={
            "reservation_id": rid,
            "confirm_resend": bool(confirm_resend),
        },
        secret=SECRET_KEY,
        now=now,
    )
    if idem.replayed:
        db.rollback()
        return JSONResponse(idem.response, headers={"Cache-Control": "no-store"})

    current_state = reservation_communication.current_confirmation_state(db, rid)
    if current_state in {"queued", "processing", "retry"}:
        db.rollback()
        raise reservation_service.ReservationError(
            409,
            "COMMUNICATION_ALREADY_PENDING",
            "Potwierdzenie jest już w kolejce albo trwa jego wysyłanie.",
            rule="communication",
        )
    if current_state == "uncertain":
        db.rollback()
        raise reservation_service.ReservationError(
            409,
            "COMMUNICATION_RECONCILIATION_REQUIRED",
            "Najpierw uzgodnij niepewny wynik poprzedniej wysyłki.",
            rule="communication",
        )
    if current_state == "failed":
        db.rollback()
        raise reservation_service.ReservationError(
            409,
            "COMMUNICATION_RETRY_REQUIRED",
            "Poprzednia wysyłka nie powiodła się. Użyj akcji Ponów przy wiadomości.",
            rule="communication",
        )
    if current_state == "sent" and not confirm_resend:
        db.rollback()
        raise reservation_service.ReservationError(
            409,
            "COMMUNICATION_RESEND_CONFIRMATION_REQUIRED",
            "Ponowna wysyłka wymaga jawnego potwierdzenia operatora.",
            rule="communication",
        )
    dedupe_key = f"reservation:{rid}:confirmation:manual:{idem.record.key_hash}"
    messages = reservation_communication.enqueue_reservation(
        db, t, "confirmation", dedupe_key=dedupe_key, actor=user,
    )
    if not messages:
        db.rollback()
        raise HTTPException(
            400,
            "Brak dostępnego kanału kontaktu albo komunikacja operacyjna jest wyłączona.",
        )
    # Together with actor_user_id this is the durable, PII-free audit of the
    # operator action.  The resend variant can only be produced after the
    # explicit acknowledgement checked above.
    template_key = (
        "confirmation_manual_resend"
        if current_state == "sent"
        else "confirmation_manual_initial"
    )
    for message in messages:
        message.template_key = template_key
    db.flush()
    response_body = {
        "queued": len(messages),
        "messages": [reservation_communication.message_dict(row) for row in messages],
    }
    reservation_service.complete_idempotency(
        idem.record,
        response=response_body,
        http_status=200,
        termin_id=t.id,
        now=now,
    )
    db.commit()
    return JSONResponse(response_body, headers={"Cache-Control": "no-store"})


@app.get(
    "/api/rezerwacje-stolik/{rid}/komunikacja",
    response_model=schemas.RezerwacjaKomunikacjaHistoriaOut,
    dependencies=[Depends(_wymagaj_modul_rezerwacje)],
)
def historia_komunikacji_rezerwacji(
    rid: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    if not _ma_dostep_rezerwacji(user, "rezerwacje.dane_kontaktowe"):
        raise HTTPException(403, "Brak uprawnienia do danych kontaktowych gości.")
    reservation = db.get(models.Termin, rid)
    if reservation is None or reservation.rodzaj != "stolik":
        raise HTTPException(404, "Brak rezerwacji.")
    summaries = reservation_communication.summaries_for_reservations(db, [rid])
    manual_state = reservation_communication.current_confirmation_state(db, rid)
    return JSONResponse(
        {
            "reservation_id": rid,
            "summary": summaries.get(rid),
            "manual_confirmation_state": manual_state,
            "manual_confirmation_resend_required": manual_state == "sent",
            "messages": reservation_communication.reservation_history(db, rid),
        },
        headers={"Cache-Control": "no-store"},
    )


def _wymagaj_dostepu_do_wiadomosci(user, message) -> None:
    if message.termin_id is not None and message.waitlist_id is None:
        owner_kind = "reservation"
    elif message.waitlist_id is not None and message.termin_id is None:
        owner_kind = "waitlist"
    else:
        raise HTTPException(404, "Brak wiadomości.")
    requirement = reservation_access.communication_owner_requirement(owner_kind)
    if not reservation_access.user_satisfies(user, requirement):
        raise HTTPException(403, "Brak uprawnień.")


@app.post(
    "/api/rezerwacje/komunikacja/{message_id}/retry",
    response_model=schemas.RezerwacjaWiadomoscOut,
    dependencies=[Depends(_wymagaj_modul_rezerwacje)],
)
def ponow_komunikacje_rezerwacji(
    message_id: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    message = reservation_communication.lock_message(db, message_id)
    if message is None:
        raise HTTPException(404, "Brak wiadomości.")
    _wymagaj_dostepu_do_wiadomosci(user, message)
    teraz = utcnow_naive()
    try:
        reservation_communication.retry_failed(
            db, message, actor=user, now=teraz,
        )
    except ValueError as exc:
        code = str(exc)
        if code == "UNCERTAIN_REQUIRES_RECONCILIATION":
            raise HTTPException(
                409,
                "Wynik jest niepewny. Najpierw uzgodnij go, aby uniknąć duplikatu.",
            ) from exc
        if code == "MESSAGE_EXPIRED":
            raise HTTPException(
                409, "Termin ważności wiadomości minął; nie można jej już wysłać.",
            ) from exc
        if code in {
            "MESSAGE_SUPERSEDED", "MESSAGE_OWNER_MISSING", "MESSAGE_OWNER_NOT_CURRENT",
        }:
            raise reservation_service.ReservationError(
                409,
                code,
                "Wiadomość nie dotyczy już bieżącego stanu rezerwacji.",
                rule="communication",
            ) from exc
        raise HTTPException(409, "Tylko niedostarczoną wiadomość można ponowić.") from exc
    db.add(models.AuditLog(
        ts=teraz,
        user_id=user.id,
        login=user.login,
        akcja="rezerwacje_komunikacja_retry",
        zasob=f"message:{message.id}",
    ))
    db.commit(); db.refresh(message)
    return JSONResponse(
        reservation_communication.message_dict(message),
        headers={"Cache-Control": "no-store"},
    )


@app.post(
    "/api/rezerwacje/komunikacja/{message_id}/reconcile",
    response_model=schemas.RezerwacjaWiadomoscOut,
    dependencies=[Depends(_wymagaj_modul_rezerwacje)],
)
def uzgodnij_komunikacje_rezerwacji(
    message_id: int,
    dane: schemas.RezerwacjaWiadomoscReconcileIn,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    message = reservation_communication.lock_message(db, message_id)
    if message is None:
        raise HTTPException(404, "Brak wiadomości.")
    _wymagaj_dostepu_do_wiadomosci(user, message)
    teraz = utcnow_naive()
    previous_error_code = message.last_error_code
    stale_delivery_acknowledged = bool(
        dane.wynik == "sent"
        and message.waitlist_id is not None
        and message.sent_at is not None
        and previous_error_code
        == reservation_communication.WAITLIST_STALE_DELIVERED_CODE
    )
    try:
        reservation_communication.reconcile_uncertain(
            db,
            message,
            outcome=dane.wynik,
            note=dane.notatka,
            actor=user,
            now=teraz,
        )
    except ValueError as exc:
        code = str(exc)
        if code == "MESSAGE_EXPIRED":
            raise HTTPException(
                409, "Termin ważności wiadomości minął; nie można jej już wysłać.",
            ) from exc
        if code == "STALE_DELIVERY_REQUIRES_ACKNOWLEDGEMENT":
            raise HTTPException(
                409,
                "Ta wiadomość została już dostarczona po zmianie oferty. "
                "Możesz tylko potwierdzić zdarzenie i zamknąć alert.",
            ) from exc
        if code in {
            "MESSAGE_SUPERSEDED", "MESSAGE_OWNER_MISSING", "MESSAGE_OWNER_NOT_CURRENT",
        }:
            raise reservation_service.ReservationError(
                409,
                code,
                "Wiadomość nie dotyczy już bieżącego stanu rezerwacji.",
                rule="communication",
            ) from exc
        raise HTTPException(
            409, "Tylko wiadomość z niepewnym wynikiem można uzgodnić.",
        ) from exc
    db.add(models.AuditLog(
        ts=teraz,
        user_id=user.id,
        login=user.login,
        akcja="rezerwacje_komunikacja_reconcile",
        zasob=f"message:{message.id}",
        szczegoly=reservation_service.canonical_json({
            "outcome": dane.wynik,
            "previous_error_code": previous_error_code,
            "stale_delivery_acknowledged": stale_delivery_acknowledged,
        }),
    ))
    db.commit(); db.refresh(message)
    return JSONResponse(
        reservation_communication.message_dict(message),
        headers={"Cache-Control": "no-store"},
    )


# ── Lista oczekujących (waitlist) ────────────────────────────────────────────
def _stoliki_zadania(stolik_id=None, stoliki=None) -> list[int]:
    """Normalizuje kompatybilny pojedynczy stół i nowy zestaw wielostołowy."""
    raw = []
    if stolik_id is not None:
        raw.append(stolik_id)
    raw.extend(stoliki or [])
    result = []
    for value in raw:
        try:
            table_id = int(value)
        except (TypeError, ValueError) as exc:
            raise HTTPException(400, "Nieprawidłowy stolik w zestawie.") from exc
        if table_id <= 0:
            raise HTTPException(400, "Nieprawidłowy stolik w zestawie.")
        if table_id not in result:
            result.append(table_id)
    return result


def _kandydat_zestawu(result, table_ids):
    wanted = set(table_ids)
    return next(
        (
            candidate for candidate in result.candidates
            if set(candidate.table_ids) == wanted
        ),
        None,
    )


def _wyczysc_hold_waitlisty(w):
    w.hold_stolik_id = None
    w.hold_stoliki_dodatkowe = None
    w.hold_godz_od = None
    w.hold_godz_do = None
    w.hold_bufor_min = None
    w.hold_do = None
    w.offer_auto_przydzielony = None
    w.offer_override_authorized = None
    w.offer_override_note = None
    w.offer_sala_id = None
    w.offer_kanal = None


def _audit_waitlisty(db, w, action, *, user=None, now=None, details=None):
    """PII-free lifecycle audit; raw idempotency keys never reach this record."""
    db.add(models.AuditLog(
        ts=now or utcnow_naive(),
        user_id=getattr(user, "id", None),
        login=getattr(user, "login", None),
        akcja=f"waitlist_{action}",
        zasob=f"waitlist:{w.id}",
        szczegoly=reservation_service.canonical_json(details or {}),
    ))


_WAITLIST_OVERRIDE_REASON_CODES = {
    "guest_request",
    "large_group_confirmed",
    "event_exception",
    "operational_decision",
    "walk_in",
    "other",
    "legacy_confirmation",
}
_WAITLIST_OVERRIDE_AUDIT_REASONS = {
    "capacity_override", "pacing_override", "other",
}
_WAITLIST_GRANDFATHER_RULES = frozenset({
    "pacing_reservations",
    "pacing_covers",
    "concurrent_reservations",
    "concurrent_covers",
})


def _waitlist_override_audit_payload(evaluation, override_context):
    if not override_context or evaluation.decision != "override_required":
        return None
    details = _szczegoly_nadpisania_r3(
        evaluation,
        # Każda nowa oferta zapisuje pełną, typowaną tożsamość naruszeń.
        legacy=False,
    )
    if not details:
        raise RuntimeError("WAITLIST_OVERRIDE_AUDIT_INCOMPLETE")
    details = dict(details)
    details["operator_reason_code"] = override_context["reason_code"]
    return {
        "authorized": True,
        "reason_code": override_context["reason_code"],
        "audit_reason": _powod_audytu_nadpisania(evaluation),
        "details": details,
    }


def _frozen_waitlist_override_context(db, w, offer_version):
    if not w.offer_override_authorized:
        return None
    row = db.query(models.AuditLog).filter_by(
        akcja="waitlist_offered",
        zasob=f"waitlist:{w.id}",
    ).order_by(models.AuditLog.id.desc()).first()
    encrypted_context = db.query(
        models.WaitlistOfferOverrideContext,
    ).filter_by(
        waitlist_id=w.id,
        offer_version=int(offer_version),
    ).one_or_none()
    try:
        payload = json.loads(row.szczegoly) if row is not None else None
        frozen = payload.get("override") if isinstance(payload, dict) else None
        reason_code = frozen.get("reason_code")
        audit_reason = frozen.get("audit_reason")
        details = frozen.get("details")
        if (
            int(payload.get("offer_version")) != int(offer_version)
            or frozen.get("authorized") is not True
            or reason_code not in _WAITLIST_OVERRIDE_REASON_CODES
            or audit_reason not in _WAITLIST_OVERRIDE_AUDIT_REASONS
            or not isinstance(details, dict)
            or encrypted_context is None
            or encrypted_context.reason_code != reason_code
        ):
            raise ValueError("invalid frozen override")
        # Reuse the strict PII-free reservation-audit validator before any hold
        # is released. Free-form text stays in the encrypted per-generation row.
        reservation_audit._normalise_override_details(details)
    except (AttributeError, TypeError, ValueError, json.JSONDecodeError) as exc:
        raise reservation_service.ReservationError(
            409,
            "WAITLIST_OFFER_OVERRIDE_AUDIT_INCOMPLETE",
            "Nie można potwierdzić audytu autoryzacji tej oferty.",
            rule="waitlist_offer",
        ) from exc
    return {
        "reason_code": reason_code,
        "note": encrypted_context.note,
        "audit_reason": audit_reason,
        "details": details,
    }


def _waitlist_accept_enforcement(evaluation, frozen_override_context):
    """Authorize only the frozen promise, never a fresh arbitrary override.

    Capacity/pacing violations are grandfathered because the complete active
    offer was already counted as a synthetic owner. Other overrideable rules
    must match the typed rule/code captured when that offer was published.
    """
    violations = tuple(evaluation.violations or ())
    if not violations:
        return False, ()
    if any(not item.overrideable_by_operator for item in violations):
        raise reservation_rules.evaluation_to_reservation_error(evaluation)

    grandfathered = tuple(sorted({
        item.rule for item in violations
        if item.rule in _WAITLIST_GRANDFATHER_RULES
    }))
    frozen_rows = (
        frozen_override_context.get("details", {}).get("violations", [])
        if frozen_override_context else []
    )
    for violation in violations:
        if violation.rule in _WAITLIST_GRANDFATHER_RULES:
            continue
        expected_identity = reservation_service.canonical_json({
            "rule": violation.rule,
            "code": violation.code,
            "scope": dict(violation.scope or {}),
            "source": dict(violation.source or {}),
        })
        if not any(
            isinstance(row, dict)
            and reservation_service.canonical_json({
                "rule": row.get("rule"),
                "code": row.get("code"),
                "scope": dict(row.get("scope") or {}),
                "source": dict(row.get("source") or {}),
            }) == expected_identity
            for row in frozen_rows
        ):
            raise reservation_rules.evaluation_to_reservation_error(evaluation)
    return True, grandfathered


def _wiadomosci_biezacej_oferty(db, w):
    current_dedupe = reservation_communication.current_waitlist_offer_dedupe(w)
    if current_dedupe is None:
        return []
    latest_query = db.query(models.RezerwacjaWiadomoscOutbox).filter_by(
        waitlist_id=w.id,
        typ_zdarzenia="table_ready",
        dedupe_key=current_dedupe,
    )
    latest = latest_query.order_by(
        models.RezerwacjaWiadomoscOutbox.id.desc(),
    ).first()
    if latest is None:
        return []
    return db.query(models.RezerwacjaWiadomoscOutbox).filter_by(
        waitlist_id=w.id,
        typ_zdarzenia="table_ready",
        dedupe_key=latest.dedupe_key,
    ).order_by(models.RezerwacjaWiadomoscOutbox.id).all()


def _odpowiedz_oferty(db, w, user):
    messages = _wiadomosci_biezacej_oferty(db, w)
    summaries = reservation_communication.summaries_for_waitlists(db, [w.id])
    visible_messages = (
        messages
        if _ma_dostep_rezerwacji(user, "rezerwacje.dane_kontaktowe")
        else []
    )
    return {
        "queued": any(
            row.stan in {"queued", "processing", "retry"} for row in messages
        ),
        "messages": [
            reservation_communication.message_dict(row)
            for row in visible_messages
        ],
        "wpis": _lista_out(w, user, summaries.get(w.id)),
    }


def _wymagaj_wersji_oferty(w, expected_version):
    current = int(w.offer_version or 0)
    if int(expected_version) != current:
        raise reservation_service.ReservationError(
            409,
            "WAITLIST_OFFER_VERSION_CONFLICT",
            f"Oferta zmieniła się w innym widoku. Bieżąca wersja: {current}.",
            rule="waitlist_offer",
        )


def _odswiez_wygasniecie_waitlisty(db, w, guards, *, teraz):
    was_offered = w.status == "zaoferowano"
    reservation_service.cleanup_expired_holds(
        db, teraz, dates=[w.data],
    )
    db.flush()
    if was_offered and w.status == "wygasla":
        _commit_zapis_rezerwacji(db, guards)
        raise reservation_service.ReservationError(
            409,
            "WAITLIST_OFFER_EXPIRED",
            "Oferta wygasła. Stoły zostały bezpiecznie zwolnione.",
            rule="waitlist_offer",
        )


def _czy_kompletny_hold_oferty(db, w, *, teraz):
    if (
        w.status != "zaoferowano"
        or w.hold_stolik_id is None
        or w.hold_godz_od is None
        or w.hold_godz_do is None
        or w.hold_do is None
        or w.hold_do <= teraz
    ):
        return False
    held_ids = _stoliki_zadania(w.hold_stolik_id, w.hold_stoliki_dodatkowe)
    if not held_ids:
        return False
    try:
        start_minute, blocked_end = reservation_service.claim_minute_window(
            w.hold_godz_od, w.hold_godz_do, int(w.hold_bufor_min or 0),
        )
    except reservation_service.ReservationError:
        return False
    expected_slots = {
        (table_id, minute)
        for table_id in held_ids
        for minute in range(start_minute, blocked_end)
    }
    claims = db.query(
        models.RezerwacjaStolikClaim.stolik_id,
        models.RezerwacjaStolikClaim.minute,
        models.RezerwacjaStolikClaim.data,
        models.RezerwacjaStolikClaim.expires_at,
    ).filter_by(waitlist_id=w.id).all()
    return (
        len(claims) == len(expected_slots)
        and {(row.stolik_id, row.minute) for row in claims} == expected_slots
        and all(
            row.data == w.data
            and row.expires_at is not None
            and row.expires_at > teraz
            and row.expires_at == w.hold_do
            for row in claims
        )
    )


def _zaloz_hold_waitlisty(
    db, w, dane, *, teraz, request, user, override_context,
):
    """Select and claim one exact allocator-approved table configuration."""
    godz = dane.godz_od or w.godz_od
    if godz is None:
        raise HTTPException(400, "Ustaw godzinę, aby utworzyć czasową ofertę.")
    existing_ids = []
    if w.hold_do is not None and w.hold_do > teraz and w.hold_stolik_id is not None:
        existing_ids = _stoliki_zadania(
            w.hold_stolik_id, w.hold_stoliki_dodatkowe,
        )
    requested_ids = _stoliki_zadania(dane.stolik_id, dane.stoliki) or existing_ids
    runtime_table_ids = {table["id"] for table in _stoly_do_seating(db)}
    if requested_ids and not set(requested_ids) <= runtime_table_ids:
        raise HTTPException(400, "Nieznany lub nieaktywny stolik.")

    reservation_service.release_waitlist_hold(db, w.id)
    _wyczysc_hold_waitlisty(w)
    db.flush()
    target_channel = "online" if w.kanal == "online" else "wewnetrzna"
    result = _ocen_przydzial_rezerwacji(
        db,
        data=w.data,
        godz_od=godz,
        osoby=w.liczba_osob or 1,
        kanal=target_channel,
        intent="quote",
        alternative_limit=9999,
        now=teraz,
    )
    candidate = (
        _kandydat_zestawu(result, requested_ids)
        if requested_ids else result.selected
    )
    if candidate is None:
        _wymagaj_dozwolonego_przydzialu(
            result,
            override=bool(
                override_context
                and result.evaluation.decision == "override_required"
            ),
        )
        raise reservation_service.ReservationError(
            409,
            "INVALID_TABLE_COMBINATION",
            "Wybrany zestaw nie jest dostępną, zatwierdzoną konfiguracją.",
            rule="table_hold",
        )
    table_ids = list(candidate.table_ids)
    locked_tables = reservation_service.lock_tables(db, table_ids)
    if len(locked_tables) != len(table_ids):
        raise HTTPException(400, "Nieznany lub nieaktywny stolik.")
    # Publikacja planu blokuje te same rekordy Stolik. Po zdobyciu locków
    # powtarzamy pełną ocenę, aby nie materializować wariantu ze starej wersji.
    revalidated_result = _ocen_przydzial_rezerwacji(
        db,
        data=w.data,
        godz_od=godz,
        osoby=w.liczba_osob or 1,
        kanal=target_channel,
        intent="quote",
        alternative_limit=9999,
        now=teraz,
    )
    revalidated_candidate = _kandydat_zestawu(
        revalidated_result, table_ids,
    )
    if revalidated_candidate is None:
        raise reservation_service.ReservationError(
            409,
            "WAITLIST_OFFER_PLAN_CHANGED",
            "Układ sali zmienił się podczas tworzenia oferty. Odśwież propozycje.",
            rule="table_hold",
        )
    result = revalidated_result
    candidate = revalidated_candidate
    evaluation = _ocen_reguly_slotu(
        db,
        data=w.data,
        godz_od=godz,
        liczba_osob=w.liczba_osob or 1,
        kanal=target_channel,
        sala_id=candidate.room_id,
        intent="create",
        now=teraz,
    )
    override_active = bool(
        override_context and evaluation.decision == "override_required"
    )
    _wymagaj_reautoryzacji_nadpisania(
        db, request, user, override_active,
    )
    reservation_rules.enforce_rule_evaluation(
        evaluation,
        override=override_active,
        can_override=override_active,
    )
    godz_do = evaluation.godz_do
    if godz_do is None:
        raise HTTPException(400, "Nie udało się wyznaczyć końca wizyty.")
    hold_do = teraz + timedelta(minutes=max(1, int(dane.minuty or 15)))
    reservation_service.replace_waitlist_hold(
        db,
        waitlist_id=w.id,
        table_ids=table_ids,
        data=w.data,
        expires_at=hold_do,
        now=teraz,
        start=godz,
        end=godz_do,
        buffer_min=evaluation.buffer_min,
        cleanup_holds=False,
    )
    w.godz_od = godz
    w.hold_stolik_id = table_ids[0]
    w.hold_stoliki_dodatkowe = table_ids[1:] or None
    w.hold_godz_od = godz
    w.hold_godz_do = godz_do
    w.hold_bufor_min = evaluation.buffer_min
    w.hold_do = hold_do
    w.offer_sala_id = candidate.room_id
    w.offer_kanal = target_channel
    recommended_ids = (
        set(result.selected.table_ids) if result.selected is not None else set()
    )
    offer_auto_przydzielony = bool(
        recommended_ids and set(table_ids) == recommended_ids
    )
    return (
        table_ids,
        hold_do,
        offer_auto_przydzielony,
        override_active,
        evaluation,
    )


def _lista_out(w: models.ListaOczekujacych, user=None, communication_summary=None) -> dict:
    kontakt = _ma_dostep_rezerwacji(user, "rezerwacje.dane_kontaktowe")
    notatki = _ma_dostep_rezerwacji(user, "rezerwacje.notatki_wewnetrzne")
    summary = dict(communication_summary) if communication_summary else None
    if summary is not None and not kontakt:
        summary["channel"] = None
    ukryte = []
    if not kontakt:
        ukryte.extend(("nazwisko", "telefon", "email"))
    if not notatki:
        ukryte.append("notatka")
    return {
        "id": w.id,
        "data": str(w.data),
        "godz_od": _hm(w.godz_od),
        "liczba_osob": w.liczba_osob,
        "nazwisko": w.nazwisko if kontakt else "Gość",
        "telefon": w.telefon if kontakt else None,
        "email": w.email if kontakt else None,
        "kanal_komunikacji": w.kanal_komunikacji if kontakt else None,
        "notatka": w.notatka if notatki else None,
        "status": w.status,
        "priorytet": int(w.priorytet or 0),
        "offer_version": int(w.offer_version or 0),
        "offer_auto_przydzielony": w.offer_auto_przydzielony,
        "offer_override_authorized": w.offer_override_authorized,
        "zaoferowano_at": w.zaoferowano_at.isoformat() if w.zaoferowano_at else None,
        "oferta_wygasa_at": (
            w.oferta_wygasa_at.isoformat() if w.oferta_wygasa_at else None
        ),
        "zaakceptowano_at": (
            w.zaakceptowano_at.isoformat() if w.zaakceptowano_at else None
        ),
        "wygasla_at": w.wygasla_at.isoformat() if w.wygasla_at else None,
        "anulowano_at": w.anulowano_at.isoformat() if w.anulowano_at else None,
        "termin_id": w.termin_id,
        "kanal": w.kanal,
        "powiadomiono_at": w.powiadomiono_at.isoformat() if w.powiadomiono_at else None,
        "communication_summary": summary,
        "hold_stolik_id": w.hold_stolik_id,
        "hold_stoliki_dodatkowe": list(w.hold_stoliki_dodatkowe or []),
        "hold_godz_od": _hm(w.hold_godz_od),
        "hold_godz_do": _hm(w.hold_godz_do),
        "hold_bufor_min": w.hold_bufor_min,
        "hold_do": w.hold_do.isoformat() if w.hold_do else None,
        "ukryte_pola": ukryte,
    }


@app.get("/api/lista-oczekujacych", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def get_lista_oczekujacych(
    data: date = Query(...),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    rows = (db.query(models.ListaOczekujacych)
            .filter(models.ListaOczekujacych.data == data)
            .order_by(
                models.ListaOczekujacych.status,
                models.ListaOczekujacych.priorytet.desc(),
                models.ListaOczekujacych.utworzono_at,
                models.ListaOczekujacych.id,
            ).all())
    summaries = reservation_communication.summaries_for_waitlists(
        db, (row.id for row in rows),
    )
    return {
        "lista": [
            _lista_out(row, user, summaries.get(row.id)) for row in rows
        ],
    }


@app.post("/api/lista-oczekujacych", status_code=201, dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def dodaj_lista_oczekujacych(
    dane: schemas.ListaOczekujacychIn,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    if not dane.nazwisko or not dane.nazwisko.strip():
        raise HTTPException(400, "Podaj nazwisko / klienta.")
    # Leniwy singleton konfiguracji może commitować; musi powstać przed blokadą.
    get_lokal_config(db)
    guards = reservation_service.begin_locked_write(db, [dane.data])
    classification = _klasyfikacja_popytu_rezerwacji(
        db,
        data=dane.data,
        godz_od=dane.godz_od,
        osoby=dane.liczba_osob,
        kanal="wewnetrzna",
    )
    now = utcnow_naive()
    w = models.ListaOczekujacych(
        data=dane.data, godz_od=dane.godz_od, liczba_osob=dane.liczba_osob,
        nazwisko=dane.nazwisko.strip(), telefon=dane.telefon, email=dane.email,
        kanal_komunikacji=dane.kanal_komunikacji,
        notatka=(dane.notatka if _ma_dostep_rezerwacji(
            user, "rezerwacje.notatki_wewnetrzne") else None),
        status="oczekuje", utworzono_at=now,
        demand_reason_code=classification.reason_code,
        demand_resource_kind=classification.resource_kind,
    )
    db.add(w)
    db.flush()
    if dane.liczba_osob is not None:
        reservation_demand.record_internal_waitlist_event(
            db,
            requested_date=dane.data,
            requested_time=dane.godz_od,
            party_size=dane.liczba_osob,
            classification=classification,
            secret=SECRET_KEY,
            captured_at=now,
        )
    _commit_zapis_rezerwacji(db, guards)
    db.refresh(w)
    return _lista_out(w, user)


@app.delete("/api/lista-oczekujacych/{wid}", status_code=204, dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def usun_lista_oczekujacych(
    wid: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    w, guards = _zablokuj_waitliste(db, wid)
    if w:
        reservation_service.release_waitlist_hold(db, w.id)
        usun_outbox_przed_usunieciem_pii(db, waitlist_ids=[w.id])
        db.delete(w); _commit_zapis_rezerwacji(db, guards)


def _anuluj_lista_oczekujacych(
    wid, *, expected_version, legacy_plain_only=False, db, user,
):
    w, guards = _zablokuj_waitliste(db, wid)
    if not w:
        raise HTTPException(404, "Brak wpisu.")
    if legacy_plain_only and w.status != "oczekuje":
        db.rollback()
        raise reservation_service.ReservationError(
            409,
            "WAITLIST_LEGACY_ALIAS_NOT_ALLOWED",
            "Ta operacja dotyczy wyłącznie wpisu bez aktywnej oferty.",
            rule="waitlist_offer",
        )
    teraz = utcnow_naive()
    _odswiez_wygasniecie_waitlisty(db, w, guards, teraz=teraz)
    if expected_version is not None:
        _wymagaj_wersji_oferty(w, expected_version)
    if w.status == "anulowano":
        db.rollback()
        return _lista_out(w, user)
    if w.status in {"zaakceptowano", "wygasla"}:
        db.rollback()
        raise reservation_service.ReservationError(
            409,
            "WAITLIST_TERMINAL_STATE",
            "Zakończonego wpisu waitlisty nie można anulować ponownie.",
            rule="waitlist_offer",
        )
    previous_version = int(w.offer_version or 0)
    reservation_service.release_waitlist_hold(db, w.id)
    _wyczysc_hold_waitlisty(w)
    w.status = "anulowano"
    w.anulowano_at = teraz
    w.offer_version = previous_version + 1
    reservation_communication.cancel_waitlist_pending(db, w.id)
    _audit_waitlisty(
        db, w, "cancelled", user=user, now=teraz,
        details={
            "offer_version": previous_version,
            "next_offer_version": w.offer_version,
        },
    )
    _commit_zapis_rezerwacji(db, guards)
    db.refresh(w)
    return _lista_out(w, user)


@app.post("/api/lista-oczekujacych/{wid}/anuluj", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def anuluj_lista_oczekujacych(
    wid: int,
    dane: schemas.WaitlistAnulujIn,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    return _anuluj_lista_oczekujacych(
        wid,
        expected_version=dane.expected_offer_version,
        db=db,
        user=user,
    )


@app.post("/api/lista-oczekujacych/{wid}/odwolaj", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def odwolaj_lista_oczekujacych(
    wid: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Legacy alias. New clients use /anuluj with optimistic version."""
    return _anuluj_lista_oczekujacych(
        wid,
        expected_version=None,
        legacy_plain_only=True,
        db=db,
        user=user,
    )


@app.post(
    "/api/lista-oczekujacych/{wid}/oferta",
    response_model=schemas.WaitlistOfertaOut,
    dependencies=[Depends(_wymagaj_modul_rezerwacje)],
)
def zaoferuj_lista_oczekujacych(
    wid: int,
    dane: schemas.WaitlistOfertaIn,
    request: Request,
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Atomically freeze one approved table set and optionally notify the guest."""
    override_context = _jawne_nadpisanie_limitow(
        user, dane.przekrocz_limity, dane.nadpisanie_limitow,
    )
    identity = reservation_service.required_idempotency_identity(
        operation="waitlist.offer:v1",
        raw_key=idempotency_key,
        payload={"waitlist_id": wid, "request": dane.model_dump(mode="json")},
        secret=SECRET_KEY,
    )
    w, guards = _zablokuj_waitliste(db, wid)
    if not w:
        raise HTTPException(404, "Brak wpisu.")
    teraz = utcnow_naive()
    _odswiez_wygasniecie_waitlisty(db, w, guards, teraz=teraz)

    if w.offer_key_hash == identity.key_hash:
        if not w.offer_request_fingerprint or not secrets.compare_digest(
            w.offer_request_fingerprint, identity.request_fingerprint,
        ):
            db.rollback()
            raise reservation_service.ReservationError(
                409,
                "IDEMPOTENCY_KEY_REUSED",
                "Ten klucz Idempotency-Key został użyty z innymi danymi.",
                rule="idempotency",
            )
        if w.status == "zaoferowano" and w.hold_do and w.hold_do > teraz:
            response = _odpowiedz_oferty(db, w, user)
            db.rollback()
            return response
        db.rollback()
        raise reservation_service.ReservationError(
            409,
            "IDEMPOTENCY_KEY_REUSED",
            "Ta generacja oferty została już zakończona.",
            rule="idempotency",
        )

    _wymagaj_wersji_oferty(w, dane.expected_offer_version)
    if w.status != "oczekuje":
        db.rollback()
        raise reservation_service.ReservationError(
            409,
            "WAITLIST_OFFER_NOT_AVAILABLE",
            "Ofertę można utworzyć wyłącznie dla oczekującego wpisu.",
            rule="waitlist_offer",
        )
    ambiguous_delivery = db.query(models.RezerwacjaWiadomoscOutbox.id).filter(
        models.RezerwacjaWiadomoscOutbox.waitlist_id == w.id,
        models.RezerwacjaWiadomoscOutbox.typ_zdarzenia == "table_ready",
        models.RezerwacjaWiadomoscOutbox.stan.in_(("processing", "uncertain")),
    ).first()
    if ambiguous_delivery is not None:
        db.rollback()
        raise reservation_service.ReservationError(
            409,
            "WAITLIST_DELIVERY_RECONCILIATION_REQUIRED",
            "Poprzednie powiadomienie wymaga zakończenia lub uzgodnienia przed nową ofertą.",
            rule="waitlist_offer",
        )

    (
        table_ids,
        deadline,
        offer_auto_przydzielony,
        offer_override_authorized,
        offer_evaluation,
    ) = _zaloz_hold_waitlisty(
        db,
        w,
        dane,
        teraz=teraz,
        request=request,
        user=user,
        override_context=override_context,
    )
    previous_version = int(w.offer_version or 0)
    w.status = "zaoferowano"
    w.offer_version = previous_version + 1
    w.offer_auto_przydzielony = offer_auto_przydzielony
    w.offer_override_authorized = offer_override_authorized
    w.offer_override_note = (
        override_context.get("note")
        if offer_override_authorized and override_context
        else None
    )
    w.offer_key_hash = identity.key_hash
    w.offer_request_fingerprint = identity.request_fingerprint
    w.zaoferowano_at = teraz
    w.oferta_wygasa_at = deadline
    w.powiadomiono_at = None
    db.flush()

    override_audit_payload = _waitlist_override_audit_payload(
        offer_evaluation, override_context,
    )
    if offer_override_authorized:
        if not override_context or override_audit_payload is None:
            raise RuntimeError("WAITLIST_OVERRIDE_CONTEXT_INCOMPLETE")
        db.add(models.WaitlistOfferOverrideContext(
            waitlist_id=w.id,
            offer_version=w.offer_version,
            reason_code=override_context["reason_code"],
            note=override_context.get("note"),
            created_at=teraz,
        ))

    messages = []
    if _ma_dostep_rezerwacji(user, "rezerwacje.dane_kontaktowe"):
        messages = reservation_communication.enqueue_table_ready(
            db,
            w,
            dedupe_key=(
                f"waitlist:{w.id}:offer:{w.offer_version}:{identity.key_hash}"
            ),
            actor=user,
            now=teraz,
        )
    _audit_waitlisty(
        db,
        w,
        "offered",
        user=user,
        now=teraz,
        details={
            "offer_version": w.offer_version,
            "previous_offer_version": previous_version,
            "table_ids": table_ids,
            "auto_przydzielony": offer_auto_przydzielony,
            "override_authorized": offer_override_authorized,
            "override": override_audit_payload,
            "deadline": deadline.isoformat(),
            "queued_channels": sorted({row.kanal for row in messages}),
        },
    )
    _commit_zapis_rezerwacji(db, guards)
    db.refresh(w)
    return _odpowiedz_oferty(db, w, user)


@app.post(
    "/api/lista-oczekujacych/{wid}/wycofaj-oferte",
    dependencies=[Depends(_wymagaj_modul_rezerwacje)],
)
def wycofaj_oferte_lista_oczekujacych(
    wid: int,
    dane: schemas.WaitlistOfferVersionIn,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    w, guards = _zablokuj_waitliste(db, wid)
    if not w:
        raise HTTPException(404, "Brak wpisu.")
    teraz = utcnow_naive()
    _odswiez_wygasniecie_waitlisty(db, w, guards, teraz=teraz)
    _wymagaj_wersji_oferty(w, dane.offer_version)
    if w.status != "zaoferowano":
        db.rollback()
        raise reservation_service.ReservationError(
            409,
            "WAITLIST_OFFER_NOT_ACTIVE",
            "Nie ma aktywnej oferty do wycofania.",
            rule="waitlist_offer",
        )
    previous_version = int(w.offer_version or 0)
    reservation_service.release_waitlist_hold(db, w.id)
    _wyczysc_hold_waitlisty(w)
    w.status = "oczekuje"
    w.offer_version = previous_version + 1
    reservation_communication.cancel_waitlist_pending(db, w.id, now=teraz)
    _audit_waitlisty(
        db,
        w,
        "withdrawn",
        user=user,
        now=teraz,
        details={
            "offer_version": previous_version,
            "next_offer_version": w.offer_version,
        },
    )
    _commit_zapis_rezerwacji(db, guards)
    db.refresh(w)
    return _lista_out(w, user)


@app.post(
    "/api/lista-oczekujacych/{wid}/priorytet",
    dependencies=[Depends(_wymagaj_modul_rezerwacje)],
)
def ustaw_priorytet_lista_oczekujacych(
    wid: int,
    dane: schemas.WaitlistPriorytetIn,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    w, guards = _zablokuj_waitliste(db, wid)
    if not w:
        raise HTTPException(404, "Brak wpisu.")
    teraz = utcnow_naive()
    _odswiez_wygasniecie_waitlisty(db, w, guards, teraz=teraz)
    _wymagaj_wersji_oferty(w, dane.expected_offer_version)
    if w.status != "oczekuje":
        db.rollback()
        raise reservation_service.ReservationError(
            409,
            "WAITLIST_PRIORITY_NOT_AVAILABLE",
            "Priorytet można zmienić tylko przed utworzeniem oferty.",
            rule="waitlist_offer",
        )
    previous_priority = int(w.priorytet or 0)
    if previous_priority == dane.priorytet:
        response = _lista_out(w, user)
        db.rollback()
        return response
    previous_version = int(w.offer_version or 0)
    w.priorytet = dane.priorytet
    w.offer_version = previous_version + 1
    _audit_waitlisty(
        db,
        w,
        "priority_changed",
        user=user,
        now=teraz,
        details={
            "previous_priority": previous_priority,
            "priority": dane.priorytet,
            "offer_version": previous_version,
            "next_offer_version": w.offer_version,
        },
    )
    _commit_zapis_rezerwacji(db, guards)
    db.refresh(w)
    return _lista_out(w, user)


def _zaakceptuj_lista_oczekujacych(
    wid: int,
    dane,
    request: Request,
    *,
    expected_version,
    idempotency_key,
    legacy_direct=False,
    db,
    user,
):
    """Convert exactly one unexpired offered table set into a reservation."""
    if legacy_direct:
        override_context = _jawne_nadpisanie_limitow(
            user, dane.przekrocz_limity, dane.nadpisanie_limitow,
        )
    else:
        if dane.przekrocz_limity or dane.nadpisanie_limitow is not None:
            raise reservation_service.ReservationError(
                409,
                "WAITLIST_ACCEPT_OVERRIDE_NOT_ALLOWED",
                "Wyjątek limitu musi być autoryzowany przy tworzeniu tej oferty.",
                rule="waitlist_offer",
            )
        override_context = None
    if not legacy_direct:
        reservation_service.required_idempotency_identity(
            operation="reservation.create.waitlist:v2",
            raw_key=idempotency_key,
            payload={
                "waitlist_id": wid,
                "expected_offer_version": expected_version,
                "request": dane.model_dump(mode="json"),
            },
            secret=SECRET_KEY,
        )
    w, guards = _zablokuj_waitliste(db, wid)
    if not w:
        raise HTTPException(404, "Brak wpisu.")
    teraz = utcnow_naive()
    if legacy_direct and w.status == "zaoferowano":
        db.rollback()
        raise reservation_service.ReservationError(
            409,
            "WAITLIST_LEGACY_ALIAS_NOT_ALLOWED",
            "Aktywną ofertę można zaakceptować tylko z jej bieżącą wersją.",
            rule="waitlist_offer",
        )
    if not legacy_direct and w.status == "zaoferowano":
        _odswiez_wygasniecie_waitlisty(db, w, guards, teraz=teraz)
        _wymagaj_wersji_oferty(w, expected_version)
    idem = reservation_service.begin_idempotency(
        db,
        operation=(
            "reservation.create.waitlist:v1"
            if legacy_direct else "reservation.create.waitlist:v2"
        ),
        raw_key=idempotency_key,
        payload=(
            {"waitlist_id": wid, **dane.model_dump(mode="json")}
            if legacy_direct else {
                "waitlist_id": wid,
                "expected_offer_version": expected_version,
                "request": dane.model_dump(mode="json"),
            }
        ),
        secret=SECRET_KEY,
        now=teraz,
    )
    if idem.replayed:
        replayed = _termin_z_replay_idempotencji(db, idem)
        current_waitlist = db.get(models.ListaOczekujacych, wid)
        if current_waitlist is None:
            raise reservation_service.ReservationError(
                409,
                "IDEMPOTENCY_TARGET_GONE",
                "Wpis listy oczekujących już nie istnieje.",
                rule="idempotency",
            )
        return {
            "rezerwacja": _rezerwacja_out(replayed, user),
            "wpis": _lista_out(current_waitlist, user),
        }
    required_status = "oczekuje" if legacy_direct else "zaoferowano"
    if w.status != required_status:
        raise reservation_service.ReservationError(
            409,
            (
                "WAITLIST_LEGACY_DIRECT_NOT_AVAILABLE"
                if legacy_direct else "WAITLIST_OFFER_NOT_ACTIVE"
            ),
            (
                "Bezpośrednio można zrealizować wyłącznie oczekujący wpis."
                if legacy_direct
                else "Akceptacja wymaga aktywnej, niewygasłej oferty."
            ),
            rule="waitlist_offer",
        )
    explicit_ids = _stoliki_zadania(dane.stolik_id, dane.stoliki)
    held_ids = []
    frozen_start = None
    frozen_end = None
    frozen_buffer = 0
    frozen_deadline = None
    frozen_offer_auto_przydzielony = None
    frozen_offer_override_authorized = None
    frozen_offer_sala_id = None
    frozen_offer_kanal = None
    frozen_override_context = None
    if legacy_direct:
        if w.hold_do is not None and w.hold_do > teraz and w.hold_stolik_id is not None:
            held_ids = _stoliki_zadania(
                w.hold_stolik_id, w.hold_stoliki_dodatkowe,
            )
        requested_ids = explicit_ids or held_ids
        godz = dane.godz_od or w.godz_od
    else:
        held_ids = _stoliki_zadania(
            w.hold_stolik_id, w.hold_stoliki_dodatkowe,
        )
        frozen_start = w.hold_godz_od
        frozen_end = w.hold_godz_do
        frozen_buffer = int(w.hold_bufor_min or 0)
        frozen_deadline = w.hold_do
        frozen_offer_auto_przydzielony = w.offer_auto_przydzielony
        frozen_offer_override_authorized = w.offer_override_authorized
        frozen_offer_sala_id = w.offer_sala_id
        frozen_offer_kanal = w.offer_kanal
        if (
            not held_ids
            or frozen_start is None
            or frozen_end is None
            or frozen_deadline is None
            or frozen_deadline <= teraz
            or w.oferta_wygasa_at != frozen_deadline
            or frozen_offer_auto_przydzielony is None
            or frozen_offer_override_authorized is None
            or frozen_offer_kanal not in {"online", "wewnetrzna"}
        ):
            raise reservation_service.ReservationError(
                409,
                "WAITLIST_OFFER_HOLD_INCOMPLETE",
                "Oferta nie ma kompletnego, aktywnego zestawu stołów.",
                rule="waitlist_offer",
            )
        if frozen_offer_override_authorized:
            frozen_override_context = _frozen_waitlist_override_context(
                db, w, expected_version,
            )
        if explicit_ids and set(explicit_ids) != set(held_ids):
            raise reservation_service.ReservationError(
                409,
                "WAITLIST_OFFER_TABLE_MISMATCH",
                "Akceptacja musi użyć dokładnie zestawu zamrożonego w ofercie.",
                rule="waitlist_offer",
            )
        if dane.godz_od is not None and dane.godz_od != frozen_start:
            raise reservation_service.ReservationError(
                409,
                "WAITLIST_OFFER_TIME_MISMATCH",
                "Akceptacja musi zachować godzinę zamrożoną w ofercie.",
                rule="waitlist_offer",
            )
        claims = db.query(models.RezerwacjaStolikClaim).filter_by(
            waitlist_id=w.id,
        ).all()
        start_minute, blocked_end = reservation_service.claim_minute_window(
            frozen_start,
            frozen_end,
            frozen_buffer,
        )
        expected_claims = len(held_ids) * max(0, blocked_end - start_minute)
        expected_claim_slots = {
            (table_id, minute)
            for table_id in held_ids
            for minute in range(start_minute, blocked_end)
        }
        if (
            len(claims) != expected_claims
            or {(row.stolik_id, row.minute) for row in claims} != expected_claim_slots
            or any(
                row.data != w.data
                or row.expires_at is None
                or row.expires_at <= teraz
                or row.expires_at != frozen_deadline
                for row in claims
            )
        ):
            raise reservation_service.ReservationError(
                409,
                "WAITLIST_OFFER_CLAIMS_INCOMPLETE",
                "Oferta nie ma kompletnego atomowego zajęcia zasobów.",
                rule="waitlist_offer",
            )
        requested_ids = held_ids
        godz = frozen_start
    reservation_service.release_waitlist_hold(db, w.id)
    _wyczysc_hold_waitlisty(w)
    db.flush()
    if godz is None:
        raise HTTPException(400, "Ustaw godzinę realizacji wpisu.")
    is_walk_in = dane.tryb == "walk_in"
    reservation_channel = (
        ("online" if frozen_offer_kanal == "online" else "reczna")
        if not legacy_direct
        else (
            "walk_in"
            if is_walk_in
            else ("online" if w.kanal == "online" else "reczna")
        )
    )
    if legacy_direct and explicit_ids:
        _waliduj_przydzial_rezerwacji(
            db,
            w.data,
            godz,
            None,
            explicit_ids,
            w.liczba_osob,
        )
    result = _ocen_przydzial_rezerwacji(
        db,
        data=w.data,
        godz_od=godz,
        osoby=w.liczba_osob or 1,
        kanal=reservation_channel,
        godz_do=(frozen_end if not legacy_direct else None),
        intent=("create" if legacy_direct else "quote"),
        preserve_explicit_interval=not legacy_direct,
        physical_buffer_min=(frozen_buffer if not legacy_direct else None),
        alternative_limit=9999,
        now=teraz,
    )
    candidate = (
        _kandydat_zestawu(result, requested_ids)
        if requested_ids else result.selected
    )
    if candidate is None:
        _wymagaj_dozwolonego_przydzialu(
            result,
            override=bool(
                (legacy_direct and override_context)
                or (not legacy_direct and frozen_offer_override_authorized)
            ),
        )
        raise reservation_service.ReservationError(
            409,
            "INVALID_TABLE_COMBINATION",
            "Wybrany zestaw nie jest dostępną, zatwierdzoną konfiguracją.",
            rule="table",
        )
    if not legacy_direct and candidate.room_id != frozen_offer_sala_id:
        raise reservation_service.ReservationError(
            409,
            "WAITLIST_OFFER_PLAN_CHANGED",
            "Układ sali zmienił się od utworzenia oferty.",
            rule="table_hold",
        )
    evaluation = _ocen_reguly_slotu(
        db,
        data=w.data,
        godz_od=godz,
        godz_do=(frozen_end if not legacy_direct else None),
        liczba_osob=w.liczba_osob or 1,
        kanal=reservation_channel,
        sala_id=candidate.room_id,
        intent="create",
        preserve_explicit_interval=not legacy_direct,
        now=teraz,
    )
    grandfathered_capacity_rules = ()
    if legacy_direct:
        enforcement_override = bool(
            override_context and evaluation.decision == "override_required"
        )
        ledger_override = enforcement_override
    else:
        enforcement_override, grandfathered_capacity_rules = (
            _waitlist_accept_enforcement(evaluation, frozen_override_context)
        )
        ledger_override = bool(frozen_offer_override_authorized)
    if legacy_direct:
        _wymagaj_reautoryzacji_nadpisania(
            db, request, user, enforcement_override,
        )
    reservation_rules.enforce_rule_evaluation(
        evaluation,
        override=enforcement_override,
        can_override=enforcement_override,
    )
    table_ids = list(candidate.table_ids)
    locked_tables = reservation_service.lock_tables(db, table_ids)
    if len(locked_tables) != len(table_ids):
        raise HTTPException(400, "Nieznany lub nieaktywny stolik.")
    godz_do = frozen_end if not legacy_direct else evaluation.godz_do
    if godz_do is None:
        raise HTTPException(400, "Nie udało się wyznaczyć końca wizyty.")
    t = models.Termin(
        data=w.data, nazwisko=w.nazwisko, telefon=w.telefon, email=w.email,
        kanal_komunikacji=w.kanal_komunikacji,
        liczba_osob=w.liczba_osob, notatka=w.notatka, status="potwierdzona", zadatek=0.0,
        utworzono_at=teraz, godz_od=godz, godz_do=godz_do, stolik_id=table_ids[0],
        stoliki_dodatkowe=(table_ids[1:] or None),
        auto_przydzielony=(
            not bool(explicit_ids)
            if legacy_direct
            else bool(frozen_offer_auto_przydzielony)
        ),
        rodzaj="stolik", kanal=reservation_channel,
        faza_hosta=("posadzony" if is_walk_in else None),
        host_arrived_at=(teraz if is_walk_in else None),
        host_seated_at=(teraz if is_walk_in else None),
        potwierdzono_at=teraz,
    )
    _ustaw_proweniencje_przydzialu(t, _kandydat_legacy_z_alokacji(candidate))
    try:
        db.add(t); db.flush()
        if legacy_direct and enforcement_override:
            override_details = _szczegoly_nadpisania_r3(
                evaluation,
                legacy=bool(override_context and override_context["legacy"]),
            )
            override_audit_reason = _powod_audytu_nadpisania(evaluation)
            override_reason_code = override_context["reason_code"]
            override_note = override_context["note"]
        elif frozen_override_context is not None:
            override_details = frozen_override_context["details"]
            override_audit_reason = frozen_override_context["audit_reason"]
            override_reason_code = frozen_override_context["reason_code"]
            override_note = frozen_override_context["note"]
        else:
            override_details = None
            override_audit_reason = None
            override_reason_code = None
            override_note = None
        _zastap_ledger_terminu(
            db,
            t,
            # Ta sama typowana ewaluacja została już wyegzekwowana powyżej.
            # Ponowne sprawdzenie myliłoby wewnętrzny transfer obietnicy
            # pojemności z ręcznym nadpisaniem i odrzucało ważną ofertę.
            enforce_pacing=False,
            evaluation=evaluation,
            override=ledger_override,
            intent="create",
            buffer_override=(frozen_buffer if not legacy_direct else None),
            preserve_interval=not legacy_direct,
            now=teraz,
            candidates=[{"stoliki": table_ids}],
            alternatives=result.to_dict(expose_exact=True).get("alternatives") or (),
        )
        accepted_offer_version = int(w.offer_version or 0)
        w.status = "zaakceptowano"
        w.zrealizowano_at = teraz
        w.zaakceptowano_at = teraz
        w.termin_id = t.id
        w.offer_version = accepted_offer_version + 1
        reservation_communication.cancel_waitlist_pending(db, w.id, now=teraz)
        _audit_waitlisty(
            db, w, "accepted", user=user, now=teraz,
            details={
                "offer_version": accepted_offer_version,
                "next_offer_version": w.offer_version,
                "reservation_id": t.id,
                "table_ids": table_ids,
                "auto_przydzielony": t.auto_przydzielony,
                "override_authorized": ledger_override,
                "grandfathered_capacity_rules": list(
                    grandfathered_capacity_rules
                ),
                "tryb": dane.tryb,
            },
        )
        odpowiedz = {
            "rezerwacja": _rezerwacja_out(t, user),
            "wpis": _lista_out(w, user),
        }
        reservation_audit.add_reservation_audit(
            db,
            termin=t,
            action="create",
            actor=user,
            after=t,
            pii_changed=_utworzone_pii(t),
        )
        if override_details:
            reservation_audit.add_reservation_audit(
                db,
                termin=t,
                action="override",
                actor=user,
                reason=override_audit_reason,
                after=t,
                override_details=override_details,
                override_reason_code=override_reason_code,
                override_note=override_note,
            )
        if not is_walk_in:
            reservation_communication.enqueue_reservation(
                db,
                t,
                "confirmation",
                dedupe_key=(
                    f"reservation:{t.id}:confirmation:waitlist:"
                    f"{secrets.token_hex(16)}"
                ),
                actor=user,
            )
            reservation_communication.schedule_reminder(db, t, actor=user)
        reservation_service.complete_idempotency(
            idem.record, response=odpowiedz, http_status=200, termin_id=t.id, now=teraz,
        )
        _commit_zapis_rezerwacji(db, guards)
    except IntegrityError as exc:
        db.rollback()
        raise reservation_service.translate_integrity_error(exc) from exc
    db.refresh(t)
    return odpowiedz


@app.post("/api/lista-oczekujacych/{wid}/zaakceptuj", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def zaakceptuj_lista_oczekujacych(
    wid: int,
    dane: schemas.WaitlistZaakceptujIn,
    request: Request,
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    return _zaakceptuj_lista_oczekujacych(
        wid,
        dane,
        request,
        expected_version=dane.offer_version,
        idempotency_key=idempotency_key,
        legacy_direct=False,
        db=db,
        user=user,
    )


@app.post("/api/lista-oczekujacych/{wid}/zrealizuj", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def zrealizuj_lista_oczekujacych(
    wid: int,
    dane: schemas.ZrealizujIn,
    request: Request,
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Legacy direct path for a plain waiting entry; never consumes an offer."""
    return _zaakceptuj_lista_oczekujacych(
        wid,
        dane,
        request,
        expected_version=None,
        idempotency_key=idempotency_key,
        legacy_direct=True,
        db=db,
        user=user,
    )


@app.post(
    "/api/lista-oczekujacych/{wid}/powiadom",
    response_model=schemas.WaitlistPowiadomOut,
    dependencies=[Depends(_wymagaj_modul_rezerwacje)],
)
def powiadom_lista_oczekujacych(
    wid: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Atomowo kolejkuje „stolik gotowy”; worker stempluje dopiero przyjęcie przez provider."""
    if not _ma_dostep_rezerwacji(user, "rezerwacje.dane_kontaktowe"):
        raise HTTPException(403, "Brak uprawnienia do danych kontaktowych gości.")
    w, guards = _zablokuj_waitliste(db, wid)
    if not w:
        raise HTTPException(404, "Brak wpisu.")
    teraz = utcnow_naive()
    _odswiez_wygasniecie_waitlisty(db, w, guards, teraz=teraz)
    if w.status != "zaoferowano":
        db.rollback()
        raise reservation_service.ReservationError(
            409,
            "WAITLIST_OFFER_NOT_ACTIVE",
            "Powiadomienie wymaga aktywnej oferty z zajętym zestawem stołów.",
            rule="waitlist_offer",
        )
    held_ids = _stoliki_zadania(w.hold_stolik_id, w.hold_stoliki_dodatkowe)
    if (
        not held_ids
        or w.hold_godz_od is None
        or w.hold_godz_do is None
        or w.hold_do is None
        or w.hold_do <= teraz
    ):
        db.rollback()
        raise reservation_service.ReservationError(
            409,
            "WAITLIST_OFFER_HOLD_INCOMPLETE",
            "Oferta nie ma kompletnego, aktywnego zestawu stołów.",
            rule="waitlist_offer",
        )
    start_minute, blocked_end = reservation_service.claim_minute_window(
        w.hold_godz_od, w.hold_godz_do, int(w.hold_bufor_min or 0),
    )
    claims = db.query(models.RezerwacjaStolikClaim).filter_by(waitlist_id=w.id).all()
    expected_claim_slots = {
        (table_id, minute)
        for table_id in held_ids
        for minute in range(start_minute, blocked_end)
    }
    if (
        len(claims) != len(expected_claim_slots)
        or {(row.stolik_id, row.minute) for row in claims} != expected_claim_slots
        or any(
            row.data != w.data
            or row.expires_at is None
            or row.expires_at <= teraz
            or row.expires_at != w.hold_do
            for row in claims
        )
    ):
        db.rollback()
        raise reservation_service.ReservationError(
            409,
            "WAITLIST_OFFER_CLAIMS_INCOMPLETE",
            "Oferta nie ma kompletnego atomowego zajęcia zasobów.",
            rule="waitlist_offer",
        )
    group = _wiadomosci_biezacej_oferty(db, w)
    if group:
        summary = reservation_communication.summaries_for_waitlists(db, [w.id]).get(w.id)
        return {
            "queued": any(row.stan in {"queued", "processing", "retry"} for row in group),
            "juz_powiadomiony": bool(group) and all(row.stan == "sent" for row in group),
            "legacy_delivery": False,
            "messages": [reservation_communication.message_dict(row) for row in group],
            "wpis": _lista_out(w, user, summary),
        }
    messages = reservation_communication.enqueue_table_ready(
        db,
        w,
        dedupe_key=(
            f"waitlist:{w.id}:offer:{w.offer_version}:{w.offer_key_hash}"
        ),
        actor=user,
        now=teraz,
    )
    if not messages:
        db.rollback()
        raise HTTPException(
            400,
            "Brak dostępnego kanału kontaktu albo komunikacja operacyjna jest wyłączona.",
        )
    _audit_waitlisty(
        db,
        w,
        "notification_queued",
        user=user,
        now=teraz,
        details={
            "offer_version": int(w.offer_version or 0),
            "queued_channels": sorted({row.kanal for row in messages}),
        },
    )
    _commit_zapis_rezerwacji(db, guards)
    summary = reservation_communication.summaries_for_waitlists(db, [w.id]).get(w.id)
    return {
        "queued": True,
        "juz_powiadomiony": False,
        "legacy_delivery": False,
        "messages": [reservation_communication.message_dict(row) for row in messages],
        "wpis": _lista_out(w, user, summary),
    }


@app.get(
    "/api/lista-oczekujacych/{wid}/komunikacja",
    response_model=schemas.WaitlistKomunikacjaHistoriaOut,
    dependencies=[Depends(_wymagaj_modul_rezerwacje)],
)
def historia_komunikacji_waitlisty(
    wid: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    if not _ma_dostep_rezerwacji(user, "rezerwacje.dane_kontaktowe"):
        raise HTTPException(403, "Brak uprawnienia do danych kontaktowych gości.")
    waitlist = db.get(models.ListaOczekujacych, wid)
    if waitlist is None:
        raise HTTPException(404, "Brak wpisu.")
    teraz = utcnow_naive()
    summary = reservation_communication.summaries_for_waitlists(db, [wid]).get(wid)
    return JSONResponse(
        {
            "waitlist_id": wid,
            "summary": summary,
            "legacy_delivery": bool(summary and summary.get("legacy_delivery")),
            "messages": reservation_communication.waitlist_history(
                db, wid, now=teraz,
            ),
        },
        headers={"Cache-Control": "no-store"},
    )


@app.post("/api/lista-oczekujacych/{wid}/hold", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def hold_lista_oczekujacych(
    wid: int,
    dane: schemas.HoldIn,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Trzyma w jednym oknie pełny zestaw stołów wybrany przez wspólny allocator."""
    w, guards = _zablokuj_waitliste(db, wid)
    if not w:
        raise HTTPException(404, "Brak wpisu.")
    if w.status != "oczekuje":
        raise HTTPException(409, "Wpis nie oczekuje już na stolik.")
    godz = dane.godz_od or w.godz_od
    if godz is None:
        raise HTTPException(400, "Ustaw godzinę, aby utworzyć czasowy hold.")
    teraz = utcnow_naive()
    reservation_service.cleanup_expired_holds(db, teraz, dates=[w.data])
    reservation_service.release_waitlist_hold(db, w.id)
    db.flush()
    requested_ids = _stoliki_zadania(dane.stolik_id, dane.stoliki)
    runtime_table_ids = {table["id"] for table in _stoly_do_seating(db)}
    if requested_ids and not set(requested_ids) <= runtime_table_ids:
        raise HTTPException(400, "Nieznany lub nieaktywny stolik.")
    result = _ocen_przydzial_rezerwacji(
        db,
        data=w.data,
        godz_od=godz,
        osoby=w.liczba_osob or 1,
        kanal="wewnetrzna",
        intent="quote",
        alternative_limit=9999,
    )
    candidate = (
        _kandydat_zestawu(result, requested_ids)
        if requested_ids else result.selected
    )
    if candidate is None:
        _wymagaj_dozwolonego_przydzialu(result)
        raise reservation_service.ReservationError(
            409,
            "INVALID_TABLE_COMBINATION",
            "Wybrany zestaw nie jest dostępną, zatwierdzoną konfiguracją.",
            rule="table_hold",
        )
    evaluation = _ocen_reguly_slotu(
        db,
        data=w.data,
        godz_od=godz,
        liczba_osob=w.liczba_osob or 1,
        kanal="wewnetrzna",
        sala_id=candidate.room_id,
        intent="create",
    )
    if evaluation.decision != "allow":
        reservation_rules.enforce_rule_evaluation(evaluation)
    table_ids = list(candidate.table_ids)
    locked_tables = reservation_service.lock_tables(db, table_ids)
    if len(locked_tables) != len(table_ids):
        raise HTTPException(400, "Nieznany lub nieaktywny stolik.")
    godz_do = evaluation.godz_do
    if godz_do is None:
        raise HTTPException(400, "Nie udało się wyznaczyć końca wizyty.")
    minuty = max(1, int(dane.minuty or 15))
    hold_do = teraz + timedelta(minutes=minuty)
    reservation_service.replace_waitlist_hold(
        db,
        waitlist_id=w.id,
        table_ids=table_ids,
        data=w.data,
        expires_at=hold_do,
        now=teraz,
        start=godz,
        end=godz_do,
        buffer_min=evaluation.buffer_min,
    )
    w.godz_od = godz
    w.hold_stolik_id = table_ids[0]
    w.hold_stoliki_dodatkowe = table_ids[1:] or None
    w.hold_godz_od = godz
    w.hold_godz_do = godz_do
    w.hold_bufor_min = evaluation.buffer_min
    w.hold_do = hold_do
    _commit_zapis_rezerwacji(db, guards); db.refresh(w)
    return _lista_out(w, user)


@app.post("/api/lista-oczekujacych/{wid}/zwolnij-hold", dependencies=[Depends(_wymagaj_modul_rezerwacje)])
def zwolnij_hold_lista_oczekujacych(
    wid: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Zwalnia wcześniejszy HOLD (stół wraca do puli)."""
    w, guards = _zablokuj_waitliste(db, wid)
    if not w:
        raise HTTPException(404, "Brak wpisu.")
    if w.status != "oczekuje":
        db.rollback()
        raise reservation_service.ReservationError(
            409,
            "WAITLIST_LEGACY_HOLD_NOT_ALLOWED",
            "Aktywną ofertę można zwolnić tylko przez wycofanie oferty.",
            rule="waitlist_offer",
        )
    reservation_service.release_waitlist_hold(db, w.id)
    _wyczysc_hold_waitlisty(w)
    _commit_zapis_rezerwacji(db, guards); db.refresh(w)
    return _lista_out(w, user)


# ═══════════════════════════════════════════════════════════════════════════
# REZERWACJE ONLINE (publiczny widget — bez logowania, za flagą rezerwacje_online)
# ═══════════════════════════════════════════════════════════════════════════

ONLINE_LIMIT_DZIENNY = 5   # anty-spam: maks. aktywnych rezerwacji online/dzień po telefonie/e-mailu
ONLINE_LIMIT_IP_DZIENNY = 15   # anty-DoS: maks. rezerwacji online/dzień z jednego IP (niezależnie od kontaktu)
PUBLIC_HOLD_TTL_SECONDS = 8 * 60
PUBLIC_PRIVACY_NOTICE_VERSION = reservation_service.PUBLIC_PRIVACY_NOTICE_VERSION
PUBLIC_MARKETING_CONSENT_VERSION = "reservation-marketing-2026-07-v1"
PUBLIC_SENSITIVE_CONSENT_VERSION = "reservation-sensitive-2026-07-v1"
PUBLIC_MANAGEMENT_COOKIE = "lokalo_reservation_capability"
PUBLIC_MANAGEMENT_COOKIE_PATH = "/api/online/zarzadzanie"


def _public_client_ip(request: Request) -> str:
    # Nie ufamy X-Forwarded-For bez jawnej konfiguracji trusted proxy po stronie
    # serwera. W bazie i tak zapisujemy wyłącznie domenowo rozdzielony HMAC.
    return request.client.host if request.client else "unknown-client"


def _wymagaj_publiczny_naglowek(value: Optional[str], *, nazwa: str) -> str:
    raw = (value or "").strip()
    if len(raw) < 16 or len(raw) > 128 or any(ord(char) < 33 or ord(char) > 126 for char in raw):
        raise HTTPException(400, f"Brak lub niepoprawny nagłówek {nazwa}.")
    return raw


def _ustaw_publiczne_cookie_zarzadzania(
    response: Response,
    raw_token: str,
    *,
    expires_at: datetime,
) -> None:
    """Utrwala capability wyłącznie jako zaszyfrowane cookie niedostępne dla JS."""
    if not szyfrowanie.aktywne():
        # Produkcja nie startuje bez ENCRYPTION_KEY. W celowo nieutwardzonym dev
        # zachowujemy nagłówkowy fallback, ale nigdy nie zapisujemy plaintext cookie.
        return
    encrypted = szyfrowanie.szyfruj(raw_token)
    if not isinstance(encrypted, str) or encrypted == raw_token:
        return
    expires_utc = expires_at.replace(tzinfo=timezone.utc)
    max_age = max(0, int((expires_at - utcnow_naive()).total_seconds()))
    response.set_cookie(
        key=PUBLIC_MANAGEMENT_COOKIE,
        value=encrypted,
        max_age=max_age,
        expires=expires_utc,
        path=PUBLIC_MANAGEMENT_COOKIE_PATH,
        secure=not app_settings.IS_DEV,
        httponly=True,
        samesite="lax",
    )


def _usun_publiczne_cookie_zarzadzania(response: Response) -> None:
    response.delete_cookie(
        key=PUBLIC_MANAGEMENT_COOKIE,
        path=PUBLIC_MANAGEMENT_COOKIE_PATH,
        secure=not app_settings.IS_DEV,
        httponly=True,
        samesite="lax",
    )


def _wymagaj_cookie_same_origin(
    request: Request,
    reservation_session: Optional[str],
) -> None:
    """Cookie-auth mutation requires a non-simple header and browser same-origin proof."""
    _wymagaj_publiczny_naglowek(
        reservation_session, nazwa="X-Reservation-Session",
    )
    expected_origin = f"{request.url.scheme}://{request.url.netloc}"
    origin = (request.headers.get("origin") or "").rstrip("/")
    fetch_site = (request.headers.get("sec-fetch-site") or "").strip().lower()
    configured_origins = {value.rstrip("/") for value in ALLOWED_ORIGINS}
    # Sec-Fetch-Site jest forbidden/browser-controlled i pozostaje wiarygodne za
    # reverse proxy, gdzie wewnętrzny request.url może mieć inny scheme niż Origin.
    if (
        origin == expected_origin
        or origin in configured_origins
        or fetch_site == "same-origin"
    ):
        return
    raise HTTPException(403, "Nie można potwierdzić pochodzenia żądania.")


def _publiczny_token_zarzadzania(
    request: Request,
    reservation_token: Optional[str],
    *,
    mutation: bool = False,
    reservation_session: Optional[str] = None,
) -> str:
    """Nowy klient używa cookie; jawny nagłówek pozostaje fallbackiem R5a."""
    # Jawnie podany credential zachowuje kontrakt istniejących klientów API.
    # Frontend przeglądarkowy nie wysyła go po powrocie od providera i wtedy
    # źródłem prawdy jest zaszyfrowane cookie.
    if reservation_token:
        return _wymagaj_publiczny_naglowek(
            reservation_token, nazwa="X-Reservation-Token",
        )
    encrypted = request.cookies.get(PUBLIC_MANAGEMENT_COOKIE)
    if encrypted:
        if not encrypted.startswith("enc:v1:") or not szyfrowanie.aktywne():
            raise HTTPException(400, "Niepoprawna sesja zarządzania rezerwacją.")
        raw_token = szyfrowanie.odszyfruj(encrypted)
        if not isinstance(raw_token, str) or raw_token == encrypted:
            raise HTTPException(400, "Niepoprawna sesja zarządzania rezerwacją.")
        token = _wymagaj_publiczny_naglowek(
            raw_token, nazwa="sesji zarządzania rezerwacją",
        )
        if mutation:
            _wymagaj_cookie_same_origin(request, reservation_session)
        return token
    return _wymagaj_publiczny_naglowek(None, nazwa="X-Reservation-Token")


def _odswiez_publiczne_cookie_z_odpowiedzi(
    response: Response,
    payload: dict,
    db: Session,
) -> None:
    raw_token = payload.get("management_token") or payload.get("token")
    if not raw_token:
        return
    record = reservation_service.lookup_management_token(
        db, raw_token, secret=SECRET_KEY,
    )
    if record is not None:
        _ustaw_publiczne_cookie_zarzadzania(
            response, raw_token, expires_at=record.expires_at,
        )


def _limit_publiczny(scope: str, limit: int, window_seconds: int):
    """Zwraca zależność zużywającą współdzieloną, bazodanową kwotę publiczną."""
    def consume(request: Request, db: Session = Depends(get_db)):
        try:
            reservation_service.consume_public_quota(
                db,
                scope=scope,
                raw_client=_public_client_ip(request),
                secret=SECRET_KEY,
                now=utcnow_naive(),
                limit=limit,
                window_seconds=window_seconds,
            )
            # Kwota ma przetrwać także walidacyjny błąd późniejszego handlera;
            # inaczej nieudane próby zgadywania tokenów byłyby darmowe.
            db.commit()
        except Exception:
            db.rollback()
            raise

    return consume


_limit_widget_config = _limit_publiczny("widget-config", 60, 60)
_limit_dostepnosc = _limit_publiczny("availability", 60, 60)
_limit_alternatywy = _limit_publiczny("alternatives", 30, 60)
_limit_hold = _limit_publiczny("hold", 12, 600)
_limit_create_online = _limit_publiczny("reservation-create", 20, 600)
_limit_waitlist_online = _limit_publiczny("waitlist-create", 15, 600)
_limit_demand_online = _limit_publiczny("demand-rejected", 30, 600)
_limit_management = _limit_publiczny("reservation-management", 30, 600)
_limit_payment_status = _limit_publiczny("payment-status", 90, 600)
_limit_payment_action = _limit_publiczny("payment-action", 10, 600)


def _wymagaj_rezerwacje_online(db: Session = Depends(get_db)):
    if not (modul_aktywny(db, "modul_rezerwacje") and modul_aktywny(db, "rezerwacje_online")):
        raise HTTPException(404, "Rezerwacje online są niedostępne.")


def _widget_v2_gotowy(cfg) -> bool:
    return bool(
        getattr(cfg, "rezerwacje_widget_v2", False)
        and (getattr(cfg, "rezerwacje_rodo_kontakt", None) or "").strip()
        and (getattr(cfg, "rezerwacje_rodo_adres", None) or "").strip()
    )


def _wymagaj_widget_v2(db: Session = Depends(get_db)):
    _wymagaj_rezerwacje_online(db)
    cfg = get_lokal_config(db)
    if not _widget_v2_gotowy(cfg):
        raise HTTPException(
            404,
            "Nowy widget nie jest jeszcze skonfigurowany dla tego lokalu.",
        )


def _waliduj_prywatnosc_widgetu(dane, cfg) -> None:
    if not dane.privacy_notice_acknowledged:
        raise HTTPException(400, "Potwierdź zapoznanie się z informacją o przetwarzaniu danych.")
    if dane.privacy_notice_version != PUBLIC_PRIVACY_NOTICE_VERSION:
        raise HTTPException(409, "Informacja o prywatności została zaktualizowana. Przeczytaj ją ponownie.")
    if dane.marketing_consent and dane.marketing_consent_version != PUBLIC_MARKETING_CONSENT_VERSION:
        raise HTTPException(409, "Treść zgody marketingowej została zaktualizowana.")
    if dane.sensitive_data and (
        not dane.sensitive_data_consent
        or dane.sensitive_data_consent_version != PUBLIC_SENSITIVE_CONSENT_VERSION
    ):
        raise HTTPException(400, "Informacje o alergiach lub potrzebach wymagają osobnej, aktualnej zgody.")
    if not ((dane.telefon or "").strip() or (dane.email or "").strip()):
        raise HTTPException(400, "Podaj telefon lub e-mail do kontaktu w sprawie rezerwacji.")


def _zapisz_publiczna_prywatnosc(
    db,
    *,
    dane,
    cfg,
    request: Request,
    now: datetime,
    termin_id: Optional[int] = None,
    waitlist_id: Optional[int] = None,
):
    retention_days = max(30, min(int(cfg.rezerwacje_retencja_dni or 365), 3650))
    owner = (
        db.get(models.Termin, termin_id)
        if termin_id is not None
        else db.get(models.ListaOczekujacych, waitlist_id)
    )
    # Termin jest datą biznesową lokalu. Dowód przechowuje niezmienny deadline
    # liczony od końca dnia wizyty (lub co najmniej od zebrania zgody), zapisany
    # w repozytoryjnej konwencji naive UTC. Późniejsza zmiana konfiguracji nie
    # może tego terminu wydłużyć.
    retention_until = now + timedelta(days=retention_days)
    if owner is not None and owner.data is not None:
        local_deadline = datetime.combine(
            owner.data + timedelta(days=retention_days + 1),
            time.min,
            tzinfo=ZoneInfo("Europe/Warsaw"),
        )
        retention_until = max(
            retention_until,
            local_deadline.astimezone(timezone.utc).replace(tzinfo=None),
        )
    consent = models.RezerwacjaZgodaPubliczna(
        termin_id=termin_id,
        waitlist_id=waitlist_id,
        notice_version=PUBLIC_PRIVACY_NOTICE_VERSION,
        notice_ack_at=now,
        marketing=bool(dane.marketing_consent),
        # Wersja jest zapisywana również dla odmowy, aby dowieść, jaki wybór pokazano.
        marketing_version=PUBLIC_MARKETING_CONSENT_VERSION,
        marketing_at=now,
        sensitive=bool(dane.sensitive_data),
        sensitive_version=(PUBLIC_SENSITIVE_CONSENT_VERSION if dane.sensitive_data else None),
        sensitive_at=(now if dane.sensitive_data else None),
        sensitive_data=(dane.sensitive_data if dane.sensitive_data else None),
        retention_until=retention_until,
        ip_hash=reservation_service.hash_public_client(
            _public_client_ip(request),
            secret=SECRET_KEY,
            purpose="reservation-consent-ip",
        ),
        subject_hash=(_identity_hash_crm(owner) if owner is not None else None),
        created_at=now,
    )
    db.add(consent)
    return consent


def _sloty_dnia(db, data: date):
    """Kompatybilny adapter do siatki ofert R3 (krok ≠ czas wizyty)."""
    return list(reservation_rules.sloty_dnia(db, data))


def _pacing_pelny(db, data, godz_od, serwis, osoby, pomin_id=None) -> bool:
    """Czy limit coverów serwisu jest wyczerpany dla slotu o godz_od (dołożenie rezerwacji na
    'osoby' przekroczyłoby pacing_max_rez lub pacing_max_osob). Liczy aktywne rezerwacje
    stolikowe startujące w oknie [godz_od, godz_od+okno). Brak limitów → nigdy pełny."""
    if serwis is None:
        return False
    max_rez = serwis.pacing_max_rez if (serwis.pacing_max_rez or 0) > 0 else None
    max_osob = serwis.pacing_max_osob if (serwis.pacing_max_osob or 0) > 0 else None
    if not max_rez and not max_osob:
        return False
    okno = serwis.pacing_okno_min or serwis.dlugosc_slotu_min or DOMYSLNY_SLOT_MIN
    return reservation_service.pacing_status(
        db,
        data=data,
        start=godz_od,
        window_min=okno,
        party_size=osoby,
        max_reservations=max_rez,
        max_covers=max_osob,
        exclude_termin_id=pomin_id,
    )["full"]


def _stolik_zajety(db, data, stolik_id, godz_od, godz_do) -> bool:
    return stolik_id in _zajete_stoly(db, data, godz_od, godz_do)


def _zajete_stoly(db, data, godz_od, godz_do, pomin_id=None) -> set:
    """Zbiór stołów zajętych według kanonicznego ledgera R0b."""
    return reservation_service.occupied_table_ids(
        db,
        data=data,
        start=godz_od,
        end=godz_do,
        buffer_min=get_lokal_config(db).rez_bufor_min or 0,
        exclude_termin_id=pomin_id,
        now=utcnow_naive(),
    )


def _wersjonowane_stoliki_ids(db) -> set[int]:
    """Stoły należące do sal, które weszły już do wersjonowanego workflow.

    Brak published nie przywraca legacy fallbacku: pierwszy szkic również nie może
    domieszać nieopublikowanych danych do działającego serwisu.
    """
    rooms = db.query(
        models.SalaRezerwacyjna.id,
        models.SalaRezerwacyjna.nazwa,
    ).join(
        models.PlanSali,
        models.PlanSali.sala_id == models.SalaRezerwacyjna.id,
    ).all()
    room_ids = {room_id for room_id, _name in rooms}
    room_names = {(name or "").strip().casefold() for _room_id, name in rooms}
    return {
        stolik.id
        for stolik in db.query(models.Stolik).all()
        if stolik.sala_id in room_ids
        or (
            stolik.sala_id is None
            and (stolik.strefa or "").strip().casefold() in room_names
        )
    }


_SNAPSHOT_REQUIRED_FALLBACK_FIELDS = frozenset({"nazwa", "kolejnosc", "pojemnosc"})


def _snapshot_value(pozycja, stolik, field):
    value = getattr(pozycja, field, None)
    if value is None and field in _SNAPSHOT_REQUIRED_FALLBACK_FIELDS:
        return getattr(stolik, field, None)
    return value


def _opublikowane_pozycje_stolikow(db):
    return (
        db.query(
            models.PozycjaStolikaPlanu,
            models.Stolik,
            models.SalaRezerwacyjna,
        )
        .join(
            models.WersjaPlanuSali,
            models.WersjaPlanuSali.id == models.PozycjaStolikaPlanu.wersja_id,
        )
        .join(models.PlanSali, models.PlanSali.id == models.WersjaPlanuSali.plan_id)
        .join(
            models.SalaRezerwacyjna,
            models.SalaRezerwacyjna.id == models.PlanSali.sala_id,
        )
        .join(models.Stolik, models.Stolik.id == models.PozycjaStolikaPlanu.stolik_id)
        .filter(models.WersjaPlanuSali.status == "published")
        .all()
    )


def _nieaktywne_sale_runtime(db):
    """Sale wyłączone z nowych przydziałów, także dla legacy fallbacku ``strefa``.

    ``SalaRezerwacyjna.aktywna`` steruje dostępnością operacyjną całej sali, nie
    widocznością jej konfiguracji ani historii. Nazwę legacy uznajemy za wyłączoną
    tylko wtedy, gdy nie istnieje równocześnie aktywna sala o tej samej nazwie.
    """
    sale = db.query(
        models.SalaRezerwacyjna.id,
        models.SalaRezerwacyjna.nazwa,
        models.SalaRezerwacyjna.aktywna,
    ).all()
    aktywne_nazwy = {
        (nazwa or "").strip().casefold()
        for _sala_id, nazwa, aktywna in sale
        if aktywna
    }
    nieaktywne_nazwy = {
        (nazwa or "").strip().casefold()
        for _sala_id, nazwa, aktywna in sale
        if not aktywna
    } - aktywne_nazwy
    return (
        {sala_id for sala_id, _nazwa, aktywna in sale if not aktywna},
        nieaktywne_nazwy,
    )


def _snapshot_stolika_do_odczytu(pozycja, stolik, sala):
    """Stary kontrakt ``StolikOut`` zasilony kanonicznym published snapshotem."""
    return {
        "id": stolik.id,
        "nazwa": _snapshot_value(pozycja, stolik, "nazwa"),
        "sala_id": sala.id,
        "strefa": sala.nazwa,
        "pojemnosc": _snapshot_value(pozycja, stolik, "pojemnosc"),
        # ``laczy_sie`` i ``rewir_nr`` nie są dziś częścią wersji planu.
        "laczy_sie": stolik.laczy_sie,
        "aktywny": pozycja.aktywny_w_planie,
        "kolejnosc": _snapshot_value(pozycja, stolik, "kolejnosc"),
        "rewir_nr": stolik.rewir_nr,
        "pojemnosc_min": _snapshot_value(pozycja, stolik, "pojemnosc_min"),
        "ksztalt": _snapshot_value(pozycja, stolik, "ksztalt"),
        "cechy": _snapshot_value(pozycja, stolik, "cechy"),
        "priorytet": _snapshot_value(pozycja, stolik, "priorytet"),
        "sekcja": _snapshot_value(pozycja, stolik, "sekcja"),
        "wersja_id": pozycja.wersja_id,
    }


def _stoliki_do_odczytu(db):
    """Published snapshot dla sal wersjonowanych, pełny legacy odczyt poza nimi.

    Endpoint konfiguracyjny historycznie zwracał również nieaktywne stoły, dlatego
    nie filtrujemy ``aktywny_w_planie``. Rekord istniejący wyłącznie w draftcie nie
    pojawia się aż do publikacji.
    """
    out = [
        _snapshot_stolika_do_odczytu(pozycja, stolik, sala)
        for pozycja, stolik, sala in _opublikowane_pozycje_stolikow(db)
    ]
    versioned_ids = _wersjonowane_stoliki_ids(db)
    out.extend(
        {
            **schemas.StolikOut.model_validate(stolik).model_dump(),
            "wersja_id": None,
        }
        for stolik in db.query(models.Stolik).all()
        if stolik.id not in versioned_ids
    )
    out.sort(key=lambda row: (row.get("kolejnosc") or 0, row["id"]))
    # Nie rozszerzamy publicznego kontraktu o techniczne ``wersja_id``.
    return [schemas.StolikOut.model_validate(row).model_dump() for row in out]


def _stoly_do_seating(db):
    """Kandydaci do nowych przydziałów: aktywne sale i stoły z published snapshotu."""
    out = []
    for pozycja, stolik, sala in _opublikowane_pozycje_stolikow(db):
        if not sala.aktywna or not pozycja.aktywny_w_planie:
            continue
        snapshot = _snapshot_stolika_do_odczytu(pozycja, stolik, sala)
        out.append({
            **snapshot,
            "sala_nazwa": sala.nazwa,
            "pojemnosc": snapshot["pojemnosc"] or 0,
            "cechy": snapshot["cechy"] or [],
            "priorytet": snapshot["priorytet"] or 0,
            "sekcja": snapshot["sekcja"] or snapshot["strefa"],
            "kolejnosc": snapshot["kolejnosc"] or 0,
            "strategia_zapelniania": sala.strategia_zapelniania or "preferuj",
            "priorytet_sali": sala.priorytet or 0,
            "kolejnosc_sali": sala.kolejnosc or 0,
        })

    versioned_ids = _wersjonowane_stoliki_ids(db)
    inactive_room_ids, inactive_room_names = _nieaktywne_sale_runtime(db)
    rooms = db.query(models.SalaRezerwacyjna).all()
    rooms_by_id = {room.id: room for room in rooms}
    rooms_by_name = {
        (room.nazwa or "").strip().casefold(): room
        for room in rooms if room.aktywna
    }
    for stolik in (
        db.query(models.Stolik)
        .filter(models.Stolik.aktywny.is_(True))
        .order_by(models.Stolik.kolejnosc, models.Stolik.id)
        .all()
    ):
        if stolik.id in versioned_ids:
            continue
        if stolik.sala_id in inactive_room_ids or (
            stolik.sala_id is None
            and (stolik.strefa or "").strip().casefold() in inactive_room_names
        ):
            continue
        room = rooms_by_id.get(stolik.sala_id)
        if room is None and stolik.sala_id is None:
            room = rooms_by_name.get((stolik.strefa or "").strip().casefold())
        out.append({
            "id": stolik.id,
            "nazwa": stolik.nazwa,
            "pojemnosc": stolik.pojemnosc or 0,
            "pojemnosc_min": stolik.pojemnosc_min,
            "ksztalt": stolik.ksztalt,
            "cechy": stolik.cechy or [],
            "priorytet": stolik.priorytet or 0,
            "strefa": stolik.strefa,
            "sekcja": stolik.sekcja or stolik.strefa,
            "kolejnosc": stolik.kolejnosc or 0,
            "sala_id": (room.id if room else stolik.sala_id),
            "sala_nazwa": (room.nazwa if room else stolik.strefa),
            "strategia_zapelniania": (
                room.strategia_zapelniania if room else "preferuj"
            ),
            "priorytet_sali": (room.priorytet if room else 0),
            "kolejnosc_sali": (room.kolejnosc if room else 0),
            "rewir_nr": stolik.rewir_nr,
            "wersja_id": None,
        })
    return sorted(out, key=lambda row: (row["kolejnosc"], row["id"]))


def _sasiedztwo_do_seating(db):
    """Wyłącznie legacy graf dla stołów poza wersjonowanymi planami.

    Published adjacency służy walidacji/edytorowi. Runtime nigdy nie generuje z
    niego dowolnych podgrafów; dla wersjonowanej sali wymagane są jawne kombinacje.
    """
    versioned_ids = _wersjonowane_stoliki_ids(db)
    return [
        (k.stolik_a, k.stolik_b)
        for k in db.query(models.SasiedztwoStolow).all()
        if k.stolik_a not in versioned_ids and k.stolik_b not in versioned_ids
    ]


def _sasiedztwo_do_odczytu(db):
    """Bieżące published krawędzie + legacy wyłącznie poza wersjonowanymi salami."""
    out = [
        {
            # Ujemny identyfikator jest tylko namespace'em adaptera legacy.
            "id": -edge.id,
            "stolik_a": edge.stolik_a_id,
            "stolik_b": edge.stolik_b_id,
        }
        for edge in (
            db.query(models.KrawedzSasiedztwaPlanu)
            .join(
                models.WersjaPlanuSali,
                models.WersjaPlanuSali.id == models.KrawedzSasiedztwaPlanu.wersja_id,
            )
            .filter(models.WersjaPlanuSali.status == "published")
            .all()
        )
    ]
    versioned_ids = _wersjonowane_stoliki_ids(db)
    out.extend(
        {"id": edge.id, "stolik_a": edge.stolik_a, "stolik_b": edge.stolik_b}
        for edge in db.query(models.SasiedztwoStolow).all()
        if edge.stolik_a not in versioned_ids and edge.stolik_b not in versioned_ids
    )
    return sorted(out, key=lambda row: row["id"])


def _kombinacje_snapshotu_do_odczytu(db):
    published_ids = {
        version_id
        for (version_id,) in db.query(models.WersjaPlanuSali.id).filter_by(
            status="published",
        ).all()
    }
    if not published_ids:
        return []
    combinations = (
        db.query(models.KombinacjaStolowPlanu)
        .filter(models.KombinacjaStolowPlanu.wersja_id.in_(published_ids))
        .all()
    )
    members = defaultdict(list)
    combination_ids = [combination.id for combination in combinations]
    if combination_ids:
        for combination_id, table_id in (
            db.query(
                models.SkladnikKombinacjiPlanu.kombinacja_id,
                models.SkladnikKombinacjiPlanu.stolik_id,
            )
            .filter(models.SkladnikKombinacjiPlanu.kombinacja_id.in_(combination_ids))
            .order_by(
                models.SkladnikKombinacjiPlanu.kombinacja_id,
                models.SkladnikKombinacjiPlanu.stolik_id,
            )
            .all()
        ):
            members[combination_id].append(table_id)
    return [
        {
            "id": combination.id,
            "wersja_id": combination.wersja_id,
            "nazwa": combination.nazwa,
            "stoliki": members.get(combination.id, []),
            "pojemnosc_min": combination.pojemnosc_min,
            "pojemnosc_max": combination.pojemnosc_max,
            "aktywna": combination.aktywna_w_planie,
            "priorytet": combination.priorytet or 0,
            "kanal": combination.kanal,
        }
        for combination in combinations
    ]


def _kombinacje_do_odczytu(db):
    """Stary kontrakt listy, bez domieszania legacy danych wersjonowanych sal."""
    out = [
        {**combination, "id": -combination["id"]}
        for combination in _kombinacje_snapshotu_do_odczytu(db)
    ]
    versioned_ids = _wersjonowane_stoliki_ids(db)
    out.extend(
        {
            **schemas.KombinacjaStolowOut.model_validate(combination).model_dump(),
            "wersja_id": None,
            "kanal": "oba",
        }
        for combination in db.query(models.KombinacjaStolow).all()
        if not (_ids_stolikow(combination.stoliki) & versioned_ids)
    )
    out.sort(key=lambda row: (row["priorytet"], row["id"]))
    # Kanał i wersja są danymi technicznymi nowego snapshotu, a legacy endpoint
    # zachowuje dotychczasowy kształt odpowiedzi.
    return [
        schemas.KombinacjaStolowOut.model_validate(row).model_dump()
        for row in out
    ]


def _obciazenie_sekcji(db, data, godz_od, godz_do, pomin_id=None):
    """Ile stołów zajętych per sekcja kelnerska w oknie — do balansu obłożenia w silniku."""
    zajete = _zajete_stoly(db, data, godz_od, godz_do, pomin_id=pomin_id)
    if not zajete:
        return {}
    obc = {}
    by_id = {stolik["id"]: stolik for stolik in _stoly_do_seating(db)}
    for stolik_id in zajete:
        stolik = by_id.get(stolik_id)
        if stolik is None:
            continue
        sek = stolik["sekcja"] or stolik["strefa"]
        if sek:
            obc[sek] = obc.get(sek, 0) + 1
    return obc


def _kombinacje_do_seating(db, kanal=None):
    """Jawne published kombinacje oraz legacy fallback poza wersjonowanymi salami."""
    runtime_table_ids = {stolik["id"] for stolik in _stoly_do_seating(db)}
    out = []
    for combination in _kombinacje_snapshotu_do_odczytu(db):
        if not combination["aktywna"]:
            continue
        if kanal and combination["kanal"] not in ("oba", kanal):
            continue
        table_ids = combination["stoliki"]
        if len(table_ids) < 2 or not set(table_ids) <= runtime_table_ids:
            continue
        out.append({
            key: value for key, value in combination.items()
            if key != "aktywna"
        })

    versioned_ids = _wersjonowane_stoliki_ids(db)
    for combination in (
        db.query(models.KombinacjaStolow)
        .filter_by(aktywna=True)
        .order_by(models.KombinacjaStolow.priorytet, models.KombinacjaStolow.id)
        .all()
    ):
        table_ids = _ids_stolikow(combination.stoliki)
        if not table_ids or table_ids & versioned_ids:
            continue
        out.append({
            "id": combination.id,
            "wersja_id": None,
            "nazwa": combination.nazwa,
            "stoliki": combination.stoliki or [],
            "pojemnosc_min": combination.pojemnosc_min,
            "pojemnosc_max": combination.pojemnosc_max,
            "priorytet": combination.priorytet or 0,
            "kanal": "oba",
        })
    return out


def _ustaw_proweniencje_przydzialu(termin, kandydat=None):
    """Ustawia lub czyści snapshot razem z każdą zmianą zasobu rezerwacji."""
    kandydat = kandydat or {}
    version_id = kandydat.get("wersja_planu_id")
    combination_id = kandydat.get("kombinacja_planu_id")
    termin.przydzial_wersja_planu_id = version_id
    termin.przydzial_kombinacja_planu_id = (
        combination_id if version_id is not None else None
    )


def _wybierz_wolny_stolik(db, data, godz_od, godz_do, osoby, pomin_id=None):
    """Najmniejszy wolny aktywny stolik mieszczący 'osoby' w oknie [godz_od, godz_do]. None gdy brak.
    pomin_id = id rezerwacji, której NIE wliczać do zajętości (edycja własnej rezerwacji)."""
    zajete = _zajete_stoly(db, data, godz_od, godz_do, pomin_id=pomin_id)
    kandydaci = sorted(
        [
            stolik for stolik in _stoly_do_seating(db)
            if stolik["pojemnosc"] >= max(1, osoby or 1)
        ],
        key=lambda stolik: (stolik["pojemnosc"], stolik["id"]),
    )
    for stolik in kandydaci:
        if stolik["id"] not in zajete:
            return stolik
    return None


def _wybierz_wolny_przydzial(
    db, data, godz_od, godz_do, osoby, pomin_id=None, kanal="wewnetrzna",
):
    """Najlepszy pojedynczy stół lub dozwolona kombinacja ze wspólnego silnika sadzania."""
    wyniki = _wolne_przydzialy(
        db, data, godz_od, godz_do, osoby, pomin_id=pomin_id, kanal=kanal,
    )
    return wyniki[0] if wyniki else None


def _wolne_przydzialy(
    db, data, godz_od, godz_do, osoby, pomin_id=None, kanal="wewnetrzna",
):
    """Wszystkie legalne zasobowo przydziały, w kolejności preferencji silnika."""
    zajete = _zajete_stoly(db, data, godz_od, godz_do, pomin_id=pomin_id)
    wyniki = seating.dopasuj(
        max(1, osoby or 1),
        _stoly_do_seating(db),
        _kombinacje_do_seating(db, kanal=kanal),
        zajete=zajete,
        limit=0,
        sasiedztwo=_sasiedztwo_do_seating(db),
        obciazenie_sekcji=_obciazenie_sekcji(
            db, data, godz_od, godz_do, pomin_id=pomin_id,
        ),
        respect_room_fill=False,
    )
    wyniki.sort(key=lambda candidate: (
        0 if candidate.get("strategia_zapelniania") == "wypelniaj_kolejno" else 1,
        candidate.get("priorytet_sali") or 0,
        candidate.get("kolejnosc_sali") or 0,
        candidate.get("koszt") or 0,
        len(candidate.get("stoliki") or ()),
        candidate.get("stoliki") or (),
    ))
    return wyniki


def _porownaj_allocator_shadow(
    db,
    *,
    canonical,
    data,
    godz_od,
    osoby,
    kanal,
    pomin_id,
    intent,
):
    """Opcjonalny shadow-read starej ścieżki bez wpływu na zapis.

    Flaga środowiskowa służy etapowemu rolloutowi i raportuje wyłącznie różnice
    decyzji/zasobu — bez danych gościa. Wynik kanoniczny zawsze pozostaje źródłem
    odpowiedzi i zapisu po cutoverze R4.
    """
    enabled = os.environ.get("RESERVATION_ALLOCATOR_SHADOW_COMPARE", "").strip().lower()
    if enabled not in {"1", "true", "yes", "on"}:
        return
    base = canonical.evaluation
    legacy_decision = base.decision
    legacy_tables = ()
    if base.godz_do is not None:
        first_blocked = None
        candidates = _wolne_przydzialy(
            db,
            data,
            godz_od,
            base.godz_do,
            osoby,
            pomin_id=pomin_id,
            kanal=kanal,
        )
        rule_intent = "edit" if intent in {"assign", "reoptimize"} else intent
        for candidate in candidates:
            evaluation = _ocen_reguly_slotu(
                db,
                data=data,
                godz_od=godz_od,
                liczba_osob=osoby,
                kanal=kanal,
                sala_id=candidate.get("sala_id"),
                existing_termin_id=pomin_id,
                intent=rule_intent,
            )
            if evaluation.decision == "allow":
                legacy_decision = "allow"
                legacy_tables = tuple(sorted(candidate.get("stoliki") or ()))
                break
            if first_blocked is None:
                first_blocked = evaluation
        else:
            if first_blocked is not None:
                legacy_decision = first_blocked.decision

    canonical_tables = tuple(sorted(
        canonical.selected.table_ids if canonical.selected is not None else ()
    ))
    if canonical.decision == legacy_decision and canonical_tables == legacy_tables:
        return
    logger.warning(
        "reservation_allocator_shadow_diff",
        extra={
            "reservation_allocator_shadow": {
                "data": str(data),
                "godz_od": _hm(godz_od),
                "kanal": reservation_rules.normalise_channel(kanal),
                "intent": intent,
                "party_size": max(1, int(osoby or 1)),
                "canonical_decision": canonical.decision,
                "legacy_decision": legacy_decision,
                "canonical_tables": canonical_tables,
                "legacy_tables": legacy_tables,
            },
        },
    )


def _kandydat_legacy_z_alokacji(candidate):
    """Adapter R4 dla starszych zapisow oczekujacych slownika z ``seating``."""
    if candidate is None:
        return None
    return {
        "stoliki": list(candidate.table_ids),
        "sala_id": candidate.room_id,
        "nazwa": candidate.name,
        "suma_pojemnosci": candidate.capacity,
        "nadmiar_miejsc": candidate.unused_seats,
        "kombinacja": candidate.combination,
        "wersja_planu_id": candidate.plan_version_id,
        "kombinacja_planu_id": candidate.plan_combination_id,
    }


def _ocen_przydzial_rezerwacji(
    db,
    *,
    data,
    godz_od,
    osoby,
    kanal,
    godz_do=None,
    pomin_id=None,
    intent="quote",
    preferowana_sala_id=None,
    preferowana_strefa=None,
    preferowane_cechy=(),
    zachowaj_nieaktywny_przydzial=False,
    preserve_explicit_interval=False,
    physical_buffer_min=None,
    alternative_limit=3,
    now=None,
):
    """Jedyny evaluator zasobu dla widgetu, recepcji, hosta i symulatora.

    Zajetosc liczymy osobno dla kazdej sali, bo R3 pozwala jej miec wlasny
    bufor. Krotszy bufor jednej sali nie oslabia wiec ochrony drugiej.
    """
    lifecycle_now = (
        now.astimezone(timezone.utc).replace(tzinfo=None)
        if now is not None and now.tzinfo is not None
        else (now.replace(tzinfo=None) if now is not None else utcnow_naive())
    )
    rule_now = lifecycle_now.replace(tzinfo=timezone.utc).astimezone(
        ZoneInfo("Europe/Warsaw"),
    )
    tables = _stoly_do_seating(db)
    combinations = _kombinacje_do_seating(
        db,
        kanal=reservation_service.normalise_reservation_channel(kanal),
    )
    rule_intent = "assign" if intent in {"assign", "reoptimize"} else intent
    base = _ocen_reguly_slotu(
        db,
        data=data,
        godz_od=godz_od,
        godz_do=godz_do,
        liczba_osob=osoby,
        kanal=kanal,
        existing_termin_id=pomin_id,
        intent=rule_intent,
        preserve_existing_room_access=zachowaj_nieaktywny_przydzial,
        preserve_explicit_interval=preserve_explicit_interval,
        now=lifecycle_now,
    )
    occupied = set()
    table_ids_by_room = defaultdict(set)
    for table in tables:
        table_ids_by_room[table.get("sala_id")].add(int(table["id"]))
    if base is not None and base.godz_do is not None:
        for room_id, room_table_ids in table_ids_by_room.items():
            evaluation = base if room_id is None else _ocen_reguly_slotu(
                db,
                data=data,
                godz_od=godz_od,
                godz_do=godz_do,
                liczba_osob=osoby,
                kanal=kanal,
                sala_id=room_id,
                existing_termin_id=pomin_id,
                intent=rule_intent,
                preserve_existing_room_access=zachowaj_nieaktywny_przydzial,
                preserve_explicit_interval=preserve_explicit_interval,
                now=lifecycle_now,
            )
            if evaluation is None or evaluation.godz_do is None:
                continue
            occupied |= room_table_ids & reservation_service.occupied_table_ids(
                db,
                data=data,
                start=godz_od,
                end=evaluation.godz_do,
                # The rule evaluation remains current and drives diagnostics.
                # A verified waitlist offer may, however, transfer an older
                # physical promise whose claim interval has a frozen buffer.
                buffer_min=(
                    evaluation.buffer_min
                    if physical_buffer_min is None
                    else max(0, int(physical_buffer_min or 0))
                ),
                exclude_termin_id=pomin_id,
                now=lifecycle_now,
            )
    section_end = base.godz_do if base is not None else godz_do
    section_load = (
        _obciazenie_sekcji(db, data, godz_od, section_end, pomin_id=pomin_id)
        if section_end is not None else {}
    )
    result = reservation_allocator.evaluate_allocation(
        db,
        reservation_allocator.AllocationRequest(
            data=data,
            godz_od=godz_od,
            godz_do=godz_do,
            liczba_osob=max(1, int(osoby or 1)),
            kanal=kanal,
            intent=intent,
            existing_termin_id=pomin_id,
            preferred_room_id=preferowana_sala_id,
            preferred_zone=preferowana_strefa,
            preferred_features=tuple(preferowane_cechy or ()),
            preserve_existing_room_access=zachowaj_nieaktywny_przydzial,
            preserve_explicit_interval=preserve_explicit_interval,
        ),
        tables=tables,
        combinations=combinations,
        occupied_table_ids=occupied,
        adjacency=_sasiedztwo_do_seating(db),
        section_load=section_load,
        now=rule_now,
        alternative_limit=alternative_limit,
    )
    _porownaj_allocator_shadow(
        db,
        canonical=result,
        data=data,
        godz_od=godz_od,
        osoby=osoby,
        kanal=kanal,
        pomin_id=pomin_id,
        intent=intent,
    )
    return result


def _klasyfikacja_popytu_rezerwacji(
    db,
    *,
    data,
    godz_od,
    osoby,
    kanal,
):
    """Wylicza kategorię R7.2 wyłącznie z kanonicznego wyniku R3/R4."""
    if isinstance(osoby, bool) or not isinstance(osoby, int) or osoby < 1:
        return reservation_demand.DemandClassification("other", "unknown")
    godziny = [godz_od] if godz_od is not None else [
        slot for slot, _serwis in _sloty_dnia(db, data)
    ]
    if not godziny:
        return reservation_demand.DemandClassification(
            "service_closed", "policy",
        )

    classifications = []
    for slot in godziny:
        result = _ocen_przydzial_rezerwacji(
            db,
            data=data,
            godz_od=slot,
            osoby=osoby,
            kanal=kanal,
            intent="quote",
            alternative_limit=0,
        )
        classification = reservation_demand.classify_allocation(result)
        if classification.reason_code == "operator_decision":
            return classification
        classifications.append(classification)

    # Dla zapytania bez godziny wybieramy najczęstszy powód dnia, a przy remisie
    # stałą kolejność produktową. Nie kopiujemy kolejności/tekstu z klienta.
    priority = {
        "service_closed": 0,
        "channel_unavailable": 1,
        "booking_window": 2,
        "party_policy": 3,
        "pacing_limit": 4,
        "concurrent_limit": 5,
        "resource_occupied": 6,
        "no_capacity_match": 7,
        "other": 8,
    }
    counts = defaultdict(int)
    for classification in classifications:
        counts[classification] += 1
    return min(
        classifications,
        key=lambda item: (-counts[item], priority.get(item.reason_code, 99)),
    )


class _AllocationAvailabilityAdapter:
    """Laczy pelny kontrakt regul R3 z kandydatem i alternatywami R4."""

    def __init__(self, result):
        self.result = result

    def to_dict(self):
        return self.result.to_dict(expose_exact=True)


def _wymagaj_dozwolonego_przydzialu(result, *, override=False):
    """Zamienia wynik quote/allocate na stabilny błąd domenowy zapisu."""
    if result.evaluation.decision == "deny" or (
        result.evaluation.decision == "override_required" and not override
    ):
        error = reservation_rules.evaluation_to_reservation_error(result.evaluation)
        error.availability = _AllocationAvailabilityAdapter(result)
        raise error
    if result.selected is None or (
        result.decision != "allow"
        and not (override and result.decision == "override_required")
    ):
        payload = result.to_dict(expose_exact=True)
        raise reservation_service.ReservationError(
            409,
            result.code or "NO_TABLE_CANDIDATE",
            result.message or "Brak wolnego stołu dla tej grupy w tym czasie.",
            rule=result.rule or "table",
            candidates=payload.get("candidates") or (),
            alternatives=payload.get("alternatives") or (),
            decision=result.decision,
            service=payload.get("service"),
            krok_slotu_min=payload.get("krok_slotu_min"),
            turn_time_min=payload.get("turn_time_min"),
            godz_do=payload.get("visit_end"),
            violations=payload.get("violations") or (),
            checks=payload.get("checks") or (),
            resource_allocation=payload.get("resource_allocation"),
        )
    return result.selected


def _przydzial_zgodny_z_regulami(
    db,
    *,
    data,
    godz_od,
    osoby,
    kanal,
    pomin_id=None,
    intent="quote",
    enforce=True,
):
    """Kompatybilny adapter starego API do wspólnego allocatora R4."""
    result = _ocen_przydzial_rezerwacji(
        db,
        data=data,
        godz_od=godz_od,
        osoby=osoby,
        kanal=kanal,
        pomin_id=pomin_id,
        intent=intent,
    )
    evaluation = result.evaluation
    if enforce and result.decision != "allow" and result.selected is not None:
        reservation_rules.enforce_rule_evaluation(evaluation)
    if enforce and result.selected is None and evaluation.decision != "allow":
        reservation_rules.enforce_rule_evaluation(evaluation)
    if result.decision != "allow" or result.selected is None:
        return None, evaluation
    return _kandydat_legacy_z_alokacji(result.selected), evaluation


@app.post(
    "/api/rezerwacje/reguly/symuluj",
    dependencies=[Depends(_wymagaj_modul_rezerwacje)],
)
def symuluj_przydzial_rezerwacji(
    dane: schemas.SymulacjaRegulRezerwacjiIn,
    db: Session = Depends(get_db),
):
    """Podglad R4: reguly, zasob, powod wyboru i bezpieczne alternatywy."""
    if (
        dane.sala_id is not None
        and db.get(models.SalaRezerwacyjna, dane.sala_id) is None
    ):
        raise HTTPException(404, "Brak sali.")
    result = _ocen_przydzial_rezerwacji(
        db,
        data=dane.data,
        godz_od=dane.godz_od,
        osoby=dane.liczba_osob,
        kanal=dane.kanal,
        intent="simulate",
        preferowana_sala_id=dane.sala_id,
    )
    return result.to_dict(expose_exact=True)


def _pierwszy_wolny_slot(db, data, osoby, cfg, teraz_lok):
    """Pierwszy slot zgodny z pełnym zestawem reguł i zasobem stołów."""
    for g, serwis in _sloty_dnia(db, data):                     # blackout → [] (pusto, sam znika)
        przydzial, evaluation = _przydzial_zgodny_z_regulami(
            db,
            data=data,
            godz_od=g,
            osoby=osoby,
            kanal="online",
            intent="quote",
            enforce=False,
        )
        if przydzial and evaluation and evaluation.decision == "allow":
            return (g, serwis)
    return None


def _sprawdz_okno_anulacji(cfg, t, teraz_lok):
    """Egzekwuje rez_anulacja_do_h: zmiany/anulacji online nie później niż X h przed terminem (0 = zawsze)."""
    if cfg.rez_anulacja_do_h and t.godz_od:
        if (datetime.combine(t.data, t.godz_od) - teraz_lok) < timedelta(hours=cfg.rez_anulacja_do_h):
            raise HTTPException(400, f"Zmiany/anulacji można dokonać najpóźniej {cfg.rez_anulacja_do_h} h przed terminem.")


def _online_rez_out(t: models.Termin, stolik=None) -> dict:
    return {"data": str(t.data), "godz_od": _hm(t.godz_od), "godz_do": _hm(t.godz_do),
            "liczba_osob": t.liczba_osob, "nazwisko": t.nazwisko, "status": t.status,
            # Dokładny zasób jest informacją operacyjną. Gość dostaje potwierdzenie
            # miejsca, ale nie układ sali ani nazwę stołu używaną przez obsługę.
            "stolik": None}


def _rez_po_tokenie(db, token: str, *, scope: str = "view"):
    record = reservation_service.validate_management_token(
        db,
        token,
        scope=scope,
        secret=SECRET_KEY,
        now=utcnow_naive(),
    )
    termin = db.get(models.Termin, record.termin_id)
    return termin if termin is not None and termin.kanal == "online" else None


def _odpowiedz_z_obroconym_tokenem(t, raw_token: str, stolik=None) -> dict:
    return {
        **_online_rez_out(t, stolik),
        "management_token": raw_token,
    }


@app.get(
    "/api/online/widget-config",
    dependencies=[Depends(_wymagaj_widget_v2), Depends(_limit_widget_config)],
)
def online_widget_config(db: Session = Depends(get_db)):
    """Publiczny, zredagowany kontrakt rollout'u i informacji dla gościa."""
    cfg = get_lokal_config(db)
    retention_days = max(30, min(int(getattr(cfg, "rezerwacje_retencja_dni", 365) or 365), 3650))
    controller = cfg.nazwa_lokalu or "Lokal"
    address = (getattr(cfg, "rezerwacje_rodo_adres", None) or "").strip() or None
    contact = (getattr(cfg, "rezerwacje_rodo_kontakt", None) or "").strip() or None
    notice_text = None
    if address and contact:
        notice_text = (
            f"Administratorem danych jest {controller}, {address}; kontakt: {contact}. "
            "Dane wykorzystamy do obsługi rezerwacji i kontaktu w jej sprawie. "
            "Przechowujemy je do realizacji wizyty, a następnie przez maksymalnie "
            f"{retention_days} dni, z zastrzeżeniem "
            "obowiązków prawnych. Masz prawo dostępu, sprostowania, usunięcia, "
            "ograniczenia, przenoszenia, sprzeciwu i skargi do PUODO."
        )
    return {
        "version": 2,
        "ready": True,
        "hold_ttl_seconds": PUBLIC_HOLD_TTL_SECONDS,
        "privacy": {
            "notice_version": PUBLIC_PRIVACY_NOTICE_VERSION,
            "notice_label": "Zapoznałem/am się z informacją o przetwarzaniu danych.",
            "notice_text": notice_text,
            "controller": controller,
            "address": address,
            "contact": contact,
            "retention_days": retention_days,
        },
        "marketing": {
            "version": PUBLIC_MARKETING_CONSENT_VERSION,
            "label": "Chcę otrzymywać informacje i oferty tego lokalu.",
            "optional": True,
        },
        "sensitive": {
            "version": PUBLIC_SENSITIVE_CONSENT_VERSION,
            "label": (
                "Zgadzam się na wykorzystanie podanych informacji o alergiach, "
                "diecie lub potrzebach dostępności wyłącznie do obsługi tej wizyty."
            ),
            "optional": True,
        },
    }


@app.post(
    "/api/online/hold",
    status_code=201,
    dependencies=[Depends(_wymagaj_widget_v2), Depends(_limit_hold)],
)
def online_utworz_hold(
    dane: schemas.PublicznyHoldIn,
    request: Request,
    reservation_session: Optional[str] = Header(None, alias="X-Reservation-Session"),
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
    db: Session = Depends(get_db),
):
    """Atomowo trzyma cały zestaw stołów wybrany przez wspólny allocator."""
    raw_session = _wymagaj_publiczny_naglowek(
        reservation_session, nazwa="X-Reservation-Session",
    )
    raw_idempotency = _wymagaj_publiczny_naglowek(
        idempotency_key, nazwa="Idempotency-Key",
    )
    teraz_lok = _teraz_lokalnie() or datetime.now(timezone.utc).replace(tzinfo=None)
    if dane.data < teraz_lok.date():
        raise HTTPException(400, "Nie można rezerwować wstecz.")
    # Singleton konfiguracji może wykonać commit, więc pobieramy go przed blokadą dnia.
    get_lokal_config(db)
    guards = reservation_service.begin_locked_write(db, [dane.data])
    now = utcnow_naive()
    try:
        przydzial, evaluation = _przydzial_zgodny_z_regulami(
            db,
            data=dane.data,
            godz_od=dane.godz_od,
            osoby=dane.liczba_osob,
            kanal="online",
            intent="create",
            enforce=True,
        )
        if not przydzial or evaluation is None or evaluation.godz_do is None:
            raise reservation_service.ReservationError(
                409,
                "NO_TABLE_CANDIDATE",
                "Wybrany termin nie jest już dostępny.",
                rule="table",
            )
        issued = reservation_service.create_public_hold(
            db,
            data=dane.data,
            start=dane.godz_od,
            end=evaluation.godz_do,
            table_ids=przydzial["stoliki"],
            allocation_snapshot=przydzial,
            party_size=dane.liczba_osob,
            buffer_min=evaluation.buffer_min,
            expires_at=now + timedelta(seconds=PUBLIC_HOLD_TTL_SECONDS),
            raw_session=raw_session,
            raw_ip=_public_client_ip(request),
            secret=SECRET_KEY,
            now=now,
            idempotency_key=raw_idempotency,
        )
        _commit_zapis_rezerwacji(db, guards)
    except IntegrityError as exc:
        db.rollback()
        raise reservation_service.translate_integrity_error(exc) from exc
    db.refresh(issued.record)
    return {
        "hold_token": issued.raw_token,
        "expires_at": issued.record.expires_at.isoformat(),
        "replayed": issued.replayed,
        "rezerwacja": {
            "data": str(issued.record.data),
            "godz_od": _hm(issued.record.godz_od),
            "godz_do": _hm(issued.record.godz_do),
            "liczba_osob": issued.record.liczba_osob,
        },
    }


@app.delete(
    "/api/online/hold",
    dependencies=[Depends(_wymagaj_widget_v2), Depends(_limit_hold)],
)
def online_zwolnij_hold(
    public_hold_token: Optional[str] = Header(None, alias="X-Reservation-Hold"),
    reservation_session: Optional[str] = Header(None, alias="X-Reservation-Session"),
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
    db: Session = Depends(get_db),
):
    """Jawnie zwalnia inventory przy powrocie, zmianie slotu lub zamknięciu flow."""
    hold_token = _wymagaj_publiczny_naglowek(
        public_hold_token, nazwa="X-Reservation-Hold",
    )
    raw_session = _wymagaj_publiczny_naglowek(
        reservation_session, nazwa="X-Reservation-Session",
    )
    _wymagaj_publiczny_naglowek(idempotency_key, nazwa="Idempotency-Key")
    current = reservation_service.lookup_public_hold(db, hold_token, secret=SECRET_KEY)
    if current is None:
        raise HTTPException(404, "Nie znaleziono holdu tej sesji.")
    guards = reservation_service.begin_locked_write(db, [current.data])
    released = reservation_service.release_public_hold(
        db,
        hold_token,
        raw_session=raw_session,
        secret=SECRET_KEY,
        now=utcnow_naive(),
    )
    _commit_zapis_rezerwacji(db, guards)
    return {"status": released.state}


@app.get(
    "/api/online/dostepnosc",
    dependencies=[Depends(_wymagaj_rezerwacje_online), Depends(_limit_dostepnosc)],
)
def online_dostepnosc(data: date = Query(...), osoby: int = 2, db: Session = Depends(get_db)):
    """Publicznie: dostępność slotów bez ujawniania zapasu ani układu sali."""
    osoby = max(1, osoby)
    pary = _sloty_dnia(db, data)                                       # blackout → [] (znika sam)
    out = []
    for g, serwis in pary:
        przydzial, evaluation = _przydzial_zgodny_z_regulami(
            db,
            data=data,
            godz_od=g,
            osoby=osoby,
            kanal="online",
            intent="quote",
            enforce=False,
        )
        ma_przydzial = bool(przydzial and evaluation and evaluation.decision == "allow")
        violations = list(evaluation.violations) if evaluation is not None else []
        pacing_pelny = any(
            item.rule in {"pacing_reservations", "pacing_covers"}
            for item in violations
        )
        reguly_blokuja = bool(evaluation and evaluation.decision != "allow")
        # Pola liczbowe pozostają zgodne ze starszym widgetem, ale są teraz wyłącznie
        # flagą 0/1. Publiczny kanał nie poznaje liczby ani rodzaju wolnych zasobów.
        dostepny = 1 if ma_przydzial else 0
        out.append({"godz_od": _hm(g), "wolne": dostepny,
                    "wolne_stoly": dostepny, "dostepny": bool(dostepny),
                    "pacing_pelny": pacing_pelny,
                    "reguly_blokuja": reguly_blokuja,
                    "decision": (evaluation.decision if evaluation else "deny"),
                    "serwis": (serwis.nazwa if serwis and serwis.nazwa else None)})
    return {"data": str(data), "osoby": osoby, "sloty": out}


@app.get(
    "/api/online/alternatywy",
    dependencies=[Depends(_wymagaj_rezerwacje_online), Depends(_limit_alternatywy)],
)
def online_alternatywy(
    data: date = Query(...),
    osoby: int = 2,
    limit: int = Query(3, ge=1, le=6),
    db: Session = Depends(get_db),
):
    """Najbliższe bezpieczne propozycje bez ujawniania układu ani zapasu sali."""
    osoby = max(1, min(int(osoby or 2), 500))
    teraz_lok = _teraz_lokalnie() or datetime.now(timezone.utc).replace(tzinfo=None)
    start = max(data, teraz_lok.date())
    alternatives = []
    for offset in range(0, 31):
        current = start + timedelta(days=offset)
        for godz_od, serwis in _sloty_dnia(db, current):
            przydzial, evaluation = _przydzial_zgodny_z_regulami(
                db,
                data=current,
                godz_od=godz_od,
                osoby=osoby,
                kanal="online",
                intent="quote",
                enforce=False,
            )
            if przydzial and evaluation and evaluation.decision == "allow":
                alternatives.append({
                    "data": str(current),
                    "godz_od": _hm(godz_od),
                    "serwis": serwis.nazwa if serwis and serwis.nazwa else None,
                })
                if len(alternatives) >= limit:
                    return {"osoby": osoby, "alternatywy": alternatives}
    return {"osoby": osoby, "alternatywy": alternatives}


@app.get(
    "/api/online/najblizszy-termin",
    dependencies=[Depends(_wymagaj_rezerwacje_online), Depends(_limit_alternatywy)],
)
def online_najblizszy_termin(osoby: int = 2, od: Optional[date] = None, dni: int = 14,
                             db: Session = Depends(get_db)):
    """Publicznie: pierwszy dzień (od 'od', domyślnie dziś) z wolnym stołem dla 'osoby'. Skanuje do
    'dni' w przód, respektując okno wyprzedzenia. Zwraca {data, slot:{godz_od,serwis}} lub {data:null}."""
    osoby = max(1, osoby)
    cfg = get_lokal_config(db)
    teraz_lok = _teraz_lokalnie() or datetime.now(timezone.utc).replace(tzinfo=None)
    start = max(od or teraz_lok.date(), teraz_lok.date())          # nigdy wstecz
    limit_dni = max(1, min(int(dni or 14), 60))                    # twardy limit skanowania
    for i in range(limit_dni):
        d = start + timedelta(days=i)
        slot = _pierwszy_wolny_slot(db, d, osoby, cfg, teraz_lok)
        if slot:
            g, serwis = slot
            return {"data": str(d), "osoby": osoby,
                    "slot": {"godz_od": _hm(g), "serwis": (serwis.nazwa if serwis and serwis.nazwa else None)}}
    return {"data": None, "osoby": osoby, "slot": None}


@app.post(
    "/api/online/rezerwacja",
    status_code=201,
    dependencies=[Depends(_wymagaj_widget_v2), Depends(_limit_create_online)],
)
def online_rezerwacja(
    dane: schemas.OnlineRezerwacjaIn,
    request: Request,
    http_response: Response,
    reservation_session: Optional[str] = Header(None, alias="X-Reservation-Session"),
    reservation_hold: Optional[str] = Header(None, alias="X-Reservation-Hold"),
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
    db: Session = Depends(get_db),
):
    """Publicznie tworzy rezerwację, token hash-only i ledger w jednej transakcji."""
    if not dane.nazwisko or not dane.nazwisko.strip():
        raise HTTPException(400, "Podaj imię/nazwisko.")
    if (dane.liczba_osob or 0) < 1:
        raise HTTPException(400, "Liczba osób musi być dodatnia.")
    teraz_lok = _teraz_lokalnie() or datetime.now(timezone.utc).replace(tzinfo=None)
    dzis_lokalnie = teraz_lok.date()
    if dane.data < dzis_lokalnie:
        raise HTTPException(400, "Nie można rezerwować wstecz.")

    # Singleton może zostać utworzony leniwie i wykonać commit; robimy to przed blokadą dnia.
    cfg = get_lokal_config(db)
    raw_session = _wymagaj_publiczny_naglowek(
        reservation_session, nazwa="X-Reservation-Session",
    )
    raw_hold = _wymagaj_publiczny_naglowek(
        reservation_hold, nazwa="X-Reservation-Hold",
    )
    _wymagaj_publiczny_naglowek(idempotency_key, nazwa="Idempotency-Key")
    _waliduj_prywatnosc_widgetu(dane, cfg)
    guards = reservation_service.begin_locked_write(db, [dane.data])
    teraz = utcnow_naive()
    operation = "reservation.create.online:v2"
    try:
        idem = reservation_service.begin_idempotency(
            db,
            operation=operation,
            raw_key=idempotency_key,
            payload=dane.model_dump(mode="json"),
            secret=SECRET_KEY,
            now=teraz,
        )
        if idem.replayed:
            replayed = _termin_z_replay_idempotencji(db, idem)
            replay_table = db.get(models.Stolik, replayed.stolik_id) if replayed.stolik_id else None
            issued = reservation_service.create_management_token(
                db,
                termin_id=replayed.id,
                scopes=(
                    "view", "confirm", "edit", "cancel", "payment:retry",
                    "data:export", "data:delete",
                ),
                secret=SECRET_KEY,
                now=teraz,
                expires_at=max(
                    teraz + timedelta(days=30),
                    datetime.combine(replayed.data + timedelta(days=30), time.min),
                ),
                idempotency_key=idempotency_key,
                operation=operation,
            )
            response = {
                "management_token": issued.raw_token,
                # Alias przejściowy dla starszych klientów; token nie trafia do kolumny Termin.
                "token": issued.raw_token,
                "rezerwacja": _online_rez_out(replayed, replay_table),
            }
            payment = (
                db.query(models.Platnosc)
                .filter_by(termin_id=replayed.id)
                .order_by(models.Platnosc.id.desc())
                .first()
            )
            if payment is not None:
                response["platnosc"] = _publiczna_platnosc_rezerwacji(
                    payment, replayed,
                )
            cookie_expires_at = issued.record.expires_at
            db.rollback()
            _ustaw_publiczne_cookie_zarzadzania(
                http_response,
                issued.raw_token,
                expires_at=cookie_expires_at,
            )
            return response

        ip = request.client.host if request.client else "?"
        if not ratelimit.zuzyj_kwote(
            f"online-rez:{ip}", str(dzis_lokalnie), ONLINE_LIMIT_IP_DZIENNY,
        ):
            raise HTTPException(429, "Przekroczono dzienny limit rezerwacji online z tego adresu.")
        if dane.telefon or dane.email:
            dzisiaj_online = db.query(models.Termin).filter(
                models.Termin.kanal == "online",
                models.Termin.data == dane.data,
                models.Termin.status.in_(REZ_AKTYWNE),
            ).all()
            ile = sum(
                1 for rezerwacja in dzisiaj_online
                if (dane.telefon and rezerwacja.telefon == dane.telefon)
                or (dane.email and rezerwacja.email == dane.email)
            )
            if ile >= ONLINE_LIMIT_DZIENNY:
                raise HTTPException(429, "Przekroczono dzienny limit rezerwacji online.")

        hold = reservation_service.validate_public_hold(
            db,
            raw_hold,
            raw_session=raw_session,
            secret=SECRET_KEY,
            now=teraz,
        )
        if (
            hold.data != dane.data
            or hold.godz_od != dane.godz_od
            or int(hold.liczba_osob) != int(dane.liczba_osob)
        ):
            raise reservation_service.ReservationError(
                409,
                "PUBLIC_HOLD_REQUEST_MISMATCH",
                "Dane rezerwacji nie odpowiadają trzymanemu terminowi.",
                rule="public_hold",
            )
        stoliki = [hold.stolik_id, *(hold.stoliki_dodatkowe or [])]
        godz_do = hold.godz_do
        evaluation = _ocen_reguly_slotu(
            db,
            data=dane.data,
            godz_od=dane.godz_od,
            godz_do=godz_do,
            liczba_osob=dane.liczba_osob,
            kanal="online",
            sala_id=_sala_dla_stolikow(db, set(stoliki)),
            intent="create",
        )
        reservation_rules.enforce_rule_evaluation(evaluation)
        przydzial = getattr(hold, "allocation_snapshot", None) or {
            "stoliki": stoliki,
            "typ": "kombinacja" if len(stoliki) > 1 else "stolik",
        }
        stolik = db.get(models.Stolik, stoliki[0])
        status = "potwierdzona" if cfg.rezerwacje_auto_potwierdzenie else "rezerwacja"
        t = models.Termin(
            data=dane.data,
            nazwisko=dane.nazwisko.strip(),
            telefon=dane.telefon,
            email=dane.email,
            kanal_komunikacji=dane.kanal_komunikacji,
            liczba_osob=dane.liczba_osob,
            notatka=dane.notatka,
            status=status,
            zadatek=0.0,
            utworzono_at=teraz,
            godz_od=dane.godz_od,
            godz_do=godz_do,
            stolik_id=stoliki[0],
            stoliki_dodatkowe=(stoliki[1:] or None),
            auto_przydzielony=True,
            rodzaj="stolik",
            kanal="online",
            token_potwierdzenia=None,
            potwierdzono_at=(teraz if status == "potwierdzona" else None),
        )
        _ustaw_proweniencje_przydzialu(t, przydzial)
        db.add(t); db.flush()
        reservation_service.consume_public_hold(
            db,
            raw_hold,
            raw_session=raw_session,
            termin_id=t.id,
            secret=SECRET_KEY,
            now=teraz,
        )
        _zastap_ledger_terminu(
            db,
            t,
            enforce_pacing=True,
            candidates=[{"stoliki": stoliki}],
            evaluation=evaluation,
            intent="create",
        )
        reservation_audit.add_reservation_audit(
            db,
            termin=t,
            action="create",
            actor_kind="guest",
            after=t,
            pii_changed=_utworzone_pii(t),
        )
        _zapisz_publiczna_prywatnosc(
            db,
            dane=dane,
            cfg=cfg,
            request=request,
            now=teraz,
            termin_id=t.id,
        )
        issued = reservation_service.create_management_token(
            db,
            termin_id=t.id,
            scopes=(
                "view", "confirm", "edit", "cancel", "payment:retry",
                "data:export", "data:delete",
            ),
            secret=SECRET_KEY,
            now=teraz,
            expires_at=max(
                teraz + timedelta(days=30),
                datetime.combine(t.data + timedelta(days=30), time.min),
            ),
            idempotency_key=idempotency_key,
            operation=operation,
        )
        odp = {
            "management_token": issued.raw_token,
            "token": issued.raw_token,
            "rezerwacja": _online_rez_out(t, stolik),
        }
        payment_policy = reservation_payments.resolve_policy(
            db,
            dane.data,
            evaluation.service_id,
            dane.liczba_osob,
            "online",
        )
        payment_provider = (
            integracje.provider_platnosci_wymaganej()
            if payment_policy is not None and payment_policy.required
            else "sandbox"
        )
        p, payment_command = reservation_payments.create_payment_for_reservation(
            db,
            t,
            payment_policy,
            provider=payment_provider,
            now=teraz,
            business_today=dzis_lokalnie,
            service_id=evaluation.service_id,
            operation_key="initial",
            actor_kind="guest",
        )
        if p is not None:
            if payment_provider == "sandbox" and payment_command is not None:
                # Tryb demonstracyjny nie udaje odpowiedzi providera. Udostępnia
                # jedynie lokalny powrót; opłacenie może potwierdzić operator.
                p.link = "/?platnosc=sandbox&rezerwuj"
                payment_command.stan = "succeeded"
                payment_command.finished_at = teraz
                payment_command.updated_at = teraz
            odp["platnosc"] = _publiczna_platnosc_rezerwacji(p, t)
        reservation_communication.enqueue_reservation(
            db,
            t,
            "confirmation",
            cfg=cfg,
            dedupe_key=(
                f"reservation:{t.id}:confirmation:online:"
                f"{secrets.token_hex(16)}"
            ),
            actor_kind="guest",
        )
        reservation_communication.schedule_reminder(
            db, t, actor_kind="guest",
        )
        reservation_service.complete_idempotency(
            idem.record,
            # Raw management token nigdy nie trafia nawet do szyfrowanego replay cache.
            response={key: value for key, value in odp.items() if key not in {"token", "management_token"}},
            http_status=201,
            termin_id=t.id,
            now=teraz,
        )
        _commit_zapis_rezerwacji(db, guards)
    except IntegrityError as exc:
        db.rollback()
        raise reservation_service.translate_integrity_error(exc) from exc

    db.refresh(t)
    wyslij_push_do_adminow(
        db,
        "Rezerwacja online",
        f"{t.nazwisko} — {t.data} {_hm(t.godz_od) or ''}".strip(),
        url="/",
    )
    _ustaw_publiczne_cookie_zarzadzania(
        http_response,
        issued.raw_token,
        expires_at=issued.record.expires_at,
    )
    return odp


def _platnosc_rezerwacji(db: Session, termin_id: int) -> Optional[models.Platnosc]:
    return (
        db.query(models.Platnosc)
        .filter(models.Platnosc.termin_id == termin_id)
        .order_by(models.Platnosc.id.desc())
        .first()
    )


def _publiczna_platnosc_rezerwacji(
    payment: Optional[models.Platnosc],
    reservation: models.Termin,
) -> dict:
    return reservation_payments.payment_public_dict(
        payment,
        reservation_active=reservation.status in REZ_AKTYWNE,
    )


@app.get(
    "/api/online/zarzadzanie/platnosc",
    dependencies=[Depends(_wymagaj_rezerwacje_online), Depends(_limit_payment_status)],
)
def online_zarzadzanie_platnosc(
    request: Request,
    reservation_token: Optional[str] = Header(None, alias="X-Reservation-Token"),
    db: Session = Depends(get_db),
):
    """Status jest projekcją webhooka; powrót z Checkout nigdy nie oznacza sukcesu."""
    token = _publiczny_token_zarzadzania(request, reservation_token)
    record = reservation_service.validate_management_token(
        db,
        token,
        scope="view",
        secret=SECRET_KEY,
        now=utcnow_naive(),
    )
    t = db.get(models.Termin, record.termin_id)
    if t is None or t.kanal != "online":
        raise HTTPException(404, "Nie znaleziono rezerwacji.")
    payment = _platnosc_rezerwacji(db, t.id)
    return {
        "rezerwacja": _online_rez_out(
            t, db.get(models.Stolik, t.stolik_id) if t.stolik_id else None,
        ),
        "platnosc": _publiczna_platnosc_rezerwacji(payment, t),
    }


@app.post(
    "/api/online/zarzadzanie/platnosc/ponow",
    dependencies=[Depends(_wymagaj_rezerwacje_online), Depends(_limit_payment_action)],
)
def online_zarzadzanie_platnosc_ponow(
    request: Request,
    http_response: Response,
    reservation_token: Optional[str] = Header(None, alias="X-Reservation-Token"),
    reservation_session: Optional[str] = Header(None, alias="X-Reservation-Session"),
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
    db: Session = Depends(get_db),
):
    token = _publiczny_token_zarzadzania(
        request,
        reservation_token,
        mutation=True,
        reservation_session=reservation_session,
    )
    key = _wymagaj_publiczny_naglowek(
        idempotency_key, nazwa="Idempotency-Key",
    )
    t, guards = _zablokuj_rezerwacje_online(db, token)
    if t is None:
        raise HTTPException(404, "Nie znaleziono rezerwacji.")
    now = utcnow_naive()
    issued = reservation_service.consume_and_rotate_management_token(
        db,
        token,
        operation="payment:retry",
        idempotency_key=key,
        payload={},
        secret=SECRET_KEY,
        now=now,
    )
    if issued.replayed:
        payment = _platnosc_rezerwacji(db, t.id)
        cookie_expires_at = issued.record.expires_at
        db.rollback()
        response = {
            "management_token": issued.raw_token,
            "token": issued.raw_token,
            "rezerwacja": _online_rez_out(
                t, db.get(models.Stolik, t.stolik_id) if t.stolik_id else None,
            ),
            "platnosc": _publiczna_platnosc_rezerwacji(payment, t),
        }
        _ustaw_publiczne_cookie_zarzadzania(
            http_response,
            issued.raw_token,
            expires_at=cookie_expires_at,
        )
        return response
    previous = _platnosc_rezerwacji(db, t.id)
    if previous is None:
        raise reservation_payments.PaymentDomainError(
            "PAYMENT_NOT_FOUND", "Ta rezerwacja nie ma płatności do ponowienia.",
        )
    retried, command = reservation_payments.retry_payment_for_reservation(
        db,
        previous,
        t,
        operation_key=key,
        now=now,
        business_today=(
            (_teraz_lokalnie() or datetime.now(timezone.utc).replace(tzinfo=None)).date()
        ),
        actor_kind="guest",
    )
    if retried.provider == "sandbox":
        retried.link = "/?platnosc=sandbox&rezerwuj"
        command.stan = "succeeded"
        command.finished_at = now
        command.updated_at = now
    _commit_zapis_rezerwacji(db, guards)
    db.refresh(retried)
    response = {
        "management_token": issued.raw_token,
        "token": issued.raw_token,
        "rezerwacja": _online_rez_out(
            t, db.get(models.Stolik, t.stolik_id) if t.stolik_id else None,
        ),
        "platnosc": _publiczna_platnosc_rezerwacji(retried, t),
    }
    _ustaw_publiczne_cookie_zarzadzania(
        http_response,
        issued.raw_token,
        expires_at=issued.record.expires_at,
    )
    return response


@app.post("/api/online/platnosci/stripe/webhook")
async def stripe_reservation_payment_webhook(
    request: Request,
    stripe_signature: Optional[str] = Header(None, alias="Stripe-Signature"),
):
    """Weryfikuje podpis na surowym body i zapisuje wyłącznie minimalny inbox."""
    if not stripe_signature:
        raise HTTPException(400, "Brak podpisu webhooka.")
    raw_body = await request.body()
    if not raw_body or len(raw_body) > 256 * 1024:
        raise HTTPException(413, "Niepoprawny rozmiar webhooka.")
    try:
        result = reservation_payment_worker.ingest_payment_webhook(
            raw_body,
            stripe_signature,
            received_at=utcnow_naive(),
        )
    except reservation_payment_worker.PaymentIntegrationDisabled as exc:
        raise HTTPException(503, "Integracja płatnicza nie jest aktywna.") from exc
    except reservation_payments.PaymentDomainError:
        raise
    except Exception as exc:
        logger.warning("Odrzucono webhook Stripe R5c: %s", type(exc).__name__)
        raise HTTPException(400, "Niepoprawny webhook płatniczy.") from exc
    return {
        "received": True,
        "duplicate": result.duplicate,
        "state": result.state,
    }


def _nalicz_no_show_fee(db, t, *, commit=True):
    """Opłata za no-show jako należność ledger, bez udawania obciążenia karty."""
    fee = get_lokal_config(db).no_show_fee or 0
    if fee > 0 and t is not None and t.rodzaj == "stolik":
        # To należność księgowa, nie pozorna transakcja online. Dzięki osobnemu
        # providerowi produkcja nie pokazuje linku sandbox ani nie uruchamia Stripe.
        payment = platnosci.utworz_platnosc(
            db,
            t.id,
            fee,
            commit=False,
            provider_override="ledger",
        )
        amount_minor = int(round(float(fee) * 100))
        payment.kwota_minor = amount_minor
        payment.przechwycono_minor = 0
        payment.zwrocono_minor = 0
        payment.waluta = "PLN"
        payment.rodzaj = "no_show"
        payment.refund_status = "brak"
        payment.tryb_przechwycenia = "automatic"
        payment.link = None
        payment.zaktualizowano_at = utcnow_naive()
        payment.version = 0
        if commit:
            db.commit()
            db.refresh(payment)
        return payment
    return None


@app.post(
    "/api/online/popyt/odrzucony",
    status_code=201,
    dependencies=[Depends(_wymagaj_rezerwacje_online), Depends(_limit_demand_online)],
)
def online_odrzucony_popyt(
    dane: schemas.OdrzuconyPopytIn,
    request: Request,
    reservation_session: Optional[str] = Header(None, alias="X-Reservation-Session"),
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
    db: Session = Depends(get_db),
):
    """Anonimowo rejestruje wyłącznie zweryfikowany przez serwer brak dostępności."""
    _wymagaj_publiczny_naglowek(
        reservation_session, nazwa="X-Reservation-Session",
    )
    raw_key = _wymagaj_publiczny_naglowek(
        idempotency_key, nazwa="Idempotency-Key",
    )
    now_local = _teraz_lokalnie() or datetime.now(timezone.utc).replace(tzinfo=None)
    if dane.data < now_local.date():
        raise HTTPException(400, "Nie można rejestrować popytu wstecz.")
    safe_payload = dane.model_dump(mode="json")
    identity = reservation_demand.event_identity(
        source_kind="availability",
        raw_key=raw_key,
        payload=safe_payload,
        secret=SECRET_KEY,
    )
    # Singleton konfiguracji może commitować; pobieramy go przed blokadą dnia.
    get_lokal_config(db)
    guards = reservation_service.begin_locked_write(db, [dane.data])
    existing = db.query(models.ReservationDemandEvent).filter(
        models.ReservationDemandEvent.source_kind == "availability",
        models.ReservationDemandEvent.event_key_hash == identity.key_hash,
    ).one_or_none()
    if existing is not None:
        _event, replayed = reservation_demand.record_event(
            db,
            source_kind="availability",
            channel="online",
            requested_date=existing.requested_date,
            requested_time=existing.requested_time,
            party_size=existing.party_size,
            classification=reservation_demand.DemandClassification(
                existing.reason_code, existing.resource_kind,
            ),
            identity=identity,
            captured_at=existing.captured_at,
        )
        db.rollback()
        return {"status": "zapisane", "replayed": replayed}

    classification = _klasyfikacja_popytu_rezerwacji(
        db,
        data=dane.data,
        godz_od=dane.godz_od,
        osoby=dane.liczba_osob,
        kanal="online",
    )
    if classification.reason_code == "operator_decision":
        db.rollback()
        raise reservation_service.ReservationError(
            409,
            "DEMAND_AVAILABILITY_EXISTS",
            "Dla tego zapytania nadal istnieje dostępny termin.",
            rule="availability",
        )
    try:
        _event, replayed = reservation_demand.record_event(
            db,
            source_kind="availability",
            channel="online",
            requested_date=dane.data,
            requested_time=dane.godz_od,
            party_size=dane.liczba_osob,
            classification=classification,
            identity=identity,
            captured_at=utcnow_naive(),
        )
        reservation_service.touch_days(guards)
        db.commit()
    except IntegrityError as exc:
        # Równoległy retry może wygrać UNIQUE między query i flush. Cała lokalna
        # transakcja jest cofana, a odpowiedź odtwarzamy z anonimowego owner row.
        db.rollback()
        owner = db.query(models.ReservationDemandEvent).filter(
            models.ReservationDemandEvent.source_kind == "availability",
            models.ReservationDemandEvent.event_key_hash == identity.key_hash,
        ).one_or_none()
        if owner is None:
            raise reservation_service.translate_integrity_error(exc) from exc
        if not secrets.compare_digest(
            owner.request_fingerprint, identity.request_fingerprint,
        ):
            raise reservation_service.ReservationError(
                409,
                "DEMAND_IDEMPOTENCY_KEY_REUSED",
                "Ten klucz Idempotency-Key został użyty z innymi danymi.",
                rule="idempotency",
            ) from exc
        replayed = True
    return {"status": "zapisane", "replayed": replayed}


def _publiczny_wpis_waitlisty(w, *, replayed=False):
    return {
        "replayed": bool(replayed),
        "wpis": {
            "data": str(w.data),
            "godz_od": _hm(w.godz_od),
            "liczba_osob": w.liczba_osob,
            "nazwisko": w.nazwisko,
            "status": w.status,
        },
    }


def _publiczny_replay_waitlisty_po_konflikcie(db, owner_identity, exc):
    db.rollback()
    if owner_identity is None:
        raise reservation_service.translate_integrity_error(exc) from exc
    owner = db.query(models.ListaOczekujacych).filter(
        models.ListaOczekujacych.create_key_hash == owner_identity.key_hash,
    ).one_or_none()
    if owner is None:
        raise reservation_service.translate_integrity_error(exc) from exc
    if not secrets.compare_digest(
        owner.create_request_fingerprint,
        owner_identity.request_fingerprint,
    ):
        raise reservation_service.ReservationError(
            409,
            "WAITLIST_CREATE_KEY_REUSED",
            "Ten klucz Idempotency-Key został użyty z innymi danymi.",
            rule="idempotency",
        ) from exc
    return _publiczny_wpis_waitlisty(owner, replayed=True)


@app.post(
    "/api/online/lista-oczekujacych",
    status_code=201,
    dependencies=[Depends(_wymagaj_widget_v2), Depends(_limit_waitlist_online)],
)
def online_lista_oczekujacych(
    dane: schemas.ListaOczekujacychIn,
    request: Request,
    reservation_session: Optional[str] = Header(None, alias="X-Reservation-Session"),
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
    db: Session = Depends(get_db),
):
    """Publiczny zapis z replayem odbudowywanym z wiersza właściciela."""
    if not dane.nazwisko or not dane.nazwisko.strip():
        raise HTTPException(400, "Podaj imię/nazwisko.")
    teraz_lok = _teraz_lokalnie() or datetime.now(timezone.utc).replace(tzinfo=None)
    dzis_lokalnie = teraz_lok.date()
    if dane.data < dzis_lokalnie:
        raise HTTPException(400, "Nie można zapisać się wstecz.")
    cfg = get_lokal_config(db)
    _wymagaj_publiczny_naglowek(
        reservation_session, nazwa="X-Reservation-Session",
    )
    _wymagaj_publiczny_naglowek(
        idempotency_key, nazwa="Idempotency-Key",
    )
    _waliduj_prywatnosc_widgetu(dane, cfg)

    operation = "waitlist.create.online:v2"
    owner_identity = reservation_service.required_idempotency_identity(
        operation=operation,
        raw_key=idempotency_key,
        payload=dane.model_dump(mode="json"),
        secret=SECRET_KEY,
    )

    guards = reservation_service.begin_locked_write(db, [dane.data])
    existing = db.query(models.ListaOczekujacych).filter(
        models.ListaOczekujacych.create_key_hash == owner_identity.key_hash,
    ).one_or_none()
    if existing is not None:
        if not secrets.compare_digest(
            existing.create_request_fingerprint,
            owner_identity.request_fingerprint,
        ):
            db.rollback()
            raise reservation_service.ReservationError(
                409,
                "WAITLIST_CREATE_KEY_REUSED",
                "Ten klucz Idempotency-Key został użyty z innymi danymi.",
                rule="idempotency",
            )
        response = _publiczny_wpis_waitlisty(existing, replayed=True)
        db.rollback()
        return response

    ip = request.client.host if request.client else "?"
    if not ratelimit.zuzyj_kwote(
        f"online-wait:{ip}", str(dzis_lokalnie), ONLINE_LIMIT_IP_DZIENNY,
    ):
        db.rollback()
        raise HTTPException(429, "Przekroczono dzienny limit zapisów z tego adresu.")
    if dane.telefon or dane.email:
        dzisiaj = db.query(models.ListaOczekujacych).filter(
            models.ListaOczekujacych.data == dane.data,
            models.ListaOczekujacych.status.in_(
                reservation_service.WAITLIST_ACTIVE_STATUSES
            ),
        ).all()
        ile = sum(
            1 for row in dzisiaj
            if (dane.telefon and row.telefon == dane.telefon)
            or (dane.email and row.email == dane.email)
        )
        if ile >= ONLINE_LIMIT_DZIENNY:
            db.rollback()
            raise HTTPException(429, "Jesteś już na liście oczekujących na ten dzień.")

    classification = _klasyfikacja_popytu_rezerwacji(
        db,
        data=dane.data,
        godz_od=dane.godz_od,
        osoby=dane.liczba_osob,
        kanal="online",
    )
    now = utcnow_naive()
    w = models.ListaOczekujacych(
        data=dane.data,
        godz_od=dane.godz_od,
        liczba_osob=dane.liczba_osob,
        nazwisko=dane.nazwisko.strip(),
        telefon=dane.telefon,
        email=dane.email,
        notatka=dane.notatka,
        kanal_komunikacji=dane.kanal_komunikacji,
        status="oczekuje",
        kanal="online",
        token=None,
        utworzono_at=now,
        create_key_hash=(owner_identity.key_hash if owner_identity else None),
        create_request_fingerprint=(
            owner_identity.request_fingerprint if owner_identity else None
        ),
        demand_reason_code=classification.reason_code,
        demand_resource_kind=classification.resource_kind,
    )
    db.add(w)
    try:
        db.flush()
    except IntegrityError as exc:
        return _publiczny_replay_waitlisty_po_konflikcie(
            db, owner_identity, exc,
        )
    if dane.liczba_osob is not None:
        safe_event_payload = {
            "data": dane.data,
            "godz_od": dane.godz_od,
            "liczba_osob": dane.liczba_osob,
            "kanal": "online",
        }
        event_identity = reservation_demand.event_identity(
            source_kind="waitlist",
            raw_key=(idempotency_key or secrets.token_urlsafe(32)),
            payload=safe_event_payload,
            secret=SECRET_KEY,
        )
        reservation_demand.record_event(
            db,
            source_kind="waitlist",
            channel="online",
            requested_date=dane.data,
            requested_time=dane.godz_od,
            party_size=dane.liczba_osob,
            classification=classification,
            identity=event_identity,
            captured_at=now,
        )
    _zapisz_publiczna_prywatnosc(
        db,
        dane=dane,
        cfg=cfg,
        request=request,
        now=now,
        waitlist_id=w.id,
    )
    reservation_service.touch_days(guards)
    try:
        db.commit()
    except IntegrityError as exc:
        return _publiczny_replay_waitlisty_po_konflikcie(
            db, owner_identity, exc,
        )
    db.refresh(w)
    wyslij_push_do_adminow(
        db,
        "Nowy wpis na liście oczekujących",
        f"{w.nazwisko} — {w.data} {_hm(w.godz_od) or ''}".strip(),
        url="/",
    )
    return _publiczny_wpis_waitlisty(w)


def _online_rezerwacja_get_core(token: str, db: Session):
    t = _rez_po_tokenie(db, token)
    if not t:
        raise HTTPException(404, "Nie znaleziono rezerwacji.")
    return _online_rez_out(t, db.get(models.Stolik, t.stolik_id) if t.stolik_id else None)


@app.get(
    "/api/online/zarzadzanie/rezerwacja",
    dependencies=[Depends(_wymagaj_rezerwacje_online), Depends(_limit_management)],
)
def online_zarzadzanie_rezerwacja_get(
    request: Request,
    reservation_token: Optional[str] = Header(None, alias="X-Reservation-Token"),
    db: Session = Depends(get_db),
):
    token = _publiczny_token_zarzadzania(request, reservation_token)
    return _online_rezerwacja_get_core(token, db)


def _online_rezerwacja_potwierdz_core(
    token: str,
    idempotency_key: Optional[str],
    db: Session,
):
    t, guards = _zablokuj_rezerwacje_online(db, token)
    if not t:
        raise HTTPException(404, "Nie znaleziono rezerwacji.")
    now = utcnow_naive()
    issued = reservation_service.consume_and_rotate_management_token(
        db,
        token,
        operation="confirm",
        idempotency_key=idempotency_key,
        payload={},
        secret=SECRET_KEY,
        now=now,
    )
    if issued.replayed:
        db.rollback()
        return _odpowiedz_z_obroconym_tokenem(
            t, issued.raw_token, db.get(models.Stolik, t.stolik_id) if t.stolik_id else None,
        )
    if t.status == "rezerwacja":
        before = reservation_audit.reservation_snapshot(t)
        t.status = "potwierdzona"; t.potwierdzono_at = now
        reservation_audit.add_reservation_audit(
            db, termin=t, action="status", actor_kind="guest", before=before, after=t,
        )
    _commit_zapis_rezerwacji(db, guards); db.refresh(t)
    return _odpowiedz_z_obroconym_tokenem(
        t, issued.raw_token, db.get(models.Stolik, t.stolik_id) if t.stolik_id else None,
    )


@app.post(
    "/api/online/zarzadzanie/potwierdz",
    dependencies=[Depends(_wymagaj_rezerwacje_online), Depends(_limit_management)],
)
def online_zarzadzanie_potwierdz(
    request: Request,
    http_response: Response,
    reservation_token: Optional[str] = Header(None, alias="X-Reservation-Token"),
    reservation_session: Optional[str] = Header(None, alias="X-Reservation-Session"),
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
    db: Session = Depends(get_db),
):
    token = _publiczny_token_zarzadzania(
        request,
        reservation_token,
        mutation=True,
        reservation_session=reservation_session,
    )
    payload = _online_rezerwacja_potwierdz_core(token, idempotency_key, db)
    _odswiez_publiczne_cookie_z_odpowiedzi(http_response, payload, db)
    return payload


def _online_rezerwacja_odwolaj_core(
    token: str,
    idempotency_key: Optional[str],
    db: Session,
):
    t, guards = _zablokuj_rezerwacje_online(db, token)
    if not t:
        raise HTTPException(404, "Nie znaleziono rezerwacji.")
    now = utcnow_naive()
    issued = reservation_service.consume_and_rotate_management_token(
        db,
        token,
        operation="cancel",
        idempotency_key=idempotency_key,
        payload={},
        secret=SECRET_KEY,
        now=now,
    )
    if issued.replayed:
        payment = _platnosc_rezerwacji(db, t.id)
        response = _odpowiedz_z_obroconym_tokenem(
            t, issued.raw_token, db.get(models.Stolik, t.stolik_id) if t.stolik_id else None,
        )
        if payment is not None:
            response["platnosc"] = _publiczna_platnosc_rezerwacji(payment, t)
        db.rollback()
        return response
    stolik_wolny = set()
    okno = None
    if t.status in REZ_AKTYWNE:
        before = reservation_audit.reservation_snapshot(t)
        teraz_lok = _teraz_lokalnie() or datetime.now(timezone.utc).replace(tzinfo=None)
        _sprawdz_okno_anulacji(get_lokal_config(db), t, teraz_lok)
        stolik_wolny = _stoly_terminu(t)
        okno = (t.data, t.godz_od, _koniec_okna(db, t)) if t.godz_od else None
        t.status = "odwolana"; t.odwolano_at = now
        reservation_service.release_termin_allocation(db, t.id)
        reservation_payments.request_reservation_cancellation_settlement(
            db,
            t,
            now=now,
            actor_kind="guest",
        )
        reservation_audit.add_reservation_audit(
            db, termin=t, action="cancel", actor_kind="guest", before=before, after=t,
        )
        reservation_communication.cancel_pending(db, t.id)
        reservation_communication.enqueue_reservation(
            db, t, "cancellation", actor_kind="guest",
        )
    _commit_zapis_rezerwacji(db, guards); db.refresh(t)
    if stolik_wolny and okno:
        _bezpiecznie_po_zwolnieniu_stolu(db, *okno)
    response = _odpowiedz_z_obroconym_tokenem(
        t, issued.raw_token, db.get(models.Stolik, t.stolik_id) if t.stolik_id else None,
    )
    payment = _platnosc_rezerwacji(db, t.id)
    if payment is not None:
        response["platnosc"] = _publiczna_platnosc_rezerwacji(payment, t)
    return response


@app.post(
    "/api/online/zarzadzanie/odwolaj",
    dependencies=[Depends(_wymagaj_rezerwacje_online), Depends(_limit_management)],
)
def online_zarzadzanie_odwolaj(
    request: Request,
    http_response: Response,
    reservation_token: Optional[str] = Header(None, alias="X-Reservation-Token"),
    reservation_session: Optional[str] = Header(None, alias="X-Reservation-Session"),
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
    db: Session = Depends(get_db),
):
    token = _publiczny_token_zarzadzania(
        request,
        reservation_token,
        mutation=True,
        reservation_session=reservation_session,
    )
    payload = _online_rezerwacja_odwolaj_core(token, idempotency_key, db)
    _odswiez_publiczne_cookie_z_odpowiedzi(http_response, payload, db)
    return payload


def _online_rezerwacja_edytuj_core(
    token: str,
    dane: schemas.OnlineEdytujIn,
    idempotency_key: Optional[str],
    db: Session,
):
    """Idempotentna zmiana terminu/liczby osób z rotacją capability tokenu."""
    t, guards = _zablokuj_rezerwacje_online(db, token, [dane.data])
    if not t:
        raise HTTPException(404, "Nie znaleziono rezerwacji.")
    now = utcnow_naive()
    issued = reservation_service.consume_and_rotate_management_token(
        db,
        token,
        operation="edit",
        idempotency_key=idempotency_key,
        payload=dane.model_dump(mode="json"),
        secret=SECRET_KEY,
        now=now,
    )
    if issued.replayed:
        payment = _platnosc_rezerwacji(db, t.id)
        response = _odpowiedz_z_obroconym_tokenem(
            t, issued.raw_token, db.get(models.Stolik, t.stolik_id) if t.stolik_id else None,
        )
        if payment is not None:
            response["platnosc"] = _publiczna_platnosc_rezerwacji(payment, t)
        db.rollback()
        return response
    before = reservation_audit.reservation_snapshot(t)
    before_public_details = (t.data, t.godz_od, t.liczba_osob)
    if t.status not in REZ_AKTYWNE:
        raise HTTPException(409, "Tej rezerwacji nie można już zmienić.")
    if t.faza_hosta is not None:
        raise HTTPException(409, "Rezerwacja jest w trakcie obsługi na sali.")
    cfg = get_lokal_config(db)
    teraz_lok = _teraz_lokalnie() or datetime.now(timezone.utc).replace(tzinfo=None)
    _sprawdz_okno_anulacji(cfg, t, teraz_lok)                        # za późno na zmianę → 400
    data = dane.data or t.data
    godz_od = dane.godz_od or t.godz_od
    osoby = max(1, dane.liczba_osob or t.liczba_osob or 1)
    payment_before_edit = _platnosc_rezerwacji(db, t.id)
    financial_context_changed = (
        (t.data, t.godz_od, int(t.liczba_osob or 1))
        != (data, godz_od, osoby)
    )
    if (
        financial_context_changed
        and payment_before_edit is not None
        and payment_before_edit.status in {"oczekuje", "autoryzowana", "oplacona"}
    ):
        db.rollback()
        raise reservation_payments.PaymentDomainError(
            "PAYMENT_SETTLEMENT_REQUIRED_BEFORE_EDIT",
            "Najpierw dokończ lub rozlicz aktywną płatność. Zmianę terminu albo liczby gości pomoże bezpiecznie wykonać lokal.",
        )
    if godz_od is None:
        raise HTTPException(400, "Rezerwacja bez godziny — nie można zmienić.")
    if data < teraz_lok.date():
        raise HTTPException(400, "Nie można przenieść rezerwacji wstecz.")
    przydzial, evaluation = _przydzial_zgodny_z_regulami(
        db,
        data=data,
        godz_od=godz_od,
        osoby=osoby,
        pomin_id=t.id,
        kanal="online",
        intent="edit",
        enforce=True,
    )
    if not przydzial:
        raise reservation_service.ReservationError(
            409,
            "NO_TABLE_CANDIDATE",
            "Brak wolnego stołu dla nowego terminu.",
            rule="table",
        )
    godz_do = evaluation.godz_do
    stoliki = przydzial["stoliki"]
    stolik = db.get(models.Stolik, stoliki[0])
    stary_okno = (t.data, t.godz_od, _koniec_okna(db, t)) if t.godz_od else None
    stare_stoliki = _stoly_terminu(t)
    t.data = data; t.godz_od = godz_od; t.godz_do = godz_do; t.liczba_osob = osoby
    t.stolik_id = stoliki[0]; t.stoliki_dodatkowe = (stoliki[1:] or None)
    t.auto_przydzielony = True
    _ustaw_proweniencje_przydzialu(t, przydzial)
    try:
        db.flush()
        _zastap_ledger_terminu(
            db,
            t,
            enforce_pacing=True,
            candidates=[{"stoliki": stoliki}],
            evaluation=evaluation,
            intent="edit",
        )
        reservation_audit.add_reservation_audit(
            db, termin=t, action="edit", actor_kind="guest", before=before, after=t,
        )
        if (
            financial_context_changed
            and payment_before_edit is not None
            and payment_before_edit.status in {
                "nieudana", "wygasla", "anulowana", "zwrocona",
            }
        ):
            if payment_before_edit.status != "zwrocona":
                reservation_payments.mark_payment_superseded(
                    db,
                    payment_before_edit,
                    now=now,
                    actor_kind="guest",
                )
            payment_before_edit.termin_id = None
        if (
            financial_context_changed
            and (
                payment_before_edit is None
                or payment_before_edit.status in {
                    "nieudana", "wygasla", "anulowana", "zwrocona",
                }
            )
        ):
            payment_policy = reservation_payments.resolve_policy(
                db,
                t.data,
                evaluation.service_id if evaluation is not None else None,
                int(t.liczba_osob or 1),
                "online",
            )
            payment_provider = (
                integracje.provider_platnosci_wymaganej()
                if payment_policy is not None and payment_policy.required
                else "sandbox"
            )
            previous_ref = payment_before_edit.id if payment_before_edit is not None else 0
            payment, payment_command = reservation_payments.create_payment_for_reservation(
                db,
                t,
                payment_policy,
                provider=payment_provider,
                now=now,
                business_today=teraz_lok.date(),
                service_id=(evaluation.service_id if evaluation is not None else None),
                operation_key=f"public-edit:{previous_ref}:{issued.record.id}",
                actor_kind="guest",
            )
            if payment is not None and payment_provider == "sandbox" and payment_command is not None:
                payment.link = "/?platnosc=sandbox&rezerwuj"
                payment_command.stan = "succeeded"
                payment_command.finished_at = now
                payment_command.updated_at = now
        if before_public_details != (t.data, t.godz_od, t.liczba_osob):
            reservation_communication.cancel_pending(
                db,
                t.id,
                event_types=("confirmation", "change"),
            )
            reservation_communication.enqueue_reservation(
                db, t, "change", actor_kind="guest",
            )
            reservation_communication.schedule_reminder(
                db, t, actor_kind="guest", force_new=True,
            )
        _commit_zapis_rezerwacji(db, guards)
    except IntegrityError as exc:
        db.rollback()
        raise reservation_service.translate_integrity_error(exc) from exc
    db.refresh(t)
    if stary_okno and (
        (stary_okno[0], _hm(stary_okno[1])) != (data, _hm(godz_od))
        or stare_stoliki != set(stoliki)
    ):
        _bezpiecznie_po_zwolnieniu_stolu(db, *stary_okno)            # stary termin zwolniony → re-optymalizacja
    response = _odpowiedz_z_obroconym_tokenem(t, issued.raw_token, stolik)
    payment = _platnosc_rezerwacji(db, t.id)
    if payment is not None:
        response["platnosc"] = _publiczna_platnosc_rezerwacji(payment, t)
    return response


@app.post(
    "/api/online/zarzadzanie/edytuj",
    dependencies=[Depends(_wymagaj_rezerwacje_online), Depends(_limit_management)],
)
def online_zarzadzanie_edytuj(
    dane: schemas.OnlineEdytujIn,
    request: Request,
    http_response: Response,
    reservation_token: Optional[str] = Header(None, alias="X-Reservation-Token"),
    reservation_session: Optional[str] = Header(None, alias="X-Reservation-Session"),
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
    db: Session = Depends(get_db),
):
    token = _publiczny_token_zarzadzania(
        request,
        reservation_token,
        mutation=True,
        reservation_session=reservation_session,
    )
    payload = _online_rezerwacja_edytuj_core(token, dane, idempotency_key, db)
    _odswiez_publiczne_cookie_z_odpowiedzi(http_response, payload, db)
    return payload


@app.get(
    "/api/online/zarzadzanie/dane",
    dependencies=[Depends(_wymagaj_rezerwacje_online), Depends(_limit_management)],
)
def online_zarzadzanie_dane(
    request: Request,
    reservation_token: Optional[str] = Header(None, alias="X-Reservation-Token"),
    db: Session = Depends(get_db),
):
    """Self-service eksportuje wyłącznie dane powiązane z jedną capability rezerwacji."""
    token = _publiczny_token_zarzadzania(request, reservation_token)
    record = reservation_service.validate_management_token(
        db,
        token,
        scope="data:export",
        secret=SECRET_KEY,
        now=utcnow_naive(),
    )
    t = db.get(models.Termin, record.termin_id)
    if t is None or t.kanal != "online":
        raise HTTPException(404, "Nie znaleziono rezerwacji.")
    consents = (
        db.query(models.RezerwacjaZgodaPubliczna)
        .filter_by(termin_id=t.id)
        .order_by(models.RezerwacjaZgodaPubliczna.created_at)
        .all()
    )
    subject_phone_ref, subject_email_ref = (
        reservation_communication.subject_refs_for_owner(t)
    )
    return {
        "rezerwacja": {
            **_online_rez_out(t),
            "telefon": t.telefon,
            "email": t.email,
            "notatka": t.notatka,
            "kanal_komunikacji": t.kanal_komunikacji,
        },
        "prywatnosc": [
            {
                "notice_version": consent.notice_version,
                "notice_ack_at": consent.notice_ack_at.isoformat(),
                "marketing": bool(consent.marketing),
                "marketing_version": consent.marketing_version,
                "marketing_at": consent.marketing_at.isoformat(),
                "sensitive": bool(consent.sensitive),
                "sensitive_data": consent.sensitive_data if consent.sensitive else None,
                "sensitive_at": consent.sensitive_at.isoformat() if consent.sensitive_at else None,
                "retention_until": consent.retention_until.isoformat(),
            }
            for consent in consents
        ],
        "komunikacja_operacyjna": eksport_komunikacji_operacyjnej(
            db,
            subject_phone_refs=(subject_phone_ref,) if subject_phone_ref else (),
            subject_email_refs=(subject_email_ref,) if subject_email_ref else (),
            termin_ids=(t.id,),
            require_owner_scope=True,
        ),
    }


@app.post(
    "/api/online/zarzadzanie/dane/usun",
    dependencies=[Depends(_wymagaj_rezerwacje_online), Depends(_limit_management)],
)
def online_zarzadzanie_usun_dane(
    dane: schemas.PubliczneUsuniecieDanychIn,
    request: Request,
    http_response: Response,
    reservation_token: Optional[str] = Header(None, alias="X-Reservation-Token"),
    reservation_session: Optional[str] = Header(None, alias="X-Reservation-Session"),
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
    db: Session = Depends(get_db),
):
    """Anuluje aktywną wizytę i usuwa PII/szczególne dane w jednej transakcji."""
    token = _publiczny_token_zarzadzania(
        request,
        reservation_token,
        mutation=True,
        reservation_session=reservation_session,
    )
    t, guards = _zablokuj_rezerwacje_online(db, token)
    if not t:
        raise HTTPException(404, "Nie znaleziono rezerwacji.")
    now = utcnow_naive()
    issued = reservation_service.consume_and_rotate_management_token(
        db,
        token,
        operation="data:delete",
        idempotency_key=idempotency_key,
        payload=dane.model_dump(mode="json"),
        secret=SECRET_KEY,
        now=now,
        # Usunięcie danych jest operacją terminalną. Retry ma zwrócić bezpieczny
        # receipt, mimo że następca został od razu unieważniony.
        allow_revoked_successor_replay=True,
    )
    if issued.replayed:
        db.rollback()
        _usun_publiczne_cookie_zarzadzania(http_response)
        return {"status": "usuniete"}
    before = reservation_audit.reservation_snapshot(t)
    old_window = (t.data, t.godz_od, _koniec_okna(db, t)) if t.godz_od else None
    had_tables = bool(_stoly_terminu(t))
    if t.status in REZ_AKTYWNE:
        t.status = "odwolana"
        t.odwolano_at = now
        reservation_service.release_termin_allocation(db, t.id)
        reservation_payments.request_reservation_cancellation_settlement(
            db,
            t,
            now=now,
            actor_kind="guest",
            operation_key=f"reservation-erasure:{t.id}",
        )
    source_token = reservation_service.lookup_management_token(
        db, token, secret=SECRET_KEY,
    )
    keep_token_ids = {issued.record.id}
    if source_token is not None:
        keep_token_ids.add(source_token.id)
    # Wspólny cleanup z pełną ścieżką RODO usuwa także profil CRM, dane
    # szczególne, wątek ustaleń i wolny tekst zadatku, zanim wyczyścimy klucz.
    usun_powiazane_pii_rezerwacji(db, [t])
    t.nazwisko = "[anonimizacja RODO]"
    t.telefon = None
    t.email = None
    t.notatka = None
    t.token_potwierdzenia = None
    usun_powiazane_publiczne_sekrety(
        db,
        [t.id],
        preserve_management_token_ids=keep_token_ids,
    )
    db.query(models.RezerwacjaTokenZarzadzania).filter(
        models.RezerwacjaTokenZarzadzania.id.in_(keep_token_ids),
    ).update({"revoked_at": now}, synchronize_session=False)
    reservation_audit.add_reservation_audit(
        db,
        termin=t,
        action="edit",
        actor_kind="guest",
        reason="guest_request",
        before=before,
        after=t,
        pii_changed={"nazwisko", "telefon", "email", "notatka"},
    )
    _commit_zapis_rezerwacji(db, guards)
    if had_tables and old_window:
        _bezpiecznie_po_zwolnieniu_stolu(db, *old_window)
    _usun_publiczne_cookie_zarzadzania(http_response)
    return {"status": "usuniete"}


# --- PUBLIKACJA GRAFIKU (admin — chronione middleware) ---

@app.get("/api/grafik/publikacja", status_code=200)
def status_publikacji(start: date = Query(...), end: date = Query(...), db: Session = Depends(get_db)):
    p = db.query(models.PublikacjaGrafiku).filter_by(start=start, koniec=end).first()
    return {"opublikowany": bool(p), "opublikowano_at": p.opublikowano_at.isoformat() if p else None}

@app.post("/api/grafik/publikuj", status_code=200)
def publikuj_grafik(start: date = Query(...), end: date = Query(...), cisza: bool = False, db: Session = Depends(get_db)):
    """Publikuje grafik tygodnia. cisza=true -> bez powiadomien push (np. dla starych tygodni)."""
    teraz = utcnow_naive()
    p = db.query(models.PublikacjaGrafiku).filter_by(start=start, koniec=end).first()
    if p:
        p.opublikowano_at = teraz
    else:
        db.add(models.PublikacjaGrafiku(start=start, koniec=end, opublikowano_at=teraz))
    db.commit()
    # AUTO „zamyka lokal": ustaw zamykającego dla każdego dnia tygodnia (też backfill
    # wcześniej wprowadzonych dni, zanim automat działał).
    dzien = start
    while dzien <= end:
        _przelicz_zamykajacego(db, dzien)
        dzien += timedelta(days=1)
    wyslano = 0
    if not cisza:
        wyslano = wyslij_push(
            db,
            "Grafik udostępniony",
            f"Twój grafik na tydzień {start.strftime('%d.%m')}–{end.strftime('%d.%m')} jest gotowy.",
            url="/",
        )
    return {"opublikowano_at": teraz.isoformat(), "push_wyslano": wyslano}

@app.delete("/api/grafik/publikuj", status_code=204)
def cofnij_publikacje(start: date = Query(...), end: date = Query(...), db: Session = Depends(get_db)):
    db.query(models.PublikacjaGrafiku).filter_by(start=start, koniec=end).delete()
    db.commit()


@app.get("/api/stanowiska", response_model=List[schemas.StanowiskoOut])
def get_stanowiska(db: Session = Depends(get_db)):
    _ensure_kwalifikacje_techniczne(db)  # Sprzątaczka/Stróż dostępne do nadania w Pracownikach
    return db.query(models.Stanowisko).all()

@app.post("/api/stanowiska", response_model=schemas.StanowiskoOut, status_code=201)
def create_stanowisko(data: schemas.StanowiskoCreate, db: Session = Depends(get_db)):
    if db.query(models.Stanowisko).filter_by(nazwa=data.nazwa).first():
        raise HTTPException(400, "Stanowisko o tej nazwie już istnieje.")
    s = models.Stanowisko(**data.model_dump())
    s.grupa_widocznosci = (s.grupa_widocznosci or "").strip() or None  # pusty string -> brak grupy
    s.rola = (s.rola or "").strip() or None
    db.add(s); db.commit(); db.refresh(s)
    return s

@app.put("/api/stanowiska/{sid}", response_model=schemas.StanowiskoOut)
def update_stanowisko(sid: int, data: schemas.StanowiskoCreate, db: Session = Depends(get_db)):
    s = db.get(models.Stanowisko, sid)
    if not s:
        raise HTTPException(404, "Nie znaleziono.")
    s.nazwa = data.nazwa
    s.tylko_weekend = data.tylko_weekend
    s.widoczny_dla_wszystkich = data.widoczny_dla_wszystkich
    s.grupa_widocznosci = (data.grupa_widocznosci or "").strip() or None
    s.rola = (data.rola or "").strip() or None
    s.daje_dostep_zamowien = bool(data.daje_dostep_zamowien)
    db.commit(); db.refresh(s)
    return s

@app.delete("/api/stanowiska/{sid}", status_code=204)
def delete_stanowisko(sid: int, db: Session = Depends(get_db)):
    s = db.get(models.Stanowisko, sid)
    if not s:
        raise HTTPException(404, "Nie znaleziono.")
    # WymaganiaDnia mają FK do stanowiska BEZ kaskady ORM/ondelete → kasujemy ręcznie, inaczej na
    # PostgreSQL (produkcja) delete rzuca IntegrityError 500, a na SQLite zostają sieroty.
    db.query(models.WymaganiaDnia).filter(models.WymaganiaDnia.stanowisko_id == sid).delete(synchronize_session=False)
    db.delete(s); db.commit()

@app.post("/api/stanowiska/{sid}/podkategorie", response_model=schemas.PodkategoriaOut)
def create_podkategoria(sid: int, data: schemas.PodkategoriaCreate, db: Session = Depends(get_db)):
    p = models.Podkategoria(**data.model_dump(), stanowisko_id=sid)
    db.add(p); db.commit(); db.refresh(p)
    return p

@app.delete("/api/podkategorie/{pid}", status_code=204)
def delete_podkategoria(pid: int, db: Session = Depends(get_db)):
    p = db.get(models.Podkategoria, pid)
    if p:
        db.delete(p); db.commit()


# ═══════════════════════════════════════════════════════════════════════════
# PRACOWNICY
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/pracownicy", response_model=List[schemas.PracownikOut])
def get_pracownicy(db: Session = Depends(get_db)):
    return db.query(models.Pracownik).order_by(models.Pracownik.kolejnosc, models.Pracownik.id).all()


@app.get("/api/grafik/kuchnia-stanowisko")
def kuchnia_stanowisko_info(db: Session = Depends(get_db)):
    """Zwraca (tworząc w razie potrzeby) ukryte stanowisko kuchni — front używa jego id do
    grafiku kuchni i stawki kuchni. Tylko admin (wymusza middleware)."""
    s = _kuchnia_stanowisko(db)
    return {"id": s.id, "nazwa": s.nazwa}


@app.get("/api/grafik/techniczny-stanowisko")
def techniczny_stanowisko_info(db: Session = Depends(get_db)):
    """Ukryte stanowisko techniczne — front używa jego id do stawki technicznej. Tylko admin."""
    s = _techniczny_stanowisko(db)
    return {"id": s.id, "nazwa": s.nazwa}

def _ustaw_stawki(db, p, stawki):
    """Nadpisuje stawki godzinowe pracownika (per stanowisko). Zapisuje tylko dodatnie."""
    db.query(models.StawkaPracownika).filter_by(pracownik_id=p.id).delete()
    for s in (stawki or []):
        if s.stanowisko_id and s.stawka and s.stawka > 0:
            db.add(models.StawkaPracownika(
                pracownik_id=p.id, stanowisko_id=s.stanowisko_id, stawka=float(s.stawka)
            ))


@app.post("/api/pracownicy", response_model=schemas.PracownikOut, status_code=201)
def create_pracownik(data: schemas.PracownikCreate, db: Session = Depends(get_db)):
    ostatni = db.query(models.Pracownik).order_by(models.Pracownik.kolejnosc.desc()).first()
    p = models.Pracownik(imie=data.imie, nazwisko=data.nazwisko, aktywny=data.aktywny,
                         kolor=data.kolor, dzial=(data.dzial or "obsluga"),
                         kolejnosc=(ostatni.kolejnosc + 1 if ostatni else 0))
    if data.kwalifikacje_ids:
        p.kwalifikacje = db.query(models.Stanowisko).filter(
            models.Stanowisko.id.in_(data.kwalifikacje_ids)
        ).all()
    db.add(p); db.commit(); db.refresh(p)
    _ustaw_stawki(db, p, data.stawki); db.commit(); db.refresh(p)
    _przypisz_odbicia_do_pracownika(db, p)  # podlinkuj zalegle odbicia RCP od razu
    return p

@app.put("/api/pracownicy/kolejnosc", status_code=200)
def ustaw_kolejnosc(data: schemas.KolejnoscIn, db: Session = Depends(get_db)):
    """Ręczna kolejność wyświetlania pracowników: kolejnosc = pozycja na liście ids."""
    for idx, pid in enumerate(data.ids):
        p = db.get(models.Pracownik, pid)
        if p:
            p.kolejnosc = idx
    db.commit()
    return {"ok": True}


@app.put("/api/pracownicy/{pid}", response_model=schemas.PracownikOut)
def update_pracownik(pid: int, data: schemas.PracownikCreate, db: Session = Depends(get_db)):
    p = db.get(models.Pracownik, pid)
    if not p:
        raise HTTPException(404, "Nie znaleziono.")
    p.imie = data.imie
    p.nazwisko = data.nazwisko
    p.aktywny = data.aktywny
    p.kolor = data.kolor
    p.dzial = data.dzial or "obsluga"
    p.kwalifikacje = db.query(models.Stanowisko).filter(
        models.Stanowisko.id.in_(data.kwalifikacje_ids)
    ).all()
    _ustaw_stawki(db, p, data.stawki)
    db.commit(); db.refresh(p)
    _przypisz_odbicia_do_pracownika(db, p)  # nazwisko moglo sie zmienic -> sprobuj podlinkowac
    return p

@app.delete("/api/pracownicy/{pid}", status_code=204)
def delete_pracownik(pid: int, db: Session = Depends(get_db)):
    p = db.get(models.Pracownik, pid)
    if not p:
        raise HTTPException(404, "Nie znaleziono.")
    # Rekordy zależne BEZ kaskady ORM/ondelete kasujemy jawnie (przez ORM — by odpaliły zagnieżdżone
    # kaskady, np. RozliczenieImprezy→pozycje). Inaczej na PostgreSQL FK RESTRICT rzuca 500, a na
    # SQLite zostają SIEROTY — m.in. wiersz RozliczenieKelner zawyżający utarg dnia bez atrybucji.
    # Twardy delete pracownika = pełne usunięcie jego danych (miękkie usunięcie = aktywny=False).
    for Model in (models.RozliczenieKelner, models.RozliczenieImprezy, models.RozliczenieGastro,
                  models.Urlop, models.ZamowienieSprzataczki, models.SprzatanieOdhaczenie):
        for row in db.query(Model).filter(Model.pracownik_id == pid).all():
            db.delete(row)
    db.delete(p); db.commit()


# ═══════════════════════════════════════════════════════════════════════════
# WYMAGANIA DNIA
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/wymagania", response_model=List[schemas.WymaganiaOut])
def get_wymagania(
    start: Optional[date] = None,
    end: Optional[date] = None,
    db: Session = Depends(get_db)
):
    if start and end:
        _odswiez_wymagania_imprez(db, start, end)   # świeże wymagania imprez przed odczytem widoku
    q = db.query(models.WymaganiaDnia)
    if start: q = q.filter(models.WymaganiaDnia.data >= start)
    if end:   q = q.filter(models.WymaganiaDnia.data <= end)
    return q.all()

@app.post("/api/wymagania", response_model=schemas.WymaganiaOut, status_code=201)
def create_wymagania(data: schemas.WymaganiaCreate, db: Session = Depends(get_db)):
    existing = db.query(models.WymaganiaDnia).filter_by(
        data=data.data,
        stanowisko_id=data.stanowisko_id,
        godz_od=data.godz_od,
        rewir=data.rewir
    ).first()
    
    if existing:
        existing.liczba_osob = data.liczba_osob
        db.commit(); db.refresh(existing)
        return existing
        
    w = models.WymaganiaDnia(**data.model_dump())
    db.add(w); db.commit(); db.refresh(w)
    return w

@app.delete("/api/wymagania/{wid}", status_code=204)
def delete_wymagania(wid: int, db: Session = Depends(get_db)):
    w = db.get(models.WymaganiaDnia, wid)
    if not w:
        raise HTTPException(404, "Nie znaleziono.")
    db.delete(w); db.commit()

def _data_z_body(body: dict, klucz: str) -> date:
    """Bezpieczne parsowanie daty ISO z surowego body: 400 (czytelny) zamiast 500,
    gdy klucz brakuje lub format jest zły."""
    try:
        return date.fromisoformat(body[klucz])
    except (KeyError, TypeError, ValueError):
        raise HTTPException(400, f"Nieprawidłowa lub brakująca data: {klucz}.")


@app.post("/api/wymagania/kopiuj", status_code=200)
def kopiuj_wymagania(body: dict, db: Session = Depends(get_db)):
    source = _data_z_body(body, "source_date")
    start  = _data_z_body(body, "start_date")
    end    = _data_z_body(body, "end_date")

    source_reqs = db.query(models.WymaganiaDnia).filter_by(data=source).all()
    if not source_reqs:
        raise HTTPException(404, "Brak wymagań dla dnia źródłowego.")

    count = 0
    current = start
    while current <= end:
        if current == source:
            current += timedelta(days=1)
            continue
        for req in source_reqs:
            existing = db.query(models.WymaganiaDnia).filter_by(
                data=current,
                stanowisko_id=req.stanowisko_id,
                godz_od=req.godz_od,
                rewir=req.rewir
            ).first()
            
            if existing:
                existing.liczba_osob = req.liczba_osob
            else:
                db.add(models.WymaganiaDnia(
                    data=current,
                    stanowisko_id=req.stanowisko_id,
                    godz_od=req.godz_od,
                    rewir=req.rewir,
                    liczba_osob=req.liczba_osob,
                ))
            count += 1
        current += timedelta(days=1)
    db.commit()
    return {"skopiowano": count}


@app.post("/api/wymagania/kopiuj-tydzien", status_code=200)
def kopiuj_wymagania_tydzien(body: dict, db: Session = Depends(get_db)):
    """Kopiuje wszystkie wymagania z tygodnia źródłowego na docelowy (dzień w dzień,
    przez offset dat). Tygodnie środa→wtorek mają ten sam układ dni tygodnia."""
    src_start = _data_z_body(body, "source_start")
    dst_start = _data_z_body(body, "target_start")
    offset = (dst_start - src_start).days
    if offset == 0:
        raise HTTPException(400, "Tydzień źródłowy i docelowy są takie same.")
    src_reqs = db.query(models.WymaganiaDnia).filter(
        models.WymaganiaDnia.data >= src_start,
        models.WymaganiaDnia.data <= src_start + timedelta(days=6),
    ).all()
    if not src_reqs:
        raise HTTPException(404, "Brak wymagań w tygodniu źródłowym.")
    count = 0
    for req in src_reqs:
        new_date = req.data + timedelta(days=offset)
        existing = db.query(models.WymaganiaDnia).filter_by(
            data=new_date,
            stanowisko_id=req.stanowisko_id,
            godz_od=req.godz_od,
            rewir=req.rewir,
        ).first()
        if existing:
            existing.liczba_osob = req.liczba_osob
        else:
            db.add(models.WymaganiaDnia(
                data=new_date,
                stanowisko_id=req.stanowisko_id,
                godz_od=req.godz_od,
                rewir=req.rewir,
                liczba_osob=req.liczba_osob,
            ))
        count += 1
    db.commit()
    return {"skopiowano": count}


# ═══════════════════════════════════════════════════════════════════════════
# DYSPOZYCJE (ZAKŁADKA IMPORtowania
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/dyspozycje", response_model=List[schemas.DyspozycjaOut])
def get_dyspozycje(
    start: Optional[date] = None,
    end: Optional[date] = None,
    db: Session = Depends(get_db)
):
    q = db.query(models.Dyspozycja)
    if start: q = q.filter(models.Dyspozycja.data >= start)
    if end:   q = q.filter(models.Dyspozycja.data <= end)
    return q.all()

@app.post("/api/dyspozycje", response_model=schemas.DyspozycjaOut, status_code=201)
def create_dyspozycja(data: schemas.DyspozycjaCreate, db: Session = Depends(get_db)):
    existing = db.query(models.Dyspozycja).filter_by(
        pracownik_id=data.pracownik_id, data=data.data
    ).first()
    if existing:
        for k, v in data.model_dump().items():
            setattr(existing, k, v)
        db.commit(); db.refresh(existing)
        return existing
    d = models.Dyspozycja(**data.model_dump())
    db.add(d); db.commit(); db.refresh(d)
    return d

@app.delete("/api/dyspozycje/{did}", status_code=204)
def delete_dyspozycja(did: int, db: Session = Depends(get_db)):
    """Czyści dyspozycyjność (admin) — wraca do stanu „brak zgłoszenia"."""
    d = db.get(models.Dyspozycja, did)
    if not d:
        raise HTTPException(404, "Nie znaleziono.")
    db.delete(d); db.commit()


# ═══════════════════════════════════════════════════════════════════════════
# PRZYDZIAŁY (ZAKTUALIZOWANE DLA WIELU ZMIAN CZASOWYCH)
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/przydzialy", response_model=List[schemas.PrzydzialOut])
def get_przydzialy(
    start: Optional[date] = None,
    end: Optional[date] = None,
    db: Session = Depends(get_db)
):
    q = db.query(models.PrzydzialZmiany)
    if start: q = q.filter(models.PrzydzialZmiany.data >= start)
    if end:   q = q.filter(models.PrzydzialZmiany.data <= end)
    return q.all()


def _powiadom_kuchnie_o_zmianie(db, pracownik_id, tytul, tresc):
    """Push do pracownika KUCHNI o KAŻDEJ zmianie w jego grafiku (dodanie/zmiana/wykreślenie).
    Dla obsługi nic nie robi (ci dostają push przy publikacji). Best-effort — błędy łykamy."""
    p = db.get(models.Pracownik, pracownik_id)
    if p and p.dzial == "kuchnia":
        try:
            wyslij_push_do_pracownika(db, pracownik_id, tytul, tresc, url="/")
        except Exception:
            pass


def _sprawdz_prawo_pracy(db, pracownik_id: int, data: date, godz_od, pomin_id: int = None):
    """Strażnik prawa pracy: odpoczynek + limit dni w tygodniu/miesiącu (parametry z LokalConfig).
    Rzuca HTTPException 400 z konkretnym komunikatem, gdy ręczny przydział łamie limit."""
    cfg = get_lokal_config(db)
    q = db.query(models.PrzydzialZmiany).filter(models.PrzydzialZmiany.pracownik_id == pracownik_id)
    if pomin_id is not None:
        q = q.filter(models.PrzydzialZmiany.id != pomin_id)
    inne = [(p.data, p.godz_od) for p in q.all()]
    blad = prawo_pracy.sprawdz(
        inne, data, godz_od,
        min_odpoczynek_h=cfg.praca_min_odpoczynek_h or 0,
        max_dni_tydzien=cfg.praca_max_dni_tydzien or 0,
        max_dni_miesiac=cfg.praca_max_dni_miesiac or 0,
    )
    if blad:
        raise HTTPException(400, blad)


@app.post("/api/przydzialy", response_model=schemas.PrzydzialOut, status_code=201)
def create_przydział(data: schemas.PrzydzialCreate, db: Session = Depends(get_db)):
    stan = db.get(models.Stanowisko, data.stanowisko_id)
    if stan and stan.tylko_weekend and data.data.weekday() < 5:
        raise HTTPException(400, f"Stanowisko '{stan.nazwa}' jest aktywne tylko w weekendy.")

    # Zasada biznesowa: maksymalnie JEDNA zmiana na pracownika w danym dniu.
    istniejaca = db.query(models.PrzydzialZmiany).filter_by(
        data=data.data,
        pracownik_id=data.pracownik_id,
    ).first()

    if istniejaca:
        raise HTTPException(400, "Pracownik ma już przydzieloną zmianę w tym dniu (maks. 1 zmiana dziennie).")

    _sprawdz_prawo_pracy(db, data.pracownik_id, data.data, data.godz_od)

    a = models.PrzydzialZmiany(**data.model_dump())
    db.add(a); db.commit(); db.refresh(a)
    _przelicz_zamykajacego(db, a.data)   # AUTO „zamyka lokal": najpóźniejszy na parkiecie tego dnia
    db.refresh(a)
    _powiadom_kuchnie_o_zmianie(db, a.pracownik_id, "Nowa zmiana w grafiku", "Dodano Ci zmianę w grafiku kuchni.")
    return a

@app.put("/api/przydzialy/{aid}", response_model=schemas.PrzydzialOut)
def update_przydział(aid: int, data: schemas.PrzydzialCreate, db: Session = Depends(get_db)):
    a = db.get(models.PrzydzialZmiany, aid)
    if not a:
        raise HTTPException(404, "Nie znaleziono.")
    stan = db.get(models.Stanowisko, data.stanowisko_id)
    if stan and stan.tylko_weekend and data.data.weekday() < 5:
        raise HTTPException(400, f"Stanowisko '{stan.nazwa}' jest aktywne tylko w weekendy.")

    # Maks. 1 zmiana/dzień — blokuj przeniesienie na dzień, gdzie pracownik ma już inną zmianę.
    kolizja = db.query(models.PrzydzialZmiany).filter(
        models.PrzydzialZmiany.data == data.data,
        models.PrzydzialZmiany.pracownik_id == data.pracownik_id,
        models.PrzydzialZmiany.id != aid,
    ).first()
    if kolizja:
        raise HTTPException(400, "Pracownik ma już przydzieloną zmianę w tym dniu (maks. 1 zmiana dziennie).")

    _sprawdz_prawo_pracy(db, data.pracownik_id, data.data, data.godz_od, pomin_id=aid)

    stara_data = a.data
    for k, v in data.model_dump().items():
        setattr(a, k, v)
    db.commit(); db.refresh(a)
    _przelicz_zamykajacego(db, a.data)              # AUTO „zamyka lokal" dla (nowego) dnia
    if stara_data != a.data:
        _przelicz_zamykajacego(db, stara_data)      # i dla dnia, z którego zmiana zniknęła
    db.refresh(a)
    _powiadom_kuchnie_o_zmianie(db, a.pracownik_id, "Zmiana w grafiku", "Zaktualizowano Twoją zmianę w grafiku kuchni.")
    return a

@app.delete("/api/przydzialy/{aid}", status_code=204)
def delete_przydział(aid: int, db: Session = Depends(get_db)):
    a = db.get(models.PrzydzialZmiany, aid)
    if not a:
        raise HTTPException(404, "Nie znaleziono.")
    pid = a.pracownik_id
    dzien = a.data
    db.delete(a); db.commit()
    _przelicz_zamykajacego(db, dzien)   # AUTO „zamyka lokal": po usunięciu zmiany przelicz dzień
    _powiadom_kuchnie_o_zmianie(db, pid, "Zmiana w grafiku", "Wykreślono Cię ze zmiany w grafiku kuchni.")

@app.put("/api/przydzialy/{aid}/zamyka", status_code=200)
def ustaw_zamykajacego(aid: int, payload: dict, db: Session = Depends(get_db)):
    """Ręczne nadpisanie zamykającego (domyślnie automat wybiera najpóźniejszego z parkietu).
      • {"reczny": true}  → TEN przydział zamyka lokal ręcznie; automat go nie zmienia,
        reszta dnia ma zamyka=False.
      • {"reczny": false} → zdejmij ręczne ustawienie i wróć do automatu dla tego dnia."""
    a = db.get(models.PrzydzialZmiany, aid)
    if not a:
        raise HTTPException(404, "Nie znaleziono.")
    if bool(payload.get("reczny")):
        for r in db.query(models.PrzydzialZmiany).filter(models.PrzydzialZmiany.data == a.data).all():
            r.zamyka = r.id == aid
            r.zamyka_reczny = r.id == aid
        db.commit()
    else:
        a.zamyka_reczny = False
        db.commit()
        _przelicz_zamykajacego(db, a.data)   # powrót do automatu
    db.refresh(a)
    return {"id": a.id, "zamyka": bool(a.zamyka), "zamyka_reczny": bool(a.zamyka_reczny)}

@app.delete("/api/przydzialy", status_code=204)
def clear_przydzialy(
    start: date = Query(...),
    end: date = Query(...),
    dzial: Optional[str] = Query(None),   # 'obsluga' | 'kuchnia' — czyść TYLKO grafik tego działu
    db: Session = Depends(get_db)
):
    q = db.query(models.PrzydzialZmiany).filter(
        models.PrzydzialZmiany.data >= start,
        models.PrzydzialZmiany.data <= end,
    )
    if dzial:
        # Czyść tylko grafik wskazanego działu (np. obsługa) — drugi grafik (kuchnia) zostaje.
        prac_ids = [pid for (pid,) in db.query(models.Pracownik.id)
                    .filter(models.Pracownik.dzial == dzial).all()]
        q = q.filter(models.PrzydzialZmiany.pracownik_id.in_(prac_ids))
    # Najpierw skasuj powiązane oferty giełdy: bulk-delete omija ORM, a ondelete=CASCADE nie jest
    # egzekwowane na SQLite → inaczej oferta zostałaby sierotą na nieistniejącym przydziale, a przy
    # odzysku id przydziału (SQLite) akceptacja takiej oferty mogłaby przepiąć CUDZĄ zmianę.
    przydz_ids = [row[0] for row in q.with_entities(models.PrzydzialZmiany.id).all()]
    if przydz_ids:
        db.query(models.OfertaZmiany).filter(
            models.OfertaZmiany.przydzial_id.in_(przydz_ids)).delete(synchronize_session=False)
    q.delete(synchronize_session=False)
    db.commit()
    # Po wyczyszczeniu przelicz zamykającego dla każdego dnia zakresu (zniknęli kandydaci).
    dzien = start
    while dzien <= end:
        _przelicz_zamykajacego(db, dzien)
        dzien += timedelta(days=1)


# ═══════════════════════════════════════════════════════════════════════════
# AUTO-ASSIGN
# ═══════════════════════════════════════════════════════════════════════════

@app.post("/api/auto-assign", response_model=schemas.AutoAssignResult)
def auto_assign_endpoint(
    start: date = Query(...),
    end: date = Query(...),
    db: Session = Depends(get_db)
):
    # Świeże wymagania pod imprezy z aktualnej tabeli — żeby auto-przydział miał co obsadzać.
    _odswiez_wymagania_imprez(db, start, end)
    result = _auto_assign(db, start, end)
    # Auto-przydział TYLKO SZKICUJE: cofamy publikację tygodnia, żeby zmiany NIE trafiły od razu
    # do obsługi. Admin sprawdza grafik i publikuje ręcznie („Udostępnij pracownikom").
    db.query(models.PublikacjaGrafiku).filter(
        models.PublikacjaGrafiku.start <= end,
        models.PublikacjaGrafiku.koniec >= start,
    ).delete(synchronize_session=False)
    db.commit()
    # AUTO „zamyka lokal": po automatycznym ułożeniu grafiku przelicz zamykającego per dzień.
    dzien = start
    while dzien <= end:
        _przelicz_zamykajacego(db, dzien)
        dzien += timedelta(days=1)
    return result


# ═══════════════════════════════════════════════════════════════════════════
# EKSPORT CSV
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/eksport-csv")
def eksport_csv(start: date, end: date, db: Session = Depends(get_db)):
    # 1. Pobieramy wszystkie przydziały z wybranego zakresu dat
    przydzialy = db.query(models.PrzydzialZmiany).filter(
        models.PrzydzialZmiany.data >= start,
        models.PrzydzialZmiany.data <= end
    ).all()
    
    # 2. Pobieramy wyłącznie pracowników, którzy są w grafiku "Aktywni"
    pracownicy = db.query(models.Pracownik).filter(models.Pracownik.aktywny == True).all()
    
    # 3. Pobieramy nazwy stanowisk ze słownika (żeby nie wyświetlać samych numerów ID)
    stanowiska_db = db.query(models.Stanowisko).all()
    stanowiska_slownik = {s.id: s.nazwa for s in stanowiska_db}
    
    # 4. Matematyka dat: Tworzymy listę wszystkich dni pomiędzy 'start' i 'end'
    ilosc_dni = (end - start).days
    lista_dat = [start + timedelta(days=i) for i in range(ilosc_dni + 1)]
    
    # 5. Tworzymy specjalny słownik do sortowania przydziałów w konkretne "kratki" tabeli
    # Kluczem jest (pracownik_id, data), a wartością lista tekstów ze zmianami
    przydzialy_w_kratkach = defaultdict(list)
    
    for p in przydzialy:
        # Odczytanie nazwy stanowiska
        nazwa_stanowiska = stanowiska_slownik.get(p.stanowisko_id, "Nieznane")
        
        # Przygotowanie ładnego formatu godzin
        g_od = p.godz_od.strftime("%H:%M") if p.godz_od else ""
        
        if g_od:
            tekst_zmiany = f"{nazwa_stanowiska} ({g_od})"
        else:
            tekst_zmiany = f"{nazwa_stanowiska} (Cały dzień)"
        
        przydzialy_w_kratkach[(p.pracownik_id, p.data)].append(tekst_zmiany)

    output = io.StringIO()
    # Używamy średnika! Polski Excel często psuje układ, jeśli użyje się standardowego przecinka
    writer = csv.writer(output, delimiter=';')
    
    # ---- TWORZENIE NAGŁÓWKA TABELI ----
    # Wygląda tak: ["Pracownik", "2026-06-01", "2026-06-02", "2026-06-03"...]
    naglowek = ["Pracownik"] + [d.strftime("%d.%m.%Y") for d in lista_dat]
    writer.writerow(naglowek)
    
    # ---- TWORZENIE WIERSZY PRACOWNIKÓW ----
    for pracownik in pracownicy:
        wiersz = [f"{pracownik.imie} {pracownik.nazwisko}"]
        
        for d in lista_dat:
            # Szukamy, czy pracownik ma w danym dniu wpisane zmiany
            zmiany = przydzialy_w_kratkach.get((pracownik.id, d), [])
            
            if zmiany:
                # Łączymy wszystkie zmiany pracownika z tego dnia znakiem |
                # (Zabezpieczenie na wypadek, gdyby pracował np. rano na Barze, a po południu na Sali)
                wiersz.append(" | ".join(zmiany))
            else:
                # Jeśli w tym dniu ma wolne, zostawiamy pustą komórkę
                wiersz.append("")
                
        writer.writerow(wiersz)
        
    # Przygotowanie pliku do wysłania
    output.seek(0)
    
    headers = {
        "Content-Disposition": f"attachment; filename=grafik_{start}_do_{end}.csv"
    }
    
    # Konwersja na bajty z ukrytym znacznikiem 'utf-8-sig'. 
    # Bez tego polskie znaki takie jak ą, ę, ł wyglądałyby w Excelu jak błędy (np. "krzaczki").
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8-sig")]),
        media_type="text/csv",
        headers=headers
    )


@app.get("/api/eksport/wyplaty")
def eksport_wyplaty(request: Request, rok: int = Query(..., ge=2000, le=2100),
                    miesiac: int = Query(..., ge=1, le=12),
                    user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Miesięczny raport wypłat jako sformatowany plik .xlsx dla księgowej: per pracownik rozbicie
    na stanowiska (godziny × stawka) + wiersz RAZEM i suma wszystkich. Dostęp do płac → audyt (RODO)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    raport = raporty.raport_godzin_miesiac(db, rok, miesiac)
    zapisz_audyt(db, user, "eksport_wyplaty", zasob=f"{rok}-{miesiac:02d}", request=request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Wyplaty {rok}-{miesiac:02d}"
    naglowek = ["Pracownik", "Stanowisko", "Godziny", "Stawka (zł/h)", "Do wypłaty (zł)"]
    ws.append(naglowek)
    naglowek_fill = PatternFill("solid", fgColor="1C1C1E")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = naglowek_fill
        c.alignment = Alignment(horizontal="center")

    suma_godzin = suma_kwota = 0.0
    for p in raport["pracownicy"]:
        for s in p["stanowiska"]:
            ws.append([p["pracownik"], s["stanowisko"], s["godziny"], s["stawka"], s["kwota"]])
        ws.append([p["pracownik"], "RAZEM", p["suma_godzin"], None, p["do_wyplaty"]])
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)
        suma_godzin += p["suma_godzin"]
        suma_kwota += p["do_wyplaty"]

    ws.append([])
    ws.append(["WSZYSCY RAZEM", None, round(suma_godzin, 2), None, round(suma_kwota, 2)])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, color="1F7A5C")
    for col, szer in zip("ABCDE", (26, 22, 10, 13, 16)):
        ws.column_dimensions[col].width = szer

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nazwa = f"wyplaty_{rok}_{miesiac:02d}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nazwa}"'})


# ═══════════════════════════════════════════════════════════════════════════
# IMPREZY Z SERWERA NAS (ZINTEGROWANE Z AUTOMATYKĄ WYMAGAŃ)
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/api/imprezy", response_model=List[schemas.ImprezaOut], dependencies=[Depends(_wymagaj_modul_imprezy)])
def get_imprezy(start: date = Query(...), end: date = Query(...), db: Session = Depends(get_db)):
    return db.query(models.Impreza).filter(models.Impreza.data >= start, models.Impreza.data <= end).order_by(models.Impreza.data.asc()).all()


def _imprezy_stanowisko(db):
    """Stanowisko imprez — najpierw po roli 'imprezy', potem fallback po nazwie (zaczyna się od
    „imprez": łapie „Impreza"/„Imprezy"). Zwraca pierwsze pasujące stanowisko albo None."""
    s = db.query(models.Stanowisko).filter_by(rola="imprezy").first()
    if s:
        return s
    for s in db.query(models.Stanowisko).all():
        if (s.nazwa or "").strip().lower().startswith("imprez"):
            return s
    return None


def _imprezy_wymagania_warning(db):
    """Ostrzeżenie dla obsadzania imprez przez auto-przydział: brak stanowiska imprez
    albo brak AKTYWNEGO pracownika z tą kwalifikacją (wtedy auto-przydział nie obsadzi imprez)."""
    stan = _imprezy_stanowisko(db)
    if not stan:
        return 'Brak stanowiska imprez (np. „Impreza"/„Imprezy") — utwórz je w zakładce Stanowiska, inaczej imprezy nie zostaną obsadzone.'
    ma_ktos = db.query(models.Pracownik).filter(
        models.Pracownik.aktywny == True,
        models.Pracownik.kwalifikacje.any(models.Stanowisko.id == stan.id),
    ).first()
    if not ma_ktos:
        return f'Żaden aktywny pracownik nie ma kwalifikacji „{stan.nazwa}" — nadaj ją w zakładce Pracownicy, inaczej auto-przydział nie obsadzi imprez.'
    return None


def _impreza_params(cfg) -> dict:
    """Parametry obsady imprez z konfiguracji lokalu (LokalConfig) → dict dla
    algorithm.przelicz_imprezy_na_wymagania. `impreza_sale_min2` to lista po przecinku."""
    sale = [s.strip() for s in (cfg.impreza_sale_min2 or "").split(",") if s.strip()]
    return {
        "osoby_na_obsluge": cfg.impreza_osoby_na_obsluge,
        "wyprzedzenie_min": cfg.impreza_wyprzedzenie_min,
        "najwczesniej": cfg.impreza_najwczesniej,
        "sale_min2": sale,
    }


def _odswiez_wymagania_imprez(db, start, end):
    """Przelicza wymagania imprez dla [start, end] z AKTUALNEJ tabeli imprez: kasuje stare
    auto-wymagania imprez (jest_impreza=True) i tworzy świeże na stanowisku imprez. Dzięki temu
    auto-przydział i widok „Wymagania" ZAWSZE mają aktualne wymagania pod imprezy — bez ręcznego
    re-syncu. Bez stanowiska imprez („Impreza"/„Imprezy") nic nie robi."""
    stan = _imprezy_stanowisko(db)
    if not stan:
        return
    imprezy = db.query(models.Impreza).filter(
        models.Impreza.data >= start, models.Impreza.data <= end
    ).all()
    nowe = przelicz_imprezy_na_wymagania(imprezy, _impreza_params(get_lokal_config(db)))
    db.query(models.WymaganiaDnia).filter(
        models.WymaganiaDnia.data >= start,
        models.WymaganiaDnia.data <= end,
        models.WymaganiaDnia.jest_impreza == True,
    ).delete(synchronize_session=False)
    for w in nowe:
        db.add(models.WymaganiaDnia(**w, stanowisko_id=stan.id))
    db.commit()


# ═══════════════════════════════════════════════════════════════════════════
# IMPREZY — IMPORT Z PLIKU .ics (eksport z iCloud / Apple Calendar)
#   Admin wgrywa plik .ics; backend tworzy z każdego wydarzenia Termin (kalendarz imprez)
#   ORAZ Imprezę (źródło obsady). „iCloud tylko dodaje" — istniejące (po UID) pomijamy.
#   Godzina z kalendarza jest NIEISTOTNA (impreza dostaje 'Brak'); obsada liczy się z osób.
# ═══════════════════════════════════════════════════════════════════════════
@app.post("/api/imprezy/import-ics", dependencies=[Depends(_wymagaj_modul_imprezy)])
def import_imprez_ics(payload: dict, db: Session = Depends(get_db)):
    """Import imprez z .ics. Body: {"ics": "<zawartość pliku>"}. Admin-only (middleware).
    Dla każdego VEVENT (jeśli jeszcze nie ma — dedup po UID):
      • Termin  — wpis w kalendarzu imprez (zadatki KP dopną się po nazwisku+dacie),
      • Imprezę — źródło wymagań obsady (godzina='Brak').
    Istniejące (po UID) POMIJAMY — ręcznych zmian w aplikacji nie nadpisujemy."""
    tekst = (payload or {}).get("ics") or ""
    if not isinstance(tekst, str) or not tekst.strip():
        raise HTTPException(400, "Pusty albo nieprawidłowy plik .ics.")

    rekordy = ical_import.wczytaj_imprezy_z_ics(tekst)
    dodano_terminy = dodano_imprezy = pominieto = bez_daty = bez_uid = bez_osob = 0
    daty = []

    for r in rekordy:
        uid = (r.get("uid") or "").strip()
        d = r.get("data")
        if not d:
            bez_daty += 1
            continue
        if not uid:
            bez_uid += 1   # bez UID nie ma jak bezpiecznie deduplikować — pomijamy
            continue
        daty.append(d)
        nazwa = (r.get("nazwisko") or "").strip() or "(bez nazwy)"
        liczba = r.get("liczba_osob")
        if not liczba:
            bez_osob += 1

        # --- Termin (kalendarz imprez) ---
        ist_t = db.query(models.Termin).filter(models.Termin.ical_uid == uid).first()
        if ist_t is None:
            db.add(models.Termin(
                data=d, nazwisko=nazwa, typ=r.get("typ"),
                liczba_osob=liczba, telefon=r.get("telefon"), sala=r.get("sala"),
                notatka=r.get("notatka"), status="rezerwacja",
                zadatek=float(r.get("zadatek") or 0), ical_uid=uid, kanal="ical",
                utworzono_at=utcnow_naive(),
            ))
            dodano_terminy += 1
        else:
            pominieto += 1

        # --- Impreza (źródło obsady; godzina nieistotna -> 'Brak') ---
        klucz = f"ical:{uid}"
        ist_i = db.query(models.Impreza).filter(models.Impreza.sciezka_pliku == klucz).first()
        if ist_i is None:
            db.add(models.Impreza(
                data=d, klient=nazwa, liczba_osob=(liczba or 0),
                godzina="Brak", sala=(r.get("sala") or "Brak"), sciezka_pliku=klucz,
            ))
            dodano_imprezy += 1

    db.commit()

    # Świeże wymagania obsady dla zakresu importu + próba dopięcia zadatków ze skrzynki
    # (nowo dodane terminy mogą pasować do wcześniej nieprzypisanych KP).
    if daty:
        _odswiez_wymagania_imprez(db, min(daty), max(daty))
    for z in db.query(models.KpZadatek).filter(models.KpZadatek.termin_id.is_(None)).all():
        _dopasuj_zadatek(db, z)
    db.commit()

    ostrzezenia = []
    w = _imprezy_wymagania_warning(db)
    if w:
        ostrzezenia.append(w)
    if bez_osob:
        ostrzezenia.append(f"{bez_osob} imprez bez liczby osób — dla nich obsada nie zostanie policzona (uzupełnij liczbę osób w kalendarzu).")
    if bez_daty:
        ostrzezenia.append(f"{bez_daty} wydarzeń bez poprawnej daty — pominięto.")
    if bez_uid:
        ostrzezenia.append(f"{bez_uid} wydarzeń bez UID — pominięto (brak klucza do deduplikacji).")

    return {
        "dodano_terminy": dodano_terminy,
        "dodano_imprezy": dodano_imprezy,
        "pominieto": pominieto,
        "bez_daty": bez_daty,
        "ostrzezenia": ostrzezenia,
    }


# ═══════════════════════════════════════════════════════════════════════════
# RCP — ODBICIA (przyjmowane od lokalnego agenta) + GODZINY/STANOWISKO
#   VPS nie łączy się z bazą RCP/Gastro. Lokalny agent (na serwerze Gastro) czyta
#   odbicia read-only (NOLOCK) i WYPYCHA je tutaj. Autoryzacja: stały token X-RCP-Token.
# ═══════════════════════════════════════════════════════════════════════════
import unicodedata

RCP_INGEST_TOKEN = os.environ.get("RCP_INGEST_TOKEN", "")

# Push wysylamy tylko dla SWIEZYCH zdarzen (wejscie/wyjscie w ostatnich N minutach),
# zeby pierwszy ingest / restart agenta nie zasypal pracownikow powiadomieniami o starych
# zmianach. Dane i tak zapisujemy zawsze — bramka dotyczy WYLACZNIE powiadomien.
RCP_POWIADOM_OKNO = timedelta(minutes=int(os.environ.get("RCP_POWIADOM_OKNO_MIN", "60")))


# Litery 'ł/Ł' (i kilka innych) NIE rozkładają się przez NFKD — mapujemy je ręcznie,
# inaczej wypadłyby z dopasowania imion (np. „Łukasz" → „ukasz").
_PL_SPEC = str.maketrans({"ł": "l", "Ł": "L", "ø": "o", "Ø": "O", "đ": "d", "Đ": "D"})


def _norm_nazwa(s: str) -> str:
    s = (s or "").translate(_PL_SPEC)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().lower()
    return " ".join(s.split())


def _przypisz_odbicia_do_pracownika(db, prac) -> int:
    """Po dodaniu/edycji pracownika podlinkuj wczesniej NIEDOPASOWANE odbicia RCP
    (pracownik_id IS NULL), ktorych znormalizowane imie+nazwisko pasuje do tego pracownika.
    Dzieki temu godziny z RCP trafiaja na konto od razu, bez czekania na okno agenta
    (agent dosyla tylko ostatnie OKNO_DNI dni). Zwraca liczbe podlinkowanych odbic."""
    cele = {
        _norm_nazwa(f"{prac.imie} {prac.nazwisko}"),
        _norm_nazwa(f"{prac.nazwisko} {prac.imie}"),
    }
    cele.discard("")
    if not cele:
        return 0
    n = 0
    for o in db.query(models.OdbicieRcp).filter(models.OdbicieRcp.pracownik_id.is_(None)).all():
        if _norm_nazwa(o.imie_nazwisko or "") in cele:
            o.pracownik_id = prac.id
            n += 1
    if n:
        db.commit()
    return n


def _parse_dt(v):
    if not v:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v)
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None


@app.post("/api/rcp/ingest")
def rcp_ingest(payload: dict, request: Request, db: Session = Depends(get_db)):
    """Przyjmuje paczkę odbić od lokalnego agenta i robi upsert po `rcp_id`.
    Wykrywa wejście (push „start zmiany") i wyjście (push „koniec zmiany" + godziny).
    Idempotentne — flagi powiadomień zapobiegają dublom."""
    if not token_agenta_ok(request, db):
        raise HTTPException(401, "Nieprawidłowy lub brakujący token agenta RCP.")

    odbicia = payload.get("odbicia", []) if isinstance(payload, dict) else []
    zrodlo = (payload.get("zrodlo") if isinstance(payload, dict) else None) or ""
    mapa = {}
    for p in db.query(models.Pracownik).all():
        mapa.setdefault(_norm_nazwa(f"{p.imie} {p.nazwisko}"), p.id)
        mapa.setdefault(_norm_nazwa(f"{p.nazwisko} {p.imie}"), p.id)
    # Jawne mapowanie POS→Lokalo (kreator „Integracja POS") ma PIERWSZEŃSTWO nad imieniem —
    # per źródło, żeby ten sam pos_id z dwóch systemów się nie mylił.
    mapa_pos = {m.pos_id: m.pracownik_id for m in
                db.query(models.PracownikPosId).filter_by(zrodlo=zrodlo).all()} if zrodlo else {}

    nowe = zakonczone = powiadomienia = 0
    teraz = _teraz_lokalnie()  # do bramki swiezosci powiadomien
    for o in odbicia:
        try:
            rcp_id = str(o["rcp_id"])
            nazwa = (o.get("imie_nazwisko") or "").strip()
            d = date.fromisoformat(str(o["data"])[:10])
        except (KeyError, ValueError, TypeError):
            continue
        wejscie = _parse_dt(o.get("wejscie"))
        wyjscie = _parse_dt(o.get("wyjscie"))
        pos_pid = o.get("pos_pracownik_id")
        # jawne mapowanie POS→Lokalo najpierw, potem dopasowanie po imieniu
        pid = (mapa_pos.get(str(pos_pid)) if pos_pid is not None else None) or mapa.get(_norm_nazwa(nazwa))

        rec = db.query(models.OdbicieRcp).filter_by(rcp_id=rcp_id).first()
        if rec is None:
            rec = models.OdbicieRcp(
                rcp_id=rcp_id, imie_nazwisko=nazwa, pracownik_id=pid, data=d,
                wejscie=wejscie, wyjscie=wyjscie,
                pos_pracownik_id=(str(pos_pid) if pos_pid is not None else None),
                zrodlo=zrodlo or None,
            )
            db.add(rec)
            nowe += 1
        else:
            if nazwa:
                rec.imie_nazwisko = nazwa
            if pid is not None:
                rec.pracownik_id = pid
            if pos_pid is not None:
                rec.pos_pracownik_id = str(pos_pid)
            if zrodlo:
                rec.zrodlo = zrodlo
            if wejscie:
                rec.wejscie = wejscie
            if wyjscie:
                rec.wyjscie = wyjscie
        if rec.wejscie and rec.wyjscie:
            rec.godziny = round((rec.wyjscie - rec.wejscie).total_seconds() / 3600.0, 2)
        rec.zaktualizowano_at = utcnow_naive()
        db.flush()

        if rec.wejscie and rec.pracownik_id and not rec.powiadomiono_wejscie:
            if teraz is None or rec.wejscie >= teraz - RCP_POWIADOM_OKNO:
                powiadomienia += wyslij_push_do_pracownika(
                    db, rec.pracownik_id, "Rozpoczęto zmianę",
                    f"Odbicie o {rec.wejscie.strftime('%H:%M')} — miłej pracy!", url="/",
                )
            rec.powiadomiono_wejscie = True  # zaznaczamy obsluzone (stare = bez spamu)
        if rec.wyjscie and rec.pracownik_id and not rec.powiadomiono_wyjscie:
            if teraz is None or rec.wyjscie >= teraz - RCP_POWIADOM_OKNO:
                powiadomienia += wyslij_push_do_pracownika(
                    db, rec.pracownik_id, "Zakończono zmianę",
                    f"Przepracowano {rec.godziny:.2f} h — dopisano do konta.", url="/",
                )
            rec.powiadomiono_wyjscie = True
            zakonczone += 1

    db.commit()
    return {"przyjeto": len(odbicia), "nowe": nowe, "zakonczone": zakonczone, "powiadomienia": powiadomienia}


def _trwajace_zmiany(db):
    """Pracownicy aktualnie NA ZMIANIE: wejscie jest, wyjscia brak, rozpoczete w ~18h.
    Dopasowani -> imie+nazwisko z konta; niedopasowani -> nazwa z RCP."""
    teraz = _teraz_lokalnie()
    prac = {p.id: f"{p.imie} {p.nazwisko}" for p in db.query(models.Pracownik).all()}
    out = []
    for o in (
        db.query(models.OdbicieRcp)
        .filter(models.OdbicieRcp.wyjscie.is_(None), models.OdbicieRcp.wejscie.isnot(None))
        .order_by(models.OdbicieRcp.wejscie.desc())
        .all()
    ):
        if teraz is not None and o.wejscie < teraz - timedelta(hours=18):
            continue
        out.append({
            "pracownik_id": o.pracownik_id,
            "pracownik": prac.get(o.pracownik_id) or o.imie_nazwisko,
            "wejscie": o.wejscie.isoformat(),
            "dopasowany": o.pracownik_id is not None,
        })
    return out


def _raport_godzin_bez_wyplat(raport: dict) -> dict:
    """Allowlista pól raportu czasu bez stawek, kwot, zaliczek i kosztowych agregatów."""
    bezpieczne_pola = (
        "rok", "miesiac", "poza_grafikiem", "duze_ciecia", "male_ciecia",
        "niedopasowani_rcp", "na_zmianie",
    )
    bez_finansow = {k: raport[k] for k in bezpieczne_pola if k in raport}
    bez_finansow["pracownicy"] = [
        {
            "pracownik_id": p.get("pracownik_id"),
            "pracownik": p.get("pracownik"),
            "dzial": p.get("dzial"),
            "suma_godzin": p.get("suma_godzin", 0),
            "stanowiska": [
                {"stanowisko": s.get("stanowisko"), "godziny": s.get("godziny", 0)}
                for s in p.get("stanowiska", [])
            ],
            "zaoszczedzone_godziny": p.get("zaoszczedzone_godziny", 0),
        }
        for p in raport.get("pracownicy", [])
    ]
    bez_finansow["zaoszczedzone"] = {
        "godziny": (raport.get("zaoszczedzone") or {}).get("godziny", 0),
    }
    bez_finansow["stanowiska_podsumowanie"] = [
        {"stanowisko": s.get("stanowisko"), "godziny": s.get("godziny", 0)}
        for s in raport.get("stanowiska_podsumowanie", [])
    ]
    return bez_finansow


@app.get("/api/raporty/godziny", status_code=200)
def raport_godzin(request: Request, rok: int = Query(..., ge=2000, le=2100), miesiac: int = Query(..., ge=1, le=12),
                  user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Raport godzin wszystkich pracowników (admin + szef — wymusza middleware).
    Dorzuca `na_zmianie` (kto teraz na zmianie) oraz cięcia godzin (duze/male) — widzą je
    admin i szef (szef_kuchni ma osobny endpoint /api/szefkuchni/godziny, bez cięć).
    Dostęp do danych płacowych jest zapisywany w dzienniku audytu (RODO)."""
    raport = raporty.raport_godzin_miesiac(db, rok, miesiac)
    raport["na_zmianie"] = _trwajace_zmiany(db)
    if uprawnienia.ma_user(user, "wyplaty.podglad"):
        zapisz_audyt(db, user, "raport_godzin", zasob=f"{rok}-{miesiac:02d}", request=request)
        return raport
    return _raport_godzin_bez_wyplat(raport)


def _kuchnia_pids(db):
    """Id pracowników KUCHNI — po DZIALE (tak jak grafik kuchni), niezależnie od roli konta.
    Dzięki temu kucharz dział=kuchnia z kontem 'employee' też jest widziany przez szefa kuchni."""
    return {p.id for p in db.query(models.Pracownik).filter(models.Pracownik.dzial == "kuchnia").all()}


@app.get("/api/szefkuchni/godziny", status_code=200)
def raport_godzin_kuchnia(rok: int = Query(..., ge=2000, le=2100), miesiac: int = Query(..., ge=1, le=12), db: Session = Depends(get_db)):
    """Godziny pracowników KUCHNI (dział „kuchnia") — BEZ kwot wypłaty. Dla szefa kuchni.
    Zwraca te same godziny/stanowiska co raport admina, ale OBCINA pola finansowe
    (stawka/kwota/do_wyplaty). Dorzuca `na_zmianie`: kto z kuchni jest teraz na zmianie (live)."""
    kuchnia_pids = _kuchnia_pids(db)
    raport = raporty.raport_godzin_miesiac(db, rok, miesiac)
    pracownicy = [
        {
            "pracownik_id": p["pracownik_id"],
            "pracownik": p["pracownik"],
            "suma_godzin": p["suma_godzin"],
            # tylko stanowisko + godziny — bez stawki i kwoty
            "stanowiska": [{"stanowisko": s["stanowisko"], "godziny": s["godziny"]} for s in p["stanowiska"]],
        }
        for p in raport["pracownicy"]
        if p["pracownik_id"] in kuchnia_pids
    ]
    na_zmianie = [z for z in _trwajace_zmiany(db) if z.get("pracownik_id") in kuchnia_pids]
    return {"rok": rok, "miesiac": miesiac, "pracownicy": pracownicy, "na_zmianie": na_zmianie}


# --- Szef kuchni: KOREKTY grafiku kuchni (tylko pracownicy działu kuchnia) ---
# Grafik publikuje admin, ale szef kuchni może na bieżąco poprawiać zmiany kuchni.
# Każda zmiana jest natychmiast widoczna kucharzowi (grafik kuchni jest „żywy") i wysyła push.

def _kuchnia_pracownik_lub_403(db, pracownik_id):
    p = db.get(models.Pracownik, pracownik_id)
    if not p:
        raise HTTPException(404, "Nie znaleziono pracownika.")
    if p.dzial != "kuchnia":
        raise HTTPException(403, "Szef kuchni może edytować tylko pracowników kuchni.")
    return p


@app.post("/api/szefkuchni/przydzialy", response_model=schemas.PrzydzialOut, status_code=201)
def szefkuchni_dodaj_przydzial(data: schemas.PrzydzialCreate,
                               user: models.User = Depends(get_current_user),
                               db: Session = Depends(get_db)):
    """Szef kuchni dodaje zmianę pracownikowi kuchni (stanowisko zawsze = Kuchnia)."""
    _kuchnia_pracownik_lub_403(db, data.pracownik_id)
    a = models.PrzydzialZmiany(
        data=data.data, stanowisko_id=_kuchnia_stanowisko(db).id, pracownik_id=data.pracownik_id,
        godz_od=data.godz_od, rewir=data.rewir, zamyka=data.zamyka,
    )
    db.add(a); db.commit(); db.refresh(a)
    _powiadom_kuchnie_o_zmianie(db, a.pracownik_id, "Nowa zmiana w grafiku", "Dodano Ci zmianę w grafiku kuchni.")
    return a


@app.put("/api/szefkuchni/przydzialy/{aid}", response_model=schemas.PrzydzialOut)
def szefkuchni_edytuj_przydzial(aid: int, data: schemas.PrzydzialCreate,
                                user: models.User = Depends(get_current_user),
                                db: Session = Depends(get_db)):
    """Szef kuchni zmienia istniejącą zmianę kuchni (godzina / rewir / kto zamyka)."""
    a = db.get(models.PrzydzialZmiany, aid)
    if not a:
        raise HTTPException(404, "Nie znaleziono.")
    _kuchnia_pracownik_lub_403(db, a.pracownik_id)
    a.godz_od = data.godz_od
    a.rewir = data.rewir
    a.zamyka = data.zamyka
    db.commit(); db.refresh(a)
    _powiadom_kuchnie_o_zmianie(db, a.pracownik_id, "Zmiana w grafiku", "Zaktualizowano Twoją zmianę w grafiku kuchni.")
    return a


@app.delete("/api/szefkuchni/przydzialy/{aid}", status_code=204)
def szefkuchni_usun_przydzial(aid: int,
                              user: models.User = Depends(get_current_user),
                              db: Session = Depends(get_db)):
    """Szef kuchni wykreśla kucharza ze zmiany."""
    a = db.get(models.PrzydzialZmiany, aid)
    if not a:
        raise HTTPException(404, "Nie znaleziono.")
    pid = a.pracownik_id
    _kuchnia_pracownik_lub_403(db, pid)
    db.delete(a); db.commit()
    _powiadom_kuchnie_o_zmianie(db, pid, "Zmiana w grafiku", "Wykreślono Cię ze zmiany w grafiku kuchni.")


@app.get("/api/szefkuchni/grafik", status_code=200)
def szefkuchni_grafik(start: date = Query(...), end: date = Query(...),
                      user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Dane edytowalnego grafiku kuchni dla szefa kuchni: pracownicy kuchni + ich przydziały."""
    kucharze = (db.query(models.Pracownik)
                .filter(models.Pracownik.dzial == "kuchnia")
                .order_by(models.Pracownik.kolejnosc, models.Pracownik.id).all())
    pracownicy = [{"id": p.id, "imie": p.imie, "nazwisko": p.nazwisko,
                   "kolor": p.kolor, "aktywny": bool(p.aktywny)} for p in kucharze]
    kuchnia_pids = [p.id for p in kucharze]
    przydzialy = []
    if kuchnia_pids:
        rows = (db.query(models.PrzydzialZmiany)
                .filter(models.PrzydzialZmiany.data >= start, models.PrzydzialZmiany.data <= end,
                        models.PrzydzialZmiany.pracownik_id.in_(kuchnia_pids)).all())
        przydzialy = [{"id": a.id, "data": str(a.data), "pracownik_id": a.pracownik_id,
                       "godz_od": a.godz_od.strftime("%H:%M") if a.godz_od else None,
                       "rewir": a.rewir, "zamyka": bool(a.zamyka)} for a in rows]
    # Kto z kuchni jest teraz na zmianie (live) — szef kuchni widzi to na głównej zakładce „Grafik".
    na_zmianie = [z for z in _trwajace_zmiany(db) if z.get("pracownik_id") in set(kuchnia_pids)]
    return {"pracownicy": pracownicy, "przydzialy": przydzialy, "na_zmianie": na_zmianie}


# ═══════════════════════════════════════════════════════════════════════════
# STOŁY (live z Gastro) — osobna, addytywna ścieżka. NIE dotyka RCP/godzin.
# Mapowanie rewirów (NGastroUzytkownik.Numer) na widok:
STOLY_WEWNATRZ = [(42, "Parter"), (52, "Góra"), (56, "Zielona"), (57, "Kryształowa")]
STOLY_ZEWNATRZ = [54, 55, 53, 108]   # TARAS, STRZECHA, ABCD+FLINSTONY, Zetka+Ka → suma
STOLY_WYNOS = 46
STOLY_KUCHNIA = -1          # pseudo-rewir: zamówienia „do wydania" na kuchni (firingi KDS, kierunek Kuchnia)
STOLY_KUCHNIA_POZYCJE = -2  # pseudo-rewir: liczba pozycji (dań) na tych niewydanych bloczkach


@app.post("/api/gastro/stoly")
def gastro_stoly_ingest(payload: dict, request: Request, db: Session = Depends(get_db)):
    """Snapshot zajętości stołów od agenta (X-RCP-Token). Upsert per rewir. NIE dotyka RCP."""
    if not token_agenta_ok(request, db):
        raise HTTPException(401, "Nieprawidłowy lub brakujący token agenta.")
    teraz = utcnow_naive()
    for it in (payload.get("stoly") or []):
        try:
            nr = int(it["rewir_nr"])
            ile = int(it.get("otwarte") or 0)
        except (KeyError, ValueError, TypeError):
            continue
        rec = db.get(models.StanStolow, nr)
        if rec is None:
            db.add(models.StanStolow(rewir_nr=nr, otwarte=ile, zaktualizowano_at=teraz))
        else:
            rec.otwarte = ile
            rec.zaktualizowano_at = teraz
    db.commit()
    return {"ok": True}


@app.get("/api/gastro/stoly")
def gastro_stoly(db: Session = Depends(get_db)):
    """Aktualny stan stołów (admin + szef): wewnątrz (sale osobno) + na zewnątrz (suma) + wynos.
    Mapowanie rewirów POS na widok z konfiguracji lokalu (NULL = stałe historyczne)."""
    mapa = get_lokal_config(db).pos_mapa_rewirow or {}
    wewnatrz_def = [(int(nr), str(nazwa)) for nr, nazwa in (mapa.get("wewnatrz") or STOLY_WEWNATRZ)]
    zewnatrz_nr = [int(x) for x in (mapa.get("zewnatrz") or STOLY_ZEWNATRZ)]
    wynos_nr = int(mapa["wynos"]) if "wynos" in mapa else STOLY_WYNOS

    stan = {s.rewir_nr: s.otwarte for s in db.query(models.StanStolow).all()}
    last = db.query(models.StanStolow).order_by(models.StanStolow.zaktualizowano_at.desc()).first()
    wewnatrz = [{"nazwa": nazwa, "liczba": stan.get(nr, 0)} for nr, nazwa in wewnatrz_def]
    return {
        "wewnatrz": wewnatrz,
        "wewnatrz_suma": sum(w["liczba"] for w in wewnatrz),
        "na_zewnatrz": sum(stan.get(nr, 0) for nr in zewnatrz_nr),
        "wynos": stan.get(wynos_nr, 0),
        "kuchnia": stan.get(STOLY_KUCHNIA, 0),
        "kuchnia_pozycje": stan.get(STOLY_KUCHNIA_POZYCJE, 0),
        # Znacznik UTC (zapis przez utcnow_naive()) — z offsetem, żeby przeglądarka
        # przeliczyła na czas lokalny (bez tego pokazywało −2h: UTC czytane jako lokalny).
        "zaktualizowano_at": (last.zaktualizowano_at.replace(tzinfo=timezone.utc).isoformat()
                              if last and last.zaktualizowano_at else None),
    }


@app.post("/api/gastro/stoly-historia")
def gastro_stoly_historia_ingest(payload: dict, request: Request, db: Session = Depends(get_db)):
    """Dzienna historia liczby stolików od agenta (X-RCP-Token). Upsert per dzień. NIE dotyka RCP."""
    if not token_agenta_ok(request, db):
        raise HTTPException(401, "Nieprawidłowy lub brakujący token agenta.")
    teraz = utcnow_naive()
    for it in (payload.get("dni") or []):
        try:
            d = date.fromisoformat(str(it["data"])[:10])
            ile = int(it.get("liczba") or 0)
        except (KeyError, ValueError, TypeError):
            continue
        rec = db.get(models.StolikiHistoria, d)
        if rec is None:
            db.add(models.StolikiHistoria(data=d, liczba=ile, zaktualizowano_at=teraz))
        else:
            rec.liczba = ile
            rec.zaktualizowano_at = teraz
    db.commit()
    return {"ok": True}


@app.get("/api/gastro/stoly-historia")
def gastro_stoly_historia(db: Session = Depends(get_db)):
    """Historia liczby obsłużonych stolików na dzień (admin + szef) — ostatnie 30 dni."""
    od = date.today() - timedelta(days=30)
    rows = (
        db.query(models.StolikiHistoria)
        .filter(models.StolikiHistoria.data >= od)
        .order_by(models.StolikiHistoria.data.asc())
        .all()
    )
    return {"dni": [{"data": r.data.isoformat(), "liczba": r.liczba} for r in rows]}


@app.post("/api/gastro/rozliczenia")
def gastro_rozliczenia_ingest(payload: dict, request: Request, db: Session = Depends(get_db)):
    """Pozycje rozliczeń kelnerów z Gastro od agenta (X-RCP-Token). Upsert po poz_id,
    mapowanie kelnera po imieniu i nazwisku (jak RCP). NIE dotyka RCP — osobna gałąź."""
    if not token_agenta_ok(request, db):
        raise HTTPException(401, "Nieprawidłowy lub brakujący token agenta.")
    mapa = {}
    for p in db.query(models.Pracownik).all():
        mapa.setdefault(_norm_nazwa(f"{p.imie} {p.nazwisko}"), p.id)
        mapa.setdefault(_norm_nazwa(f"{p.nazwisko} {p.imie}"), p.id)
    teraz = utcnow_naive()
    n = 0
    dni_batch = set()
    for it in (payload.get("pozycje") or []):
        try:
            poz_id = str(it["poz_id"])
            roz_id = str(it["rozliczenie_id"])
            d = date.fromisoformat(str(it["data"])[:10])
        except (KeyError, ValueError, TypeError):
            continue
        dni_batch.add(d)
        nazwa = (it.get("imie_nazwisko") or "").strip()
        try:
            zamkniete = bool(int(it.get("zamkniete") or 0))
        except (ValueError, TypeError):
            zamkniete = bool(it.get("zamkniete"))
        rec = db.get(models.RozliczenieGastro, poz_id)
        if rec is None:
            rec = models.RozliczenieGastro(poz_id=poz_id)
            db.add(rec)
        rec.rozliczenie_id = roz_id
        if nazwa:
            rec.imie_nazwisko = nazwa
        pid = mapa.get(_norm_nazwa(nazwa))
        if pid is not None:
            rec.pracownik_id = pid
        rec.data = d
        rec.zamknieto = _parse_dt(it.get("zamknieto"))
        rec.zamkniete = zamkniete
        rec.forma = (it.get("forma") or "").strip()
        rec.sprzedaz = float(it.get("sprzedaz") or 0)
        rec.deklarowane = float(it.get("deklarowane") or 0)
        rec.zaktualizowano_at = teraz
        n += 1
    db.commit()
    # Push „raport oczekuje" do kelnerów sali z ZAMKNIĘTYM rozliczeniem Gastro (wpisane w komputer)
    # — raz na kelnera/dzień (push_oczekuje_at). Po przesłaniu raportu (potwierdzone) przycisk znika.
    for d in dni_batch:
        if ROZLICZENIA_START and d < ROZLICZENIA_START:
            continue   # nie powiadamiaj o zmianach sprzed startu systemu (zaległe)
        try:
            roz = _zbuduj_rozliczenie(db, d)
        except Exception:
            continue
        for k in roz.kelnerzy:
            if k.potwierdzone or k.push_oczekuje_at is not None:
                continue
            zamk = (db.query(models.RozliczenieGastro)
                    .filter_by(pracownik_id=k.pracownik_id, data=d, zamkniete=True).first())
            if zamk:
                wyslij_push_do_pracownika(db, k.pracownik_id, "Rozliczenie zmiany",
                                          "Twój raport oczekuje na przesłanie", url="/")
                k.push_oczekuje_at = teraz
        db.commit()
    return {"ok": True, "pozycje": n}


_RE_NAZW = re.compile(r"\bp\.?\s*([A-ZĄĆĘŁŃÓŚŹŻ][a-ząćęłńóśźż]+)")
_RE_DATA = re.compile(r"(\d{1,2})\.(\d{1,2})\.(\d{4})|(\d{4})\.(\d{1,2})\.(\d{1,2})")


def _parsuj_zadatek(opis):
    """Z opisu KP wyciąga nazwisko (słowo po „p.") i datę imprezy (DD.MM.RRRR lub RRRR.MM.DD)."""
    if not opis:
        return None, None
    mn = _RE_NAZW.search(opis)
    nazwisko = mn.group(1) if mn else None
    data_imp = None
    for m in _RE_DATA.finditer(opis):
        try:
            if m.group(3):
                d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            else:
                y, mo, d = int(m.group(4)), int(m.group(5)), int(m.group(6))
            data_imp = date(y, mo, d)
            break
        except (ValueError, TypeError):
            continue
    return nazwisko, data_imp


def _dopasuj_zadatek(db, z) -> bool:
    """Auto-dopasowanie: termin z tą datą imprezy + nazwiskiem (zawiera). Tylko gdy DOKŁADNIE 1 trafienie."""
    if z.termin_id or not z.nazwisko or not z.data_imprezy:
        return False
    naz = z.nazwisko.lower()
    kand = [t for t in db.query(models.Termin).filter(models.Termin.data == z.data_imprezy).all()
            if naz in (t.nazwisko or "").lower()]
    if len(kand) == 1:
        z.termin_id = kand[0].id
        return True
    return False


@app.post("/api/gastro/zadatki")
def gastro_zadatki_ingest(payload: dict, request: Request, db: Session = Depends(get_db)):
    """Zadatki (KP „Kasa przyjęła") z Gastro od agenta (X-RCP-Token). Upsert po id. Parsuje opis
    (nazwisko + data imprezy) i próbuje auto-dopasować do terminu w kalendarzu."""
    if not token_agenta_ok(request, db):
        raise HTTPException(401, "Nieprawidłowy lub brakujący token agenta.")
    teraz = utcnow_naive()
    n = 0
    for it in (payload.get("zadatki") or []):
        try:
            zid = str(it["id"])
            d = date.fromisoformat(str(it["data"])[:10])
        except (KeyError, ValueError, TypeError):
            continue
        rec = db.get(models.KpZadatek, zid)
        if rec is None:
            rec = models.KpZadatek(id=zid); db.add(rec)
        rec.numer = (it.get("numer") or None)
        rec.kwota = float(it.get("kwota") or 0)
        rec.opis = (it.get("opis") or None)
        rec.data = d
        rec.nazwisko, rec.data_imprezy = _parsuj_zadatek(rec.opis)
        rec.zaktualizowano_at = teraz
        n += 1
    db.commit()
    for z in db.query(models.KpZadatek).filter(models.KpZadatek.termin_id.is_(None)).all():
        _dopasuj_zadatek(db, z)
    db.commit()
    return {"ok": True, "zadatki": n}


@app.get("/api/gastro/rozliczenia")
def gastro_rozliczenia(start: date = Query(...), end: date = Query(...), db: Session = Depends(get_db)):
    """Podgląd zebranych rozliczeń kelnerów (admin) — zgrupowane per rozliczenie.
    Posłuży jako prefill formularzy rozliczeń dnia (Etap D2). REPREZENTACJA nie wchodzi
    do utargu — front pokaże ją osobno."""
    rows = (
        db.query(models.RozliczenieGastro)
        .filter(models.RozliczenieGastro.data >= start, models.RozliczenieGastro.data <= end)
        .order_by(models.RozliczenieGastro.data.desc(), models.RozliczenieGastro.imie_nazwisko.asc(),
                  models.RozliczenieGastro.forma.asc())
        .all()
    )
    prac_map = {p.id: f"{p.imie} {p.nazwisko}" for p in db.query(models.Pracownik).all()}
    return {"pozycje": [{
        "poz_id": r.poz_id, "rozliczenie_id": r.rozliczenie_id,
        "imie_nazwisko": r.imie_nazwisko, "pracownik_id": r.pracownik_id,
        "pracownik": prac_map.get(r.pracownik_id), "data": str(r.data),
        "zamknieto": r.zamknieto.isoformat() if r.zamknieto else None,
        "zamkniete": bool(r.zamkniete), "forma": r.forma,
        "sprzedaz": r.sprzedaz, "deklarowane": r.deklarowane,
    } for r in rows]}


# ═══════════════════════════════════════════════════════════════════════════
# REZERWACJE — wspólny, kanoniczny agregat dla wszystkich ról i kanałów.
@app.get("/api/rezerwacje")
def get_rezerwacje(db: Session = Depends(get_db)):
    """Admin + szef: rezerwacje na 30 dni — per dzień z rozbiciem per godzina."""
    return {"dni": rezerwacje.czytaj_rezerwacje(db, 30)}


# ── SERWOWANIE FRONTENDU (zbudowany React z frontend/dist) ─────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Ścieżkę statyk można nadpisać przez FRONTEND_DIST (aplikacja desktopowa/Electron podaje
# katalog zasobów, bo względne „../frontend/dist" nie działa w spakowanej appce/PyInstaller).
FRONTEND_DIST = os.environ.get("FRONTEND_DIST") or os.path.abspath(os.path.join(BASE_DIR, "..", "frontend", "dist"))
if os.path.isdir(FRONTEND_DIST):
    app.mount("/", StaticFiles(directory=FRONTEND_DIST, html=True), name="frontend")
else:
    # Tryb deweloperski: frontend serwuje Vite (:5173), backend udostępnia tylko /api.
    logger.warning("frontend/dist nie istnieje — pomijam serwowanie frontu. Uruchom 'npm --prefix frontend run build' lub 'npm run dev'.")
